"""Benchmark alternative LLM models for Pebble Excel import analysis.

Tests Qwen models via Together AI, DashScope (Alibaba), and DeepInfra
against Claude Sonnet as baseline. Evaluates:
1. JSON parse success rate
2. Indicator detection accuracy (count, names, hierarchy)
3. Formula quality (correct indicator references)
4. Response time and cost

Usage:
    python bench_llm_import.py [--sheets N] [--models MODEL1,MODEL2]
"""
import asyncio
import json
import os
import re
import sys
import time
import hashlib
from pathlib import Path
from datetime import datetime

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import load_workbook
from backend.routers.import_excel import (
    _extract_sheet_text,
    PEBBLE_SYSTEM_PROMPT,
    SHEET_ANALYSIS_PROMPT,
    _parse_claude_json,
)

# Enhanced prompt suffix for models that struggle with name disambiguation
QWEN_PROMPT_SUFFIX = """

CRITICAL ADDITIONAL RULES (read carefully):

A) NAME DISAMBIGUATION IS MANDATORY. When the same indicator name appears under different product groups (e.g. "количество выдач" under "Потребительский кредит" and "BNPL"), you MUST append the group name in parentheses to make each name globally unique:
   - "количество выдач (потребительский)" under Потребительский кредит
   - "количество выдач (BNPL)" under BNPL
   - "количество выдач (факторинг)" under Факторинг
   DO NOT use bare "количество выдач" — it is ambiguous and will break formulas!
   The suffix should be SHORT — abbreviate the group name to the key word.

B) CROSS-SHEET REFERENCES use :: separator with the sheet's display_name:
   CORRECT: [параметры::количество партнеров]
   WRONG:   [количество партнеров]('0'::периоды="текущий")

   The syntax is: [SheetDisplayName::indicator_name]
   Nothing else. No parentheses, no period modifiers on cross-sheet refs.

C) formula_first: When a formula uses (периоды="предыдущий") and there IS no previous period (first column), formula_first must handle this:
   - If the formula computes a delta: formula_first = "0"
   - If the formula computes an average with prev: formula_first should use only current period
   - NEVER just copy the main formula as formula_first if it references "предыдущий"
"""

# ── LLM Provider configs ──

PROVIDERS = {
    "claude-sonnet": {
        "type": "anthropic",
        "model": "claude-sonnet-4-20250514",
        "api_key_env": "ANTHROPIC_API_KEY",
        "cost_per_1m_input": 3.0,
        "cost_per_1m_output": 15.0,
    },
    "qwen3-235b-together": {
        "type": "openai",
        "model": "Qwen/Qwen3-235B-A22B-Instruct-2507-tput",
        "base_url": "https://api.together.xyz/v1",
        "api_key": "fb64c5f9af4418fa785aebcc1dd47b0d1462691be8a1e04d0c84dec490c4d18c",
        "cost_per_1m_input": 0.30,
        "cost_per_1m_output": 0.50,
    },
    "qwen2.5-7b-together": {
        "type": "openai",
        "model": "Qwen/Qwen2.5-7B-Instruct-Turbo",
        "base_url": "https://api.together.xyz/v1",
        "api_key": "fb64c5f9af4418fa785aebcc1dd47b0d1462691be8a1e04d0c84dec490c4d18c",
        "cost_per_1m_input": 0.30,
        "cost_per_1m_output": 0.30,
    },
    "qwen3.5-397b-together": {
        "type": "openai",
        "model": "Qwen/Qwen3.5-397B-A17B",
        "base_url": "https://api.together.xyz/v1",
        "api_key": "fb64c5f9af4418fa785aebcc1dd47b0d1462691be8a1e04d0c84dec490c4d18c",
        "cost_per_1m_input": 0.60,
        "cost_per_1m_output": 3.60,
    },
    "qwen3-coder-30b-together": {
        "type": "openai",
        "model": "Qwen/Qwen3-Coder-30B-A3B-Instruct",
        "base_url": "https://api.together.xyz/v1",
        "api_key": "fb64c5f9af4418fa785aebcc1dd47b0d1462691be8a1e04d0c84dec490c4d18c",
        "cost_per_1m_input": 0.00,
        "cost_per_1m_output": 0.00,
    },
    "qwen-plus-dashscope": {
        "type": "openai",
        "model": "qwen-plus",
        "base_url": "https://dashscope-us.aliyuncs.com/compatible-mode/v1",
        "api_key": "sk-dabdbef906214e208fff4c9520e1d8d4",
        "cost_per_1m_input": 0.80,
        "cost_per_1m_output": 2.00,
    },
    "qwen-max-dashscope": {
        "type": "openai",
        "model": "qwen-max",
        "base_url": "https://dashscope-us.aliyuncs.com/compatible-mode/v1",
        "api_key": "sk-dabdbef906214e208fff4c9520e1d8d4",
        "cost_per_1m_input": 2.40,
        "cost_per_1m_output": 9.60,
    },
    "qwen-turbo-dashscope": {
        "type": "openai",
        "model": "qwen-turbo",
        "base_url": "https://dashscope-us.aliyuncs.com/compatible-mode/v1",
        "api_key": "sk-dabdbef906214e208fff4c9520e1d8d4",
        "cost_per_1m_input": 0.30,
        "cost_per_1m_output": 0.60,
    },
    "qwen2.5-72b-deepinfra": {
        "type": "openai",
        "model": "Qwen/Qwen2.5-72B-Instruct",
        "base_url": "https://api.deepinfra.com/v1/openai",
        "api_key": "LDZ5mfAZGByrpArb42lPSBs8i3mYFMkI",
        "cost_per_1m_input": 0.23,
        "cost_per_1m_output": 0.40,
    },
}


# ── Cache ──

CACHE_DIR = Path(__file__).parent / ".bench_cache"
CACHE_DIR.mkdir(exist_ok=True)

def _cache_key(provider: str, sheet_text: str) -> str:
    h = hashlib.sha256(f"{provider}:{sheet_text}".encode()).hexdigest()[:16]
    return str(CACHE_DIR / f"{provider}_{h}.json")

def _cache_get(provider: str, sheet_text: str):
    path = _cache_key(provider, sheet_text)
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return None

def _cache_set(provider: str, sheet_text: str, data: dict):
    path = _cache_key(provider, sheet_text)
    with open(path, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── LLM call implementations ──

def _strip_think_tags(text: str) -> str:
    """Remove <think>...</think> blocks from Qwen3 responses."""
    return re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL).strip()

def call_anthropic(config: dict, sheet_text: str) -> dict:
    """Call Anthropic API (Claude)."""
    import anthropic

    api_key = os.environ.get(config["api_key_env"])
    if not api_key:
        return {"error": f"Missing {config['api_key_env']}", "raw": ""}

    client = anthropic.Anthropic(api_key=api_key)
    t0 = time.time()
    message = client.messages.create(
        model=config["model"],
        max_tokens=16384,
        system=PEBBLE_SYSTEM_PROMPT + "\n\nIMPORTANT: Return ONLY valid JSON. No markdown fences, no comments, no trailing commas.",
        messages=[
            {"role": "user", "content": SHEET_ANALYSIS_PROMPT + sheet_text},
            {"role": "assistant", "content": "{"},
        ],
    )
    elapsed = time.time() - t0
    raw_text = message.content[0].text
    input_tokens = message.usage.input_tokens
    output_tokens = message.usage.output_tokens
    return {
        "raw": "{" + raw_text,
        "elapsed": elapsed,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
    }


def call_openai_compat(config: dict, sheet_text: str) -> dict:
    """Call OpenAI-compatible API (Together, DashScope, DeepInfra)."""
    import httpx

    headers = {
        "Authorization": f"Bearer {config['api_key']}",
        "Content-Type": "application/json",
    }

    # For Qwen3 models, add /no_think suffix or use extra parameters
    extra_params = {}
    model = config["model"]
    if "Qwen3" in model or "qwen3" in model.lower():
        # Disable thinking for Qwen3 to get clean JSON output
        extra_params["extra_body"] = {"chat_template_kwargs": {"enable_thinking": False}}

    # Use enhanced prompt for non-Claude models
    system_prompt = PEBBLE_SYSTEM_PROMPT + "\n\nIMPORTANT: Return ONLY valid JSON. No markdown fences, no comments, no trailing commas. No <think> tags."
    user_prompt = SHEET_ANALYSIS_PROMPT + sheet_text + QWEN_PROMPT_SUFFIX

    body = {
        "model": model,
        "max_tokens": 16384,
        "temperature": 0.1,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    }

    url = config["base_url"].rstrip("/") + "/chat/completions"
    t0 = time.time()
    with httpx.Client(timeout=300) as client:
        resp = client.post(url, json=body, headers=headers)
        resp.raise_for_status()
    elapsed = time.time() - t0
    data = resp.json()

    choice = data["choices"][0]
    raw_text = choice["message"]["content"]
    usage = data.get("usage", {})

    # Strip <think> tags if present
    raw_text = _strip_think_tags(raw_text)

    return {
        "raw": raw_text,
        "elapsed": elapsed,
        "input_tokens": usage.get("prompt_tokens", 0),
        "output_tokens": usage.get("completion_tokens", 0),
    }


def call_llm(provider_name: str, sheet_text: str) -> dict:
    """Call LLM and return raw response + metadata. Uses cache."""
    cached = _cache_get(provider_name, sheet_text)
    if cached:
        cached["from_cache"] = True
        return cached

    config = PROVIDERS[provider_name]
    try:
        if config["type"] == "anthropic":
            result = call_anthropic(config, sheet_text)
        else:
            result = call_openai_compat(config, sheet_text)
    except Exception as e:
        result = {"error": str(e), "raw": "", "elapsed": 0, "input_tokens": 0, "output_tokens": 0}

    result["from_cache"] = False
    if "error" not in result:
        _cache_set(provider_name, sheet_text, result)
    return result


# ── Evaluation ──

def parse_response(raw: str) -> dict | None:
    """Try to parse JSON from LLM response."""
    try:
        return _parse_claude_json(raw)
    except:
        pass
    # Try extracting JSON from text
    try:
        match = re.search(r'\{[\s\S]*\}', raw)
        if match:
            text = match.group()
            text = re.sub(r',\s*([}\]])', r'\1', text)
            return json.loads(text)
    except:
        pass
    return None


def count_indicators(node: dict, depth: int = 0) -> tuple[int, int, int]:
    """Count total indicators, groups, and leaves in a tree."""
    groups = 1 if node.get("is_group") else 0
    leaves = 0 if node.get("is_group") else 1
    total = 1
    for child in node.get("children", []):
        ct, cg, cl = count_indicators(child, depth + 1)
        total += ct
        groups += cg
        leaves += cl
    return total, groups, leaves


def collect_names(indicators: list) -> set[str]:
    """Collect all indicator names recursively."""
    names = set()
    for ind in indicators:
        if ind.get("name"):
            names.add(ind["name"].strip().lower())
        for child in ind.get("children", []):
            names.add(child.get("name", "").strip().lower())
            # recurse
            for n in collect_names(ind.get("children", [])):
                names.add(n)
    return names - {""}


def collect_formulas(indicators: list) -> list[dict]:
    """Collect all formula indicators."""
    formulas = []
    for ind in indicators:
        if ind.get("rule") == "formula" and ind.get("formula"):
            formulas.append({"name": ind["name"], "formula": ind["formula"]})
        for child in ind.get("children", []):
            if child.get("rule") == "formula" and child.get("formula"):
                formulas.append({"name": child["name"], "formula": child["formula"]})
            formulas.extend(collect_formulas(child.get("children", [])))
    return formulas


def check_formula_refs(formula: str, all_names: set[str]) -> tuple[int, int]:
    """Check how many [ref] references in a formula match known indicator names."""
    refs = re.findall(r'\[([^\]]+)\]', formula)
    if not refs:
        return 0, 0
    matched = 0
    for ref in refs:
        # Strip sheet prefix
        clean = ref.split("::")[-1].strip().lower()
        # Remove period suffix
        clean = re.sub(r'\(периоды=.*\)', '', clean).strip()
        if clean in all_names:
            matched += 1
    return matched, len(refs)


def evaluate(parsed: dict, baseline_parsed: dict | None = None) -> dict:
    """Evaluate parsed result quality."""
    indicators = parsed.get("indicators", [])
    total = sum(count_indicators(ind)[0] for ind in indicators)
    groups = sum(count_indicators(ind)[1] for ind in indicators)
    leaves = sum(count_indicators(ind)[2] for ind in indicators)
    names = collect_names(indicators)
    formulas = collect_formulas(indicators)

    # Formula reference check
    total_refs = 0
    matched_refs = 0
    for f in formulas:
        m, t = check_formula_refs(f["formula"], names)
        matched_refs += m
        total_refs += t
    ref_accuracy = matched_refs / total_refs if total_refs > 0 else 1.0

    result = {
        "total_indicators": total,
        "groups": groups,
        "leaves": leaves,
        "formulas": len(formulas),
        "formula_ref_accuracy": round(ref_accuracy, 3),
        "has_display_name": bool(parsed.get("display_name")),
        "has_data_start_col": bool(parsed.get("data_start_col")),
    }

    # Compare to baseline if available
    if baseline_parsed:
        bl_indicators = baseline_parsed.get("indicators", [])
        bl_names = collect_names(bl_indicators)
        bl_total = sum(count_indicators(ind)[0] for ind in bl_indicators)
        name_overlap = len(names & bl_names) / max(len(bl_names), 1)
        result["vs_baseline_name_overlap"] = round(name_overlap, 3)
        result["vs_baseline_count_diff"] = total - bl_total

    return result


# ── Main ──

def extract_test_sheets(excel_path: str, max_sheets: int = 3) -> dict[str, str]:
    """Extract sheet texts from Excel file."""
    wb = load_workbook(excel_path)
    texts = {}
    for sn in wb.sheetnames[:max_sheets]:
        ws = wb[sn]
        texts[sn] = _extract_sheet_text(ws, sn)
    return texts


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--sheets", type=int, default=2, help="Number of sheets to test per Excel file")
    parser.add_argument("--models", type=str, default=None, help="Comma-separated list of model names to test")
    parser.add_argument("--excel", type=str, default=None, help="Path to Excel file")
    parser.add_argument("--no-cache", action="store_true", help="Ignore cache")
    args = parser.parse_args()

    # Select models to test
    if args.models:
        model_names = [m.strip() for m in args.models.split(",")]
    else:
        model_names = list(PROVIDERS.keys())

    # Select Excel files
    excel_files = []
    if args.excel:
        excel_files = [args.excel]
    else:
        docs = Path("/Users/mac/Documents")
        for f in ["MIS model-1.xlsx", "Doscredobank FinModel BaaS 2026-2028 v.12.xlsx"]:
            p = docs / f
            if p.exists():
                excel_files.append(str(p))

    if not excel_files:
        print("No Excel files found!")
        return

    # Run benchmarks
    all_results = []
    for excel_path in excel_files:
        print(f"\n{'='*60}")
        print(f"Excel: {Path(excel_path).name}")
        print(f"{'='*60}")
        sheets = extract_test_sheets(excel_path, args.sheets)

        for sheet_name, sheet_text in sheets.items():
            print(f"\n--- Sheet: {sheet_name} (text={len(sheet_text)} chars) ---\n")

            baseline_parsed = None
            results = {}

            for model_name in model_names:
                if args.no_cache:
                    # Clear cache for this
                    path = _cache_key(model_name, sheet_text)
                    if os.path.exists(path):
                        os.remove(path)

                print(f"  {model_name}...", end=" ", flush=True)
                resp = call_llm(model_name, sheet_text)

                if "error" in resp:
                    print(f"ERROR: {resp['error'][:80]}")
                    results[model_name] = {"error": resp["error"]}
                    continue

                parsed = parse_response(resp["raw"])
                if parsed is None:
                    print(f"PARSE FAIL (raw={resp['raw'][:100]}...)")
                    results[model_name] = {"parse_error": True, "raw_preview": resp["raw"][:200]}
                    continue

                # Use Claude as baseline
                if model_name == "claude-sonnet" and parsed:
                    baseline_parsed = parsed

                eval_result = evaluate(parsed, baseline_parsed if model_name != "claude-sonnet" else None)
                cost_in = resp["input_tokens"] / 1_000_000 * PROVIDERS[model_name]["cost_per_1m_input"]
                cost_out = resp["output_tokens"] / 1_000_000 * PROVIDERS[model_name]["cost_per_1m_output"]

                eval_result.update({
                    "elapsed": round(resp["elapsed"], 1) if not resp.get("from_cache") else "cached",
                    "input_tokens": resp["input_tokens"],
                    "output_tokens": resp["output_tokens"],
                    "cost_usd": round(cost_in + cost_out, 4),
                    "from_cache": resp.get("from_cache", False),
                })

                cached_str = " (cached)" if resp.get("from_cache") else ""
                elapsed_str = f"{resp['elapsed']:.1f}s" if not resp.get("from_cache") else "cached"
                print(f"OK {elapsed_str}{cached_str} | {eval_result['total_indicators']} ind, "
                      f"{eval_result['formulas']} formulas, ref_acc={eval_result['formula_ref_accuracy']}, "
                      f"${eval_result['cost_usd']:.4f}")

                if "vs_baseline_name_overlap" in eval_result:
                    print(f"    vs Claude: name_overlap={eval_result['vs_baseline_name_overlap']}, "
                          f"count_diff={eval_result['vs_baseline_count_diff']:+d}")

                results[model_name] = eval_result

            all_results.append({
                "excel": Path(excel_path).name,
                "sheet": sheet_name,
                "results": results,
            })

    # Summary table
    print(f"\n\n{'='*80}")
    print("SUMMARY")
    print(f"{'='*80}")
    print(f"{'Model':<30} {'Ind':>5} {'Form':>5} {'RefAcc':>7} {'Cost':>8} {'Time':>8} {'NameOvl':>8}")
    print("-" * 80)

    # Aggregate across sheets
    model_agg: dict[str, dict] = {}
    for r in all_results:
        for model_name, data in r["results"].items():
            if "error" in data or "parse_error" in data:
                continue
            if model_name not in model_agg:
                model_agg[model_name] = {"ind": [], "form": [], "ref_acc": [], "cost": [], "time": [], "overlap": []}
            agg = model_agg[model_name]
            agg["ind"].append(data["total_indicators"])
            agg["form"].append(data["formulas"])
            agg["ref_acc"].append(data["formula_ref_accuracy"])
            agg["cost"].append(data["cost_usd"])
            if isinstance(data.get("elapsed"), (int, float)):
                agg["time"].append(data["elapsed"])
            if "vs_baseline_name_overlap" in data:
                agg["overlap"].append(data["vs_baseline_name_overlap"])

    for model_name, agg in model_agg.items():
        avg_ind = sum(agg["ind"]) / len(agg["ind"])
        avg_form = sum(agg["form"]) / len(agg["form"])
        avg_ref = sum(agg["ref_acc"]) / len(agg["ref_acc"])
        total_cost = sum(agg["cost"])
        avg_time = sum(agg["time"]) / len(agg["time"]) if agg["time"] else 0
        avg_overlap = sum(agg["overlap"]) / len(agg["overlap"]) if agg["overlap"] else 1.0
        print(f"{model_name:<30} {avg_ind:>5.0f} {avg_form:>5.0f} {avg_ref:>7.3f} "
              f"${total_cost:>7.4f} {avg_time:>7.1f}s {avg_overlap:>7.3f}")

    # Save results
    results_path = Path(__file__).parent / "bench_results.json"
    with open(results_path, "w") as f:
        json.dump(all_results, f, indent=2, ensure_ascii=False, default=str)
    print(f"\nDetailed results saved to {results_path}")


if __name__ == "__main__":
    main()
