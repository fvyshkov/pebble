#!/usr/bin/env python3
"""Apply Excel formulas to BaaS model via the translator.

Steps:
1. Build row→name maps from Excel rows matched to DB record names
2. Translate all formulas using excel_formula_translator
3. Write formulas to cell_data
4. Recalculate model
5. Compare with Excel data_only values
"""

import sqlite3
import json
import asyncio
from openpyxl import load_workbook
from backend.excel_formula_translator import translate_excel_formula

EXCEL_PATH = "models.xlsx"
DB_PATH = "pebble.db"

# ── Sheet configs ─────────────────────────────────────────────────────────
# Maps Excel sheet name → {data_start_col, label_col, row_range, db_sheet_name}

SHEET_CONFIGS = {
    "0": {
        "data_start_col": 4,  # D
        "label_col": 1,       # A
        "db_sheet_name": "BaaS - параметры модели",
    },
    "BaaS.1": {
        "data_start_col": 4,  # D
        "label_col": 1,       # A
        "db_sheet_name": "BaaS - Онлайн кредитование",
    },
    "BaaS.2": {
        "data_start_col": 4,  # D
        "label_col": 1,       # A
        "db_sheet_name": "BaaS - Онлайн депозит",
    },
    "BaaS.3": {
        "data_start_col": 4,  # D
        "label_col": 1,       # A
        "db_sheet_name": "BaaS - Онлайн транзакционный бизнес",
    },
    "BS": {
        "data_start_col": 3,  # C
        "label_col": 2,       # B
        "db_sheet_name": "Баланс BaaS",
    },
    "PL": {
        "data_start_col": 3,  # C
        "label_col": 2,       # B
        "db_sheet_name": "Финансовый результат BaaS",
    },
    "OPEX+CAPEX": {
        "data_start_col": 5,  # E
        "label_col": 3,       # C (Статья расходов)
        "db_sheet_name": "Операционные расходы и Инвестиции в BaaS",
    },
}


def get_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    return db


def build_excel_row_labels(ws, label_col, max_row=500):
    """Extract {row_number: label} from Excel sheet."""
    labels = {}
    for r in range(1, min(ws.max_row or 1, max_row) + 1):
        v = ws.cell(r, label_col).value
        if v is not None:
            labels[r] = str(v).strip()
    return labels


def build_db_record_names(db, sheet_name):
    """Get ordered list of (sort_order, name, record_id, parent_id) from DB."""
    sheet = db.execute("SELECT id FROM sheets WHERE name = ?", (sheet_name,)).fetchone()
    if not sheet:
        return []
    bindings = db.execute("""
        SELECT sa.analytic_id, a.is_periods
        FROM sheet_analytics sa JOIN analytics a ON a.id = sa.analytic_id
        WHERE sa.sheet_id = ? ORDER BY sa.sort_order
    """, (sheet["id"],)).fetchall()

    records = []
    for b in bindings:
        if b["is_periods"]:
            continue
        recs = db.execute(
            "SELECT id, parent_id, sort_order, data_json FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
            (b["analytic_id"],),
        ).fetchall()
        for r in recs:
            data = json.loads(r["data_json"])
            records.append({
                "sort_order": r["sort_order"],
                "name": data.get("name", ""),
                "id": r["id"],
                "parent_id": r["parent_id"],
            })
    return records


def match_excel_rows_to_db(excel_labels, db_records, excel_sheet_name):
    """Match Excel row numbers to DB record names.

    Strategy: match by Excel label → DB name (case-insensitive, strip).
    DB names may have disambiguating suffixes like "(потребительский кредит)"
    that don't appear in Excel. We match by checking if the Excel label
    is a prefix of the DB name, or an exact match.

    For sheets with duplicate names across groups, we use positional order
    within matched candidates.
    """
    row_to_name = {}
    row_to_rid = {}

    # Build lookup: lowercase_label → list of DB records
    db_by_label = {}
    for r in db_records:
        key = r["name"].lower().strip()
        db_by_label.setdefault(key, []).append(r)

    # Also index by base name (without parenthetical suffixes and #N markers)
    db_by_base = {}
    import re as _re
    for r in db_records:
        name = r["name"].strip()
        # Remove trailing " #N", " — suffix", and " (suffix)" iteratively
        base = name
        base = _re.sub(r'\s*#\d+\s*$', '', base)       # remove #2, #3 etc
        base = _re.sub(r'\s*—\s*[^—]+$', '', base)     # remove — suffix
        base = _re.sub(r'\s*\([^)]*\)\s*$', '', base)   # remove last (suffix)
        base = base.lower().strip()
        if base != name.lower().strip():
            db_by_base.setdefault(base, []).append(r)

    used_rids = set()

    # Process Excel rows in order
    for row_num in sorted(excel_labels.keys()):
        label = excel_labels[row_num].strip()
        if not label:
            continue
        label_lower = label.lower()

        # Try exact match first, then base name match
        picked = None
        for source in [db_by_label.get(label_lower, []), db_by_base.get(label_lower, [])]:
            for c in source:
                if c["id"] not in used_rids:
                    picked = c
                    break
            if picked:
                break

        if picked:
            used_rids.add(picked["id"])
            row_to_name[row_num] = picked["name"]
            row_to_rid[row_num] = picked["id"]

    return row_to_name, row_to_rid


def build_all_row_maps(wb, db):
    """Build row→name maps for all sheets."""
    all_row_maps = {}      # {excel_sheet_name: {row: db_name}}
    all_row_to_rid = {}    # {excel_sheet_name: {row: record_id}}
    sheet_display_names = {}  # {excel_sheet_name: pebble_display_name}
    sheet_data_starts = {}   # {excel_sheet_name: data_start_col}

    for excel_name, cfg in SHEET_CONFIGS.items():
        ws = wb[excel_name]
        labels = build_excel_row_labels(ws, cfg["label_col"])
        db_records = build_db_record_names(db, cfg["db_sheet_name"])
        sheet_data_starts[excel_name] = cfg["data_start_col"]

        row_to_name, row_to_rid = match_excel_rows_to_db(labels, db_records, excel_name)
        all_row_maps[excel_name] = row_to_name
        all_row_to_rid[excel_name] = row_to_rid
        sheet_display_names[excel_name] = cfg["db_sheet_name"]

    return all_row_maps, all_row_to_rid, sheet_display_names, sheet_data_starts


def translate_all_formulas(wb, all_row_maps, sheet_display_names, sheet_data_starts):
    """Translate all formulas for all sheets.

    Returns per-cell formulas: {excel_name: {row: {col: pebble_formula_or_None}}}
    None means manual (keep original value).
    """
    results = {}

    for excel_name, cfg in SHEET_CONFIGS.items():
        ws = wb[excel_name]
        row_to_name = all_row_maps[excel_name]
        data_start = cfg["data_start_col"]
        max_col = min(ws.max_column or 1, 200)

        sheet_results = {}
        for row in sorted(row_to_name.keys()):
            # Check if this row has ANY formula cells
            has_formula = False
            for c in range(data_start, min(data_start + 40, max_col + 1)):
                v = ws.cell(row, c).value
                if isinstance(v, str) and v.startswith("="):
                    has_formula = True
                    break
            if not has_formula:
                continue

            # Translate each cell individually
            row_formulas = {}
            for c in range(data_start, min(data_start + 40, max_col + 1)):
                cell_val = ws.cell(row, c).value
                is_first = (c == data_start)

                if isinstance(cell_val, str) and cell_val.startswith("="):
                    pebble = translate_excel_formula(
                        cell_val, c, data_start, row_to_name,
                        all_row_maps, sheet_display_names,
                        is_first_period=is_first,
                        sheet_data_starts=sheet_data_starts,
                    )
                    row_formulas[c] = pebble
                else:
                    # Manual value or constant — mark as None
                    row_formulas[c] = None

            sheet_results[row] = {
                "per_col": row_formulas,
                "name": row_to_name[row],
            }

        results[excel_name] = sheet_results

    return results


def restore_manual_values(db, wb_data, all_row_to_rid):
    """Restore original Excel values for manual cells (undo previous calc overwrites)."""
    from datetime import datetime
    MONTH_NAMES_RU = ["Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]

    model = db.execute("SELECT id FROM models WHERE name = 'BaaS'").fetchone()
    mid = model["id"]
    sheets = db.execute("SELECT id, name FROM sheets WHERE model_id = ?", (mid,)).fetchall()
    sheet_name_to_id = {s["name"]: s["id"] for s in sheets}

    # Get leaf periods
    period_aid = db.execute("""
        SELECT sa.analytic_id FROM sheet_analytics sa
        JOIN analytics a ON a.id = sa.analytic_id
        WHERE sa.sheet_id = ? AND a.is_periods = 1
    """, (sheets[0]["id"],)).fetchone()["analytic_id"]
    period_recs = db.execute(
        "SELECT id, parent_id, data_json FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
        (period_aid,),
    ).fetchall()
    child_map = {}
    for r in period_recs:
        if r["parent_id"]:
            child_map.setdefault(r["parent_id"], []).append(r["id"])
    leaf_periods = [r for r in period_recs if r["id"] not in child_map]
    period_names = {json.loads(r["data_json"])["name"]: r["id"] for r in leaf_periods}

    restored = 0
    for excel_name, cfg in SHEET_CONFIGS.items():
        db_sheet_name = cfg["db_sheet_name"]
        sheet_id = sheet_name_to_id.get(db_sheet_name)
        if not sheet_id:
            continue

        ws_d = wb_data[excel_name]
        data_start = cfg["data_start_col"]
        row_to_rid = all_row_to_rid.get(excel_name, {})

        # Build col→period_rid
        col_to_period = {}
        for c in range(data_start, min((ws_d.max_column or 1) + 1, 50)):
            for hr in range(1, 7):
                v = ws_d.cell(hr, c).value
                if isinstance(v, datetime):
                    pname = f"{MONTH_NAMES_RU[v.month-1]} {v.year}"
                    if pname in period_names:
                        col_to_period[c] = period_names[pname]
                    break

        for row, rid in row_to_rid.items():
            for col, period_rid in col_to_period.items():
                excel_val = ws_d.cell(row, col).value
                if excel_val is None:
                    continue
                coord_key = f"{period_rid}|{rid}"
                db.execute(
                    "UPDATE cell_data SET value = ? WHERE sheet_id = ? AND coord_key = ?",
                    (str(excel_val), sheet_id, coord_key),
                )
                restored += 1

    db.commit()
    return restored


def write_formulas_to_db(db, all_formulas, all_row_to_rid, wb_data):
    """Write translated formulas to cell_data in DB.

    all_formulas: {excel_name: {row: {per_col: {col: pebble_or_None}, name: str}}}
    """
    from datetime import datetime as dt
    MONTH_NAMES = ["Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]

    model = db.execute("SELECT id FROM models WHERE name = 'BaaS'").fetchone()
    mid = model["id"]

    sheets = db.execute("SELECT id, name FROM sheets WHERE model_id = ?", (mid,)).fetchall()
    sheet_name_to_id = {s["name"]: s["id"] for s in sheets}

    # Reset ALL cells to manual
    for s in sheets:
        db.execute(
            "UPDATE cell_data SET rule = 'manual', formula = '' WHERE sheet_id = ? AND rule = 'formula'",
            (s["id"],),
        )
    print("  Reset all cells to manual")

    # Get leaf period records
    period_aid = db.execute("""
        SELECT sa.analytic_id FROM sheet_analytics sa
        JOIN analytics a ON a.id = sa.analytic_id
        WHERE sa.sheet_id = ? AND a.is_periods = 1
    """, (sheets[0]["id"],)).fetchone()["analytic_id"]

    period_recs = db.execute(
        "SELECT id, parent_id, data_json FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
        (period_aid,),
    ).fetchall()
    child_map = {}
    for r in period_recs:
        if r["parent_id"]:
            child_map.setdefault(r["parent_id"], []).append(r["id"])
    leaf_periods = [r for r in period_recs if r["id"] not in child_map]
    period_name_to_rid = {json.loads(r["data_json"])["name"]: r["id"] for r in leaf_periods}

    total_updated = 0
    total_formulas = 0

    for excel_name, cfg in SHEET_CONFIGS.items():
        db_sheet_name = cfg["db_sheet_name"]
        sheet_id = sheet_name_to_id.get(db_sheet_name)
        if not sheet_id:
            continue

        sheet_formulas = all_formulas.get(excel_name, {})
        row_to_rid = all_row_to_rid.get(excel_name, {})
        data_start = cfg["data_start_col"]

        # Build col→period_rid mapping using data_only workbook
        ws_d = wb_data[excel_name]
        col_to_period = {}
        for c in range(data_start, min((ws_d.max_column or 1) + 1, 50)):
            for hr in range(1, 7):
                v = ws_d.cell(hr, c).value
                if isinstance(v, dt):
                    pname = f"{MONTH_NAMES[v.month-1]} {v.year}"
                    if pname in period_name_to_rid:
                        col_to_period[c] = period_name_to_rid[pname]
                    break

        for row, info in sheet_formulas.items():
            rid = row_to_rid.get(row)
            if not rid:
                continue

            per_col = info["per_col"]

            for col, pebble_formula in per_col.items():
                period_rid = col_to_period.get(col)
                if not period_rid:
                    continue

                coord_key = f"{period_rid}|{rid}"

                if pebble_formula is None:
                    # Manual cell — keep existing value
                    continue

                # Check if it's a constant
                is_constant = False
                try:
                    const_val = float(pebble_formula)
                    is_constant = True
                except (ValueError, TypeError):
                    pass

                existing = db.execute(
                    "SELECT id FROM cell_data WHERE sheet_id = ? AND coord_key = ?",
                    (sheet_id, coord_key),
                ).fetchone()

                if is_constant:
                    if existing:
                        db.execute(
                            "UPDATE cell_data SET rule = 'manual', formula = '', value = ? WHERE id = ?",
                            (str(const_val), existing["id"]),
                        )
                    total_updated += 1
                elif existing:
                    db.execute(
                        "UPDATE cell_data SET rule = 'formula', formula = ? WHERE id = ?",
                        (pebble_formula, existing["id"]),
                    )
                    total_updated += 1
                else:
                    import uuid
                    db.execute(
                        "INSERT INTO cell_data (id, sheet_id, coord_key, value, data_type, rule, formula) VALUES (?,?,?,?,?,?,?)",
                        (str(uuid.uuid4()), sheet_id, coord_key, "0", "sum", "formula", pebble_formula),
                    )
                    total_updated += 1

                total_formulas += 1

    db.commit()
    return total_updated, total_formulas


def verify_against_excel(db, wb_data, all_row_to_rid):
    """Compare calculated values with Excel data_only values."""
    model = db.execute("SELECT id FROM models WHERE name = 'BaaS'").fetchone()
    mid = model["id"]
    sheets = db.execute("SELECT id, name FROM sheets WHERE model_id = ?", (mid,)).fetchall()
    sheet_name_to_id = {s["name"]: s["id"] for s in sheets}

    # Get period mapping (month index → period_rid)
    period_aid = db.execute("""
        SELECT sa.analytic_id FROM sheet_analytics sa
        JOIN analytics a ON a.id = sa.analytic_id
        WHERE sa.sheet_id = ? AND a.is_periods = 1
    """, (sheets[0]["id"],)).fetchone()["analytic_id"]

    period_recs = db.execute(
        "SELECT id, parent_id, sort_order, data_json FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
        (period_aid,),
    ).fetchall()
    child_map = {}
    for r in period_recs:
        if r["parent_id"]:
            child_map.setdefault(r["parent_id"], []).append(r["id"])
    leaf_periods = [r for r in period_recs if r["id"] not in child_map]

    # Map leaf periods to Excel columns by month name matching
    from datetime import datetime
    MONTH_NAMES_RU = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
    ]

    total_cells = 0
    match_cells = 0
    mismatch_cells = 0
    mismatches = []

    for excel_name, cfg in SHEET_CONFIGS.items():
        db_sheet_name = cfg["db_sheet_name"]
        sheet_id = sheet_name_to_id.get(db_sheet_name)
        if not sheet_id:
            continue

        ws_d = wb_data[excel_name]
        data_start = cfg["data_start_col"]
        row_to_rid = all_row_to_rid.get(excel_name, {})

        # Build col→period_rid mapping for this sheet
        col_to_period = {}
        for c in range(data_start, min((ws_d.max_column or 1) + 1, 50)):
            # Check header row for dates
            for hr in range(1, 7):
                v = ws_d.cell(hr, c).value
                if isinstance(v, datetime):
                    month_key = f"{v.year}-{v.month:02d}"
                    for lp in leaf_periods:
                        pdata = json.loads(lp["data_json"])
                        pname = pdata.get("name", "")
                        if f"{MONTH_NAMES_RU[v.month-1]} {v.year}" == pname:
                            col_to_period[c] = lp["id"]
                            break
                    break

        # Load DB cell values
        cells = db.execute(
            "SELECT coord_key, value FROM cell_data WHERE sheet_id = ? AND rule = 'formula'",
            (sheet_id,),
        ).fetchall()
        db_cells = {c["coord_key"]: c["value"] for c in cells}

        sheet_total = 0
        sheet_match = 0

        for row, rid in row_to_rid.items():
            for col, period_rid in col_to_period.items():
                coord_key = f"{period_rid}|{rid}"
                if coord_key not in db_cells:
                    continue

                excel_val = ws_d.cell(row, col).value
                if excel_val is None:
                    continue

                try:
                    excel_num = float(excel_val)
                except (ValueError, TypeError):
                    continue

                try:
                    db_num = float(db_cells[coord_key])
                except (ValueError, TypeError):
                    db_num = 0.0

                total_cells += 1
                sheet_total += 1

                tol = abs(excel_num) * 0.01 + 0.01
                if abs(db_num - excel_num) <= tol:
                    match_cells += 1
                    sheet_match += 1
                else:
                    mismatch_cells += 1
                    if len(mismatches) < 20:
                        name = ""
                        for r2 in all_row_to_rid.get(excel_name, {}):
                            if all_row_to_rid[excel_name][r2] == rid:
                                from backend.excel_formula_translator import translate_excel_formula
                                name = row_to_rid  # just use row
                                break
                        mismatches.append({
                            "sheet": excel_name,
                            "row": row,
                            "excel": excel_num,
                            "db": db_num,
                            "coord": coord_key[:20],
                        })

        if sheet_total > 0:
            pct = sheet_match / sheet_total * 100
            print(f"  {excel_name}: {sheet_match}/{sheet_total} ({pct:.1f}%)")

    print(f"\nTotal: {match_cells}/{total_cells} ({match_cells/total_cells*100:.1f}%)" if total_cells else "No cells to compare")

    if mismatches:
        print(f"\nFirst {len(mismatches)} mismatches:")
        for m in mismatches:
            print(f"  {m['sheet']} row {m['row']}: excel={m['excel']:.4f} db={m['db']:.4f}")


def main():
    print("=== Step 1: Loading Excel and DB ===")
    wb = load_workbook(EXCEL_PATH)
    wb_data = load_workbook(EXCEL_PATH, data_only=True)
    db = get_db()

    print("=== Step 2: Building row→name maps ===")
    all_row_maps, all_row_to_rid, sheet_display_names, sheet_data_starts = build_all_row_maps(wb, db)

    for excel_name in SHEET_CONFIGS:
        rmap = all_row_maps[excel_name]
        print(f"  {excel_name}: {len(rmap)} rows mapped")

    print("\n=== Step 3: Translating formulas ===")
    all_formulas = translate_all_formulas(wb, all_row_maps, sheet_display_names, sheet_data_starts)

    for excel_name, formulas in all_formulas.items():
        print(f"  {excel_name}: {len(formulas)} formula rows")
        for row, info in list(formulas.items())[:3]:
            per_col = info["per_col"]
            formula_cols = [c for c, f in per_col.items() if f is not None]
            if formula_cols:
                first_f = per_col[formula_cols[0]]
                print(f"    row {row} [{info['name']}]: {first_f[:80] if first_f else 'N/A'}")

    print("\n=== Step 3.5: Restoring original Excel values ===")
    restored = restore_manual_values(db, wb_data, all_row_to_rid)
    print(f"  Restored {restored} cell values from Excel")

    print("\n=== Step 4: Writing formulas to DB ===")
    updated, total = write_formulas_to_db(db, all_formulas, all_row_to_rid, wb_data)
    print(f"  Updated {updated} cells, {total} total formula cells")

    print("\n=== Step 5: Recalculating model ===")
    # Need async for calculate_model
    from backend.formula_engine import calculate_model
    from backend.db import get_db as get_async_db

    async def recalc():
        import aiosqlite
        adb = await aiosqlite.connect(DB_PATH)
        adb.row_factory = aiosqlite.Row

        # Monkey-patch for our needs
        adb.execute_fetchall = lambda q, p=(): _fetchall(adb, q, p)

        model = db.execute("SELECT id FROM models WHERE name = 'BaaS'").fetchone()
        result = await calculate_model(adb, model["id"])

        # Write results back
        total_written = 0
        for sheet_id, cells in result.items():
            for coord_key, value in cells.items():
                db.execute(
                    "UPDATE cell_data SET value = ? WHERE sheet_id = ? AND coord_key = ?",
                    (value, sheet_id, coord_key),
                )
                total_written += 1
        db.commit()
        await adb.close()
        return total_written

    async def _fetchall(adb, query, params=()):
        cursor = await adb.execute(query, params)
        rows = await cursor.fetchall()
        return rows

    written = asyncio.run(recalc())
    print(f"  Recalculated and wrote {written} cells")

    print("\n=== Step 6: Verifying against Excel ===")
    # Reload DB to get fresh values
    db.close()
    db = get_db()
    verify_against_excel(db, wb_data, all_row_to_rid)

    db.close()


if __name__ == "__main__":
    main()
