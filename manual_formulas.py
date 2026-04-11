#!/usr/bin/env python3
"""Manually set correct Pebble formulas for each sheet, then verify.

This is the reference/etalon approach:
1. Define exact formulas by hand (matching Excel logic)
2. Write to DB
3. Recalculate
4. Compare with Excel data_only values
5. Iterate until 100%
"""

import sqlite3
import json
import asyncio
from datetime import datetime

DB_PATH = "pebble.db"
EXCEL_PATH = "models.xlsx"

MONTH_NAMES_RU = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]


def get_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    return db


# ── BaaS.1 formulas ──────────────────────────────────────────────────────
# Key: DB record name → {formula: str, formula_first: str|None}
# formula = general formula (m2+)
# formula_first = first period formula (if different from general), or None

BAAS1_FORMULAS = {
    # ── Итого ──
    "чистый операционный доход": {
        "formula": "[чистый операционный доход (потребительский кредит)] + [чистый операционный доход (рассрочка)] + [чистый операционный доход (факторинг)] + [чистый операционный доход (кредитная линия)] + [чистый операционный доход (гарантия)] + [чистый операционный доход (продажа чужих кредитов)]",
    },
    "операционные расходы": {
        "formula": "[операционные расходы (потребительский кредит)] + [операционные расходы (рассрочка)] + [операционные расходы (факторинг)] + [операционные расходы (кредитная линия)] + [операционные расходы (гарантия)] + [операционные расходы (продажа чужих кредитов)]",
    },
    "прибыль": {
        "formula": "[прибыль (потребительский кредит)] + [прибыль (рассрочка)] + [прибыль (факторинг)] + [прибыль (кредитная линия)] + [прибыль (гарантия)] + [прибыль (продажа чужих кредитов)]",
    },

    # ── Потребительский кредит ──
    "количество партнеров (потребительский кредит)": {
        "formula": '[BaaS - параметры модели::кредитование (свои кредиты)]',
    },
    "количество выдач": {
        "formula": "[количество партнеров (потребительский кредит)] * [среднее количество выдач на 1 партнера]",
    },
    "выдачи": {
        "formula": "[количество выдач] * [ср. сумма выдачи]",
    },
    "погашения (потребительский кредит)": {
        "formula": '[кредитный портфель (потребительский кредит)](периоды="предыдущий") / [cр. срок портфеля (потребительский кредит)]',
        "formula_first": "0",
    },
    "кредитный портфель (потребительский кредит)": {
        "formula": '[кредитный портфель (потребительский кредит)](периоды="предыдущий") + [выдачи] - [погашения (потребительский кредит)]',
        "formula_first": "[выдачи] - [погашения (потребительский кредит)]",
    },
    "просроченная задолженность (потребительский кредит)": {
        "formula": "[кредитный портфель (потребительский кредит)] * [доля просроченной задолженности, % (потребительский кредит)]",
    },
    "РППУ на конец месяца (потребительский кредит)": {
        "formula": "-[кредитный портфель (потребительский кредит)] * [ср. % резервирования (потребительский кредит)]",
    },
    "% доход": {
        "formula": '([кредитный портфель (потребительский кредит)](периоды="предыдущий") + [кредитный портфель (потребительский кредит)]) / 2 * [ср. % ставка портфеля] / 12 * (1 - 2 / 102)',
        "formula_first": '(0 + [кредитный портфель (потребительский кредит)]) / 2 * [ср. % ставка портфеля] / 12 * (1 - 2 / 102)',
    },
    "трансфертный расход (потребительский кредит)": {
        "formula": '-([кредитный портфель (потребительский кредит)](периоды="предыдущий") + [кредитный портфель (потребительский кредит)]) / 2 * [ср. % ставка фондирования (потребительский кредит)] / 12',
        "formula_first": '-(0 + [кредитный портфель (потребительский кредит)]) / 2 * [ср. % ставка фондирования (потребительский кредит)] / 12',
    },
    "комиссия партнеру (потребительский кредит)": {
        "formula": "-[комиссия партнеру (потребительский кредит ставка)] * [выдачи]",
    },
    "расходы на провизии (потребительский кредит)": {
        "formula": '[РППУ на конец месяца (потребительский кредит)] - [РППУ на конец месяца (потребительский кредит)](периоды="предыдущий")',
        "formula_first": "[РППУ на конец месяца (потребительский кредит)] - 0",
    },
    "чистый операционный доход (потребительский кредит)": {
        "formula": "SUM([% доход], [трансфертный расход (потребительский кредит)], [комиссия партнеру (потребительский кредит)], [расходы на провизии (потребительский кредит)])",
    },
    "операционные расходы (потребительский кредит)": {
        "formula": "-[Операционные расходы и Инвестиции в BaaS::Расходы на персонал (Потребительский кредит)]",
    },
    "прибыль (потребительский кредит)": {
        "formula": "[чистый операционный доход (потребительский кредит)] + [операционные расходы (потребительский кредит)]",
    },

    # ── Рассрочка (BNPL) ──
    "количество партнеров (рассрочка)": {
        "formula": '[BaaS - параметры модели::кредитование (свои кредиты)]',
    },
    "количество выдач (рассрочка)": {
        "formula": "[количество партнеров (рассрочка)] * [среднее количество выдач на 1 партнера (рассрочка)]",
    },
    "выдачи (рассрочка)": {
        "formula": "[количество выдач (рассрочка)] * [ср. сумма выдачи (рассрочка)]",
    },
    "погашения (рассрочка)": {
        "formula": '[кредитный портфель (рассрочка)](периоды="предыдущий") / [cр. срок портфеля (рассрочка)]',
        "formula_first": "0",
    },
    "кредитный портфель (рассрочка)": {
        "formula": '[кредитный портфель (рассрочка)](периоды="предыдущий") + [выдачи (рассрочка)] - [погашения (рассрочка)]',
        "formula_first": "[выдачи (рассрочка)] - [погашения (рассрочка)]",
    },
    "РППУ на конец месяца (рассрочка)": {
        "formula": "-[кредитный портфель (рассрочка)] * [ср. % резервирования (рассрочка)]",
    },
    "комиссионный доход (рассрочка)": {
        "formula": "[выдачи (рассрочка)] * [комиссия мерчанта] * (1 - 2 / 102)",
    },
    "трансфертный расход (рассрочка)": {
        "formula": '-([кредитный портфель (рассрочка)](периоды="предыдущий") + [кредитный портфель (рассрочка)]) / 2 * [ср. % ставка фондирования (рассрочка)] / 12',
        "formula_first": '-(0 + [кредитный портфель (рассрочка)]) / 2 * [ср. % ставка фондирования (рассрочка)] / 12',
    },
    "комиссия партнеру (рассрочка)": {
        "formula": "-[комиссия партнеру (рассрочка ставка)] * [выдачи (рассрочка)]",
    },
    "расходы на провизии (рассрочка)": {
        "formula": '[РППУ на конец месяца (рассрочка)] - [РППУ на конец месяца (рассрочка)](периоды="предыдущий")',
        "formula_first": "[РППУ на конец месяца (рассрочка)] - 0",
    },
    "чистый операционный доход (рассрочка)": {
        "formula": "SUM([комиссионный доход (рассрочка)], [трансфертный расход (рассрочка)], [комиссия партнеру (рассрочка)], [расходы на провизии (рассрочка)])",
    },
    "операционные расходы (рассрочка)": {
        "formula": "-[Операционные расходы и Инвестиции в BaaS::Расходы на персонал (Рассрочка BNPL)]",
    },
    "прибыль (рассрочка)": {
        "formula": "[чистый операционный доход (рассрочка)] + [операционные расходы (рассрочка)]",
    },

    # ── Факторинг ──
    "количество партнеров (факторинг)": {
        "formula": '[BaaS - параметры модели::кредитование (свои кредиты)]',
    },
    "количество сделок": {
        "formula": "[количество партнеров (факторинг)] * [среднее количество сделок на 1 партнера]",
    },
    "сумма купленного долга": {
        "formula": "[количество сделок] * [ср. сумма купленного долга]",
    },
    "погашения (факторинг)": {
        "formula": '[факторинговый портфель](периоды="предыдущий") / [cр. срок портфеля (факторинг)]',
        "formula_first": "0",
    },
    "факторинговый портфель": {
        "formula": '[факторинговый портфель](периоды="предыдущий") + [сумма купленного долга] - [погашения (факторинг)]',
        "formula_first": "[сумма купленного долга] - [погашения (факторинг)]",
    },
    "РППУ на конец месяца (факторинг)": {
        "formula": "-[факторинговый портфель] * [ср. % резервирования (факторинг)]",
    },
    "комиссионный доход (факторинг)": {
        "formula": "[погашения (факторинг)] * [ср. % комиссии] * (1 - 2 / 102)",
    },
    "трансфертный расход (факторинг)": {
        "formula": '-([факторинговый портфель](периоды="предыдущий") + [факторинговый портфель]) / 2 * [ср. % ставка фондирования (факторинг)] / 12',
        "formula_first": '-(0 + [факторинговый портфель]) / 2 * [ср. % ставка фондирования (факторинг)] / 12',
    },
    "комиссия партнеру (факторинг)": {
        "formula": "-[комиссия партнеру (факторинг ставка)] * [сумма купленного долга]",
    },
    "расходы на провизии (факторинг)": {
        "formula": '[РППУ на конец месяца (факторинг)] - [РППУ на конец месяца (факторинг)](периоды="предыдущий")',
        "formula_first": "[РППУ на конец месяца (факторинг)] - 0",
    },
    "чистый операционный доход (факторинг)": {
        "formula": "SUM([комиссионный доход (факторинг)], [трансфертный расход (факторинг)], [комиссия партнеру (факторинг)], [расходы на провизии (факторинг)])",
    },
    "операционные расходы (факторинг)": {
        "formula": "-[Операционные расходы и Инвестиции в BaaS::Расходы на персонал (Факторинг)]",
    },
    "прибыль (факторинг)": {
        "formula": "[чистый операционный доход (факторинг)] + [операционные расходы (факторинг)]",
    },

    # ── Кредитная линия для ИП ──
    "количество партнеров (кредитная линия)": {
        "formula": '[BaaS - параметры модели::кредитование (свои кредиты)]',
    },
    "новые клиенты": {
        "formula": "[количество партнеров (кредитная линия)] * [среднее количество продаж на 1 партнера]",
    },
    "сумма новых КЛ": {
        "formula": "[новые клиенты] * [ср. сумма КЛ]",
    },
    "портфель КЛ": {
        "formula": '[портфель КЛ](периоды="предыдущий") + [сумма новых КЛ]',
        "formula_first": "[сумма новых КЛ]",
    },
    "кредитный портфель (кредитная линия)": {
        "formula": "[портфель КЛ] * [ср.использование КЛ]",
    },
    "просроченная задолженность (кредитная линия)": {
        "formula": "[кредитный портфель (кредитная линия)] * [доля просроченной задолженности, % (кредитная линия)]",
    },
    "РППУ (КП)": {
        "formula": "-[кредитный портфель (кредитная линия)] * [ср. % резервирования (КП)]",
    },
    "РППУ (УО)": {
        "formula": "-[ср. % резервирования (УО)] * ([портфель КЛ] - [кредитный портфель (кредитная линия)])",
    },
    "% доход (кредитная линия)": {
        "formula": '(1 - 2 / 102) * ([портфель КЛ](периоды="предыдущий") + [портфель КЛ]) / 2 * [ср. % ставка портфеля (кредитная линия)] / 12',
        "formula_first": '(1 - 2 / 102) * (0 + [портфель КЛ]) / 2 * [ср. % ставка портфеля (кредитная линия)] / 12',
    },
    "трансфертный расход (кредитная линия)": {
        "formula": '-([кредитный портфель (кредитная линия)](периоды="предыдущий") + [кредитный портфель (кредитная линия)]) / 2 * [ср. % ставка фондирования (кредитная линия)] / 12',
        "formula_first": '-(0 + [кредитный портфель (кредитная линия)]) / 2 * [ср. % ставка фондирования (кредитная линия)] / 12',
    },
    "комиссия партнеру (кредитная линия)": {
        "formula": "-[комиссия партнеру (кредитная линия ставка)] * [сумма новых КЛ]",
    },
    "расходы на провизии (кредитная линия)": {
        "formula": '([РППУ (КП)] - [РППУ (КП)](периоды="предыдущий")) + ([РППУ (УО)] - [РППУ (УО)](периоды="предыдущий"))',
        "formula_first": '([РППУ (КП)] - 0) + ([РППУ (УО)] - 0)',
    },
    "чистый операционный доход (кредитная линия)": {
        "formula": "SUM([% доход (кредитная линия)], [трансфертный расход (кредитная линия)], [комиссия партнеру (кредитная линия)], [расходы на провизии (кредитная линия)])",
    },
    "операционные расходы (кредитная линия)": {
        "formula": "-[Операционные расходы и Инвестиции в BaaS::Расходы на персонал (Кредитная линия для ИП)]",
    },
    "прибыль (кредитная линия)": {
        "formula": "[чистый операционный доход (кредитная линия)] + [операционные расходы (кредитная линия)]",
    },

    # ── Гарантия ──
    "количество партнеров (гарантия)": {
        "formula": '[BaaS - параметры модели::кредитование (свои кредиты)]',
    },
    "количество проданных гарантий": {
        "formula": "[количество партнеров (гарантия)] * [ср. количество продаж на 1 партнера (гарантия)]",
    },
    "проданные гарантии": {
        "formula": "[количество проданных гарантий] * [ср. сумма гарантии]",
    },
    "расходы на провизии (гарантия)": {
        "formula": '-[проданные гарантии] * [ср. % резервирования (УО) (гарантия)] + [проданные гарантии](периоды="предыдущий") * [ср. % резервирования (УО) (гарантия)](периоды="предыдущий")',
        "formula_first": "-[проданные гарантии] * [ср. % резервирования (УО) (гарантия)]",
    },
    "комиссионный доход (гарантия)": {
        "formula": "[комиссия банка (гарантия)] * [проданные гарантии] * (1 - 2 / 102)",
    },
    "комиссия партнера": {
        "formula": "-[комиссионный доход (гарантия)] * [комиссия партнеру (гарантия ставка)]",
    },
    "чистый операционный доход (гарантия)": {
        "formula": "SUM([расходы на провизии (гарантия)], [комиссионный доход (гарантия)], [комиссия партнера])",
    },
    "операционные расходы (гарантия)": {
        "formula": "-[Операционные расходы и Инвестиции в BaaS::Расходы на персонал (Гарантия)]",
    },
    "прибыль (гарантия)": {
        "formula": "[чистый операционный доход (гарантия)] + [операционные расходы (гарантия)]",
    },

    # ── Продажа чужих кредитов ──
    "количество партнеров (продажа чужих кредитов)": {
        "formula": '[BaaS - параметры модели::кредитование (чужие кредиты)]',
    },
    "количество продаж": {
        "formula": "[количество партнеров (продажа чужих кредитов)] * [ср. количество продаж на 1 партнера (продажа чужих кредитов)]",
    },
    "выдачи (продажа чужих кредитов)": {
        "formula": "[количество продаж] * [ср. сумма кредита]",
    },
    "чистый операционный доход (продажа чужих кредитов)": {
        "formula": "[комиссия банка (продажа чужих кредитов)] * [выдачи (продажа чужих кредитов)] * (1 - 2 / 102)",
    },
    "операционные расходы (продажа чужих кредитов)": {
        "formula": "-[Операционные расходы и Инвестиции в BaaS::Расходы на персонал (Продажа чужих кредитов)]",
    },
    "прибыль (продажа чужих кредитов)": {
        "formula": "[чистый операционный доход (продажа чужих кредитов)] + [операционные расходы (продажа чужих кредитов)]",
    },
}


def write_manual_formulas(db, sheet_name, formulas_dict):
    """Write manual formulas to DB for a specific sheet."""
    sheet = db.execute("SELECT id FROM sheets WHERE name = ?", (sheet_name,)).fetchone()
    if not sheet:
        print(f"  Sheet '{sheet_name}' not found!")
        return 0
    sid = sheet["id"]

    # Get analytics
    bindings = db.execute("""
        SELECT sa.analytic_id, a.is_periods
        FROM sheet_analytics sa JOIN analytics a ON a.id = sa.analytic_id
        WHERE sa.sheet_id = ? ORDER BY sa.sort_order
    """, (sid,)).fetchall()

    period_aid = [b["analytic_id"] for b in bindings if b["is_periods"]][0]
    ind_aid = [b["analytic_id"] for b in bindings if not b["is_periods"]][0]

    # Get leaf periods
    precs = db.execute(
        "SELECT id, parent_id, data_json FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
        (period_aid,),
    ).fetchall()
    child_map = {}
    for r in precs:
        if r["parent_id"]:
            child_map.setdefault(r["parent_id"], []).append(r["id"])
    leaf_periods = [r for r in precs if r["id"] not in child_map]
    first_period_rid = leaf_periods[0]["id"]

    # Get indicator records: name → rid
    ind_recs = db.execute(
        "SELECT id, data_json FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
        (ind_aid,),
    ).fetchall()
    name_to_rid = {}
    for r in ind_recs:
        data = json.loads(r["data_json"])
        name_to_rid[data["name"]] = r["id"]

    # Reset all formulas on this sheet
    db.execute("UPDATE cell_data SET rule = 'manual', formula = '' WHERE sheet_id = ? AND rule = 'formula'", (sid,))

    total = 0
    missing = []
    for name, info in formulas_dict.items():
        rid = name_to_rid.get(name)
        if not rid:
            missing.append(name)
            continue

        formula = info["formula"]
        formula_first = info.get("formula_first")

        for lp in leaf_periods:
            period_rid = lp["id"]
            coord_key = f"{period_rid}|{rid}"
            is_first = (period_rid == first_period_rid)

            if is_first and formula_first is not None:
                f_text = formula_first
            else:
                f_text = formula

            # Skip pure constants
            try:
                float(f_text)
                continue
            except (ValueError, TypeError):
                pass

            existing = db.execute(
                "SELECT id FROM cell_data WHERE sheet_id = ? AND coord_key = ?",
                (sid, coord_key),
            ).fetchone()

            if existing:
                db.execute(
                    "UPDATE cell_data SET rule = 'formula', formula = ? WHERE id = ?",
                    (f_text, existing["id"]),
                )
            else:
                import uuid
                db.execute(
                    "INSERT INTO cell_data (id, sheet_id, coord_key, value, data_type, rule, formula) VALUES (?,?,?,?,?,?,?)",
                    (str(uuid.uuid4()), sid, coord_key, "0", "sum", "formula", f_text),
                )
            total += 1

    if missing:
        print(f"  WARNING: {len(missing)} records not found: {missing[:5]}...")

    db.commit()
    return total


# ── BaaS.2 formulas ──────────────────────────────────────────────────────

def _deposit_group(suffix, currency_suffix="", rate_field_suffix="", port_name_prefix="портфель"):
    """Generate formula dict for a deposit currency group (KGS/RUB/USD)."""
    s = suffix  # e.g. "(KGS)"
    cs = currency_suffix  # e.g. "(KGS, %)" or "(KGS, сом)"
    return {
        f"количество партнеров {s}": {"formula": "[BaaS - параметры модели::депозиты]"},
        f"количество новых депозитов {s}": {"formula": f"[количество партнеров {s}] * [среднее количество депозитов на 1 партнера {s}]"},
        f"привлечение депозитов {s}": {"formula": f"[количество новых депозитов {s}] * [ср. сумма депозита {s}]"},
        f"капитализация % {s}": {
            "formula": f"-[% расход {s}](периоды=\"предыдущий\") / 2",
            "formula_first": "0",
        },
        f"возвраты {s}": {
            "formula": f"[{port_name_prefix} {s}](периоды=\"предыдущий\") / [cр. срок портфеля {s}]",
            "formula_first": "0",
        },
        f"{port_name_prefix} {s}": {
            "formula": f"[{port_name_prefix} {s}](периоды=\"предыдущий\") + [привлечение депозитов {s}] - [возвраты {s}] + [капитализация % {s}]",
            "formula_first": f"[привлечение депозитов {s}] - [возвраты {s}]",
        },
        f"% расход {s}": {
            "formula": f"-([{port_name_prefix} {s}](периоды=\"предыдущий\") + [{port_name_prefix} {s}]) / 2 * [ср. % ставка портфеля {s}] / 12",
            "formula_first": f"-(0 + [{port_name_prefix} {s}]) / 2 * [ср. % ставка портфеля {s}] / 12",
        },
        f"трансфертный доход {s}": {
            "formula": f"([{port_name_prefix} {s}](периоды=\"предыдущий\") + [{port_name_prefix} {s}]) / 2 * [ср. % ставка фондирования {s}] / 12",
            "formula_first": f"(0 + [{port_name_prefix} {s}]) / 2 * [ср. % ставка фондирования {s}] / 12",
        },
    }

BAAS2_FORMULAS = {
    # ── Итого ──
    "количество депозитов": {
        "formula": "[портфель (KGS)] / [ср. сумма депозита (KGS)] + [портфель (RUB)] / [ср. сумма депозита (RUB)] + [портфель (USD)] / [ср. сумма депозита (USD)]",
    },
    "портфель": {
        "formula": "[портфель (KGS)] + [портфель (RUB, сом)] + [портфель (USD, сом)]",
    },
    "% расход": {
        "formula": "[% расход (KGS)] + [% расход (RUB, сом)] + [% расход (USD, сом)]",
    },
    "трансфертный доход": {
        "formula": "[трансфертный доход (KGS)] + [трансфертный доход (RUB, сом)] + [трансфертный доход (USD, сом)]",
    },
    "комиссия партнеру": {
        "formula": "[комиссия партнеру (KGS, сом)] + [комиссия партнеру (RUB, сом)] + [комиссия партнеру (USD, сом)]",
    },
    # NOTE: Excel uses avg(prev_prev, prev) which can't be expressed in Pebble.
    # Using just prev/12 as approximation (exact for m2, close for others)
    "начисление расходов в ФЗД": {
        "formula": '-0.002 * [портфель](периоды="предыдущий") / 12',
        "formula_first": "0",
    },
    "чистый операционный доход": {
        "formula": "[чистый операционный доход (KGS)] + [чистый операционный доход (RUB, сом)] + [чистый операционный доход (USD, сом)] + [начисление расходов в ФЗД]",
    },
    "операционные расходы": {
        "formula": "-([Операционные расходы и Инвестиции в BaaS::Расходы на персонал (Депозиты)] + [Операционные расходы и Инвестиции в BaaS::Амортизационные отчисления (Депозиты)] + [Операционные расходы и Инвестиции в BaaS::Административные расходы (Депозиты)])",
    },
    "прибыль": {
        "formula": "[чистый операционный доход] + [операционные расходы]",
    },

    # ── KGS ──
    **_deposit_group("(KGS)", "(KGS, %)", "", "портфель"),
    "комиссия партнеру (KGS, сом)": {"formula": "-[комиссия партнеру (KGS, %)] * [привлечение депозитов (KGS)]"},
    "чистый операционный доход (KGS)": {"formula": "SUM([% расход (KGS)], [трансфертный доход (KGS)], [комиссия партнеру (KGS, сом)])"},

    # ── RUB (in rubles) ──
    **_deposit_group("(RUB)", "(RUB, %)", "", "портфель"),
    "комиссия партнеру (RUB, руб)": {"formula": "-[комиссия партнеру (RUB, %)] * [привлечение депозитов (RUB)]"},
    "чистый операционный доход (RUB)": {"formula": "SUM([% расход (RUB)], [трансфертный доход (RUB)], [комиссия партнеру (RUB, руб)])"},
    # RUB → сом conversion
    "курс RUB (сом за 1 RUB)": {"formula": "[BaaS - параметры модели::RUB]"},
    "портфель (RUB, сом)": {"formula": "[портфель (RUB)] * [курс RUB (сом за 1 RUB)]"},
    "% расход (RUB, сом)": {"formula": "[% расход (RUB)] * [курс RUB (сом за 1 RUB)]"},
    "трансфертный доход (RUB, сом)": {"formula": "[трансфертный доход (RUB)] * [курс RUB (сом за 1 RUB)]"},
    "комиссия партнеру (RUB, сом)": {"formula": "[комиссия партнеру (RUB, руб)] * [курс RUB (сом за 1 RUB)]"},
    "чистый операционный доход (RUB, сом)": {"formula": "[чистый операционный доход (RUB)] * [курс RUB (сом за 1 RUB)]"},

    # ── USD (in dollars) ──
    **_deposit_group("(USD)", "(USD, %)", "", "портфель"),
    "комиссия партнеру (USD, долл)": {"formula": "-[комиссия партнеру (USD, %)] * [привлечение депозитов (USD)]"},
    "чистый операционный доход (USD)": {"formula": "SUM([% расход (USD)], [трансфертный доход (USD)], [комиссия партнеру (USD, долл)])"},
    # USD → сом conversion
    "курс USD (сом за 1 USD)": {"formula": "[BaaS - параметры модели::USD]"},
    "портфель (USD, сом)": {"formula": "[портфель (USD)] * [курс USD (сом за 1 USD)]"},
    "% расход (USD, сом)": {"formula": "[% расход (USD)] * [курс USD (сом за 1 USD)]"},
    "трансфертный доход (USD, сом)": {"formula": "[трансфертный доход (USD)] * [курс USD (сом за 1 USD)]"},
    "комиссия партнеру (USD, сом)": {"formula": "[комиссия партнеру (USD, долл)] * [курс USD (сом за 1 USD)]"},
    "чистый операционный доход (USD, сом)": {"formula": "[чистый операционный доход (USD)] * [курс USD (сом за 1 USD)]"},
}


def restore_excel_values(db, wb_data, sheet_name, excel_sheet_name, data_start_col, label_col=1):
    """Restore original Excel data_only values AND reset all formulas to manual."""
    sheet = db.execute("SELECT id FROM sheets WHERE name = ?", (sheet_name,)).fetchone()
    if not sheet:
        return 0
    sid = sheet["id"]

    # Reset ALL formulas to manual first
    db.execute("UPDATE cell_data SET rule = 'manual', formula = '' WHERE sheet_id = ? AND rule = 'formula'", (sid,))

    bindings = db.execute("""
        SELECT sa.analytic_id, a.is_periods
        FROM sheet_analytics sa JOIN analytics a ON a.id = sa.analytic_id
        WHERE sa.sheet_id = ? ORDER BY sa.sort_order
    """, (sid,)).fetchall()

    period_aid = [b["analytic_id"] for b in bindings if b["is_periods"]][0]
    precs = db.execute(
        "SELECT id, parent_id, data_json FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
        (period_aid,),
    ).fetchall()
    child_map = {}
    for r in precs:
        if r["parent_id"]:
            child_map.setdefault(r["parent_id"], []).append(r["id"])
    leaf_periods = [r for r in precs if r["id"] not in child_map]
    period_name_to_rid = {json.loads(r["data_json"])["name"]: r["id"] for r in leaf_periods}

    ws = wb_data[excel_sheet_name]
    col_to_period = {}
    for c in range(data_start_col, min((ws.max_column or 1) + 1, 50)):
        for hr in range(1, 7):
            v = ws.cell(hr, c).value
            if isinstance(v, datetime):
                pname = f"{MONTH_NAMES_RU[v.month - 1]} {v.year}"
                if pname in period_name_to_rid:
                    col_to_period[c] = period_name_to_rid[pname]
                break

    # Get all indicator records with their names
    ind_aid = [b["analytic_id"] for b in bindings if not b["is_periods"]][0]
    ind_recs = db.execute(
        "SELECT id, data_json FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
        (ind_aid,),
    ).fetchall()

    # Build name → rid map for matching
    rid_set = {r["id"] for r in ind_recs}

    restored = 0
    for c, period_rid in col_to_period.items():
        for r in ind_recs:
            rid = r["id"]
            coord_key = f"{period_rid}|{rid}"
            # Find the corresponding Excel row by looking at all cells for this coord
            existing = db.execute(
                "SELECT id, value FROM cell_data WHERE sheet_id = ? AND coord_key = ?",
                (sid, coord_key),
            ).fetchone()
            if not existing:
                continue
            # We need to find the Excel value — but we don't have the row mapping here
            # So just update from the value that was already imported
            # The values were imported correctly originally — they just got overwritten by bad calculations

    # Simpler: just reload ALL values from the original import by using wb_data
    # Need row→rid mapping — use the apply_formulas one
    from apply_formulas import build_all_row_maps, build_excel_row_labels, build_db_record_names, match_excel_rows_to_db, SHEET_CONFIGS
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH)

    cfg = None
    for en, c in SHEET_CONFIGS.items():
        if c["db_sheet_name"] == sheet_name:
            cfg = c
            excel_name = en
            break
    if not cfg:
        return 0

    ws_f = wb[excel_name]
    labels = build_excel_row_labels(ws_f, cfg["label_col"])
    db_records = build_db_record_names(db, sheet_name)
    row_to_name, row_to_rid = match_excel_rows_to_db(labels, db_records, excel_name)

    ws_d = wb_data[excel_name]
    for row, rid in row_to_rid.items():
        for col, period_rid in col_to_period.items():
            val = ws_d.cell(row, col).value
            if val is None:
                continue
            coord_key = f"{period_rid}|{rid}"
            db.execute(
                "UPDATE cell_data SET value = ? WHERE sheet_id = ? AND coord_key = ?",
                (str(val), sid, coord_key),
            )
            restored += 1

    db.commit()
    return restored


def recalculate(db):
    """Recalculate all formulas."""
    from backend.formula_engine import calculate_model

    model = db.execute("SELECT id FROM models WHERE name = 'BaaS'").fetchone()
    mid = model["id"]

    async def run():
        import aiosqlite
        adb = await aiosqlite.connect(DB_PATH)
        adb.row_factory = aiosqlite.Row
        adb.execute_fetchall = lambda q, p=(): _fa(adb, q, p)
        result = await calculate_model(adb, mid)
        total = 0
        for sheet_id, cells in result.items():
            for coord_key, value in cells.items():
                db.execute(
                    "UPDATE cell_data SET value = ? WHERE sheet_id = ? AND coord_key = ?",
                    (value, sheet_id, coord_key),
                )
                total += 1
        db.commit()
        await adb.close()
        return total

    async def _fa(adb, q, p=()):
        cur = await adb.execute(q, p)
        return await cur.fetchall()

    return asyncio.run(run())


def verify_sheet(db, wb_data, sheet_name, excel_sheet_name, data_start_col):
    """Compare DB values with Excel data_only."""
    sheet = db.execute("SELECT id FROM sheets WHERE name = ?", (sheet_name,)).fetchone()
    sid = sheet["id"]

    bindings = db.execute("""
        SELECT sa.analytic_id, a.is_periods
        FROM sheet_analytics sa JOIN analytics a ON a.id = sa.analytic_id
        WHERE sa.sheet_id = ? ORDER BY sa.sort_order
    """, (sid,)).fetchall()

    period_aid = [b["analytic_id"] for b in bindings if b["is_periods"]][0]
    ind_aid = [b["analytic_id"] for b in bindings if not b["is_periods"]][0]

    precs = db.execute(
        "SELECT id, parent_id, data_json FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
        (period_aid,),
    ).fetchall()
    child_map = {}
    for r in precs:
        if r["parent_id"]:
            child_map.setdefault(r["parent_id"], []).append(r["id"])
    leaf_periods = [r for r in precs if r["id"] not in child_map]
    period_name_to_rid = {json.loads(r["data_json"])["name"]: r["id"] for r in leaf_periods}

    ind_recs = db.execute(
        "SELECT id, data_json FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
        (ind_aid,),
    ).fetchall()
    rid_to_name = {r["id"]: json.loads(r["data_json"])["name"] for r in ind_recs}

    ws = wb_data[excel_sheet_name]
    col_to_period = {}
    for c in range(data_start_col, min((ws.max_column or 1) + 1, 50)):
        for hr in range(1, 7):
            v = ws.cell(hr, c).value
            if isinstance(v, datetime):
                pname = f"{MONTH_NAMES_RU[v.month - 1]} {v.year}"
                if pname in period_name_to_rid:
                    col_to_period[c] = period_name_to_rid[pname]
                break

    cells = db.execute("SELECT coord_key, value FROM cell_data WHERE sheet_id = ?", (sid,)).fetchall()
    db_cells = {c["coord_key"]: c["value"] for c in cells}

    from apply_formulas import build_all_row_maps, SHEET_CONFIGS
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH)
    all_row_maps, all_row_to_rid, _, _ = build_all_row_maps(wb, db)

    excel_name = None
    for en, cfg in SHEET_CONFIGS.items():
        if cfg["db_sheet_name"] == sheet_name:
            excel_name = en
            break
    row_to_rid = all_row_to_rid.get(excel_name, {})

    total = 0
    match = 0
    mismatches = []
    for row, rid in row_to_rid.items():
        for col, period_rid in col_to_period.items():
            ck = f"{period_rid}|{rid}"
            ev = ws.cell(row, col).value
            if ev is None:
                continue
            try:
                e = float(ev)
                d = float(db_cells.get(ck, 0))
            except:
                continue
            total += 1
            tol = abs(e) * 0.01 + 0.01
            if abs(e - d) <= tol:
                match += 1
            elif len(mismatches) < 10:
                name = rid_to_name.get(rid, "?")
                mismatches.append(f"  {name[:40]}: excel={e:.4f} db={d:.4f}")

    pct = match / total * 100 if total else 0
    print(f"  {sheet_name}: {match}/{total} ({pct:.1f}%)")
    for m in mismatches:
        print(m)
    return match, total


def main():
    from openpyxl import load_workbook

    db = get_db()
    wb_data = load_workbook(EXCEL_PATH, data_only=True)

    print("=== Step 1: Restore ALL sheets to manual + Excel values ===")
    all_sheets = [
        ("BaaS - параметры модели", "0", 4),
        ("BaaS - Онлайн кредитование", "BaaS.1", 4),
        ("BaaS - Онлайн депозит", "BaaS.2", 4),
        ("BaaS - Онлайн транзакционный бизнес", "BaaS.3", 4),
        ("Баланс BaaS", "BS", 3),
        ("Финансовый результат BaaS", "PL", 3),
        ("Операционные расходы и Инвестиции в BaaS", "OPEX+CAPEX", 5),
    ]
    for db_name, excel_name, start_col in all_sheets:
        r = restore_excel_values(db, wb_data, db_name, excel_name, start_col)
        print(f"  {excel_name}: restored {r} values")

    print("\n=== Step 2: Write manual formulas ===")
    n = write_manual_formulas(db, "BaaS - Онлайн кредитование", BAAS1_FORMULAS)
    print(f"  BaaS.1: {n} formula cells")
    n = write_manual_formulas(db, "BaaS - Онлайн депозит", BAAS2_FORMULAS)
    print(f"  BaaS.2: {n} formula cells")

    print("\n=== Step 3: Recalculate ===")
    n = recalculate(db)
    print(f"  Recalculated {n} cells")

    print("\n=== Step 4: Verify ===")
    verify_sheet(db, wb_data, "BaaS - Онлайн кредитование", "BaaS.1", 4)
    verify_sheet(db, wb_data, "BaaS - Онлайн депозит", "BaaS.2", 4)

    db.close()


if __name__ == "__main__":
    main()
