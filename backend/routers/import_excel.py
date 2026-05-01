"""Import an Excel workbook as a Pebble model using Claude API for intelligent
structure analysis.

Flow:
1. Extract text representation of each Excel sheet (headers, row labels, data presence)
2. Apply Knowledge Base (KB) patterns for hierarchy detection
3. Ask clarifying questions for ambiguous cases (interactive Q&A via SSE)
4. Fall back to Claude API for sheets where KB is insufficient
5. Create model + period analytic with proper year/quarter/month hierarchy
6. Create indicator analytics per sheet with hierarchical records
7. Import cell values (manual vs formula detection by theme color)
"""

import asyncio
import uuid
import json
import io
import os
import re
import logging
from datetime import datetime, date
from calendar import monthrange

from fastapi import APIRouter, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from backend.db import get_db

router = APIRouter(prefix="/api/import", tags=["import"])
log = logging.getLogger(__name__)

# ── Q&A session management ────────────────────────────────────────────────
# Each streaming import gets a unique session_id.  When the backend needs
# user input it yields a "question" SSE event and awaits an asyncio.Event.
# The answer endpoint sets the event with the user's response.

_qa_sessions: dict[str, dict] = {}  # session_id -> {questions: {qid: {event, answer}}}


def _create_qa_session() -> str:
    sid = str(uuid.uuid4())
    _qa_sessions[sid] = {"questions": {}}
    return sid


def _cleanup_qa_session(sid: str) -> None:
    _qa_sessions.pop(sid, None)


async def _ask_question(sid: str, qid: str, timeout: float = 300) -> str | None:
    """Register a question and wait for the user's answer (up to timeout seconds)."""
    evt = asyncio.Event()
    _qa_sessions[sid]["questions"][qid] = {"event": evt, "answer": None}
    try:
        await asyncio.wait_for(evt.wait(), timeout=timeout)
        return _qa_sessions[sid]["questions"][qid]["answer"]
    except asyncio.TimeoutError:
        return None
    finally:
        _qa_sessions[sid]["questions"].pop(qid, None)


@router.post("/answer/{session_id}")
async def submit_answer(session_id: str, body: dict = None):
    """Submit an answer to a pending import question."""
    if body is None:
        return {"error": "No body"}
    from fastapi import Request
    qid = body.get("question_id", "")
    answer = body.get("answer", "")
    session = _qa_sessions.get(session_id)
    if not session:
        return {"error": "Session not found or expired"}
    q = session["questions"].get(qid)
    if not q:
        return {"error": "Question not found or already answered"}
    q["answer"] = answer
    q["event"].set()
    return {"ok": True}


MONTH_NAMES_RU = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]
QUARTER_NAMES_RU = ["1-й квартал", "2-й квартал", "3-й квартал", "4-й квартал"]

# ── Claude API prompts ─────────────────────────────────────────────────────

PEBBLE_SYSTEM_PROMPT = """\
You are an assistant that analyzes Excel financial models for import into Pebble.

Pebble object model:
- Model: top-level container (e.g. "BaaS Financial Model")
- Analytic: a dimension/axis (e.g. "Периоды", "Показатели"). Has fields and hierarchical records.
  - Records are hierarchical (parent_id tree): groups contain children
- Sheet: a named view binding 2+ analytics. First = columns (periods), rest = rows
- Cell: value + rule (manual = user input, formula = computed)

Pebble formula syntax (references by indicator NAME, not cell):
- [indicator_name] — value of another indicator, same period, same group
- [indicator_name](периоды="предыдущий") — value from PREVIOUS period
- [SheetName::indicator_name] — cross-sheet reference (use :: separator, sheet display name)
- Standard math: +, -, *, /, parentheses, numbers
- SUM([a], [b], [c]) — sum of multiple indicators
- For first-period special cases (e.g. no previous period), use "0" as value

Examples:
  Excel: =D13*D14 (where R13=количество партнеров, R14=ср. количество выдач)
  Pebble: [количество партнеров] * [среднее количество выдач на 1 партнера]

  Excel: =D20/E18 (prev period's портфель / current ср.срок)
  Pebble: [кредитный портфель](периоды="предыдущий") / [cр. срок портфеля]

  Excel: =(C20+D20)/2*D24/12 (avg of prev+current портфель * rate / 12)
  Pebble: ([кредитный портфель](периоды="предыдущий") + [кредитный портфель]) / 2 * [ср. % ставка портфеля] / 12

  Excel: =D27-C27 (delta from previous period)
  Pebble: [РППУ на конец месяца] - [РППУ на конец месяца](периоды="предыдущий")

  Excel: =SUM(D28:D31) (sum of rows 28-31)
  Pebble: SUM([% доход], [трансфертный расход], [комиссия партнеру], [расходы на провизии])
"""

SHEET_ANALYSIS_PROMPT = """\
Analyze ONE sheet from an Excel financial model. The text shows:
- Header rows (dates, labels)
- Row labels with row numbers, formatting (BOLD, indent), data presence flags
- F1= formula for first period, F2= formula for second period (if different)
- INPUT(bg=color) = cell has non-default background color (yellow, beige, green, etc.)
  This usually means manual user input — numbers typed by hand, not computed.
  If a row has INPUT(bg=...) AND has numbers but NO Excel formula, it's almost certainly manual.
  If a row has INPUT(bg=...) AND ALSO has a FORMULA flag, the color overrides: treat as manual (keep the number, ignore the formula).
- FORMULA = cell contains an Excel formula (computed)

Return a JSON describing the sheet's indicator hierarchy WITH formulas.

RULES:
1. Build a hierarchical tree. BOLD rows with "ЕИ" = group headers (product types, sections).
2. Items beneath a group are its children indicators.
3. For each indicator: name, unit, row, rule (manual/formula), and for formulas: the Pebble formula.
4. Convert Excel cell references to Pebble [indicator_name] references using the ROW MAPPING.
5. If F1 and F2 differ (e.g. first period is 0, subsequent use @prev), provide TWO formulas: "formula_first" and "formula".
6. Cross-sheet references like ='0'!D10 → [SheetDisplayName::indicator_name] using :: separator and the display_name of the referenced sheet.
7. display_name: from A1/B1 title. data_start_col: first period column number.
8. Skip header rows (Показатель, ЕИ, Отв.исп.) — they are not indicators.
9. CRITICAL: Every [indicator_name] in a formula must EXACTLY match the "name" field of SOME indicator in the JSON output. If the same indicator name appears in multiple groups, append a disambiguating suffix in parentheses to BOTH the name and all formula references. Example: "портфель" in KGS group → name "портфель (KGS)", in RUB group → "портфель (RUB)".
10. For "Итого" / summary rows that SUM across groups: use the EXACT disambiguated names. E.g. formula: "[портфель (KGS)] + [портфель (RUB)] + [портфель (USD)]". NEVER write [портфель] + [портфель] + [портфель] — that's a self-reference!
11. Rows like "ВСЕГО АКТИВЫ", "ИТОГО ОБЯЗАТЕЛЬСТВА" are NOT separate indicators — they are the group header itself. The group header row IS the aggregation row.
12. GROUPING BY NAME PATTERN — MANDATORY, IGNORE MISSING FORMULA: A row is a group header whenever its label matches ANY grouping pattern below, even if the Excel cell is empty or contains no formula. NEVER mark such a row as "manual". Grouping patterns (case-insensitive):
    - ends with "в т.ч.:" / "в т.ч." / "в том числе:" / "в том числе" / "включая:" / "включая"
      EXAMPLE: "общее количество партнеров, в т.ч.:" → is_group=true, rule="sum_children", children=[all indented rows below it]
    - starts with "Итого" / "Всего" / "Всего по " / "Общее " / "Общий " / "Общая "
    - "Суммарн" / "Сумма " prefixes
    For ALL matching rows: is_group=true, rule="sum_children", NO formula. Rows below at greater indent are its children.
13. INDENTATION RULE: A row with indent=N whose immediately following rows have indent>N is a parent group header. Attach those deeper-indent rows as children, even if the header row has no bold/formula. This is a HARD rule — do not put indented rows as siblings of their parent.
14. Sheet 0 / «параметры» caveat: most rows there ARE manual input, BUT a row matching rule 12 above is still a grouping header with `sum_children`, not manual. Do not blanket-mark every row on sheet 0 as manual.

15. NAME DISAMBIGUATION IS MANDATORY. When the same indicator name appears under different product groups (e.g. "количество выдач" under both "Потребительский кредит" and "BNPL"), you MUST append the group name in parentheses to make each name globally unique:
   "количество выдач (потребительский)" under Потребительский кредит, "количество выдач (BNPL)" under BNPL.
   DO NOT leave bare duplicate names — they break formula references! Use a SHORT suffix (key word from group).
16. CROSS-SHEET REFERENCES: the ONLY valid syntax is [SheetDisplayName::indicator_name] with :: separator.
   CORRECT: [параметры::количество партнеров]
   WRONG: [количество партнеров]('0'::периоды="текущий")
17. formula_first MUST differ from formula when the formula uses (периоды="предыдущий"):
   - Delta formulas (X - X(prev)): formula_first = "0"
   - Average with prev ((X+X(prev))/2): formula_first uses only current value
   NEVER copy the main formula verbatim as formula_first if it references "предыдущий".

Return ONLY valid JSON, no markdown:
{"excel_name":"Tab","display_name":"Title","data_start_col":4,"indicators":[
  {"name":"Тип продукта","unit":"","row":12,"is_group":true,"children":[
    {"name":"показатель","unit":"тыс сом","row":13,"is_group":false,"children":[],
     "rule":"formula","formula":"[a] * [b]"},
    {"name":"ввод","unit":"%","row":14,"is_group":false,"children":[],
     "rule":"manual"}
  ]}
]}

For formulas with first-period exception:
{"rule":"formula","formula":"[портфель](периоды=\"предыдущий\") / [срок]","formula_first":"0"}

Sheet content:
"""

FORMULA_FIX_PROMPT = """\
You have a Pebble model with these sheets and indicator names.
For each formula cell, the formula must reference EXACT indicator names from this list.
Cross-sheet references use :: separator: [SheetDisplayName::indicator_name]

AVAILABLE INDICATORS PER SHEET:
{indicators_context}

FORMULA CELLS TO FIX (current formula → expected behavior based on Excel formula):
{formula_fixes}

Return ONLY a JSON array of fixes:
[{{"coord_key": "...", "formula": "corrected formula", "formula_first": "optional first-period formula"}}]
"""


# ── Excel text extraction ──────────────────────────────────────────────────

def _extract_sheet_text(ws, sheet_name: str, max_rows: int = 500) -> str:
    """Extract a compact text representation of a sheet for Claude analysis."""
    lines = [f"=== Sheet: {sheet_name} ==="]
    max_col = min(ws.max_column or 1, 200)
    max_row = min(ws.max_row or 1, max_rows)

    # Header rows (first 6)
    lines.append("--- Header rows ---")
    for r in range(1, min(7, max_row + 1)):
        row_vals = []
        for c in range(1, min(max_col + 1, 50)):
            v = ws.cell(r, c).value
            if v is not None:
                if isinstance(v, datetime):
                    row_vals.append(f"{get_column_letter(c)}:{v.strftime('%Y-%m-%d')}")
                else:
                    row_vals.append(f"{get_column_letter(c)}:{str(v)[:60]}")
        if row_vals:
            lines.append(f"  Row {r}: {' | '.join(row_vals)}")

    # Row labels with formatting and data presence
    lines.append("--- Row labels ---")
    for r in range(1, max_row + 1):
        # Collect labels from first few columns
        labels = []
        for c in range(1, min(6, max_col + 1)):
            v = ws.cell(r, c).value
            if v is not None and not isinstance(v, datetime):
                s = str(v).strip()
                if s and len(s) < 100:
                    labels.append(f"{get_column_letter(c)}='{s}'")

        if not labels:
            continue

        # Check formatting
        cell_a = ws.cell(r, 1)
        is_bold = cell_a.font and cell_a.font.bold
        indent = cell_a.alignment.indent if cell_a.alignment and cell_a.alignment.indent else 0

        # Check data presence in period columns (sample cols 4-10)
        has_data = False
        has_formula = False
        for c in range(4, min(15, max_col + 1)):
            cv = ws.cell(r, c).value
            if cv is not None:
                has_data = True
                if isinstance(cv, str) and cv.startswith("="):
                    has_formula = True
                break

        # Check cell background color (non-default = likely manual input)
        is_input = False
        cell_color = None
        for c in range(4, min(15, max_col + 1)):
            cv = ws.cell(r, c).value
            if cv is not None:
                cell_color = _get_cell_bg_color(ws.cell(r, c))
                if cell_color:
                    is_input = True
                break

        # Extract first two period formulas for formula cells
        formula_m1 = ""
        formula_m2 = ""
        if has_formula:
            for c in range(4, min(max_col + 1, 50)):
                cv = ws.cell(r, c).value
                if cv and isinstance(cv, str) and cv.startswith("="):
                    if not formula_m1:
                        formula_m1 = str(cv)[:80]
                    elif not formula_m2:
                        formula_m2 = str(cv)[:80]
                        break
                elif cv is not None and not formula_m1:
                    # First period is a constant (e.g. 0 for погашения)
                    formula_m1 = str(cv)[:20]
                    continue

        flags = []
        if is_bold:
            flags.append("BOLD")
        if indent:
            flags.append(f"indent={int(indent)}")
        if has_data:
            flags.append("HAS_DATA")
        if has_formula:
            flags.append("FORMULA")
        if is_input and cell_color:
            flags.append(f"INPUT(bg={cell_color})")

        flag_str = f" [{', '.join(flags)}]" if flags else ""
        formula_str = ""
        if formula_m1:
            formula_str = f" F1={formula_m1}"
            if formula_m2 and formula_m2 != formula_m1:
                formula_str += f" F2={formula_m2}"
        lines.append(f"  Row {r}: {' | '.join(labels)}{flag_str}{formula_str}")

    return "\n".join(lines)


def _get_cell_bg_color(cell) -> str | None:
    """Return human-readable background color description, or None if default/white."""
    fill = cell.fill
    if not fill or fill.patternType != "solid":
        return None
    fg = fill.fgColor
    if not fg:
        return None
    # Theme-based colors
    if fg.type == "theme":
        try:
            theme = int(fg.theme)
            theme_names = {
                0: None,  # white / default
                1: None,  # black text (not a bg color)
                2: None,  # light gray (often default)
                3: "dark gray", 4: "blue", 5: "orange",
                6: "green", 7: "yellow", 8: "teal", 9: "purple",
            }
            return theme_names.get(theme)
        except (TypeError, ValueError):
            pass
    # RGB-based
    if fg.type == "rgb" and isinstance(fg.rgb, str):
        raw = fg.rgb
        # Strip alpha prefix if present (e.g. "00FFFFFF" → "FFFFFF")
        hex6 = raw[-6:] if len(raw) >= 6 else raw
        if hex6.upper() in ("FFFFFF", "000000"):
            return None  # white or black (default)
        r = int(hex6[0:2], 16)
        g = int(hex6[2:4], 16)
        b = int(hex6[4:6], 16)
        if r > 240 and g > 240 and b > 240:
            return None  # near-white
        # Describe the color
        if r > 200 and g > 200 and b < 100:
            return "yellow"
        if r > 200 and g > 180 and b < 150 and b < r - 60:
            return "beige/light yellow"
        if r > 200 and g < 150 and b < 150:
            return "red/pink"
        if r < 150 and g > 200 and b < 150:
            return "green"
        if r < 150 and g < 150 and b > 200:
            return "blue"
        if r > 200 and g > 150 and b < 100:
            return "orange"
        return f"#{hex6}"
    return None


def _is_input_cell(cell) -> bool:
    """Check if cell has a non-default background color (likely manual input)."""
    return _get_cell_bg_color(cell) is not None


# ── LLM API call ──────────────────────────────────────────────────────────

# Disambiguation/cross-sheet/formula_first rules merged into SHEET_ANALYSIS_PROMPT (rules 15-17)


def _get_import_llm_provider() -> str:
    """Return 'openai-compat' (default, Qwen via Together AI) or 'claude' as fallback."""
    return os.environ.get("PEBBLE_IMPORT_LLM", "openai-compat").lower()


def _get_claude_client():
    """Create Anthropic client from environment."""
    import anthropic

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY environment variable not set")

    base_url = os.environ.get("ANTHROPIC_BASE_URL")
    kwargs = {"api_key": api_key}
    if base_url:
        kwargs["base_url"] = base_url

    return anthropic.Anthropic(**kwargs)


def _parse_claude_json(response_text: str) -> dict:
    """Parse JSON from Claude response, handling markdown fences and common issues."""
    import re
    text = response_text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1]
        if "```" in text:
            text = text[: text.rfind("```")]
        text = text.strip()
    # Fix trailing commas before } or ]
    text = re.sub(r',\s*([}\]])', r'\1', text)
    return json.loads(text)


# ── LLM cache (DB-backed with file fallback for migration) ───────────────

# Legacy file cache dir (kept for reading old entries)
_LLM_CACHE_DIR = os.path.join(
    os.environ.get("PEBBLE_DB", "").rsplit("/", 1)[0] or os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    ".llm_cache"
)


def _cache_hash(key: str) -> str:
    import hashlib
    return hashlib.sha256(key.encode()).hexdigest()[:32]


async def _llm_cache_get(key: str):
    """Get cached LLM response from DB, falling back to legacy file cache."""
    from backend.db import get_db
    h = _cache_hash(key)
    try:
        db = get_db()
        rows = await db.execute_fetchall(
            "SELECT response FROM llm_cache WHERE cache_key = ?", (h,)
        )
        if rows:
            return json.loads(rows[0]["response"])
    except Exception:
        pass
    # Fallback: legacy file cache
    import hashlib
    h16 = hashlib.sha256(key.encode()).hexdigest()[:16]
    path = os.path.join(_LLM_CACHE_DIR, f"{h16}.json")
    if os.path.exists(path):
        with open(path) as f:
            data = json.load(f)
        # Migrate to DB
        try:
            await _llm_cache_set(key, data)
        except Exception:
            pass
        return data
    return None


async def _llm_cache_set(key: str, value, provider: str = ""):
    """Save LLM response to DB cache."""
    from backend.db import get_db
    h = _cache_hash(key)
    try:
        db = get_db()
        await db.execute(
            "INSERT OR REPLACE INTO llm_cache (cache_key, response, provider) VALUES (?, ?, ?)",
            (h, json.dumps(value, ensure_ascii=False), provider),
        )
        await db.commit()
    except Exception as e:
        log.warning("Failed to cache LLM response: %s", e)


# Limit concurrent LLM requests to avoid API throttling (503s, timeouts)
_LLM_REQUEST_SEM: asyncio.Semaphore | None = None

def _get_llm_semaphore() -> asyncio.Semaphore:
    global _LLM_REQUEST_SEM
    if _LLM_REQUEST_SEM is None:
        _LLM_REQUEST_SEM = asyncio.Semaphore(5)
    return _LLM_REQUEST_SEM


async def _analyze_sheet_with_openai_compat(sheet_text: str) -> dict:
    """Analyze one sheet via OpenAI-compatible API (Together AI, etc.)."""
    import httpx, asyncio

    base_url = os.environ.get("PEBBLE_IMPORT_LLM_BASE_URL", "https://api.together.xyz/v1")
    api_key = os.environ.get("PEBBLE_IMPORT_LLM_API_KEY", "")
    model = os.environ.get("PEBBLE_IMPORT_LLM_MODEL", "Qwen/Qwen3-235B-A22B-Instruct-2507-tput")

    if not api_key:
        raise RuntimeError("PEBBLE_IMPORT_LLM_API_KEY not set")

    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    body = {
        "model": model,
        "max_tokens": 16384,
        "temperature": 0.1,
        "messages": [
            {"role": "system", "content": PEBBLE_SYSTEM_PROMPT
             + "\n\nIMPORTANT: Return ONLY valid JSON. No markdown fences, no comments, no trailing commas. No <think> tags."},
            {"role": "user", "content": SHEET_ANALYSIS_PROMPT + sheet_text},
        ],
    }
    url = base_url.rstrip("/") + "/chat/completions"

    loop = asyncio.get_event_loop()
    def _do_request():
        with httpx.Client(timeout=300) as c:
            resp = c.post(url, json=body, headers=headers)
            resp.raise_for_status()
            return resp.json()

    async with _get_llm_semaphore():
        data = await loop.run_in_executor(None, _do_request)
    raw_text = data["choices"][0]["message"]["content"]
    # Strip <think> tags from Qwen3
    raw_text = re.sub(r'<think>.*?</think>', '', raw_text, flags=re.DOTALL).strip()
    return _parse_claude_json(raw_text)


async def _analyze_sheet_with_claude_direct(client, sheet_text: str, retries: int = 3) -> dict:
    """Analyze one sheet with Claude Anthropic API."""
    import asyncio
    loop = asyncio.get_event_loop()
    for attempt in range(retries):
        try:
            message = await loop.run_in_executor(None, lambda: client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=16384,
                system=PEBBLE_SYSTEM_PROMPT + "\n\nIMPORTANT: Return ONLY valid JSON. No markdown fences, no comments, no trailing commas.",
                messages=[
                    {"role": "user", "content": SHEET_ANALYSIS_PROMPT + sheet_text},
                    {"role": "assistant", "content": "{"},  # prefill to force JSON
                ],
            ))
            return _parse_claude_json("{" + message.content[0].text)
        except Exception as e:
            if attempt < retries - 1 and ("overloaded" in str(e).lower() or "529" in str(e)):
                await asyncio.sleep(2 ** attempt)
                continue
            raise


async def _analyze_sheet_with_claude(client, sheet_text: str, retries: int = 3) -> dict:
    """Analyze one sheet with LLM. Default: Together AI (Qwen), fallback: Claude.
    Returns sheet config dict (cached in DB)."""

    # Check cache first
    cached = await _llm_cache_get(sheet_text)
    if cached:
        log.info("Cache hit for sheet analysis")
        return cached

    provider = _get_import_llm_provider()

    if provider == "openai-compat":
        import asyncio as _aio
        last_err = None
        for attempt in range(retries):
            try:
                result = await _analyze_sheet_with_openai_compat(sheet_text)
                await _llm_cache_set(sheet_text, result, provider="openai-compat")
                return result
            except Exception as e:
                last_err = e
                log.warning("OpenAI-compat attempt %d/%d failed: %s", attempt + 1, retries, e)
                if attempt < retries - 1:
                    await _aio.sleep(2 ** attempt)
        # All retries failed — try Claude fallback if available
        if os.environ.get("ANTHROPIC_API_KEY") and client is not None:
            log.warning("Falling back to Claude after %d failures", retries)
            result = await _analyze_sheet_with_claude_direct(client, sheet_text, retries)
            await _llm_cache_set(sheet_text, result, provider="claude-fallback")
            return result
        raise RuntimeError(f"LLM analysis failed after {retries} retries: {last_err}")

    # Explicit Claude mode
    result = await _analyze_sheet_with_claude_direct(client, sheet_text, retries)
    await _llm_cache_set(sheet_text, result, provider="claude")
    return result


# ── Sheet chunking for large sheets ──────────────────────────────────────

# Threshold: sheets above this size (chars) get chunked
_CHUNK_THRESHOLD = 5000


def _split_sheet_into_chunks(sheet_text: str) -> list[str]:
    """Split a large sheet text into chunks by BOLD group headers.

    Each chunk contains the header section + one BOLD group and its children.
    Small sheets (< _CHUNK_THRESHOLD chars) are returned as a single chunk.
    """
    if len(sheet_text) < _CHUNK_THRESHOLD:
        return [sheet_text]

    lines = sheet_text.split("\n")

    # Separate header section (everything before "--- Row labels ---" + first few rows)
    header_lines = []
    body_lines = []
    in_body = False
    for line in lines:
        if "--- Row labels ---" in line:
            header_lines.append(line)
            in_body = True
            continue
        if not in_body:
            header_lines.append(line)
        else:
            body_lines.append(line)

    if not body_lines:
        return [sheet_text]

    # Find BOLD row positions (group headers) in body
    groups: list[list[str]] = []
    current_group: list[str] = []

    for line in body_lines:
        is_bold = "BOLD" in line and "indent=" not in line  # top-level BOLD = group header
        if is_bold and current_group:
            groups.append(current_group)
            current_group = [line]
        else:
            current_group.append(line)
    if current_group:
        groups.append(current_group)

    # If only 1-2 groups, split by row-number gaps or fixed size
    if len(groups) <= 2:
        import re
        _MAX_ROWS_PER_CHUNK = 30
        segments: list[list[str]] = []
        seg: list[str] = []
        prev_row = 0
        for line in body_lines:
            m = re.match(r'\s*Row\s+(\d+)', line)
            cur_row = int(m.group(1)) if m else prev_row
            # Split at row gaps (skipped Excel rows = section break) if segment is big enough
            if m and cur_row > prev_row + 1 and len(seg) >= _MAX_ROWS_PER_CHUNK:
                segments.append(seg)
                seg = []
            seg.append(line)
            prev_row = cur_row
        if seg:
            segments.append(seg)
        # If still too few segments, split every _MAX_ROWS_PER_CHUNK lines
        if len(segments) < 3:
            segments = []
            for i in range(0, len(body_lines), _MAX_ROWS_PER_CHUNK):
                segments.append(body_lines[i:i + _MAX_ROWS_PER_CHUNK])
        if len(segments) >= 2:
            groups = segments
        else:
            return [sheet_text]

    # Build chunks: header + each group
    header_text = "\n".join(header_lines)
    chunks = []
    for group_lines in groups:
        chunk = header_text + "\n" + "\n".join(group_lines)
        chunks.append(chunk)

    log.info("Split sheet into %d chunks (from %d chars)", len(chunks), len(sheet_text))
    return chunks


def _merge_chunk_results(results: list[dict]) -> dict:
    """Merge multiple chunk analysis results into a single sheet config.

    Each chunk gets the sheet header in front of its body, so the LLM tends to
    re-emit header rows (e.g. an "АКТИВЫ" / "ОБЯЗАТЕЛЬСТВА" wrapper) in every
    chunk. Naively concatenating yields N copies of the same wrapper.

    We collapse duplicates by (row, name): repeated wrappers are merged into a
    single parent whose children are the union of all chunks' children. Inside
    that union, each unique excel_row is kept once (first occurrence wins).
    """
    if len(results) == 1:
        return results[0]

    merged = {
        "excel_name": results[0].get("excel_name", ""),
        "display_name": results[0].get("display_name", ""),
        "data_start_col": results[0].get("data_start_col"),
        "indicators": [],
    }

    def _key(item: dict) -> tuple:
        row = item.get("row")
        name = (item.get("name") or "").strip().lower()
        return (row if row is not None else "", name)

    # Global row tracking — each excel_row may only appear once across the
    # whole tree, since cell rows map 1:1 to rids.
    rows_seen: set[int] = set()

    def _merge_into(out: list[dict], items: list[dict]):
        index: dict[tuple, dict] = {_key(it): it for it in out}
        for item in items:
            row = item.get("row")
            k = _key(item)
            if k in index:
                existing = index[k]
                existing_children = existing.setdefault("children", []) or []
                _merge_into(existing_children, item.get("children") or [])
                continue
            if row is not None and row in rows_seen:
                # Different name for an already-classified row — drop
                continue
            new_item = dict(item)
            new_item["children"] = []
            if row is not None:
                rows_seen.add(row)
            _merge_into(new_item["children"], item.get("children") or [])
            out.append(new_item)
            index[k] = new_item

    for r in results:
        _merge_into(merged["indicators"], r.get("indicators", []))

    # Use display_name from first non-empty result
    for r in results:
        if r.get("display_name"):
            merged["display_name"] = r["display_name"]
            break

    return merged


async def _analyze_sheet_chunked(client, sheet_text: str) -> dict:
    """Analyze a sheet, chunking if large. Sends chunks in parallel."""
    import asyncio

    chunks = _split_sheet_into_chunks(sheet_text)
    if len(chunks) == 1:
        return await _analyze_sheet_with_claude(client, chunks[0])

    # Process chunks in parallel
    tasks = [_analyze_sheet_with_claude(client, chunk) for chunk in chunks]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    # Filter out errors
    good_results = []
    for i, r in enumerate(results):
        if isinstance(r, Exception):
            log.warning("Chunk %d failed: %s", i, r)
        else:
            good_results.append(r)

    if not good_results:
        raise RuntimeError("All chunks failed")

    return _merge_chunk_results(good_results)


async def _verify_import_against_excel(
    db, model_id: str, wb, created_sheets: list[dict],
    tolerance: float = 0.01,
) -> list[dict]:
    """Compare all imported Pebble cell values against the original Excel.

    Uses col_to_period_rid and row_to_rid mappings stored on each created_sheet
    entry to find the Excel cell for each Pebble cell_data row.

    Returns list of mismatch dicts: {sheet, indicator, period, excel, pebble, rel_error}.
    """
    mismatches = []

    for cs in created_sheets:
        sheet_id = cs["id"]
        sheet_name = cs["name"]
        excel_name = cs.get("excel_name", sheet_name)
        col_to_period_rid = cs.get("col_to_period_rid", {})
        row_to_rid = cs.get("row_to_rid", {})

        if not col_to_period_rid or not row_to_rid:
            continue

        # Find the Excel worksheet
        ws = None
        for sn in wb.sheetnames:
            if sn == excel_name:
                ws = wb[sn]
                break
        if ws is None:
            continue

        # Invert mappings: period_rid → col, indicator_rid → row
        period_rid_to_col = {str(rid): int(col) for col, rid in col_to_period_rid.items()}
        indicator_rid_to_row = {str(rid): int(row) for row, rid in row_to_rid.items()}

        # Get indicator names for reporting
        indicator_rid_to_name: dict[str, str] = {}
        for row, rid in row_to_rid.items():
            try:
                recs = await db.execute_fetchall(
                    "SELECT data_json FROM analytic_records WHERE id = ?", (str(rid),))
                if recs:
                    dj = json.loads(recs[0]["data_json"]) if isinstance(recs[0]["data_json"], str) else recs[0]["data_json"]
                    indicator_rid_to_name[str(rid)] = dj.get("name", "")
            except Exception:
                pass

        # Get all cells for this sheet
        cells = await db.execute_fetchall(
            "SELECT coord_key, value FROM cell_data WHERE sheet_id = ?", (sheet_id,)
        )

        checked = 0
        for cell in cells:
            parts = cell["coord_key"].split("|")
            if len(parts) < 2:
                continue

            period_rid = parts[0]
            indicator_rid = parts[-1]

            col = period_rid_to_col.get(period_rid)
            row = indicator_rid_to_row.get(indicator_rid)
            if not col or not row:
                continue

            excel_val = ws.cell(row, col).value
            if excel_val is None:
                continue
            try:
                excel_num = float(excel_val)
            except (ValueError, TypeError):
                continue
            try:
                pebble_num = float(cell["value"])
            except (ValueError, TypeError):
                continue

            checked += 1
            if excel_num == 0 and pebble_num == 0:
                continue
            rel_err = abs(pebble_num - excel_num) / max(abs(excel_num), 1e-9)

            if rel_err > tolerance:
                mismatches.append({
                    "sheet": sheet_name,
                    "indicator": indicator_rid_to_name.get(indicator_rid, "?"),
                    "period": period_rid,
                    "excel": round(excel_num, 4),
                    "pebble": round(pebble_num, 4),
                    "rel_error": round(rel_err, 4),
                })

        log.info("Verified %d cells for sheet '%s', %d mismatches", checked, sheet_name, len(mismatches))

    mismatches.sort(key=lambda m: m.get("rel_error", 0), reverse=True)
    return mismatches


async def _analyze_workbook_with_claude(sheet_texts: dict[str, str], period_dates: list) -> dict:
    """Analyze entire workbook: detect periods from dates, analyze each sheet with Claude."""
    client = _get_claude_client()

    # Period config from detected dates (deterministic, no AI needed)
    if period_dates:
        min_d = min(period_dates)
        max_d = max(period_dates)
        start = f"{min_d.year}-{min_d.month:02d}-01"
        end_m = max_d.month
        end_y = max_d.year
        end_day = monthrange(end_y, end_m)[1]
        end = f"{end_y}-{end_m:02d}-{end_day:02d}"
    else:
        start, end = "2026-01-01", "2028-12-31"

    period_config = {"period_types": ["year", "quarter", "month"], "start": start, "end": end}

    # Analyze each sheet with LLM (chunked for large sheets)
    sheets = []
    for sheet_name, text in sheet_texts.items():
        try:
            sheet_cfg = await _analyze_sheet_chunked(client, text)
            sheet_cfg["excel_name"] = sheet_name  # Ensure correct name
            sheets.append(sheet_cfg)
            log.info("Sheet '%s' analyzed: %d indicators", sheet_name,
                     len(sheet_cfg.get("indicators", [])))
        except Exception as e:
            log.warning("Claude analysis failed for sheet '%s': %s", sheet_name, e)

    return {"period_config": period_config, "sheets": sheets}


# ── Detect periods from date headers (fallback) ───────────────────────────

def _detect_periods_from_headers(ws, max_col: int, base_year: int = 2025) -> list[dict]:
    """Detect period columns from date headers.

    Scans rows 1-20, skipping empty rows. Supports:
    - datetime objects → monthly periods
    - "Январь 2026" text → monthly
    - "N мес" / "N год" → monthly / yearly
    - Q1/Q2/Q3/Q4 → quarterly (with year from adjacent row)
    - H1/H2 → half-yearly (with year from adjacent row)
    - Y0/Y1/... → yearly

    Returns list of dicts, each with:
      col: column number
      name: display name
      period_key: unique key like "2026-01", "2025-Q1", "2026-H1", "2024-Y"
      date: datetime (for monthly) or None
    """
    import re as _re_period

    _MONTH_LOWER = {name.lower(): i + 1 for i, name in enumerate(MONTH_NAMES_RU)}
    _NMES_RE = _re_period.compile(r'^(\d{1,2})\s*мес$', _re_period.IGNORECASE)
    _NGOD_RE = _re_period.compile(r'^(\d{1,2})\s*год$', _re_period.IGNORECASE)
    _QHY_RE = _re_period.compile(r'^([QHY])(\d+)$', _re_period.IGNORECASE)

    # ── Phase 1: Find all candidate header rows ──
    # Scan rows 1-20 for period identifiers (don't stop early — Q/H/Y may be below dates)
    date_row = None
    text_date_row = None
    nmes_row = None
    qhy_row = None  # row with Q1/H1/Y0 etc
    bare_year_row = None  # row with bare year numbers (2024, 2025, ...)

    for r in range(1, 21):
        row_year_count = 0
        for c in range(1, min(max_col + 1, 50)):
            v = ws.cell(r, c).value
            if v is None:
                continue

            if isinstance(v, datetime) and not date_row:
                date_row = r

            if isinstance(v, (int, float)) and 2020 <= v <= 2040 and v == int(v):
                row_year_count += 1

            if isinstance(v, str):
                stripped = v.strip()
                if not text_date_row:
                    parts = stripped.split()
                    if len(parts) == 2 and parts[0].lower() in _MONTH_LOWER:
                        try:
                            int(parts[1])
                            text_date_row = r
                        except ValueError:
                            pass
                if not nmes_row and _NMES_RE.match(stripped):
                    nmes_row = r
                if not qhy_row and _QHY_RE.match(stripped):
                    qhy_row = r

        # Bare year row: 3+ year numbers, no more specific header type found yet
        if row_year_count >= 3 and not bare_year_row:
            bare_year_row = r

    # ── Phase 2: For Q/H/Y, find the year row ──
    year_row = None
    year_row_values = {}  # {col: year_int}
    if qhy_row:
        # Look for a row with year numbers (2024, 2025, ...) near the qhy_row
        for r in range(max(1, qhy_row - 3), min(qhy_row + 3, 21)):
            if r == qhy_row:
                continue
            year_count = 0
            for c in range(1, min(max_col + 1, 50)):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and 2020 <= v <= 2040 and v == int(v):
                    year_count += 1
            if year_count >= 3:
                year_row = r
                for c in range(1, max_col + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, (int, float)) and 2020 <= v <= 2040 and v == int(v):
                        year_row_values[c] = int(v)
                break

    # ── Phase 3: No header found at all → return empty ──
    if date_row is None and text_date_row is None and nmes_row is None and qhy_row is None and bare_year_row is None:
        return []

    # ── Phase 4: Build period list ──
    # When both date_row and qhy_row exist, pick the one with more columns.
    # Q/H/Y sheets (like "Funnel QH") have a few datetime values at quarter
    # end-dates, but the actual structure is Q1/Q2/H1/H2/Y — prefer that.
    periods = []
    seen_keys = set()
    if date_row and qhy_row:
        date_count = sum(1 for c in range(1, max_col + 1) if isinstance(ws.cell(date_row, c).value, datetime))
        qhy_count = sum(1 for c in range(1, max_col + 1) if isinstance(ws.cell(qhy_row, c).value, str) and _QHY_RE.match(ws.cell(qhy_row, c).value.strip()))
        scan_row = qhy_row if qhy_count > date_count else date_row
    else:
        scan_row = date_row or text_date_row or qhy_row or nmes_row or bare_year_row
    month_counter = 0  # for "N мес" format

    for c in range(1, max_col + 1):
        v = ws.cell(scan_row, c).value
        period_key = None
        name = None
        dt = None

        if isinstance(v, datetime):
            year, month = v.year, v.month
            period_key = f"{year}-{month:02d}"
            name = f"{MONTH_NAMES_RU[month - 1]} {year}"
            dt = datetime(year, month, 1)

        elif isinstance(v, (int, float)) and 2020 <= v <= 2040 and v == int(v) and scan_row == bare_year_row:
            # Bare year number (2024, 2025, ...) on a yearly sheet
            year = int(v)
            period_key = f"{year}-Y"
            name = str(year)

        elif isinstance(v, str):
            stripped = v.strip()
            parts = stripped.split()

            # "Январь 2026"
            if len(parts) == 2 and parts[0].lower() in _MONTH_LOWER:
                try:
                    year = int(parts[1])
                    month = _MONTH_LOWER[parts[0].lower()]
                    period_key = f"{year}-{month:02d}"
                    name = f"{MONTH_NAMES_RU[month - 1]} {year}"
                    dt = datetime(year, month, 1)
                except ValueError:
                    pass

            # "N мес"
            elif _NMES_RE.match(stripped):
                month_counter += 1
                year = base_year + (month_counter - 1) // 12
                month = ((month_counter - 1) % 12) + 1
                period_key = f"{year}-{month:02d}"
                name = f"{MONTH_NAMES_RU[month - 1]} {year}"
                dt = datetime(year, month, 1)

            # "N год" — yearly total
            elif _NGOD_RE.match(stripped):
                m_god = _NGOD_RE.match(stripped)
                n = int(m_god.group(1))
                year = base_year + n - 1
                period_key = f"{year}-Y"
                name = str(year)

            # Q1/Q2/Q3/Q4, H1/H2, Y0/Y1/...
            elif _QHY_RE.match(stripped):
                m = _QHY_RE.match(stripped)
                letter = m.group(1).upper()
                num = int(m.group(2))
                # Determine year from year_row
                col_year = year_row_values.get(c)
                if col_year is None and year_row_values:
                    # Inherit from nearest column to the left that has a year
                    for cc in range(c - 1, 0, -1):
                        if cc in year_row_values:
                            col_year = year_row_values[cc]
                            break
                if col_year is None:
                    col_year = base_year + num  # fallback

                if letter == 'Q':
                    period_key = f"{col_year}-Q{num}"
                    quarter_names = ["1-й квартал", "2-й квартал", "3-й квартал", "4-й квартал"]
                    name = f"{quarter_names[num - 1]} {col_year}" if 1 <= num <= 4 else f"Q{num} {col_year}"
                elif letter == 'H':
                    period_key = f"{col_year}-H{num}"
                    name = f"{'1-е' if num == 1 else '2-е'} полугодие {col_year}"
                elif letter == 'Y':
                    # Y0=base_year-1, Y1=base_year, Y2=base_year+1...
                    period_key = f"{col_year}-Y"
                    name = str(col_year)

        if period_key and period_key not in seen_keys:
            seen_keys.add(period_key)
            periods.append({
                "col": c,
                "name": name,
                "period_key": period_key,
                "date": dt,  # None for non-monthly
            })

    return periods


# ── Detect version labels (факт / план / прогноз) in header rows ─────────

_VERSION_LABELS = {
    "факт": "факт", "fact": "факт", "actual": "факт",
    "план": "план", "plan": "план", "budget": "план",
    "прогноз": "план", "forecast": "план", "прог.": "план",
}


def _detect_version_labels(ws, max_col: int) -> dict[int, str]:
    """Scan header rows 1-20 for version labels (факт/план/прогноз).

    Returns {col_number: normalised_version} where normalised_version
    is one of "факт" or "план".  Columns that carry m1/m2/Y1/Y2 but
    no explicit label are NOT included — callers infer "план" for them
    when at least one explicit label exists on the sheet.

    When multiple rows contain version labels, the row with the most
    labels wins (covers sheets like CAPEX FM where "прог." appears on
    one row and the full факт/план set on another).
    """
    best: dict[int, str] = {}
    for r in range(1, 21):
        labels: dict[int, str] = {}
        for c in range(1, min(max_col + 1, 200)):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            key = v.strip().lower()
            if key in _VERSION_LABELS:
                labels[c] = _VERSION_LABELS[key]
        if len(labels) > len(best):
            best = labels
    return best


# ── Pre-substitute total-column references with values ───────────────────

_CELL_REF_SIMPLE = re.compile(r"(?<!')\b(\$?[A-Z]{1,3})(\$?\d+)\b")

_CELL_REF_DOLLAR = re.compile(r"(?<!['\w])(\$?[A-Z]{1,3})(\$?\d+)(?=\b)")

# External workbook refs: [11]CashCredit_dossym!D46 or [11]'Sheet Name'!D46
# External workbook refs: [11]Sheet!D46, [11]'Sheet Name'!D46, '[12]Sheet Name'!D46
_EXTERNAL_REF_RE = re.compile(
    r"(?:"
    r"'?\[(\d+)\](?:'?([^'!]+)'?)"          # [N]Sheet or [N]'Sheet' or '[N]Sheet'
    r"!(\$?[A-Z]{1,3})(\$?\d+)"
    r")",
    re.UNICODE,
)


class _ExternalRefSkip(Exception):
    """Raised when a formula has external workbook refs that can't be resolved."""
    pass


def _has_external_refs(formula: str) -> bool:
    """Check if formula contains [number]Sheet!Cell external workbook references."""
    # Quick check before running regex
    if "[" not in formula:
        return False
    return bool(_EXTERNAL_REF_RE.search(formula))


def _classify_total_leaf_cols(sheet_periods: list[dict]) -> tuple[set[int], set[int]]:
    """Classify period columns as 'total' or 'leaf' per year.

    A column is 'total' for year X if there exist finer-granularity columns
    for the same year. E.g., if 2025 has Q1-Q4, then H1 2025 and Y 2025 are totals.
    But if 2026 only has H1-H2, those H columns are leaves for 2026.
    """
    year_finest: dict[str, str] = {}
    rank = {"M": 0, "Q": 1, "H": 2, "Y": 3}
    for sp in sheet_periods:
        pk = sp.get("period_key", "")
        if not pk or "-" not in pk:
            continue
        year = pk[:4]
        if re.match(r'\d{4}-\d{2}$', pk):
            ptype = "M"
        elif "-Q" in pk:
            ptype = "Q"
        elif "-H" in pk:
            ptype = "H"
        elif pk.endswith("-Y"):
            ptype = "Y"
        else:
            continue
        if year not in year_finest or rank[ptype] < rank[year_finest[year]]:
            year_finest[year] = ptype

    total_cols: set[int] = set()
    leaf_cols: set[int] = set()
    for sp in sheet_periods:
        pk = sp.get("period_key", "")
        if not pk:
            leaf_cols.add(sp["col"])
            continue
        year = pk[:4]
        finest = year_finest.get(year, "Y")
        if re.match(r'\d{4}-\d{2}$', pk):
            ptype = "M"
        elif "-Q" in pk:
            ptype = "Q"
        elif "-H" in pk:
            ptype = "H"
        elif pk.endswith("-Y"):
            ptype = "Y"
        else:
            leaf_cols.add(sp["col"])
            continue
        if rank[ptype] > rank[finest]:
            total_cols.add(sp["col"])
        else:
            leaf_cols.add(sp["col"])
    return total_cols, leaf_cols


def _substitute_non_indicator_refs(
    formula: str, ws_data, row_to_name: dict[int, str], data_start_col: int,
    base_col: int = 0,
) -> str:
    """Replace same-sheet cell refs with Excel values when they can't be translated.

    Only substitutes refs to rows NOT in row_to_name (header rows, date rows,
    rate constants). Indicator refs (including anchored $D8, $H8 to different
    periods) are left for the formula translator to handle as period references.
    """
    from openpyxl.utils import column_index_from_string
    from datetime import datetime, date

    def replace_match(m):
        col_str = m.group(1)
        row_str = m.group(2).replace("$", "")
        row_num = int(row_str)
        col_clean = col_str.replace("$", "")
        try:
            col_num = column_index_from_string(col_clean)
        except (ValueError, KeyError):
            return m.group(0)

        if row_num in row_to_name:
            # Indicator row — let the translator handle it (including anchored refs)
            return m.group(0)

        v = ws_data.cell(row_num, col_num).value
        if v is None:
            return "0"
        # Convert datetime to ordinal for numeric operations (Excel stores dates as numbers)
        if isinstance(v, (datetime, date)):
            if isinstance(v, datetime):
                v = v.toordinal() + 2  # Excel epoch offset (1900-01-00 = day 1)
            else:
                v = v.toordinal() + 2
        try:
            fv = float(v)
            return f"{fv:.10f}".rstrip("0").rstrip(".")
        except (ValueError, TypeError):
            return m.group(0)

    # Split on cross-sheet references (both quoted and unquoted) to avoid
    # touching them — they belong to other sheets, not the current one.
    parts = re.split(r"('[^']*'!\$?[A-Z]+\$?\d+|(?<!['\w])[A-Za-z\w.+\-]+!\$?[A-Z]+\$?\d+)", formula)
    result_parts = []
    for i, part in enumerate(parts):
        if i % 2 == 1:
            result_parts.append(part)
        else:
            result_parts.append(_CELL_REF_DOLLAR.sub(replace_match, part))
    return "".join(result_parts)


def _substitute_total_col_refs(
    formula: str, ws_data, current_row: int, total_cols: set[int],
    data_start_col: int, row_to_name: dict[int, str] | None = None,
) -> str:
    """Replace same-sheet cell refs pointing to total columns with their Excel values.

    E.g. =P8 where col P (16) is a year-total → substituted with the actual value.
    Only substitutes SAME-SHEET bare refs (not 'Sheet'!XX cross-sheet refs).
    Skips indicator rows — those are handled by the formula translator as period refs.
    """
    from openpyxl.utils import column_index_from_string

    def replace_match(m):
        col_str = m.group(1).replace("$", "")
        row_str = m.group(2).replace("$", "")
        try:
            col_num = column_index_from_string(col_str)
        except (ValueError, KeyError):
            return m.group(0)
        if col_num not in total_cols:
            return m.group(0)
        row_num = int(row_str)
        # Skip indicator rows — let the translator handle them as period refs
        if row_to_name and row_num in row_to_name:
            return m.group(0)
        v = ws_data.cell(row_num, col_num).value
        if v is None:
            return "0"
        try:
            fv = float(v)
            return f"{fv:.10f}".rstrip("0").rstrip(".")
        except (ValueError, TypeError):
            return m.group(0)

    # Only substitute bare refs (no sheet prefix)
    # Split on cross-sheet refs (quoted and unquoted) to avoid touching them
    parts = re.split(r"('[^']*'!\$?[A-Z]+\$?\d+|(?<!['\w])[A-Za-z\w.+\-]+!\$?[A-Z]+\$?\d+)", formula)
    result_parts = []
    for i, part in enumerate(parts):
        if i % 2 == 1:
            result_parts.append(part)  # cross-sheet ref, keep as-is
        else:
            result_parts.append(_CELL_REF_SIMPLE.sub(replace_match, part))
    return "".join(result_parts)


# ── Cross-period-level reference substitution ────────────────────────────

_CROSS_SHEET_REF_RE = re.compile(
    r"'([^']+)'!(\$?[A-Z]{1,3})(\$?\d+)"   # quoted: 'Sheet Name'!E19
    r"|"
    r"(?<!['\w])([A-Za-z\w.+]+)!(\$?[A-Z]{1,3})(\$?\d+)",  # unquoted: Sheet!E19
    re.UNICODE,
)


def _get_sheet_period_type(sheet_periods: list[dict]) -> str:
    """Categorize sheet as 'monthly', 'qhy', or 'yearly' based on its period keys."""
    has_month = has_qhy = has_year = False
    for sp in sheet_periods:
        pk = sp.get("period_key", "")
        if re.match(r'\d{4}-\d{2}$', pk):
            has_month = True
        elif "-Q" in pk or "-H" in pk:
            has_qhy = True
        elif pk.endswith("-Y"):
            has_year = True
    if has_month:
        return "monthly"
    if has_qhy:
        return "qhy"
    if has_year:
        return "yearly"
    return "unknown"


def _substitute_cross_period_refs(
    formula: str,
    source_period_type: str,
    sheet_period_types: dict[str, str],
    wb_data,
    all_sheet_total_cols: dict[str, set[int]] | None = None,
    all_sheet_row_maps: dict[str, dict[int, str]] | None = None,
) -> str:
    """Replace cross-sheet refs with Excel values when they reference non-indicator rows.

    Indicator refs are preserved for the formula translator to handle (it uses
    absolute period keys for cross-period-type refs).
    Only substitutes non-indicator row refs with actual Excel values.
    """
    from openpyxl.utils import column_index_from_string
    all_sheet_row_maps = all_sheet_row_maps or {}

    def replace_match(m):
        if m.group(1):
            sheet_name = m.group(1)
            col_str = m.group(2).replace("$", "")
            row_str = m.group(3).replace("$", "")
        elif m.group(4):
            sheet_name = m.group(4)
            col_str = m.group(5).replace("$", "")
            row_str = m.group(6).replace("$", "")
        else:
            return m.group(0)

        row_num = int(row_str)
        target_row_map = all_sheet_row_maps.get(sheet_name, {})

        # If the target row is an indicator, let the translator handle it
        if row_num in target_row_map:
            return m.group(0)

        # Non-indicator row — substitute with actual Excel value
        try:
            if sheet_name not in wb_data.sheetnames:
                return m.group(0)
            ws = wb_data[sheet_name]
            col_num = column_index_from_string(col_str)
            v = ws.cell(row_num, col_num).value
            if v is None:
                return "0"
            fv = float(v)
            return f"{fv:.10f}".rstrip("0").rstrip(".")
        except (ValueError, TypeError, KeyError):
            return m.group(0)

    return _CROSS_SHEET_REF_RE.sub(replace_match, formula)


# ── Total column detection (year / quarter totals) ────────────────────────

import re as _re_total

_YEAR_RE = _re_total.compile(r'^\s*(\d{4})\s*$')
_QUARTER_RE = _re_total.compile(
    r'(?:\d[- ]?й?\s*)?(?:кв(?:артал)?|Q)\s*\d?\s*\d{4}|'
    r'\d{4}\s*(?:кв(?:артал)?|Q)\s*\d|'
    r'(?:итого|всего)\s+за\s+(?:год|квартал)',
    _re_total.IGNORECASE,
)
_TOTAL_KEYWORDS = _re_total.compile(
    r'(?:итого|всего|total|год|year)',
    _re_total.IGNORECASE,
)
# Simple SUM pattern: =SUM(...) where interior is a single range or comma-separated refs
_SUM_ONLY_RE = _re_total.compile(
    r'^=?\s*SUM\s*\([^)]+\)\s*$',
    _re_total.IGNORECASE,
)
# Simple addition: =A1+B1+C1+... (all same-row refs, no division/multiplication)
_SIMPLE_ADD_RE = _re_total.compile(
    r'^=?\s*\$?[A-Z]{1,3}\$?\d+(?:\s*\+\s*\$?[A-Z]{1,3}\$?\d+)+\s*$',
    _re_total.IGNORECASE,
)


def _detect_total_columns(
    ws_data, ws_formulas, col_to_period_rid: dict[int, str], max_col: int,
) -> list[dict]:
    """Detect year/quarter total columns by scanning headers.

    Returns list of {col, type: 'year'|'quarter', label}.
    Skips columns that are already in col_to_period_rid (monthly data columns).
    """
    totals = []
    period_cols = set(col_to_period_rid.keys())

    for c in range(1, min(max_col + 1, 100)):
        if c in period_cols:
            continue
        # Scan first 6 header rows for year/quarter/total markers
        for r in range(1, 7):
            val = ws_data.cell(r, c).value
            if val is None:
                continue
            sval = str(val).strip()
            # Year total: header is just a 4-digit year
            ym = _YEAR_RE.match(sval)
            if ym:
                year = int(ym.group(1))
                if 1990 <= year <= 2100:
                    totals.append({"col": c, "type": "year", "label": sval})
                    break
            # Quarter total
            if _QUARTER_RE.search(sval):
                totals.append({"col": c, "type": "quarter", "label": sval})
                break
            # Generic total keywords
            if _TOTAL_KEYWORDS.search(sval):
                totals.append({"col": c, "type": "year", "label": sval})
                break

    return totals


def _is_sum_formula(excel_formula: str) -> bool:
    """Check if an Excel formula is a plain SUM or simple addition."""
    if not excel_formula or not isinstance(excel_formula, str):
        return False
    f = excel_formula.strip()
    if not f.startswith("="):
        return False
    if _SUM_ONLY_RE.match(f):
        return True
    if _SIMPLE_ADD_RE.match(f):
        return True
    return False


# Pattern for AVERAGE(range)
_AVERAGE_RE = _re_total.compile(r'^=?\s*AVERAGE\s*\([^)]+\)\s*$', _re_total.IGNORECASE)
# Single cell ref: =O13 or =$O$13
_SINGLE_REF_RE = _re_total.compile(r'^=?\s*\$?([A-Z]{1,3})\$?(\d+)\s*$', _re_total.IGNORECASE)
# Same-column formula: refs only to cells in the same column as the total
# E.g. =AO5/AO2 where AO is the total column
_SAME_COL_FORMULA_RE = _re_total.compile(r'\$?([A-Z]{1,3})\$?(\d+)')


def _classify_consolidation_formula(
    excel_formula: str,
    target_col: int,
    row_num: int,
    row_to_name: dict[int, str],
    period_cols: set[int],
) -> str | None:
    """Classify a year/quarter total column formula and return a Pebble
    consolidation formula, or None to skip (SUM = default).

    Returns:
      - None: SUM/default, don't store
      - "AVERAGE": average consolidation
      - "LAST": take last child value (stock/balance indicator)
      - formula string: Pebble formula like "[indicator_a] / [indicator_b]"
    """
    from openpyxl.utils import column_index_from_string

    if not excel_formula or not isinstance(excel_formula, str):
        return None
    f = excel_formula.strip().lstrip("=").strip()
    if not f:
        return None

    full = excel_formula.strip()

    # 1. SUM → skip
    if _is_sum_formula(full):
        return None

    # 2. AVERAGE → store as AVERAGE consolidation (arithmetic mean of children).
    # Not ideal for weighted metrics, but better than LLM guessing wrong names.
    if _AVERAGE_RE.match(full):
        return "AVERAGE"

    # 3. Single cell reference (e.g. =O13) — points to a monthly column in same row
    sm = _SINGLE_REF_RE.match(full)
    if sm:
        ref_col = column_index_from_string(sm.group(1).replace("$", ""))
        ref_row = int(sm.group(2).replace("$", ""))
        if ref_row == row_num and ref_col in period_cols:
            return "LAST"
        # Single ref to different row in same column → just a copy, skip
        if ref_col == target_col:
            return None
        return None

    # 4. Formula with refs only in the same total column → ratio formula
    # E.g. =AO5/AO2 where AO is col 41
    refs = _SAME_COL_FORMULA_RE.findall(f)
    if refs:
        all_same_col = all(
            column_index_from_string(col_str.replace("$", "")) == target_col
            for col_str, _ in refs
        )
        if all_same_col:
            # Translate: replace each cell ref with [indicator_name]
            result = f
            for col_str, row_str in refs:
                ref_row = int(row_str.replace("$", ""))
                name = row_to_name.get(ref_row)
                if not name:
                    return None  # can't resolve → skip
                result = result.replace(f"{col_str}{row_str}", f"[{name}]", 1)
                # Also handle $ variants
                result = result.replace(f"${col_str}${row_str}", f"[{name}]")
                result = result.replace(f"${col_str}{row_str}", f"[{name}]")
                result = result.replace(f"{col_str}${row_str}", f"[{name}]")
            return result

    return None


async def _extract_and_store_consolidation_rules(
    db,
    ws_formulas,
    total_cols: list[dict],
    row_to_rid: dict[int, str],
    row_to_name: dict[int, str],
    pebble_sheet_id: str,
    period_cols: set[int] | None = None,
    sheet_row_maps: dict[str, dict[int, str]] | None = None,
    sheet_display_names: dict[str, str] | None = None,
) -> int:
    """Extract consolidation formulas from total columns and store non-SUM ones
    as indicator_formula_rules(kind='consolidation').

    Returns the number of rules written.
    """
    if not total_cols:
        return 0

    # Prefer year-level total columns over quarter-level
    year_cols = [tc for tc in total_cols if tc["type"] == "year"]
    target_cols = year_cols if year_cols else total_cols

    # Use the first suitable total column
    target_col = target_cols[0]["col"]
    pcols = period_cols or set()

    written = 0
    seen_indicators: set[str] = set()

    for row_num, indicator_rid in row_to_rid.items():
        if indicator_rid in seen_indicators:
            continue

        cell_val = ws_formulas.cell(row_num, target_col).value
        if not cell_val or not isinstance(cell_val, str) or not cell_val.startswith("="):
            continue

        pebble_formula = _classify_consolidation_formula(
            cell_val, target_col, row_num, row_to_name, pcols,
        )
        if pebble_formula is None:
            continue  # SUM or unrecognized → default

        seen_indicators.add(indicator_rid)
        await db.execute(
            "INSERT OR IGNORE INTO indicator_formula_rules "
            "(id, sheet_id, indicator_id, kind, scope_json, priority, formula) "
            "VALUES (?, ?, ?, 'consolidation', '{}', 0, ?)",
            (str(uuid.uuid4()), pebble_sheet_id, indicator_rid, pebble_formula),
        )
        written += 1

    return written


# ── Period hierarchy creation (mirrors generate_periods from analytics.py) ─

async def _create_period_hierarchy(db, analytic_id: str, period_types: list[str],
                                   start_date: date, end_date: date) -> dict:
    """Create year > quarter > month hierarchy. Returns {period_key: record_id} mapping.

    period_key formats: "2026-01" (month), "2026-Q1" (quarter), "2026-H1" (half),
                        "2026-Y" (year)
    """
    has_year = "year" in period_types
    has_quarter = "quarter" in period_types
    has_half = "half" in period_types
    has_month = "month" in period_types

    period_record_ids = {}
    sort = 0
    year = start_date.year

    while year <= end_date.year:
        year_start = max(start_date, date(year, 1, 1))
        year_end = min(end_date, date(year, 12, 31))

        year_id = None
        if has_year:
            year_id = str(uuid.uuid4())
            await db.execute(
                "INSERT INTO analytic_records (id, analytic_id, parent_id, sort_order, data_json) VALUES (?,?,?,?,?)",
                (year_id, analytic_id, None, sort, json.dumps(
                    {"name": str(year), "start": str(year_start), "end": str(year_end),
                     "period_key": f"{year}-Y"},
                    ensure_ascii=False)),
            )
            period_record_ids[f"{year}-Y"] = year_id
            sort += 1

        # Half-year records
        half_ids = {}
        if has_half:
            for h in range(1, 3):
                h_start_month = (h - 1) * 6 + 1
                h_end_month = h * 6
                h_start = date(year, h_start_month, 1)
                h_end = date(year, h_end_month, monthrange(year, h_end_month)[1])
                if h_start > end_date or h_end < start_date:
                    continue
                h_start = max(h_start, start_date)
                h_end = min(h_end, end_date)
                hid = str(uuid.uuid4())
                await db.execute(
                    "INSERT INTO analytic_records (id, analytic_id, parent_id, sort_order, data_json) VALUES (?,?,?,?,?)",
                    (hid, analytic_id, year_id, sort, json.dumps(
                        {"name": f"{'1-е' if h == 1 else '2-е'} полугодие {year}",
                         "start": str(h_start), "end": str(h_end),
                         "period_key": f"{year}-H{h}"},
                        ensure_ascii=False)),
                )
                period_record_ids[f"{year}-H{h}"] = hid
                half_ids[h] = hid
                sort += 1

        for q in range(4):
            q_start_month = q * 3 + 1
            q_end_month = q * 3 + 3
            q_start = date(year, q_start_month, 1)
            q_end = date(year, q_end_month, monthrange(year, q_end_month)[1])
            if q_start > end_date or q_end < start_date:
                continue
            q_start = max(q_start, start_date)
            q_end = min(q_end, end_date)

            # Quarter parent: half > year > None
            half_num = 1 if q < 2 else 2
            quarter_parent = half_ids.get(half_num) or year_id

            quarter_id = None
            if has_quarter:
                quarter_id = str(uuid.uuid4())
                await db.execute(
                    "INSERT INTO analytic_records (id, analytic_id, parent_id, sort_order, data_json) VALUES (?,?,?,?,?)",
                    (quarter_id, analytic_id, quarter_parent, sort, json.dumps(
                        {"name": f"{QUARTER_NAMES_RU[q]} {year}", "start": str(q_start), "end": str(q_end),
                         "period_key": f"{year}-Q{q + 1}"},
                        ensure_ascii=False)),
                )
                period_record_ids[f"{year}-Q{q + 1}"] = quarter_id
                sort += 1

            if has_month:
                for m in range(q_start_month, q_end_month + 1):
                    m_start = date(year, m, 1)
                    m_end = date(year, m, monthrange(year, m)[1])
                    if m_start > end_date or m_end < start_date:
                        continue
                    m_start = max(m_start, start_date)
                    m_end = min(m_end, end_date)
                    mid = str(uuid.uuid4())
                    parent = quarter_id if quarter_id else (half_ids.get(half_num) or year_id)
                    await db.execute(
                        "INSERT INTO analytic_records (id, analytic_id, parent_id, sort_order, data_json) VALUES (?,?,?,?,?)",
                        (mid, analytic_id, parent, sort, json.dumps(
                            {"name": f"{MONTH_NAMES_RU[m - 1]} {year}", "start": str(m_start), "end": str(m_end),
                             "period_key": f"{year}-{m:02d}"},
                            ensure_ascii=False)),
                    )
                    period_record_ids[f"{year}-{m:02d}"] = mid
                    sort += 1

        year += 1

    return period_record_ids


async def _create_sheet_period_records(
    db, analytic_id: str, detected_periods: list[dict],
) -> dict:
    """Create period records for ONLY the period_keys detected in a sheet's Excel headers.

    This avoids creating quarters for years that only have half-year data, etc.
    Builds a minimal hierarchy: if both a parent (Y) and children (H/Q) exist,
    link them; otherwise root-level.
    """
    # Collect unique period_keys and their metadata
    pk_set: set[str] = set()
    pk_meta: dict[str, dict] = {}  # period_key → {name, start, end}
    for sp in detected_periods:
        pk = sp.get("period_key", "")
        if not pk:
            continue
        pk_set.add(pk)
        # Build display name and dates from period_key
        if pk not in pk_meta:
            pk_meta[pk] = _period_key_to_meta(pk)

    if not pk_set:
        return {}

    # Sort period_keys: years first within each year, then halves, then quarters
    def _pk_sort(pk):
        m = re.match(r'(\d{4})-(Y|H\d|Q\d|\d{2})', pk)
        if not m:
            return (9999, 99)
        yr = int(m.group(1))
        suffix = m.group(2)
        # Order: Y=0, H1=1, H2=2, Q1=3, Q2=4, Q3=5, Q4=6, months=10+
        if suffix == 'Y':
            return (yr, 0)
        elif suffix.startswith('H'):
            return (yr, int(suffix[1]))
        elif suffix.startswith('Q'):
            return (yr, 2 + int(suffix[1]))
        else:
            return (yr, 10 + int(suffix))
    sorted_pks = sorted(pk_set, key=_pk_sort)

    # Create records with hierarchy
    period_record_ids: dict[str, str] = {}
    sort = 0

    # First pass: create all records, building parent mapping
    # Year records are parents of H/Q within same year
    # H records are parents of Q within same half
    year_rids: dict[int, str] = {}  # year → record_id
    half_rids: dict[str, str] = {}  # "YYYY-H1" → record_id

    for pk in sorted_pks:
        m = re.match(r'(\d{4})-(Y|H\d|Q\d|\d{2})', pk)
        if not m:
            continue
        yr = int(m.group(1))
        suffix = m.group(2)
        meta = pk_meta[pk]

        # Determine parent
        parent_id = None
        if suffix.startswith('H'):
            parent_id = year_rids.get(yr)
        elif suffix.startswith('Q'):
            q_num = int(suffix[1])
            h_num = 1 if q_num <= 2 else 2
            parent_id = half_rids.get(f"{yr}-H{h_num}") or year_rids.get(yr)
        elif suffix != 'Y' and len(suffix) == 2:  # month
            # Parent: quarter or half or year
            month_num = int(suffix)
            q_num = (month_num - 1) // 3 + 1
            h_num = 1 if q_num <= 2 else 2
            q_key = f"{yr}-Q{q_num}"
            h_key = f"{yr}-H{h_num}"
            parent_id = period_record_ids.get(q_key) or half_rids.get(h_key) or year_rids.get(yr)

        rid = str(uuid.uuid4())
        await db.execute(
            "INSERT INTO analytic_records (id, analytic_id, parent_id, sort_order, data_json) VALUES (?,?,?,?,?)",
            (rid, analytic_id, parent_id, sort, json.dumps(
                {"name": meta["name"], "start": meta["start"], "end": meta["end"],
                 "period_key": pk},
                ensure_ascii=False)),
        )
        period_record_ids[pk] = rid
        sort += 1

        if suffix == 'Y':
            year_rids[yr] = rid
        elif suffix.startswith('H'):
            half_rids[pk] = rid

    return period_record_ids


def _period_key_to_meta(pk: str) -> dict:
    """Convert a period_key like '2025-Q1' to display name + date range."""
    m = re.match(r'(\d{4})-(Y|H(\d)|Q(\d)|(\d{2}))', pk)
    if not m:
        return {"name": pk, "start": "2025-01-01", "end": "2025-12-31"}
    yr = int(m.group(1))
    suffix = m.group(2)
    if suffix == 'Y':
        return {"name": str(yr), "start": f"{yr}-01-01", "end": f"{yr}-12-31"}
    elif m.group(3):  # H1 or H2
        h = int(m.group(3))
        s_month = (h - 1) * 6 + 1
        e_month = h * 6
        return {
            "name": f"{'1-е' if h == 1 else '2-е'} полугодие {yr}",
            "start": f"{yr}-{s_month:02d}-01",
            "end": f"{yr}-{e_month:02d}-{monthrange(yr, e_month)[1]}",
        }
    elif m.group(4):  # Q1-Q4
        q = int(m.group(4))
        s_month = (q - 1) * 3 + 1
        e_month = q * 3
        return {
            "name": f"{QUARTER_NAMES_RU[q - 1]} {yr}",
            "start": f"{yr}-{s_month:02d}-01",
            "end": f"{yr}-{e_month:02d}-{monthrange(yr, e_month)[1]}",
        }
    elif m.group(5):  # month
        mo = int(m.group(5))
        return {
            "name": f"{MONTH_NAMES_RU[mo - 1]} {yr}",
            "start": f"{yr}-{mo:02d}-01",
            "end": f"{yr}-{mo:02d}-{monthrange(yr, mo)[1]}",
        }
    return {"name": pk, "start": f"{yr}-01-01", "end": f"{yr}-12-31"}


# ── Post-process: fix missing parent-child grouping ─────────────────────────

import re as _re

_GROUP_PATTERN = _re.compile(
    r'(?:в\s+т\.?\s*ч\.?\s*:?|в\s+том\s+числе\s*:?|включая\s*:?)\s*$',
    _re.IGNORECASE,
)
_GROUP_PREFIX = _re.compile(
    r'^(?:итого|всего|общее\s|общий\s|общая\s|суммарн|сумма\s)',
    _re.IGNORECASE,
)


def _enrich_with_indent(indicators: list[dict], ws) -> None:
    """Read Excel indent values and store them on each indicator dict.

    Detects which column contains indicator names (1-4) by checking where
    the majority of known indicator names actually live, then reads indent
    from that column.
    """
    # Collect all rows to detect label column
    all_rows: list[int] = []
    def _collect_rows(items):
        for item in items:
            if item.get("row"):
                all_rows.append(item["row"])
            if item.get("children"):
                _collect_rows(item["children"])
    _collect_rows(indicators)

    # Detect label column: check cols 1-4, pick the one where most names match
    label_col = 1
    if all_rows:
        best_col, best_count = 1, 0
        sample_rows = all_rows[:30]
        for c in range(1, 5):
            match_count = 0
            for r in sample_rows:
                v = ws.cell(r, c).value
                if v and str(v).strip():
                    match_count += 1
            if match_count > best_count:
                best_count = match_count
                best_col = c
        label_col = best_col

    def _walk(items):
        for item in items:
            row = item.get("row")
            if row:
                cell = ws.cell(row, label_col)
                indent = cell.alignment.indent if cell.alignment and cell.alignment.indent else 0
                outline = ws.row_dimensions[row].outline_level if hasattr(ws.row_dimensions[row], 'outline_level') else 0
                # Use outline_level as indent when cell indent is 0 but outline differs
                # outline=0 means top-level (group header), outline=1+ means nested
                if indent == 0 and outline > 0:
                    indent = outline
                item["_indent"] = int(indent)
                item["_outline"] = int(outline)
            if item.get("children"):
                _walk(item["children"])
    _walk(indicators)


def _validate_hierarchy_by_indent(indicators: list[dict]) -> list[dict]:
    """Validate Claude's hierarchy against Excel indent levels.

    If a group has ANY children at the same indent as itself, the LLM's
    grouping is wrong. Flatten ALL children of that group back to siblings
    so that _fix_indent_grouping can rebuild the hierarchy correctly.
    """
    has_indent = False
    def _check(items):
        nonlocal has_indent
        for item in items:
            if item.get("_indent") is not None:
                has_indent = True
                return
            if item.get("children"):
                _check(item["children"])
    _check(indicators)
    if not has_indent:
        return indicators

    def _flatten(items: list[dict]) -> list[dict]:
        """Recursively flatten all children into a flat list."""
        result: list[dict] = []
        for item in items:
            children = item.get("children", [])
            item["children"] = []
            item["is_group"] = False
            result.append(item)
            if children:
                result.extend(_flatten(children))
        return result

    def _walk(items: list[dict]) -> list[dict]:
        result: list[dict] = []
        for item in items:
            children = item.get("children", [])
            if not children:
                result.append(item)
                continue

            # Recursively validate children first
            children = _walk(children)

            parent_indent = item.get("_indent", 0)
            # Check if ANY child has same indent as parent → LLM grouping is wrong
            has_same_indent = any(ch.get("_indent", 0) <= parent_indent for ch in children)

            if has_same_indent:
                # Flatten ALL children — let _fix_indent_grouping rebuild correctly
                item["children"] = []
                item["is_group"] = False
                if item.get("rule") == "sum_children":
                    item["rule"] = "manual"
                result.append(item)
                result.extend(_flatten(children))
            else:
                item["children"] = children
                result.append(item)

        return result

    return _walk(indicators)


def _fix_indicator_hierarchy(indicators: list[dict]) -> list[dict]:
    """Post-process indicator list: ensure rows matching grouping name patterns
    (e.g. "в т.ч.:", "Итого") have subsequent rows nested as children.
    This is a safety net after Claude/heuristic analysis — it only promotes
    flat siblings into children when the grouping wasn't already detected."""
    result: list[dict] = []
    i = 0
    while i < len(indicators):
        item = indicators[i]
        name = (item.get("name") or "").strip()
        already_group = item.get("is_group", False) and len(item.get("children", [])) > 0

        # Recursively fix children of already-detected groups
        if item.get("children"):
            item["children"] = _fix_indicator_hierarchy(item["children"])

        # If already a group with children, keep it
        if already_group:
            result.append(item)
            i += 1
            continue

        # Check if this row matches a grouping pattern ("в т.ч.:", "Итого", etc.)
        is_group_by_name = bool(_GROUP_PATTERN.search(name)) or bool(_GROUP_PREFIX.match(name))

        if is_group_by_name and not already_group:
            # Collect subsequent flat siblings as children
            item["is_group"] = True
            if item.get("rule") in (None, "manual", ""):
                item["rule"] = "sum_children"
            existing_children = item.get("children", [])
            j = i + 1
            while j < len(indicators):
                next_item = indicators[j]
                next_name = (next_item.get("name") or "").strip()
                # Stop collecting if we hit another group header or empty
                if (bool(_GROUP_PATTERN.search(next_name)) or
                    bool(_GROUP_PREFIX.match(next_name)) or
                    (next_item.get("is_group", False) and len(next_item.get("children", [])) > 0)):
                    break
                existing_children.append(next_item)
                j += 1
            item["children"] = existing_children
            result.append(item)
            i = j
        else:
            result.append(item)
            i += 1

    # Second pass: indent-based grouping for remaining flat items.
    # If item at indent=N is followed by items at indent>N, make it a group.
    result = _fix_indent_grouping(result)

    # Hard constraint: rows with outline > 0 cannot be groups (they are children in Excel)
    result = _enforce_outline_constraint(result)

    # Re-run indent grouping to absorb orphaned children back into their parent groups
    result = _fix_indent_grouping(result)

    return result


def _regroup_by_outline(indicators: list[dict]) -> list[dict]:
    """Fix hierarchy based on Excel outline levels.

    After heuristic analysis + indent enrichment, some outline=1 rows
    may be at root level (or wrong parent) because the heuristic
    incorrectly created groups from rows whose names match group patterns
    (e.g. 'сумма ...'). This function:
    1. Absorbs root-level outline>0 items into preceding outline=0 group
    2. Flattens false groups (outline>0 with children) before absorbing
    """
    has_outline = any(item.get("_outline") is not None for item in indicators)
    if not has_outline:
        return indicators

    # Recursively fix children of existing groups first
    for item in indicators:
        if item.get("children"):
            item["children"] = _regroup_by_outline(item["children"])

    result: list[dict] = []
    last_outline0_group = None  # pointer to the last outline=0 group in result

    for item in indicators:
        outline = item.get("_outline", 0)
        children = item.get("children", [])

        if outline == 0:
            # This is a top-level item — add to result
            result.append(item)
            if children:
                last_outline0_group = item
            else:
                # Even a non-group outline=0 item resets the "absorb" target
                # (subsequent outline>0 items shouldn't jump over non-group outline=0 items)
                last_outline0_group = None
        elif last_outline0_group is not None:
            # outline > 0 at root level → absorb into last_outline0_group
            if children:
                # False group: flatten it, add self + children to parent
                item_copy = dict(item)
                item_copy["children"] = []
                item_copy["is_group"] = False
                if item_copy.get("rule") == "sum_children":
                    item_copy["rule"] = "manual"
                last_outline0_group["children"].append(item_copy)
                for ch in children:
                    last_outline0_group["children"].append(ch)
            else:
                last_outline0_group["children"].append(item)
        else:
            # outline > 0 but no preceding group — just keep as-is
            result.append(item)

    return result


def _enforce_outline_constraint(indicators: list[dict]) -> list[dict]:
    """Demote any group whose Excel outline_level > 0 back to a leaf.

    outline=0 means top-level header in Excel; outline>0 means nested/child row.
    If the LLM or indent-grouping incorrectly promoted such a row to a group,
    flatten its children back as siblings.
    """
    has_outline = any(item.get("_outline") is not None for item in indicators)
    if not has_outline:
        # Also check children
        for item in indicators:
            if item.get("children"):
                item["children"] = _enforce_outline_constraint(item["children"])
        return indicators

    result: list[dict] = []
    for item in indicators:
        outline = item.get("_outline", 0)
        children = item.get("children", [])
        if outline > 0 and children:
            # This row is nested in Excel — cannot be a group. Flatten children as siblings.
            item["children"] = []
            item["is_group"] = False
            if item.get("rule") == "sum_children":
                item["rule"] = "manual"
            result.append(item)
            result.extend(_enforce_outline_constraint(children))
        else:
            if children:
                item["children"] = _enforce_outline_constraint(children)
            result.append(item)
    return result


def _fix_indent_grouping(indicators: list[dict]) -> list[dict]:
    """Group flat items based on Excel indent levels (_indent field)."""
    # First, recursively fix children
    for item in indicators:
        if item.get("children"):
            item["children"] = _fix_indent_grouping(item["children"])

    # Check if any items have indent info
    has_indent = any(item.get("_indent") is not None for item in indicators)
    if not has_indent:
        return indicators

    result: list[dict] = []
    i = 0
    while i < len(indicators):
        item = indicators[i]
        cur_indent = item.get("_indent", 0)
        already_group = item.get("is_group", False) and len(item.get("children", [])) > 0

        if already_group:
            result.append(item)
            i += 1
            continue

        # Look ahead: if next item has deeper indent, this is a group header
        # But never promote a row with outline > 0 (it's a child row in Excel)
        can_be_group = item.get("_outline", 0) == 0
        if i + 1 < len(indicators) and can_be_group:
            next_indent = indicators[i + 1].get("_indent", 0)
            if next_indent > cur_indent and not already_group:
                # Collect all items with deeper indent as children
                item["is_group"] = True
                if item.get("rule") in (None, "manual", ""):
                    item["rule"] = "sum_children"
                existing_children = item.get("children", [])
                j = i + 1
                while j < len(indicators):
                    nxt = indicators[j]
                    nxt_indent = nxt.get("_indent", 0)
                    if nxt_indent <= cur_indent:
                        break
                    existing_children.append(nxt)
                    j += 1
                item["children"] = existing_children
                # Recursively fix the collected children too
                item["children"] = _fix_indent_grouping(item["children"])
                result.append(item)
                i = j
                continue

        result.append(item)
        i += 1

    return result


def _verify_group_rules(indicators: list[dict], ws, data_start_col: int) -> None:
    """Verify group indicator rules against actual Excel values.

    For each group with rule='sum_children' (or manual parent with children),
    check a few period columns: if parent value == sum of direct children values,
    confirm sum_children. If parent has a non-zero value that does NOT match
    sum of children, keep as manual.

    Also: if a non-group leaf has a value that matches the sum of subsequent
    same-indent items, this hints it should be a group — but we don't restructure
    here, only fix rules on already-detected groups.
    """
    max_col = min(ws.max_column or 1, 200)
    # Pick up to 3 data columns to verify sums
    check_cols: list[int] = []
    for c in range(data_start_col, max_col + 1):
        if len(check_cols) >= 3:
            break
        # Check if this column has numeric data in at least one row
        for r in range(2, min(ws.max_row or 2, 100)):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)) and v != 0:
                check_cols.append(c)
                break
    if not check_cols:
        return

    def _direct_children_rows(item: dict) -> list[int]:
        """Get row numbers of direct children (not nested grandchildren)."""
        rows = []
        for ch in item.get("children", []):
            if ch.get("row"):
                rows.append(ch["row"])
        return rows

    def _walk(items):
        for item in items:
            if item.get("children"):
                _walk(item["children"])

            if not item.get("is_group") or not item.get("children"):
                continue
            parent_row = item.get("row")
            if not parent_row:
                continue

            child_rows = _direct_children_rows(item)
            if not child_rows:
                continue

            # Check if parent value == sum of children across check columns
            sum_matches = 0
            sum_checks = 0
            for c in check_cols:
                pv = ws.cell(parent_row, c).value
                if pv is None:
                    continue
                try:
                    parent_val = float(pv)
                except (ValueError, TypeError):
                    continue

                children_sum = 0.0
                children_found = 0
                for cr in child_rows:
                    cv = ws.cell(cr, c).value
                    if cv is not None:
                        try:
                            children_sum += float(cv)
                            children_found += 1
                        except (ValueError, TypeError):
                            pass

                if children_found == 0:
                    continue
                sum_checks += 1
                if abs(parent_val) < 1e-9 and abs(children_sum) < 1e-9:
                    sum_matches += 1
                elif abs(parent_val) > 1e-9 and abs(children_sum - parent_val) / abs(parent_val) < 0.01:
                    sum_matches += 1

            if sum_checks > 0 and sum_matches == sum_checks:
                # All checked columns match → confirm sum_children
                item["rule"] = "sum_children"
            elif sum_checks > 0 and sum_matches == 0:
                # No column matches → keep as manual (don't override a formula)
                if item.get("rule") == "sum_children":
                    item["rule"] = "manual"

    _walk(indicators)


def _recover_missing_rows(indicators: list[dict], ws, data_start_col: int) -> list[dict]:
    """Scan Excel sheet for data rows that Claude's analysis missed.

    Claude sometimes omits large sections of a sheet (e.g. currency sub-groups).
    This function finds rows with a name + numeric data in period columns
    that are NOT already in the indicator list, and inserts them at the correct
    position. Then re-runs indent grouping to establish hierarchy.
    """
    # 1. Collect all row numbers already present
    known_rows: set[int] = set()
    def _collect(items):
        for item in items:
            if item.get("row"):
                known_rows.add(item["row"])
            if item.get("children"):
                _collect(item["children"])
    _collect(indicators)

    if not known_rows:
        return indicators

    min_row = min(known_rows)
    max_row_known = max(known_rows)
    # Extend scan beyond known max — Claude may have truncated early
    max_row = min(ws.max_row or 1, max_row_known + 500, 2000)

    max_col = min(ws.max_column or 1, 200)

    # 2. Determine label column (usually 1, sometimes 2)
    label_col = 1
    # Heuristic: scan data region for columns with the most text content
    max_row_scan = min(ws.max_row or 1, max_row, 500)
    col_text_counts = {}
    for _lc in (1, 2):
        _tc = 0
        for _lr in range(min_row, max_row_scan + 1):
            _lv = ws.cell(_lr, _lc).value
            if _lv is not None and isinstance(_lv, str) and len(str(_lv).strip()) > 1:
                _tc += 1
        col_text_counts[_lc] = _tc
    if col_text_counts.get(2, 0) > col_text_counts.get(1, 0) * 1.5 and col_text_counts.get(2, 0) > 10:
        label_col = 2

    # 3. Scan for missing rows
    recovered: list[dict] = []
    for r in range(min_row, max_row + 1):
        if r in known_rows:
            continue

        name_val = ws.cell(r, label_col).value
        if name_val is None:
            continue
        name = str(name_val).strip()
        if not name or len(name) > 200:
            continue

        # Check for numeric data in period columns
        has_data = False
        for c in range(data_start_col, min(data_start_col + 10, max_col + 1)):
            cv = ws.cell(r, c).value
            if cv is not None:
                if isinstance(cv, (int, float)):
                    has_data = True
                    break
                if isinstance(cv, str) and cv.startswith("="):
                    has_data = True
                    break

        # Also accept rows that look like group headers (bold, no data)
        is_bold = ws.cell(r, label_col).font and ws.cell(r, label_col).font.bold
        is_group_by_name = bool(_GROUP_PATTERN.search(name)) or bool(_GROUP_PREFIX.match(name))
        is_group_header = is_group_by_name or (is_bold and not has_data)

        if not has_data and not is_group_header:
            continue

        # Read indent
        cell = ws.cell(r, label_col)
        indent = int(cell.alignment.indent) if cell.alignment and cell.alignment.indent else 0

        # Read unit from next column
        unit_val = ws.cell(r, label_col + 1).value
        unit = str(unit_val).strip() if unit_val else ""

        item = {
            "name": name,
            "unit": unit,
            "row": r,
            "is_group": is_group_header,
            "children": [],
            "rule": "sum_children" if is_group_header else "manual",
            "_indent": indent,
            "_recovered": True,
        }
        recovered.append(item)

    if not recovered:
        return indicators

    # 4. Flatten existing indicators, merge with recovered, sort by row
    flat: list[dict] = []
    def _flatten(items):
        for item in items:
            # Strip existing hierarchy — we'll rebuild from indent
            children = item.pop("children", [])
            flat.append(item)
            if children:
                _flatten(children)
    _flatten(indicators)

    flat.extend(recovered)
    flat.sort(key=lambda x: x.get("row", 0))

    # 5. Re-enrich indent for items that don't have it yet
    for item in flat:
        if "_indent" not in item:
            r = item.get("row")
            if r:
                cell = ws.cell(r, label_col)
                indent = int(cell.alignment.indent) if cell.alignment and cell.alignment.indent else 0
                item["_indent"] = indent

    # 6. Rebuild hierarchy via indent grouping
    result = _fix_indent_grouping(flat)
    # Also run name-based grouping
    result = _fix_indicator_hierarchy(result)
    return result


# ── Indicator records creation (recursive) ─────────────────────────────────

async def _create_indicator_records(db, analytic_id: str, indicators: list[dict]) -> tuple[dict, dict, dict, dict]:
    """Create hierarchical indicator records.
    Returns (row_to_rid, rid_to_formula, row_to_name, row_to_parent_name)
    """
    row_to_rid = {}
    rid_to_formula = {}
    sum_children_rids: set[str] = set()
    sort_idx = 0

    async def insert_items(items: list[dict], parent_id: str | None):
        nonlocal sort_idx
        for item in items:
            rid = str(uuid.uuid4())
            data = {"name": item["name"]}
            if item.get("unit"):
                data["unit"] = item["unit"]
                if item["unit"] == "%":
                    data["format"] = "percent"

            excel_row = item.get("row")
            await db.execute(
                "INSERT INTO analytic_records (id, analytic_id, parent_id, sort_order, data_json, excel_row) VALUES (?,?,?,?,?,?)",
                (rid, analytic_id, parent_id, sort_idx, json.dumps(data, ensure_ascii=False), excel_row),
            )
            row_to_rid[item["row"]] = rid

            # Store formula info and track sum_children groups.
            # Only treat as sum_children if the row actually has children —
            # otherwise materialize_sums (which deletes all sum_children cells
            # then re-aggregates) would erase the row's own value with nothing
            # to replace it. Childless "group headers" with own data fall back
            # to manual/formula based on the Excel cell content.
            rule = item.get("rule", "manual")
            if rule == "sum_children" and item.get("children"):
                sum_children_rids.add(rid)
            elif rule == "formula":
                rid_to_formula[rid] = {
                    "rule": "formula",
                    "formula": item.get("formula", ""),
                    "formula_first": item.get("formula_first", ""),
                }

            sort_idx += 1

            if item.get("children"):
                await insert_items(item["children"], rid)

    # Deduplicate indicator names within each parent group only.
    # Names may repeat across groups (e.g. "комиссия партнеру" in each product)
    # — that's fine, formulas use [parent/name] to disambiguate.
    # Only add #N suffix when names collide within the SAME parent.
    def _dedup_names_per_parent(items: list[dict]):
        """Rename duplicates within each sibling list with #N suffix."""
        from collections import Counter
        names = [item.get("name", "").lower() for item in items]
        counts = Counter(names)
        seen: dict[str, int] = {}
        for item in items:
            name_lower = (item.get("name") or "").lower()
            if counts[name_lower] > 1:
                idx = seen.get(name_lower, 0) + 1
                seen[name_lower] = idx
                if idx >= 2:
                    item["name"] = f"{item['name']} #{idx}"
            # Recurse into children
            if item.get("children"):
                _dedup_names_per_parent(item["children"])
    _dedup_names_per_parent(indicators)

    await insert_items(indicators, None)

    # Build row_to_name and row_to_parent_name mappings for formula translator
    row_to_name: dict[int, str] = {}
    row_to_parent_name: dict[int, str] = {}
    def collect_names(items: list[dict], parent_name: str | None = None):
        for item in items:
            if item.get("row") and item.get("name"):
                row_to_name[item["row"]] = item["name"]
                if parent_name:
                    row_to_parent_name[item["row"]] = parent_name
            if item.get("children"):
                collect_names(item["children"], item.get("name"))
    collect_names(indicators)

    return row_to_rid, rid_to_formula, row_to_name, row_to_parent_name, sum_children_rids


# ── Main import endpoint ───────────────────────────────────────────────────

@router.post("/excel")
async def import_excel(file: UploadFile = File(...), model_name: str = Form("Imported Model")):
    db = get_db()

    # Ensure unique model name
    existing = await db.execute_fetchall("SELECT id FROM models WHERE name = ?", (model_name,))
    if existing:
        model_name = f"{model_name} ({datetime.now().strftime('%Y-%m-%d %H:%M')})"

    content = await file.read()
    wb_formulas = load_workbook(io.BytesIO(content))
    wb_data = load_workbook(io.BytesIO(content), data_only=True)

    # ── Step 1: Extract text, detect dates and period types ──
    sheet_texts = {}
    all_dates = []
    detected_period_types = set()  # "month", "quarter", "half", "year"
    _has_y0 = False  # Track if Y0 baseline period exists (shifts monthly base_year by +1)
    _qhy_re_scan = re.compile(r'^([QHY])(\d+)$', re.IGNORECASE)
    _nmes_re_scan = re.compile(r'^\d{1,2}\s*мес$', re.IGNORECASE)
    for sn in wb_formulas.sheetnames:
        ws = wb_formulas[sn]
        sheet_texts[sn] = _extract_sheet_text(ws, sn)
        # Scan rows 1-20 for dates AND period type identifiers
        ws_d_scan = wb_data[sn] if sn in wb_data.sheetnames else None
        for scan_ws in ([ws, ws_d_scan] if ws_d_scan else [ws]):
            for r in range(1, 21):
                for c in range(1, min((scan_ws.max_column or 1) + 1, 200)):
                    v = scan_ws.cell(r, c).value
                    if isinstance(v, datetime):
                        all_dates.append(datetime(v.year, v.month, 1))
                        detected_period_types.add("month")
                    elif isinstance(v, str):
                        stripped = v.strip()
                        m = _qhy_re_scan.match(stripped)
                        if m:
                            letter = m.group(1).upper()
                            num = int(m.group(2))
                            if letter == 'Q':
                                detected_period_types.add("quarter")
                            elif letter == 'H':
                                detected_period_types.add("half")
                            elif letter == 'Y':
                                detected_period_types.add("year")
                                if num == 0:
                                    _has_y0 = True
                        elif _nmes_re_scan.match(stripped):
                            detected_period_types.add("month")
                    # Detect year numbers in header rows (2024, 2025, etc.)
                    if isinstance(v, (int, float)) and 2020 <= v <= 2040 and r <= 10 and v == int(v):
                        all_dates.append(datetime(int(v), 1, 1))
                        all_dates.append(datetime(int(v), 12, 1))

    # ── Step 2: Compute period config from detected dates ──
    if all_dates:
        min_d = min(all_dates)
        max_d = max(all_dates)
        p_start = f"{min_d.year}-{min_d.month:02d}-01"
        end_m = max_d.month
        end_y = max_d.year
        end_day = monthrange(end_y, end_m)[1]
        p_end = f"{end_y}-{end_m:02d}-{end_day:02d}"
    else:
        p_start, p_end = "2026-01-01", "2028-12-31"
    period_config = {"period_types": ["year", "quarter", "month"], "start": p_start, "end": p_end}
    if detected_period_types:
        pt = set(period_config["period_types"])
        pt.update(detected_period_types)
        if any(x in pt for x in ("quarter", "half", "month")):
            pt.add("year")
        period_config["period_types"] = sorted(pt, key=lambda x: ["year", "half", "quarter", "month"].index(x) if x in ["year", "half", "quarter", "month"] else 99)

    # ── Step 2b: Analyze sheet structure with LLM ──
    try:
        analysis = await _analyze_workbook_with_claude(sheet_texts, all_dates)
        sheets_config = analysis["sheets"]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LLM analysis failed: {e}. Check PEBBLE_IMPORT_LLM_API_KEY.")

    # ── Step 3: Create model ──
    model_id = str(uuid.uuid4())
    await db.execute(
        "INSERT INTO models (id, name, description) VALUES (?, ?, ?)",
        (model_id, model_name, "Импортировано из Excel"),
    )

    # ── Step 4: Create period analytics ──
    # Pre-scan all sheets to detect per-sheet period granularity (monthly/qhy/yearly).
    # Create separate period analytics for each granularity so yearly sheets
    # don't show 114 monthly periods.
    period_start = period_config.get("start", "2026-01-01")
    period_end = period_config.get("end", "2028-12-31")
    start_d = date.fromisoformat(period_start)
    end_d = date.fromisoformat(period_end)
    all_period_types = period_config.get("period_types", ["year", "quarter", "month"])

    # Compute base_year early (needed for period detection)
    _base_year_raw = date.fromisoformat(period_config.get("start", "2026-01-01")).year
    _nmes_base_year = _base_year_raw + (1 if _has_y0 else 0)

    # Pre-scan: detect period type and specific detected periods for each sheet
    _prescan_sheet_ptypes: dict[str, str] = {}  # excel_name → "monthly"|"qhy"|"yearly"
    _prescan_sheet_periods: dict[str, list[dict]] = {}  # excel_name → detected period list
    for sheet_cfg in sheets_config:
        excel_name = sheet_cfg["excel_name"]
        if excel_name not in wb_data.sheetnames:
            continue
        ws_scan = wb_data[excel_name]
        max_col_scan = min(ws_scan.max_column or 1, 200)
        sp_scan = _detect_periods_from_headers(ws_scan, max_col_scan, base_year=_nmes_base_year)
        _prescan_sheet_ptypes[excel_name] = _get_sheet_period_type(sp_scan)
        _prescan_sheet_periods[excel_name] = sp_scan

    # Pre-scan: detect version labels (факт/план) per sheet
    _prescan_version_labels: dict[str, dict[int, str]] = {}
    _has_version_labels = False
    for sheet_cfg in sheets_config:
        excel_name = sheet_cfg["excel_name"]
        if excel_name not in wb_data.sheetnames:
            continue
        ws_scan = wb_data[excel_name]
        vl = _detect_version_labels(ws_scan, min(ws_scan.max_column or 1, 200))
        _prescan_version_labels[excel_name] = vl
        if vl:
            _has_version_labels = True

    # ── Single shared period analytic for all sheets ──
    # Always include ALL period levels (M, Q, H, Y). Per-sheet visibility is
    # controlled by min_period_level on the sheet_analytics binding.
    _needs_monthly = any(pt == "monthly" for pt in _prescan_sheet_ptypes.values())
    pt_list = list(all_period_types)
    if "year" not in pt_list:
        pt_list = ["year"] + pt_list
    # Always include all levels so the shared analytic has full hierarchy
    for lvl in ["year", "half", "quarter"]:
        if lvl not in pt_list:
            pt_list.append(lvl)
    if _needs_monthly and "month" not in pt_list:
        pt_list.append("month")
    # Sort by coarseness
    _lvl_order = {"year": 0, "half": 1, "quarter": 2, "month": 3}
    pt_list = sorted(set(pt_list), key=lambda x: _lvl_order.get(x, 99))

    pa_id = str(uuid.uuid4())
    _period_analytic_sort = 0
    await db.execute(
        """INSERT INTO analytics (id, model_id, name, code, icon, is_periods, data_type,
           period_types, period_start, period_end, sort_order)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (pa_id, model_id, "Периоды", "periods",
         "CalendarMonthOutlined", 1, "sum",
         json.dumps(pt_list), period_start, period_end, _period_analytic_sort),
    )
    _period_analytic_sort += 1
    for sort_i, (fname, fcode, ftype) in enumerate([
        ("Наименование", "name", "string"),
        ("Начало", "start", "date"),
        ("Окончание", "end", "date"),
    ]):
        await db.execute(
            "INSERT INTO analytic_fields (id, analytic_id, name, code, data_type, sort_order) VALUES (?,?,?,?,?,?)",
            (str(uuid.uuid4()), pa_id, fname, fcode, ftype, sort_i),
        )
    period_record_ids = await _create_period_hierarchy(db, pa_id, pt_list, start_d, end_d)
    period_analytic_id = pa_id

    # ── Shared version analytic (Факт / План) ──
    version_analytic_id = None
    version_record_ids: dict[str, str] = {}  # "факт" → rid, "план" → rid
    if _has_version_labels:
        version_analytic_id = str(uuid.uuid4())
        await db.execute(
            """INSERT INTO analytics (id, model_id, name, code, icon, data_type, sort_order)
               VALUES (?,?,?,?,?,?,?)""",
            (version_analytic_id, model_id, "Версия", "version",
             "SwapHorizOutlined", "sum", _period_analytic_sort),
        )
        _period_analytic_sort += 1
        await db.execute(
            "INSERT INTO analytic_fields (id, analytic_id, name, code, data_type, sort_order) VALUES (?,?,?,?,?,?)",
            (str(uuid.uuid4()), version_analytic_id, "Наименование", "name", "string", 0),
        )
        for v_sort, (v_name, v_code) in enumerate([("Факт", "fact"), ("План", "plan")]):
            v_rid = str(uuid.uuid4())
            await db.execute(
                "INSERT INTO analytic_records (id, analytic_id, parent_id, sort_order, data_json) VALUES (?,?,?,?,?)",
                (v_rid, version_analytic_id, None, v_sort,
                 json.dumps({"name": v_name, "code": v_code})),
            )
            version_record_ids[v_code.replace("fact", "факт").replace("plan", "план")] = v_rid
        log.info(f"[import] Created version analytic with records: Факт, План")

    # Detect min_period_level per sheet from its Excel headers
    _sheet_min_period_level: dict[str, str | None] = {}  # excel_name → 'M'|'Q'|'H'|'Y'|None
    for excel_name, ptype in _prescan_sheet_ptypes.items():
        if ptype == "monthly":
            _sheet_min_period_level[excel_name] = "M"
        else:
            detected = _prescan_sheet_periods.get(excel_name, [])
            has_month = has_quarter = has_half = has_year = False
            for sp in detected:
                pk = sp.get("period_key", "")
                if re.match(r'\d{4}-\d{2}$', pk):
                    has_month = True
                elif "-Q" in pk:
                    has_quarter = True
                elif "-H" in pk:
                    has_half = True
                elif pk.endswith("-Y"):
                    has_year = True
            if has_month:
                _sheet_min_period_level[excel_name] = "M"
            elif has_quarter:
                _sheet_min_period_level[excel_name] = "Q"
            elif has_half:
                _sheet_min_period_level[excel_name] = "H"
            elif has_year:
                _sheet_min_period_level[excel_name] = "Y"
            else:
                _sheet_min_period_level[excel_name] = None

    # Build per-sheet visible period record IDs:
    # - detected period_keys from Excel headers (exact match)
    # - plus parent records (to complete the tree for consolidation)
    _all_period_record_ids: dict[str, dict[str, str]] = {}
    _sheet_visible_rids: dict[str, set[str]] = {}  # excel_name → set of visible record IDs

    # Build parent_id map for the shared period analytic
    _period_rid_to_parent: dict[str, str | None] = {}
    _period_rid_to_pk: dict[str, str] = {}

    def _pk_rank(pk: str) -> int:
        if re.match(r'\d{4}-\d{2}$', pk): return 0
        if "-Q" in pk: return 1
        if "-H" in pk: return 2
        if pk.endswith("-Y"): return 3
        return -1

    for excel_name, ptype in _prescan_sheet_ptypes.items():
        if ptype == "monthly":
            # Monthly sheets: all records are visible
            _all_period_record_ids[excel_name] = dict(period_record_ids)
            _sheet_visible_rids[excel_name] = set(period_record_ids.values())
            continue

        detected = _prescan_sheet_periods.get(excel_name, [])
        detected_pks = {sp.get("period_key", "") for sp in detected if sp.get("period_key")}

        # Map detected period_keys to record_ids
        sheet_rids: dict[str, str] = {}
        visible_rids: set[str] = set()
        for pk in detected_pks:
            if pk in period_record_ids:
                sheet_rids[pk] = period_record_ids[pk]
                visible_rids.add(period_record_ids[pk])

        # Add parent records up to root (for tree completeness + consolidation)
        # We need parent_id info from the DB — but records aren't loaded yet,
        # so we infer parents from period_key structure.
        for pk in list(detected_pks):
            # Q → H parent → Y parent
            if re.match(r'^\d{4}-Q\d$', pk):
                year = pk[:4]
                q_num = int(pk[-1])
                h_num = 1 if q_num <= 2 else 2
                h_pk = f"{year}-H{h_num}"
                y_pk = f"{year}-Y"
                for parent_pk in [h_pk, y_pk]:
                    if parent_pk in period_record_ids and parent_pk not in sheet_rids:
                        sheet_rids[parent_pk] = period_record_ids[parent_pk]
                        visible_rids.add(period_record_ids[parent_pk])
            elif re.match(r'^\d{4}-H\d$', pk):
                year = pk[:4]
                y_pk = f"{year}-Y"
                if y_pk in period_record_ids and y_pk not in sheet_rids:
                    sheet_rids[y_pk] = period_record_ids[y_pk]
                    visible_rids.add(period_record_ids[y_pk])
            elif re.match(r'^\d{4}-\d{2}$', pk):
                year = pk[:4]
                month = int(pk[5:7])
                q_num = (month - 1) // 3 + 1
                h_num = 1 if q_num <= 2 else 2
                for parent_pk in [f"{year}-Q{q_num}", f"{year}-H{h_num}", f"{year}-Y"]:
                    if parent_pk in period_record_ids and parent_pk not in sheet_rids:
                        sheet_rids[parent_pk] = period_record_ids[parent_pk]
                        visible_rids.add(period_record_ids[parent_pk])

        _all_period_record_ids[excel_name] = sheet_rids
        _sheet_visible_rids[excel_name] = visible_rids

    # ── Step 5: Process each sheet (two passes: 1. create structure, 2. import cells) ──
    from backend.excel_formula_translator import translate_excel_formula

    created_sheets = []
    analytic_sort = _period_analytic_sort  # after all period analytics
    sheet_sort = 0

    # First pass: create indicator hierarchies for ALL sheets (needed for cross-sheet formula translation)
    all_sheet_row_maps: dict[str, dict[int, str]] = {}  # {excel_sheet_name: {row: indicator_name}}
    all_sheet_display_names: dict[str, str] = {}  # {excel_sheet_name: pebble_display_name}
    all_sheet_data_starts: dict[str, int] = {}  # {excel_sheet_name: data_start_col}
    all_row_to_parent_names: dict[str, dict[int, str]] = {}  # {excel_sheet_name_or___self__: {row: parent_name}}
    all_sheet_period_types: dict[str, str] = {}  # {excel_sheet_name: "monthly"|"qhy"|"yearly"}
    all_sheet_total_cols: dict[str, set[int]] = {}  # {excel_sheet_name: set of total column numbers}
    _all_col_to_pidx: dict[str, dict[int, int]] = {}  # {excel_sheet_name: {col: period_idx}}
    _all_col_to_pk: dict[str, dict[int, str]] = {}  # {excel_sheet_name: {col: period_key}}
    sheet_meta: list[dict] = []  # store per-sheet metadata for second pass
    _used_display_names: set[str] = set()  # track used names to avoid duplicates

    for sheet_cfg in sheets_config:
        excel_name = sheet_cfg["excel_name"]
        display_name = sheet_cfg.get("display_name", excel_name)
        sheet_display = display_name if display_name != excel_name else excel_name
        # Ensure unique display names (duplicate names break cross-sheet refs)
        if sheet_display in _used_display_names:
            sheet_display = excel_name  # fall back to Excel name
            if sheet_display in _used_display_names:
                sheet_display = f"{excel_name} (2)"
        _used_display_names.add(sheet_display)
        indicators = sheet_cfg.get("indicators", [])
        data_start_col = sheet_cfg.get("data_start_col", 4)

        if excel_name not in wb_formulas.sheetnames:
            continue
        if not indicators:
            continue

        ws_f = wb_formulas[excel_name]
        ws_d = wb_data[excel_name]

        # Create indicator analytic
        indicator_analytic_id = str(uuid.uuid4())
        analytic_name = f"Показатели ({excel_name})"
        await db.execute(
            """INSERT INTO analytics (id, model_id, name, code, icon, data_type, sort_order)
               VALUES (?,?,?,?,?,?,?)""",
            (indicator_analytic_id, model_id, analytic_name,
             f"indicators_{excel_name.lower().replace('.', '_').replace('+', '_')}",
             "ListAltOutlined", "sum", analytic_sort),
        )
        analytic_sort += 1

        for sort_i, (fname, fcode, ftype) in enumerate([
            ("Наименование", "name", "string"),
            ("Единица измерения", "unit", "string"),
        ]):
            fid = str(uuid.uuid4())
            await db.execute(
                "INSERT INTO analytic_fields (id, analytic_id, name, code, data_type, sort_order) VALUES (?,?,?,?,?,?)",
                (fid, indicator_analytic_id, fname, fcode, ftype, sort_i),
            )

        _enrich_with_indent(indicators, ws_d)
        indicators = _regroup_by_outline(indicators)
        indicators = _validate_hierarchy_by_indent(indicators)
        indicators = _fix_indicator_hierarchy(indicators)
        _verify_group_rules(indicators, ws_d, data_start_col)
        indicators = _recover_missing_rows(indicators, ws_d, data_start_col)
        row_to_rid, rid_to_formula, row_to_name, row_to_parent_name, sum_children_rids = await _create_indicator_records(db, indicator_analytic_id, indicators)

        # Create Pebble sheet
        pebble_sheet_id = str(uuid.uuid4())
        await db.execute(
            "INSERT INTO sheets (id, model_id, name, sort_order, excel_code) VALUES (?,?,?,?,?)",
            (pebble_sheet_id, model_id, sheet_display, sheet_sort, excel_name),
        )
        sheet_sort += 1

        # Bind the single shared period analytic with visible_record_ids
        _min_lvl = _sheet_min_period_level.get(excel_name)
        _vis_rids = _sheet_visible_rids.get(excel_name)
        _vis_json = json.dumps(sorted(_vis_rids)) if _vis_rids and len(_vis_rids) < len(period_record_ids) else None
        _analytics_to_bind = [period_analytic_id, indicator_analytic_id]
        if version_analytic_id:
            _analytics_to_bind.append(version_analytic_id)
        for bind_idx, aid in enumerate(_analytics_to_bind):
            sa_id = str(uuid.uuid4())
            is_main = 1 if aid == indicator_analytic_id else 0
            min_pl = _min_lvl if aid == period_analytic_id else None
            vis = _vis_json if aid == period_analytic_id else None
            await db.execute(
                "INSERT INTO sheet_analytics (id, sheet_id, analytic_id, sort_order, is_main, min_period_level, visible_record_ids) VALUES (?,?,?,?,?,?,?)",
                (sa_id, pebble_sheet_id, aid, bind_idx, is_main, min_pl, vis),
            )

        users = await db.execute_fetchall("SELECT id FROM users")
        for u in users:
            pid = str(uuid.uuid4())
            try:
                await db.execute(
                    "INSERT INTO sheet_permissions (id, sheet_id, user_id, can_view, can_edit) VALUES (?,?,?,1,1)",
                    (pid, pebble_sheet_id, u["id"]),
                )
            except Exception:
                pass

        # Store cross-sheet maps
        all_sheet_row_maps[excel_name] = row_to_name
        all_sheet_display_names[excel_name] = sheet_display
        all_sheet_data_starts[excel_name] = data_start_col
        all_row_to_parent_names[excel_name] = row_to_parent_name

        # Detect period type and total columns for cross-sheet substitution
        _sp_fp = _detect_periods_from_headers(ws_d, min(ws_d.max_column or 1, 200), base_year=_nmes_base_year)
        all_sheet_period_types[excel_name] = _get_sheet_period_type(_sp_fp)

        _tc, _ = _classify_total_leaf_cols(_sp_fp)
        all_sheet_total_cols[excel_name] = _tc

        # Build col→period_idx for cross-sheet refs (leaf columns only)
        _sorted_sp_fp = sorted(_sp_fp, key=lambda x: x["col"])
        _all_col_to_pidx_sheet: dict[int, int] = {}
        _leaf_idx_fp = 0
        for _spp in _sorted_sp_fp:
            if _spp["col"] in _tc:
                continue
            _all_col_to_pidx_sheet[_spp["col"]] = _leaf_idx_fp
            _leaf_idx_fp += 1
        _all_col_to_pidx[excel_name] = _all_col_to_pidx_sheet

        # Build col→period_key for cross-sheet absolute period refs
        _pk_map_fp: dict[int, str] = {}
        for _spp in _sp_fp:
            pk = _spp.get("period_key", "")
            if pk:
                _pk_map_fp[_spp["col"]] = pk
        _all_col_to_pk[excel_name] = _pk_map_fp

        sheet_meta.append({
            "excel_name": excel_name,
            "sheet_display": sheet_display,
            "pebble_sheet_id": pebble_sheet_id,
            "row_to_rid": row_to_rid,
            "rid_to_formula": rid_to_formula,
            "row_to_name": row_to_name,
            "row_to_parent_name": row_to_parent_name,
            "sum_children_rids": sum_children_rids,
            "data_start_col": data_start_col,
        })

    # Second pass: import cell data (now all sheet row maps are available for cross-sheet refs)
    for meta in sheet_meta:
        excel_name = meta["excel_name"]
        sheet_display = meta["sheet_display"]
        pebble_sheet_id = meta["pebble_sheet_id"]
        row_to_rid = meta["row_to_rid"]
        rid_to_formula = meta["rid_to_formula"]
        row_to_name = meta["row_to_name"]
        row_to_parent_name = meta.get("row_to_parent_name", {})
        sum_children_rids = meta.get("sum_children_rids", set())
        data_start_col = meta["data_start_col"]

        ws_f = wb_formulas[excel_name]
        ws_d = wb_data[excel_name]

        # ── Extract starting values (column before data_start) ──
        pre_data_values: dict[int, float] = {}
        if data_start_col > 1:
            pre_col = data_start_col - 1
            for row_num in row_to_name:
                try:
                    v = ws_d.cell(row_num, pre_col).value
                    if v is not None:
                        fv = float(v)
                        if fv != 0:
                            pre_data_values[row_num] = fv
                except (ValueError, TypeError):
                    pass

        # ── Import cell data ──
        # Build col -> period_record_id mapping from headers
        # Use the period_record_ids for THIS sheet's period analytic
        _sheet_period_rids = _all_period_record_ids.get(excel_name, period_record_ids)

        sheet_periods = _detect_periods_from_headers(ws_d, min(ws_d.max_column or 1, 200), base_year=_nmes_base_year)
        col_to_period_rid = {}
        for sp in sheet_periods:
            pkey = sp.get("period_key")
            if pkey and pkey in _sheet_period_rids:
                col_to_period_rid[sp["col"]] = _sheet_period_rids[pkey]
            elif sp.get("date"):
                # Backward compat: try year-month key from date
                d = sp["date"]
                key = f"{d.year}-{d.month:02d}"
                if key in _sheet_period_rids:
                    col_to_period_rid[sp["col"]] = _sheet_period_rids[key]

        # Determine which period is "first" (for formula_first)
        sorted_period_cols = sorted(col_to_period_rid.keys())
        first_period_rid = col_to_period_rid[sorted_period_cols[0]] if sorted_period_cols else None

        total_cols, leaf_cols = _classify_total_leaf_cols(sheet_periods)

        # Build col→period_idx mapping: sequential indices for LEAF columns only,
        # skipping total columns. This ensures cross-year-boundary
        # refs get correct period_diff (col 17→col 15 = 1 period, not 2).
        _sorted_sp = sorted(sheet_periods, key=lambda x: x["col"])
        _col_to_pidx: dict[int, int] = {}
        _leaf_idx = 0
        for _sp in _sorted_sp:
            if _sp["col"] in total_cols:
                continue  # skip total columns
            _col_to_pidx[_sp["col"]] = _leaf_idx
            _leaf_idx += 1

        # Build col→period_key mapping for formula translator (absolute period refs)
        _col_to_pk: dict[int, str] = {}
        for sp in sheet_periods:
            pk = sp.get("period_key", "")
            if pk:
                _col_to_pk[sp["col"]] = pk

        cell_count = 0
        for row_num, indicator_rid in row_to_rid.items():
            formula_info = rid_to_formula.get(indicator_rid)

            # Detect percentage number_format from the first data cell in this row
            _fmt_detected = False
            for _fc in sorted_period_cols:
                nf = ws_d.cell(row_num, _fc).number_format or ""
                if "%" in nf:
                    await db.execute(
                        "UPDATE analytic_records SET data_json = json_set(data_json, '$.format', 'percent') WHERE id = ?",
                        (indicator_rid,),
                    )
                    _fmt_detected = True
                    break
                if nf and nf != "General":
                    break  # has a non-percent format, skip

            for col_num, period_rid in col_to_period_rid.items():
                cell_val = ws_d.cell(row_num, col_num)
                val = cell_val.value
                if val is None:
                    continue

                # Skip non-numeric strings (month codes like m1/m2, labels, etc.)
                if isinstance(val, str):
                    try:
                        val = float(val.replace(",", ".").replace(" ", ""))
                    except (ValueError, AttributeError):
                        continue

                # Cell color is authoritative: yellow/beige fill (theme=7) means
                # "user input" in the source spreadsheet — treat as manual even
                # if a formula exists. We keep the numeric value Excel computed.
                is_yellow = _is_input_cell(ws_f.cell(row_num, col_num))

                # Determine rule and formula — prefer actual Excel formula translated deterministically
                if is_yellow:
                    rule = "manual"
                    formula_text = ""
                else:
                    # Try actual Excel formula first (deterministic, handles cross-sheet)
                    excel_formula = ws_f.cell(row_num, col_num).value
                    is_first = (period_rid == first_period_rid)
                    if isinstance(excel_formula, str) and excel_formula.startswith("="):
                        rule = "formula"
                        try:
                            # External workbook refs [N]Sheet!Cell → can't resolve, keep manual
                            if _has_external_refs(excel_formula):
                                rule = "manual"
                                formula_text = ""
                                raise _ExternalRefSkip()
                            # Pre-substitute: replace same-sheet refs to total columns with Excel values
                            if total_cols and col_num not in total_cols:
                                excel_formula = _substitute_total_col_refs(
                                    excel_formula, ws_d, row_num, total_cols, data_start_col, row_to_name)
                            # Pre-substitute: replace cross-sheet refs to different period types or total columns
                            src_ptype = all_sheet_period_types.get(excel_name, "unknown")
                            if src_ptype != "unknown":
                                excel_formula = _substitute_cross_period_refs(
                                    excel_formula, src_ptype, all_sheet_period_types, wb_data,
                                    all_sheet_total_cols, all_sheet_row_maps)
                            # Pre-substitute: replace refs to non-indicator rows with values
                            excel_formula = _substitute_non_indicator_refs(
                                excel_formula, ws_d, row_to_name, data_start_col, base_col=col_num)
                            # Build parent name maps: use __self__ for current sheet
                            parent_maps = dict(all_row_to_parent_names)
                            parent_maps["__self__"] = row_to_parent_name
                            formula_text = translate_excel_formula(
                                excel_formula,
                                base_col=col_num,
                                data_start_col=data_start_col,
                                row_to_name=row_to_name,
                                sheet_row_maps=all_sheet_row_maps,
                                sheet_display_names=all_sheet_display_names,
                                is_first_period=is_first,
                                sheet_data_starts=all_sheet_data_starts,
                                row_to_parent_names=parent_maps,
                                pre_data_values=pre_data_values,
                                col_to_period_idx=_col_to_pidx,
                                sheet_col_to_period_idx=_all_col_to_pidx,
                                col_to_period_key=_col_to_pk,
                                sheet_col_to_period_key=_all_col_to_pk,
                            )
                        except _ExternalRefSkip:
                            pass  # rule/formula_text already set to manual/""
                        except Exception:
                            # Fallback to Claude's formula if translator fails
                            formula_text = (formula_info or {}).get("formula", excel_formula)
                    else:
                        # Excel cell has no formula (constant) — import as manual.
                        rule = "manual"
                        formula_text = ""

                # Post-translation: if formula contains unparseable range `:` notation
                # (from partially substituted SUM/AVERAGE ranges), fall back to manual.
                # Allow `::` (cross-sheet separator like `[Sheet::name]`).
                if rule == "formula" and formula_text and re.search(r'(?<!:):(?!:)', formula_text):
                    rule = "manual"
                    formula_text = ""

                # Build coord_key: period|indicator[|version]
                if version_analytic_id:
                    _sheet_vlabels = _prescan_version_labels.get(excel_name, {})
                    _vlabel = _sheet_vlabels.get(col_num)
                    if _vlabel:
                        _ver_rid = version_record_ids[_vlabel]
                    else:
                        # Column without explicit label → "план" (budget/forecast)
                        _ver_rid = version_record_ids.get("план", "")
                    coord_key = f"{period_rid}|{indicator_rid}|{_ver_rid}"
                else:
                    coord_key = f"{period_rid}|{indicator_rid}"
                value_str = str(val)

                # First-period formula cells often reference starting balances from
                # before data_start. If our translated formula would produce 0
                # but Excel has a non-zero value, import as manual to preserve
                # the starting balance.
                if rule == "formula" and is_first and val != 0:
                    # Check if formula contains "0" literal or references that would be 0
                    ft = formula_text.strip()
                    if ft == "0" or ft.startswith("0/") or ft.startswith("0*") or \
                       'предыдущий' in ft or 'назад' in ft:
                        rule = "manual"
                        formula_text = ""

                # Group indicators → sum_children rule (engine will compute).
                # But only when we don't already have a real Excel formula —
                # explicit formulas (e.g. grand totals like =D32+D54+...) must
                # win over the sum_children heuristic.
                if indicator_rid in sum_children_rids and not (rule == "formula" and formula_text):
                    rule = "sum_children"
                    formula_text = ""

                cid = str(uuid.uuid4())
                try:
                    from backend.db import intern_formula
                    fid = await intern_formula(db, formula_text) if formula_text else None
                    await db.execute(
                        "INSERT INTO cell_data (id, sheet_id, coord_key, value, data_type, rule, formula, formula_id) VALUES (?,?,?,?,?,?,?,?)",
                        (cid, pebble_sheet_id, coord_key, value_str, "sum", rule, "" if fid else formula_text, fid),
                    )
                    cell_count += 1
                except Exception:
                    pass  # Skip duplicates

        # Extract consolidation formulas from total columns (year/quarter totals)
        total_cols = _detect_total_columns(ws_d, ws_f, col_to_period_rid, min(ws_d.max_column or 1, 200))
        if total_cols:
            n_consol = await _extract_and_store_consolidation_rules(
                db, ws_f, total_cols, row_to_rid, row_to_name, pebble_sheet_id,
                period_cols=set(col_to_period_rid.keys()),
            )
            if n_consol:
                log.info(f"[import] Extracted {n_consol} consolidation formulas from Excel totals for «{sheet_display}»")

        created_sheets.append({
                "name": sheet_display, "id": pebble_sheet_id, "cells": cell_count,
                "excel_name": excel_name,
                "col_to_period_rid": dict(col_to_period_rid),
                "row_to_rid": dict(row_to_rid),
            })

    await db.commit()

    return {
        "model_id": model_id,
        "model_name": model_name,
        "sheets": len(created_sheets),
        "sheet_list": created_sheets,
        "periods": len(period_record_ids),
        "period_hierarchy": pt_list,
    }


# ── Streaming import endpoint (SSE) ───────────────────────────────────────

@router.post("/excel-stream")
async def import_excel_stream(file: UploadFile = File(...), model_name: str = Form("Imported Model"), lang: str = Form("ru"), use_kb: str = Form("0")):
    import asyncio
    content = await file.read()
    _lang = lang if lang in ("ru", "en", "ky", "vi") else "ru"
    _use_kb = use_kb in ("1", "true", "yes")
    _session_id = _create_qa_session()

    async def generate():
        import time as _time
        _t0 = _time.time()

        # ── Import progress message translations ──
        _IM = {
            "loading":       {"ru": "Загрузка файла Excel...", "en": "Loading Excel file...", "ky": "Excel файлы жүктөлүүдө...", "vi": "Đang tải file Excel..."},
            "found_sheets":  {"ru": "Найдено {n} листов: {names}", "en": "Found {n} sheets: {names}", "ky": "{n} барак табылды: {names}", "vi": "Tìm thấy {n} trang: {names}"},
            "analyzing":     {"ru": "Анализ структуры с помощью Claude AI...", "en": "Analyzing structure with Claude AI...", "ky": "Claude AI менен структураны талдоо...", "vi": "Phân tích cấu trúc bằng Claude AI..."},
            "analyzing_n":   {"ru": "🔍 Анализ {n} листов параллельно...", "en": "🔍 Analyzing {n} sheets in parallel...", "ky": "🔍 {n} барак параллель талдоо...", "vi": "🔍 Phân tích song song {n} trang..."},
            "sheet_err":     {"ru": "   [ERR]«{name}»: ошибка — {err}", "en": "   [ERR]\"{name}\": error — {err}", "ky": "   [ERR]«{name}»: ката — {err}", "vi": "   [ERR]\"{name}\": lỗi — {err}"},
            "sheet_ok":      {"ru": "   [OK]«{name}» ({i}/{n}): {groups} групп, {inds} показателей", "en": "   [OK]\"{name}\" ({i}/{n}): {groups} groups, {inds} indicators", "ky": "   [OK]«{name}» ({i}/{n}): {groups} топ, {inds} көрсөткүч", "vi": "   [OK]\"{name}\" ({i}/{n}): {groups} nhóm, {inds} chỉ tiêu"},
            "sheet_parse_err":{"ru": "   [ERR]«{name}»: не удалось разобрать", "en": "   [ERR]\"{name}\": failed to parse", "ky": "   [ERR]«{name}»: талдоо мүмкүн эмес", "vi": "   [ERR]\"{name}\": không thể phân tích"},
            "claude_unavail":{"ru": "[WARN]Claude API недоступен, используем эвристики: {err}", "en": "[WARN]Claude API unavailable, using heuristics: {err}", "ky": "[WARN]Claude API жеткиликсиз, эвристикалар колдонулууда: {err}", "vi": "[WARN]Claude API không khả dụng, sử dụng heuristics: {err}"},
            "model_created": {"ru": "Создана модель «{name}»", "en": "Model \"{name}\" created", "ky": "«{name}» модели түзүлдү", "vi": "Đã tạo mô hình \"{name}\""},
            "periods":       {"ru": "Создана иерархия периодов: {start} — {end} ({n} периодов)", "en": "Period hierarchy created: {start} — {end} ({n} periods)", "ky": "Мезгилдер иерархиясы түзүлдү: {start} — {end} ({n} мезгил)", "vi": "Đã tạo phân cấp kỳ: {start} — {end} ({n} kỳ)"},
            "total_inds":    {"ru": "📊 Всего {n} показателей в {sheets} листах", "en": "📊 Total {n} indicators in {sheets} sheets", "ky": "📊 Жалпы {sheets} баракта {n} көрсөткүч", "vi": "📊 Tổng {n} chỉ tiêu trong {sheets} trang"},
            "creating":      {"ru": "Создаю структуру «{name}» ({done}/{total})...", "en": "Creating structure \"{name}\" ({done}/{total})...", "ky": "«{name}» структурасы түзүлүүдө ({done}/{total})...", "vi": "Đang tạo cấu trúc \"{name}\" ({done}/{total})..."},
            "importing_data":{"ru": "Структура создана. Импорт данных ({n} листов)...", "en": "Structure created. Importing data ({n} sheets)...", "ky": "Структура түзүлдү. Маалыматтар импорттолууда ({n} барак)...", "vi": "Đã tạo cấu trúc. Nhập dữ liệu ({n} trang)..."},
            "data_ok":       {"ru": "   [OK]«{name}»: {inds} показателей, {cells} ячеек{consol} ({done}/{total})", "en": "   [OK]\"{name}\": {inds} indicators, {cells} cells{consol} ({done}/{total})", "ky": "   [OK]«{name}»: {inds} көрсөткүч, {cells} уячалар{consol} ({done}/{total})", "vi": "   [OK]\"{name}\": {inds} chỉ tiêu, {cells} ô{consol} ({done}/{total})"},
            "sheets_warn":   {"ru": "[WARN]Импортировано {done}/{total} листов. Пропущены: {missing}", "en": "[WARN]Imported {done}/{total} sheets. Skipped: {missing}", "ky": "[WARN]{done}/{total} барак импорттолду. Калтырылды: {missing}", "vi": "[WARN]Đã nhập {done}/{total} trang. Bỏ qua: {missing}"},
            "zero_cells":    {"ru": "[WARN]«{name}»: 0 ячеек импортировано", "en": "[WARN]\"{name}\": 0 cells imported", "ky": "[WARN]«{name}»: 0 уячалар импорттолду", "vi": "[WARN]\"{name}\": 0 ô đã nhập"},
            "consol_rules":  {"ru": "   [OK]Claude подобрал {n} формул консолидации по периодам", "en": "   [OK]Claude matched {n} consolidation formulas", "ky": "   [OK]Claude {n} консолидация формулаларын тандады", "vi": "   [OK]Claude đã khớp {n} công thức hợp nhất"},
            "translating":   {"ru": "Перевод названий (ru/en/ky/vi)...", "en": "Translating names (ru/en/ky/vi)...", "ky": "Аттар которулууда (ru/en/ky/vi)...", "vi": "Đang dịch tên (ru/en/ky/vi)..."},
            "translated_ok": {"ru": "   [OK]Переведено {n} названий на {langs} языка", "en": "   [OK]Translated {n} names into {langs} languages", "ky": "   [OK]{n} ат {langs} тилге которулду", "vi": "   [OK]Đã dịch {n} tên sang {langs} ngôn ngữ"},
            "translate_fail":{"ru": "[WARN]Перевод не удался: {err}", "en": "[WARN]Translation failed: {err}", "ky": "[WARN]Которуу ишке ашкан жок: {err}", "vi": "[WARN]Dịch thất bại: {err}"},
            "verifying":     {"ru": "Верификация с Excel...", "en": "Verifying against Excel...", "ky": "Excel менен текшерүү...", "vi": "Xác minh với Excel..."},
            "verify_warn":   {"ru": "[WARN]Расхождения с Excel: {n} ячеек", "en": "[WARN]Mismatches with Excel: {n} cells", "ky": "[WARN]Excel менен дал келбөөлөр: {n} уячалар", "vi": "[WARN]Sai lệch với Excel: {n} ô"},
            "verify_more":   {"ru": "   ...и ещё {n}", "en": "   ...and {n} more", "ky": "   ...жана дагы {n}", "vi": "   ...và thêm {n}"},
            "verify_ok":     {"ru": "   [OK]Все проверенные значения совпадают с Excel", "en": "   [OK]All verified values match Excel", "ky": "   [OK]Бардык текшерилген маанилер Excel менен дал келет", "vi": "   [OK]Tất cả giá trị đã xác minh khớp với Excel"},
            "done":          {"ru": "[DONE]Импорт завершён! {sheets} листов, {cells} ячеек", "en": "[DONE]Import complete! {sheets} sheets, {cells} cells", "ky": "[DONE]Импорт аяктады! {sheets} барак, {cells} уячалар", "vi": "[DONE]Nhập hoàn tất! {sheets} trang, {cells} ô"},
            # KB-related messages
            "kb_loading":    {"ru": "Загрузка базы знаний импорта...", "en": "Loading import knowledge base...", "ky": "Импорт билим базасы жүктөлүүдө...", "vi": "Đang tải cơ sở kiến thức nhập..."},
            "kb_loaded":     {"ru": "   [OK]KB: {n} паттернов загружено", "en": "   [OK]KB: {n} patterns loaded", "ky": "   [OK]KB: {n} паттерн жүктөлдү", "vi": "   [OK]KB: {n} mẫu đã tải"},
            "kb_analyzing":  {"ru": "Анализ структуры по базе знаний...", "en": "Analyzing structure using knowledge base...", "ky": "Билим базасы менен структура талдоо...", "vi": "Phân tích cấu trúc bằng cơ sở kiến thức..."},
            "kb_sheet_ok":   {"ru": "   [OK]«{name}» (KB): {groups} групп, {inds} показателей", "en": "   [OK]\"{name}\" (KB): {groups} groups, {inds} indicators", "ky": "   [OK]«{name}» (KB): {groups} топ, {inds} көрсөткүч", "vi": "   [OK]\"{name}\" (KB): {groups} nhóm, {inds} chỉ tiêu"},
            "kb_question":   {"ru": "Вопрос по листу «{name}»", "en": "Question about sheet \"{name}\"", "ky": "«{name}» барагы боюнча суроо", "vi": "Câu hỏi về trang \"{name}\""},
            "kb_answer_ok":  {"ru": "   [OK]Ответ получен, паттерн сохранён", "en": "   [OK]Answer received, pattern saved", "ky": "   [OK]Жооп алынды, паттерн сакталды", "vi": "   [OK]Đã nhận câu trả lời, mẫu đã lưu"},
            "kb_answer_skip":{"ru": "   Нет ответа, используем LLM", "en": "   No answer, falling back to LLM", "ky": "   Жооп жок, LLM колдонулууда", "vi": "   Không có câu trả lời, chuyển sang LLM"},
            "kb_session":    {"ru": "Сессия: {sid}", "en": "Session: {sid}", "ky": "Сессия: {sid}", "vi": "Phiên: {sid}"},
        }

        def _m(key: str, **kwargs) -> str:
            """Get localized import message."""
            tpl = _IM.get(key, {}).get(_lang) or _IM.get(key, {}).get("ru", key)
            try:
                return tpl.format(**kwargs) if kwargs else tpl
            except (KeyError, IndexError):
                return tpl

        _ts_labels = {"ru": "с", "en": "s", "ky": "сек", "vi": "s"}
        _ts_label = _ts_labels.get(_lang, "с")

        def event(msg: str, data: dict | None = None):
            elapsed = _time.time() - _t0
            ts = f"[{int(elapsed)}{_ts_label}]"
            payload = json.dumps({"message": f"{ts} {msg}", **(data or {})}, ensure_ascii=False)
            return f"data: {payload}\n\n"

        db = get_db()
        loop = asyncio.get_event_loop()

        # Unique name
        existing = await db.execute_fetchall("SELECT id FROM models WHERE name = ?", (model_name,))
        if existing:
            model_name_final = f"{model_name} ({datetime.now().strftime('%Y-%m-%d %H:%M')})"
        else:
            model_name_final = model_name

        yield event(_m("loading"))

        # Run blocking openpyxl in executor to not block event loop
        wb_formulas = await loop.run_in_executor(None, lambda: load_workbook(io.BytesIO(content)))
        wb_data = await loop.run_in_executor(None, lambda: load_workbook(io.BytesIO(content), data_only=True))
        sheet_names = wb_formulas.sheetnames

        yield event(_m("found_sheets", n=len(sheet_names), names=', '.join(sheet_names)))

        # Extract text, dates and period types
        sheet_texts = {}
        all_dates = []
        detected_period_types = set()
        _has_y0 = False
        _qhy_re_scan2 = re.compile(r'^([QHY])(\d+)$', re.IGNORECASE)
        _nmes_re_scan2 = re.compile(r'^\d{1,2}\s*мес$', re.IGNORECASE)
        for sn in sheet_names:
            ws = wb_formulas[sn]
            sheet_texts[sn] = _extract_sheet_text(ws, sn)
            ws_d_scan = wb_data[sn] if sn in wb_data.sheetnames else None
            for scan_ws in ([ws, ws_d_scan] if ws_d_scan else [ws]):
                for r in range(1, 21):
                    for c in range(1, min((scan_ws.max_column or 1) + 1, 200)):
                        v = scan_ws.cell(r, c).value
                        if isinstance(v, datetime):
                            all_dates.append(datetime(v.year, v.month, 1))
                            detected_period_types.add("month")
                        elif isinstance(v, str):
                            stripped = v.strip()
                            m = _qhy_re_scan2.match(stripped)
                            if m:
                                letter = m.group(1).upper()
                                num = int(m.group(2))
                                if letter == 'Q':
                                    detected_period_types.add("quarter")
                                elif letter == 'H':
                                    detected_period_types.add("half")
                                elif letter == 'Y':
                                    detected_period_types.add("year")
                                    if num == 0:
                                        _has_y0 = True
                            elif _nmes_re_scan2.match(stripped):
                                detected_period_types.add("month")
                        if isinstance(v, (int, float)) and 2020 <= v <= 2040 and r <= 10 and v == int(v):
                            all_dates.append(datetime(int(v), 1, 1))
                            all_dates.append(datetime(int(v), 12, 1))

        yield event(_m("analyzing"))

        # Compute period config from detected dates BEFORE Claude analysis
        # so it's preserved even if Claude API is unavailable
        if all_dates:
            min_d = min(all_dates)
            max_d = max(all_dates)
            p_start = f"{min_d.year}-{min_d.month:02d}-01"
            end_m = max_d.month
            end_y = max_d.year
            end_day = monthrange(end_y, end_m)[1]
            p_end = f"{end_y}-{end_m:02d}-{end_day:02d}"
        else:
            p_start, p_end = "2026-01-01", "2028-12-31"
        log.info("Period detection: %d dates, unique_years=%s, range=%s..%s",
                 len(all_dates), sorted(set(d.year for d in all_dates)) if all_dates else [], p_start, p_end)

        period_config = {"period_types": ["year", "quarter", "month"], "start": p_start, "end": p_end}
        if detected_period_types:
            pt = set(period_config["period_types"])
            pt.update(detected_period_types)
            if any(x in pt for x in ("quarter", "half", "month")):
                pt.add("year")
            period_config["period_types"] = sorted(pt, key=lambda x: ["year", "half", "quarter", "month"].index(x) if x in ["year", "half", "quarter", "month"] else 99)

        # ── KB-based analysis (runs before LLM) ─────────────────────────────
        from backend.import_kb import ImportKB, extract_rows_from_worksheet, analyze_sheet_with_kb

        kb = ImportKB()
        kb_sheets_done: set[str] = set()  # sheets fully analyzed by KB
        kb_configs: dict[str, dict] = {}  # sheet_name -> config dict

        if _use_kb:
            yield event(_m("kb_loading"))
            try:
                await kb.load(db)
                yield event(_m("kb_loaded", n=len(kb.patterns)))
            except Exception as e:
                log.warning("KB load failed: %s", e)
                yield event(f"[WARN]KB load failed: {e}")

            yield event(_m("kb_analyzing"))

            # Emit session_id so frontend can send answers
            yield event(_m("kb_session", sid=_session_id), {"session_id": _session_id})

            for sn in sheet_names:
                if sn not in wb_formulas.sheetnames:
                    continue
                ws_kb = wb_formulas[sn]
                try:
                    sheet_rows = extract_rows_from_worksheet(ws_kb, sn)
                    indicators, questions = analyze_sheet_with_kb(kb, sheet_rows, sn)

                    if indicators:
                        # Convert IndicatorNode tree to dict format matching LLM output
                        ind_dicts = [ind.to_dict() for ind in indicators]
                        # Detect display_name and data_start_col from headers
                        _dn = sn
                        for r in range(1, min(4, (ws_kb.max_row or 1) + 1)):
                            v = ws_kb.cell(r, 1).value or ws_kb.cell(r, 2).value
                            if v and isinstance(v, str) and v.strip():
                                _dn = v.strip()
                                break
                        cfg = {
                            "excel_name": sn,
                            "display_name": _dn,
                            "data_start_col": 4,
                            "indicators": ind_dicts,
                            "_source": "kb",
                        }
                        kb_configs[sn] = cfg

                        # Count groups and indicators
                        def _count_g(items):
                            return sum(1 for it in items if it.get("is_group") or it.get("children"))
                        def _count_i(items):
                            return sum(1 + len(it.get("children", [])) for it in items)
                        yield event(_m("kb_sheet_ok", name=sn,
                                       groups=_count_g(ind_dicts),
                                       inds=_count_i(ind_dicts)))

                    # Handle questions (if any)
                    for q in questions:
                        # Check if we already have a session pattern for this
                        if kb.has_session_pattern(q.pattern_type):
                            continue
                        # Yield question event and wait for answer
                        yield event(_m("kb_question", name=sn), {
                            "type": "question",
                            "session_id": _session_id,
                            "question_id": q.question_id,
                            "sheet_name": sn,
                            "text": q.text,
                            "context": q.context,
                            "options": q.options,
                        })
                        answer = await _ask_question(_session_id, q.question_id, timeout=120)
                        if answer:
                            yield event(_m("kb_answer_ok"))
                            # Save as session pattern for auto-apply on subsequent sheets
                            kb.add_session_pattern(q.pattern_type, {
                                "answer": answer,
                                "sheet": sn,
                            })
                        else:
                            yield event(_m("kb_answer_skip"))

                except Exception as e:
                    log.warning("KB analysis failed for sheet %s: %s", sn, e)

        # Analyze with Claude (per-sheet with progress)
        # Create client lazily — cache may satisfy all requests without an API key
        try:
            client = _get_claude_client()
        except Exception:
            client = None

        try:
            sheets_config = []

            # For sheets already analyzed by KB, use KB results; for rest, use LLM
            _need_llm = [sn for sn in sheet_names if sn not in kb_configs]
            if _need_llm:
                yield event(_m("analyzing_n", n=len(_need_llm)))
            elif kb_configs:
                # All sheets analyzed by KB — skip LLM entirely
                for sn in sheet_names:
                    if sn in kb_configs:
                        sheets_config.append(kb_configs[sn])
                # Jump past LLM analysis
                _need_llm = []

            async def analyze_one(sn):
                """Analyze one sheet: KB → cache(full sheet) → Claude(chunked)."""
                # KB result already available?
                if sn in kb_configs:
                    return kb_configs[sn]
                full_text = sheet_texts[sn]
                # Check sheet-level cache first (survives chunking changes)
                cached = await _llm_cache_get(full_text)
                if cached:
                    cached["excel_name"] = sn
                    if len(cached.get("indicators", [])) > 0:
                        return cached
                # Try LLM analysis (chunks have their own cache inside)
                for attempt in range(2):
                    try:
                        cfg = await _analyze_sheet_chunked(client, full_text)
                        cfg["excel_name"] = sn
                        if len(cfg.get("indicators", [])) > 0:
                            # Cache merged result at sheet level for future runs
                            try:
                                await _llm_cache_set(full_text, cfg)
                            except Exception:
                                pass
                            return cfg
                    except Exception:
                        pass
                # No fallback — LLM is required
                return None

            # Only run LLM for sheets not already handled by KB
            _sheets_to_analyze = _need_llm if _need_llm else sheet_names
            _already_in_config = {sc.get("excel_name") for sc in sheets_config}
            _sheets_to_analyze = [sn for sn in _sheets_to_analyze if sn not in _already_in_config]

            if _sheets_to_analyze:
                tasks = [analyze_one(sn) for sn in _sheets_to_analyze]
                results = await asyncio.gather(*tasks, return_exceptions=True)

                for i, (sn, result) in enumerate(zip(_sheets_to_analyze, results)):
                    if isinstance(result, Exception):
                        yield event(_m("sheet_err", name=sn, err=result))
                    elif result and len(result.get("indicators", [])) > 0:
                        ind_count = len(result.get("indicators", []))
                        ch_count = sum(len(x.get("children", [])) for x in result.get("indicators", []))
                        yield event(_m("sheet_ok", name=sn, i=i+1, n=len(sheet_names), groups=ind_count, inds=ch_count))
                        sheets_config.append(result)
                    else:
                        yield event(_m("sheet_parse_err", name=sn))

            analysis = {"period_config": period_config, "sheets": sheets_config}
        except Exception as e:
            yield event(_m("claude_unavail", err=e))
            yield event(_m("error", msg=f"LLM analysis failed: {e}. Check PEBBLE_IMPORT_LLM_API_KEY."))
            return

        sheets_config = analysis["sheets"]

        # Ensure detected period types are always included
        if detected_period_types:
            pt = set(period_config.get("period_types", []))
            pt.update(detected_period_types)
            if any(x in pt for x in ("quarter", "half", "month")):
                pt.add("year")
            period_config["period_types"] = sorted(pt, key=lambda x: ["year", "half", "quarter", "month"].index(x) if x in ["year", "half", "quarter", "month"] else 99)

        # Create model
        model_id = str(uuid.uuid4())
        await db.execute(
            "INSERT INTO models (id, name, description) VALUES (?, ?, ?)",
            (model_id, model_name_final, "Импортировано из Excel"),
        )
        yield event(_m("model_created", name=model_name_final))

        # Period analytics — one per granularity (monthly/qhy/yearly)
        all_period_types = period_config.get("period_types", ["year", "quarter", "month"])
        period_start = period_config.get("start", "2026-01-01")
        period_end = period_config.get("end", "2028-12-31")
        start_d = date.fromisoformat(period_start)
        end_d = date.fromisoformat(period_end)

        # Compute base_year early (needed for period detection)
        _base_year_raw = date.fromisoformat(period_config.get("start", "2026-01-01")).year
        _nmes_base_year = _base_year_raw + (1 if _has_y0 else 0)

        # Pre-scan: detect period type and specific detected periods for each sheet
        _prescan_sheet_ptypes: dict[str, str] = {}
        _prescan_sheet_periods: dict[str, list[dict]] = {}
        for sheet_cfg in sheets_config:
            excel_name = sheet_cfg["excel_name"]
            if excel_name not in wb_data.sheetnames:
                continue
            ws_scan = wb_data[excel_name]
            max_col_scan = min(ws_scan.max_column or 1, 200)
            sp_scan = _detect_periods_from_headers(ws_scan, max_col_scan, base_year=_nmes_base_year)
            _prescan_sheet_ptypes[excel_name] = _get_sheet_period_type(sp_scan)
            _prescan_sheet_periods[excel_name] = sp_scan

        # Pre-scan: detect version labels (факт/план) per sheet
        _prescan_version_labels: dict[str, dict[int, str]] = {}
        _has_version_labels = False
        for sheet_cfg in sheets_config:
            excel_name = sheet_cfg["excel_name"]
            if excel_name not in wb_data.sheetnames:
                continue
            ws_scan = wb_data[excel_name]
            vl = _detect_version_labels(ws_scan, min(ws_scan.max_column or 1, 200))
            _prescan_version_labels[excel_name] = vl
            if vl:
                _has_version_labels = True

        # ── Single shared period analytic for all sheets ──
        _needs_monthly = any(pt == "monthly" for pt in _prescan_sheet_ptypes.values())
        pt_list = list(all_period_types)
        if "year" not in pt_list:
            pt_list = ["year"] + pt_list
        for lvl in ["year", "half", "quarter"]:
            if lvl not in pt_list:
                pt_list.append(lvl)
        if _needs_monthly and "month" not in pt_list:
            pt_list.append("month")
        _lvl_order = {"year": 0, "half": 1, "quarter": 2, "month": 3}
        pt_list = sorted(set(pt_list), key=lambda x: _lvl_order.get(x, 99))

        pa_id = str(uuid.uuid4())
        _period_analytic_sort = 0
        await db.execute(
            """INSERT INTO analytics (id, model_id, name, code, icon, is_periods, data_type,
               period_types, period_start, period_end, sort_order)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (pa_id, model_id, "Периоды", "periods",
             "CalendarMonthOutlined", 1, "sum",
             json.dumps(pt_list), period_start, period_end, _period_analytic_sort),
        )
        _period_analytic_sort += 1
        for sort_i, (fname, fcode, ftype) in enumerate([
            ("Наименование", "name", "string"),
            ("Начало", "start", "date"),
            ("Окончание", "end", "date"),
        ]):
            await db.execute(
                "INSERT INTO analytic_fields (id, analytic_id, name, code, data_type, sort_order) VALUES (?,?,?,?,?,?)",
                (str(uuid.uuid4()), pa_id, fname, fcode, ftype, sort_i),
            )
        period_record_ids = await _create_period_hierarchy(db, pa_id, pt_list, start_d, end_d)
        period_analytic_id = pa_id

        # ── Shared version analytic (Факт / План) ──
        version_analytic_id = None
        version_record_ids: dict[str, str] = {}
        if _has_version_labels:
            version_analytic_id = str(uuid.uuid4())
            await db.execute(
                """INSERT INTO analytics (id, model_id, name, code, icon, data_type, sort_order)
                   VALUES (?,?,?,?,?,?,?)""",
                (version_analytic_id, model_id, "Версия", "version",
                 "SwapHorizOutlined", "sum", _period_analytic_sort),
            )
            _period_analytic_sort += 1
            await db.execute(
                "INSERT INTO analytic_fields (id, analytic_id, name, code, data_type, sort_order) VALUES (?,?,?,?,?,?)",
                (str(uuid.uuid4()), version_analytic_id, "Наименование", "name", "string", 0),
            )
            for v_sort, (v_name, v_code) in enumerate([("Факт", "fact"), ("План", "plan")]):
                v_rid = str(uuid.uuid4())
                await db.execute(
                    "INSERT INTO analytic_records (id, analytic_id, parent_id, sort_order, data_json) VALUES (?,?,?,?,?)",
                    (v_rid, version_analytic_id, None, v_sort,
                     json.dumps({"name": v_name, "code": v_code})),
                )
                version_record_ids[v_code.replace("fact", "факт").replace("plan", "план")] = v_rid
            log.info(f"[import] Created version analytic with records: Факт, План")

        # Detect min_period_level per sheet
        _sheet_min_period_level: dict[str, str | None] = {}
        for excel_name, ptype in _prescan_sheet_ptypes.items():
            if ptype == "monthly":
                _sheet_min_period_level[excel_name] = "M"
            else:
                detected = _prescan_sheet_periods.get(excel_name, [])
                has_month = has_quarter = has_half = has_year = False
                for sp in detected:
                    pk = sp.get("period_key", "")
                    if re.match(r'\d{4}-\d{2}$', pk):
                        has_month = True
                    elif "-Q" in pk:
                        has_quarter = True
                    elif "-H" in pk:
                        has_half = True
                    elif pk.endswith("-Y"):
                        has_year = True
                if has_month:
                    _sheet_min_period_level[excel_name] = "M"
                elif has_quarter:
                    _sheet_min_period_level[excel_name] = "Q"
                elif has_half:
                    _sheet_min_period_level[excel_name] = "H"
                elif has_year:
                    _sheet_min_period_level[excel_name] = "Y"
                else:
                    _sheet_min_period_level[excel_name] = None

        _all_period_record_ids: dict[str, dict[str, str]] = {}
        _sheet_visible_rids: dict[str, set[str]] = {}

        for excel_name, ptype in _prescan_sheet_ptypes.items():
            if ptype == "monthly":
                _all_period_record_ids[excel_name] = dict(period_record_ids)
                _sheet_visible_rids[excel_name] = set(period_record_ids.values())
                continue

            detected = _prescan_sheet_periods.get(excel_name, [])
            detected_pks = {sp.get("period_key", "") for sp in detected if sp.get("period_key")}
            sheet_rids: dict[str, str] = {}
            visible_rids: set[str] = set()
            for pk in detected_pks:
                if pk in period_record_ids:
                    sheet_rids[pk] = period_record_ids[pk]
                    visible_rids.add(period_record_ids[pk])
            # Add parent records
            for pk in list(detected_pks):
                if re.match(r'^\d{4}-Q\d$', pk):
                    year = pk[:4]
                    q_num = int(pk[-1])
                    h_num = 1 if q_num <= 2 else 2
                    for parent_pk in [f"{year}-H{h_num}", f"{year}-Y"]:
                        if parent_pk in period_record_ids and parent_pk not in sheet_rids:
                            sheet_rids[parent_pk] = period_record_ids[parent_pk]
                            visible_rids.add(period_record_ids[parent_pk])
                elif re.match(r'^\d{4}-H\d$', pk):
                    y_pk = f"{pk[:4]}-Y"
                    if y_pk in period_record_ids and y_pk not in sheet_rids:
                        sheet_rids[y_pk] = period_record_ids[y_pk]
                        visible_rids.add(period_record_ids[y_pk])
                elif re.match(r'^\d{4}-\d{2}$', pk):
                    year = pk[:4]
                    month = int(pk[5:7])
                    q_num = (month - 1) // 3 + 1
                    h_num = 1 if q_num <= 2 else 2
                    for parent_pk in [f"{year}-Q{q_num}", f"{year}-H{h_num}", f"{year}-Y"]:
                        if parent_pk in period_record_ids and parent_pk not in sheet_rids:
                            sheet_rids[parent_pk] = period_record_ids[parent_pk]
                            visible_rids.add(period_record_ids[parent_pk])
            _all_period_record_ids[excel_name] = sheet_rids
            _sheet_visible_rids[excel_name] = visible_rids

        total_period_count = len(period_record_ids)
        yield event(_m("periods", start=period_start, end=period_end, n=total_period_count))

        # Count total indicators across all sheets for progress
        def _count_indicators(items):
            return sum(1 + _count_indicators(it.get("children", [])) for it in items)
        total_indicators = sum(_count_indicators(sc.get("indicators", [])) for sc in sheets_config)
        done_indicators = 0
        yield event(_m("total_inds", n=total_indicators, sheets=len(sheets_config)))

        # Process sheets (two passes: 1. create structure, 2. import cells)
        from backend.excel_formula_translator import translate_excel_formula

        created_sheets = []
        analytic_sort = _period_analytic_sort
        sheet_sort = 0
        total_cells = 0

        # First pass: create indicator hierarchies for ALL sheets
        all_sheet_row_maps: dict[str, dict[int, str]] = {}
        all_sheet_display_names: dict[str, str] = {}
        all_sheet_data_starts: dict[str, int] = {}
        all_row_to_parent_names: dict[str, dict[int, str]] = {}
        all_sheet_period_types: dict[str, str] = {}
        all_sheet_total_cols: dict[str, set[int]] = {}
        _all_col_to_pidx2: dict[str, dict[int, int]] = {}
        _all_col_to_pk2: dict[str, dict[int, str]] = {}
        sheet_meta: list[dict] = []
        _used_display_names2: set[str] = set()

        for sheet_cfg in sheets_config:
            excel_name = sheet_cfg["excel_name"]
            display_name = sheet_cfg.get("display_name", excel_name)
            sheet_display = display_name if display_name != excel_name else excel_name
            if sheet_display in _used_display_names2:
                sheet_display = excel_name
                if sheet_display in _used_display_names2:
                    sheet_display = f"{excel_name} (2)"
            _used_display_names2.add(sheet_display)
            indicators = sheet_cfg.get("indicators", [])
            data_start_col = sheet_cfg.get("data_start_col", 4)

            if excel_name not in wb_formulas.sheetnames or not indicators:
                continue

            sheet_indicators = _count_indicators(indicators)
            yield event(_m("creating", name=sheet_display, done=done_indicators, total=total_indicators))

            indicator_analytic_id = str(uuid.uuid4())
            await db.execute(
                """INSERT INTO analytics (id, model_id, name, code, icon, data_type, sort_order)
                   VALUES (?,?,?,?,?,?,?)""",
                (indicator_analytic_id, model_id, f"Показатели ({excel_name})",
                 f"indicators_{excel_name.lower().replace('.', '_').replace('+', '_')}",
                 "ListAltOutlined", "sum", analytic_sort),
            )
            analytic_sort += 1

            for sort_i, (fname, fcode, ftype) in enumerate([
                ("Наименование", "name", "string"),
                ("Единица измерения", "unit", "string"),
            ]):
                await db.execute(
                    "INSERT INTO analytic_fields (id, analytic_id, name, code, data_type, sort_order) VALUES (?,?,?,?,?,?)",
                    (str(uuid.uuid4()), indicator_analytic_id, fname, fcode, ftype, sort_i),
                )

            if excel_name in wb_data.sheetnames:
                _enrich_with_indent(indicators, wb_data[excel_name])
            indicators = _regroup_by_outline(indicators)
            indicators = _validate_hierarchy_by_indent(indicators)
            indicators = _fix_indicator_hierarchy(indicators)
            if excel_name in wb_data.sheetnames:
                _verify_group_rules(indicators, wb_data[excel_name], data_start_col)
                indicators = _recover_missing_rows(indicators, wb_data[excel_name], data_start_col)
            row_to_rid, rid_to_formula, row_to_name, row_to_parent_name, sum_children_rids = await _create_indicator_records(db, indicator_analytic_id, indicators)

            pebble_sheet_id = str(uuid.uuid4())
            await db.execute("INSERT INTO sheets (id, model_id, name, sort_order, excel_code) VALUES (?,?,?,?,?)",
                             (pebble_sheet_id, model_id, sheet_display, sheet_sort, excel_name))
            sheet_sort += 1

            # Bind the single shared period analytic with visible_record_ids
            _min_lvl = _sheet_min_period_level.get(excel_name)
            _vis_rids = _sheet_visible_rids.get(excel_name)
            _vis_json = json.dumps(sorted(_vis_rids)) if _vis_rids and len(_vis_rids) < len(period_record_ids) else None
            _analytics_to_bind = [period_analytic_id, indicator_analytic_id]
            if version_analytic_id:
                _analytics_to_bind.append(version_analytic_id)
            for bind_idx, aid in enumerate(_analytics_to_bind):
                is_main = 1 if aid == indicator_analytic_id else 0
                min_pl = _min_lvl if aid == period_analytic_id else None
                vis = _vis_json if aid == period_analytic_id else None
                await db.execute(
                    "INSERT INTO sheet_analytics (id, sheet_id, analytic_id, sort_order, is_main, min_period_level, visible_record_ids) VALUES (?,?,?,?,?,?,?)",
                    (str(uuid.uuid4()), pebble_sheet_id, aid, bind_idx, is_main, min_pl, vis),
                )

            users = await db.execute_fetchall("SELECT id FROM users")
            for u in users:
                try:
                    await db.execute(
                        "INSERT INTO sheet_permissions (id, sheet_id, user_id, can_view, can_edit) VALUES (?,?,?,1,1)",
                        (str(uuid.uuid4()), pebble_sheet_id, u["id"]),
                    )
                except Exception:
                    pass

            all_sheet_row_maps[excel_name] = row_to_name
            all_sheet_display_names[excel_name] = sheet_display
            all_sheet_data_starts[excel_name] = data_start_col
            all_row_to_parent_names[excel_name] = row_to_parent_name

            # Detect period type and total columns for cross-sheet substitution
            _sp_fp2 = _detect_periods_from_headers(
                wb_data[excel_name] if excel_name in wb_data.sheetnames else wb_formulas[excel_name],
                min((wb_data[excel_name].max_column if excel_name in wb_data.sheetnames else wb_formulas[excel_name].max_column) or 1, 200),
                base_year=_nmes_base_year)
            all_sheet_period_types[excel_name] = _get_sheet_period_type(_sp_fp2)

            _tc2, _ = _classify_total_leaf_cols(_sp_fp2)
            all_sheet_total_cols[excel_name] = _tc2

            # Build col→period_idx for cross-sheet refs (leaf columns only)
            _sorted_sp_fp2 = sorted(_sp_fp2, key=lambda x: x["col"])
            _pidx_map2: dict[int, int] = {}
            _leaf_idx_fp2 = 0
            for _spp2x in _sorted_sp_fp2:
                if _spp2x["col"] in _tc2:
                    continue
                _pidx_map2[_spp2x["col"]] = _leaf_idx_fp2
                _leaf_idx_fp2 += 1
            _all_col_to_pidx2[excel_name] = _pidx_map2

            # Build col→period_key for cross-sheet absolute period refs
            _pk_map_fp2: dict[int, str] = {}
            for _spp2 in _sp_fp2:
                pk = _spp2.get("period_key", "")
                if pk:
                    _pk_map_fp2[_spp2["col"]] = pk
            _all_col_to_pk2[excel_name] = _pk_map_fp2

            done_indicators += sheet_indicators
            sheet_meta.append({
                "excel_name": excel_name,
                "sheet_display": sheet_display,
                "pebble_sheet_id": pebble_sheet_id,
                "row_to_rid": row_to_rid,
                "rid_to_formula": rid_to_formula,
                "row_to_name": row_to_name,
                "row_to_parent_name": row_to_parent_name,
                "sum_children_rids": sum_children_rids,
                "data_start_col": data_start_col,
                "sheet_indicators": sheet_indicators,
            })

        yield event(_m("importing_data", n=len(sheet_meta)))

        # Second pass: import cells using deterministic formula translator
        for meta in sheet_meta:
            excel_name = meta["excel_name"]
            sheet_display = meta["sheet_display"]
            pebble_sheet_id = meta["pebble_sheet_id"]
            row_to_rid = meta["row_to_rid"]
            rid_to_formula = meta["rid_to_formula"]
            row_to_name = meta["row_to_name"]
            row_to_parent_name = meta.get("row_to_parent_name", {})
            sum_children_rids = meta.get("sum_children_rids", set())
            data_start_col = meta["data_start_col"]

            ws_f = wb_formulas[excel_name]
            ws_d = wb_data[excel_name]

            # Extract starting values (column before data_start)
            pre_data_values: dict[int, float] = {}
            if data_start_col > 1:
                pre_col = data_start_col - 1
                for row_num in row_to_name:
                    try:
                        v = ws_d.cell(row_num, pre_col).value
                        if v is not None:
                            fv = float(v)
                            if fv != 0:
                                pre_data_values[row_num] = fv
                    except (ValueError, TypeError):
                        pass

            # Use per-sheet period_record_ids
            _sheet_period_rids = _all_period_record_ids.get(excel_name, period_record_ids)

            sheet_periods = _detect_periods_from_headers(ws_d, min(ws_d.max_column or 1, 200), base_year=_nmes_base_year)
            col_to_period_rid = {}
            for sp in sheet_periods:
                pkey = sp.get("period_key")
                if pkey and pkey in _sheet_period_rids:
                    col_to_period_rid[sp["col"]] = _sheet_period_rids[pkey]
                elif sp.get("date"):
                    d = sp["date"]
                    key = f"{d.year}-{d.month:02d}"
                    if key in _sheet_period_rids:
                        col_to_period_rid[sp["col"]] = _sheet_period_rids[key]

            sorted_period_cols = sorted(col_to_period_rid.keys())
            first_period_rid = col_to_period_rid[sorted_period_cols[0]] if sorted_period_cols else None

            total_cols2, _ = _classify_total_leaf_cols(sheet_periods)

            # Build col→period_idx for this sheet (leaf columns only)
            _sorted_sp2 = sorted(sheet_periods, key=lambda x: x["col"])
            _col_to_pidx2: dict[int, int] = {}
            _leaf_idx2 = 0
            for _sp2 in _sorted_sp2:
                if _sp2["col"] in total_cols2:
                    continue
                _col_to_pidx2[_sp2["col"]] = _leaf_idx2
                _leaf_idx2 += 1

            # Build col→period_key for this sheet
            _col_to_pk2: dict[int, str] = {}
            for sp in sheet_periods:
                pk = sp.get("period_key", "")
                if pk:
                    _col_to_pk2[sp["col"]] = pk

            cell_count = 0
            for row_num, indicator_rid in row_to_rid.items():
                formula_info = rid_to_formula.get(indicator_rid)

                # Detect percentage number_format from the first data cell
                for _fc in sorted_period_cols:
                    nf = ws_d.cell(row_num, _fc).number_format or ""
                    if "%" in nf:
                        await db.execute(
                            "UPDATE analytic_records SET data_json = json_set(data_json, '$.format', 'percent') WHERE id = ?",
                            (indicator_rid,),
                        )
                        break
                    if nf and nf != "General":
                        break

                for col_num, period_rid in col_to_period_rid.items():
                    val = ws_d.cell(row_num, col_num).value
                    if val is None:
                        continue
                    if isinstance(val, str):
                        try:
                            val = float(val.replace(",", ".").replace(" ", ""))
                        except (ValueError, AttributeError):
                            continue
                    is_yellow = _is_input_cell(ws_f.cell(row_num, col_num))
                    if is_yellow:
                        rule = "manual"
                        formula_text = ""
                    else:
                        excel_formula = ws_f.cell(row_num, col_num).value
                        is_first = (period_rid == first_period_rid)
                        if isinstance(excel_formula, str) and excel_formula.startswith("="):
                            rule = "formula"
                            try:
                                # External workbook refs → import as manual
                                if _has_external_refs(excel_formula):
                                    rule = "manual"
                                    formula_text = ""
                                    raise _ExternalRefSkip()
                                # Pre-substitute total-column refs with values
                                if total_cols2 and col_num not in total_cols2:
                                    excel_formula = _substitute_total_col_refs(
                                        excel_formula, ws_d, row_num, total_cols2, data_start_col, row_to_name)
                                # Pre-substitute cross-sheet refs to different period types or total columns
                                src_ptype = all_sheet_period_types.get(excel_name, "unknown")
                                if src_ptype != "unknown":
                                    excel_formula = _substitute_cross_period_refs(
                                        excel_formula, src_ptype, all_sheet_period_types, wb_data,
                                        all_sheet_total_cols, all_sheet_row_maps)
                                # Pre-substitute: replace refs to non-indicator rows with values
                                excel_formula = _substitute_non_indicator_refs(
                                    excel_formula, ws_d, row_to_name, data_start_col, base_col=col_num)
                                parent_maps = dict(all_row_to_parent_names)
                                parent_maps["__self__"] = row_to_parent_name
                                formula_text = translate_excel_formula(
                                    excel_formula,
                                    base_col=col_num,
                                    data_start_col=data_start_col,
                                    row_to_name=row_to_name,
                                    sheet_row_maps=all_sheet_row_maps,
                                    sheet_display_names=all_sheet_display_names,
                                    is_first_period=is_first,
                                    sheet_data_starts=all_sheet_data_starts,
                                    row_to_parent_names=parent_maps,
                                    pre_data_values=pre_data_values,
                                    col_to_period_idx=_col_to_pidx2,
                                    sheet_col_to_period_idx=_all_col_to_pidx2,
                                    col_to_period_key=_col_to_pk2,
                                    sheet_col_to_period_key=_all_col_to_pk2,
                                )
                            except _ExternalRefSkip:
                                pass  # rule/formula_text already set
                            except Exception:
                                formula_text = (formula_info or {}).get("formula", excel_formula)
                        else:
                            # Excel cell has no formula (constant) — import as manual.
                            # Claude's formula_info is NOT used here because the Excel
                            # value is authoritative for non-formula cells.
                            rule = "manual"
                            formula_text = ""
                    # Post-translation: if formula has unparseable `:` range, fall back to manual.
                    # Allow `::` (cross-sheet separator like `[Sheet::name]`).
                    if rule == "formula" and formula_text and re.search(r'(?<!:):(?!:)', formula_text):
                        rule = "manual"
                        formula_text = ""
                    # First-period formula cells: preserve starting balances
                    if rule == "formula" and is_first and val != 0:
                        ft = formula_text.strip()
                        if ft == "0" or ft.startswith("0/") or ft.startswith("0*") or \
                           'предыдущий' in ft or 'назад' in ft:
                            rule = "manual"
                            formula_text = ""

                    # Build coord_key: period|indicator[|version]
                    if version_analytic_id:
                        _sheet_vlabels = _prescan_version_labels.get(excel_name, {})
                        _vlabel = _sheet_vlabels.get(col_num)
                        if _vlabel:
                            _ver_rid = version_record_ids[_vlabel]
                        else:
                            _ver_rid = version_record_ids.get("план", "")
                        _coord_key = f"{period_rid}|{indicator_rid}|{_ver_rid}"
                    else:
                        _coord_key = f"{period_rid}|{indicator_rid}"

                    # Group indicators → sum_children rule (engine will compute).
                    # But preserve explicit Excel formulas — totals like
                    # =D32+D54+D76+... must not be overwritten by sum_children.
                    if indicator_rid in sum_children_rids and not (rule == "formula" and formula_text):
                        rule = "sum_children"
                        formula_text = ""

                    try:
                        from backend.db import intern_formula
                        fid = await intern_formula(db, formula_text) if formula_text else None
                        await db.execute(
                            "INSERT INTO cell_data (id, sheet_id, coord_key, value, data_type, rule, formula, formula_id) VALUES (?,?,?,?,?,?,?,?)",
                            (str(uuid.uuid4()), pebble_sheet_id, _coord_key, str(val), "sum", rule, "" if fid else formula_text, fid),
                        )
                        cell_count += 1
                    except Exception:
                        pass

            # Extract consolidation formulas from total columns
            total_cols = _detect_total_columns(ws_d, ws_f, col_to_period_rid, min(ws_d.max_column or 1, 200))
            n_consol = 0
            if total_cols:
                n_consol = await _extract_and_store_consolidation_rules(
                    db, ws_f, total_cols, row_to_rid, row_to_name, pebble_sheet_id,
                    period_cols=set(col_to_period_rid.keys()),
                )

            total_cells += cell_count
            created_sheets.append({
                "name": sheet_display, "id": pebble_sheet_id, "cells": cell_count,
                "excel_name": excel_name,
                "col_to_period_rid": dict(col_to_period_rid),
                "row_to_rid": dict(row_to_rid),
            })
            consol_msg = f", {n_consol} формул консолидации из Excel" if n_consol else ""
            yield event(_m("data_ok", name=sheet_display, inds=len(row_to_rid), cells=cell_count, consol=consol_msg, done=done_indicators, total=total_indicators))

        await db.commit()

        # ── Post-import validation ──
        expected_sheets = len(sheet_names)
        actual_sheets = len(created_sheets)
        if actual_sheets < expected_sheets:
            missing = set(sheet_names) - {sc["excel_name"] for sc in sheets_config}
            yield event(_m("sheets_warn", done=actual_sheets, total=expected_sheets, missing=', '.join(missing)))

        # Log cell counts per sheet (informational, not a warning)
        for cs in created_sheets:
            if cs["cells"] == 0:
                yield event(_m("zero_cells", name=cs['name']))

        # ── Post-import: detect period-consolidation formulas for ratios/averages ──
        # Ask Claude which indicators should NOT be summed across periods
        # (e.g. interest rates, averages). Tolerant of API failures.
        if os.environ.get("ANTHROPIC_API_KEY"):
            from backend.formula_suggester import suggest_consolidations_for_sheet
            # Get the periods analytic name (used as context in the prompt).
            period_analytic_name = "Периоды"
            try:
                pa = await db.execute_fetchall(
                    "SELECT name FROM analytics WHERE id = ? LIMIT 1",
                    (period_analytic_id,),
                )
                if pa:
                    period_analytic_name = pa[0]["name"]
            except Exception:
                pass
            total_rules = 0
            async def _suggest_one(cs):
                try:
                    return await suggest_consolidations_for_sheet(db, cs["id"], period_analytic_name)
                except Exception as e:
                    print(f"[import] suggest_consolidations failed for {cs['id']}: {e}")
                    return 0
            suggest_results = await asyncio.gather(*[_suggest_one(cs) for cs in created_sheets])
            total_rules = sum(suggest_results)
            # Propagate formulas across sheets for same-named indicators
            if len(created_sheets) > 1:
                from backend.formula_suggester import propagate_consolidations_across_sheets
                try:
                    total_rules += await propagate_consolidations_across_sheets(db, model_id)
                except Exception as e:
                    print(f"[import] propagate_consolidations failed: {e}")
            if total_rules:
                yield event(_m("consol_rules", n=total_rules))

        # ── Post-import: translate all names to ru/en/ky/vi ──
        if os.environ.get("PEBBLE_SKIP_TRANSLATE"):
            yield event(_m("translating") + " (skipped)")
        else:
            yield event(_m("translating"))
            try:
                from backend.translation_service import batch_translate, save_translations, SUPPORTED_LANGS

                # Collect all names that need translation
                names_to_translate: list[str] = []
                entity_map: list[tuple[str, str, str]] = []  # (entity_type, entity_id, name)

                # Model name
                names_to_translate.append(model_name_final)
                entity_map.append(("model", model_id, model_name_final))

                # Sheet names
                sheets_rows = await db.execute_fetchall(
                    "SELECT id, name FROM sheets WHERE model_id = ?", (model_id,))
                for sr in sheets_rows:
                    if sr["name"]:
                        names_to_translate.append(sr["name"])
                        entity_map.append(("sheet", sr["id"], sr["name"]))

                # Analytic names
                analytics_rows = await db.execute_fetchall(
                    "SELECT id, name FROM analytics WHERE model_id = ?", (model_id,))
                for ar in analytics_rows:
                    if ar["name"]:
                        names_to_translate.append(ar["name"])
                        entity_map.append(("analytic", ar["id"], ar["name"]))

                # Analytic record names (indicator names, period names)
                record_rows = await db.execute_fetchall(
                    """SELECT r.id, r.data_json FROM analytic_records r
                       JOIN analytics a ON r.analytic_id = a.id
                       WHERE a.model_id = ?""", (model_id,))
                for rr in record_rows:
                    try:
                        dj = json.loads(rr["data_json"]) if isinstance(rr["data_json"], str) else rr["data_json"]
                        name = dj.get("name", "")
                        if name:
                            names_to_translate.append(name)
                            entity_map.append(("analytic_record", rr["id"], name))
                    except Exception:
                        pass

                # Batch translate (deduplicated inside)
                unique_names = list(dict.fromkeys(names_to_translate))
                # Translate in chunks of 50 to avoid token limits
                all_translations: dict[str, dict[str, str]] = {}
                for i in range(0, len(unique_names), 50):
                    chunk = unique_names[i:i+50]
                    chunk_result = await batch_translate(chunk)
                    all_translations.update(chunk_result)

                # Save translations
                for etype, eid, name in entity_map:
                    tr = all_translations.get(name, {lang: name for lang in SUPPORTED_LANGS})
                    await save_translations(etype, eid, "name", tr, db=db)

                await db.commit()
                yield event(_m("translated_ok", n=len(unique_names), langs=len(SUPPORTED_LANGS)))
            except Exception as e:
                yield event(_m("translate_fail", err=e))

        # ── Post-import: verify values against Excel ──
        try:
            yield event(_m("verifying"))
            mismatches = await _verify_import_against_excel(db, model_id, wb_data, created_sheets, tolerance=0.01)
            if mismatches:
                yield event(_m("verify_warn", n=len(mismatches)))
                for mm in mismatches[:5]:  # show first 5
                    yield event(f"   {mm['sheet']}: {mm['indicator']} / {mm['period']} — "
                                f"Excel={mm['excel']}, Pebble={mm['pebble']}")
                if len(mismatches) > 5:
                    yield event(_m("verify_more", n=len(mismatches) - 5))
            else:
                yield event(_m("verify_ok"))
        except Exception as e:
            log.warning("Verification failed: %s", e)

        yield event(_m("done", sheets=len(created_sheets), cells=total_cells),
                     {"done": True, "model_id": model_id, "model_name": model_name_final})

        # Cleanup Q&A session
        _cleanup_qa_session(_session_id)

    return StreamingResponse(generate(), media_type="text/event-stream")


# ── Fallback heuristic analysis (when Claude API unavailable) ──────────────

def _fallback_heuristic_analysis(wb) -> dict:
    """Basic heuristic analysis as fallback when Claude API is not available."""
    # Detect period range from all sheets (datetime objects + year numbers)
    dates = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, 11):
            for c in range(1, min((ws.max_column or 1) + 1, 200)):
                v = ws.cell(r, c).value
                if isinstance(v, datetime):
                    dates.append(v)
                elif isinstance(v, (int, float)) and 2020 <= v <= 2040 and v == int(v):
                    dates.append(datetime(int(v), 1, 1))
                    dates.append(datetime(int(v), 12, 1))

    if dates:
        min_d = min(dates)
        max_d = max(dates)
        start = f"{min_d.year}-{min_d.month:02d}-01"
        end_month = max_d.month
        end_year = max_d.year
        end_day = monthrange(end_year, end_month)[1]
        end = f"{end_year}-{end_month:02d}-{end_day:02d}"
    else:
        start, end = "2026-01-01", "2028-12-31"

    period_config = {"period_types": ["year", "quarter", "month"], "start": start, "end": end}

    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = min(ws.max_row or 1, 500)
        max_col = min(ws.max_column or 1, 200)

        # Find data start col
        data_start_col = 4
        for r in range(1, 7):
            for c in range(1, min(max_col + 1, 50)):
                if isinstance(ws.cell(r, c).value, datetime):
                    data_start_col = c
                    break

        # Detect display name
        title = ws.cell(1, 1).value or ws.cell(1, 2).value or sheet_name
        title = str(title).strip()

        # Detect label column — pick column with most text content
        label_col = 1
        col_text = {}
        for _lc in (1, 2):
            text_count = 0
            for _lr in range(3, max_row + 1):
                _lv = ws.cell(_lr, _lc).value
                if _lv is not None and isinstance(_lv, str) and len(str(_lv).strip()) > 1:
                    text_count += 1
            col_text[_lc] = text_count
        if col_text[2] > col_text[1] * 1.5 and col_text[2] > 10:
            label_col = 2

        # Find first data row
        data_start_row = 7
        for r in range(1, max_row + 1):
            v = ws.cell(r, label_col).value
            if v is not None and str(v).strip() not in ("", "(тыс. сом)", "(тыс сом)"):
                if r > 3:
                    data_start_row = r
                    break

        # Build indicator hierarchy with improved grouping
        indicators = []
        current_group = None

        for r in range(data_start_row, max_row + 1):
            name = ws.cell(r, label_col).value
            if name is None or str(name).strip() == "":
                if current_group:
                    indicators.append(current_group)
                    current_group = None
                continue

            name = str(name).strip()
            unit_col = label_col + 1
            unit = ws.cell(r, unit_col).value
            unit = str(unit).strip() if unit else ""

            # Check for data in period columns
            has_data = False
            for c in range(data_start_col, min(data_start_col + 5, max_col + 1)):
                if ws.cell(r, c).value is not None:
                    has_data = True
                    break

            # Check if bold
            is_bold = ws.cell(r, label_col).font and ws.cell(r, label_col).font.bold

            # Check indentation
            cell_label = ws.cell(r, label_col)
            indent = cell_label.alignment.indent if cell_label.alignment and cell_label.alignment.indent else 0

            # Group detection: name pattern, bold, or no data in period columns
            is_group_by_name = bool(_GROUP_PATTERN.search(name)) or bool(_GROUP_PREFIX.match(name))
            is_group_header = is_group_by_name or (is_bold and not has_data) or (
                not has_data and unit in ("ЕИ", "")
                and ws.cell(r, unit_col + 1).value in ("Отв.исп.", None, "")
            )

            item = {"name": name, "unit": unit, "row": r, "is_group": False, "children": []}

            if is_group_header:
                if current_group:
                    indicators.append(current_group)
                current_group = {"name": name, "unit": "", "row": r, "is_group": True, "children": []}
            elif current_group:
                current_group["children"].append(item)
            else:
                indicators.append(item)

        if current_group:
            indicators.append(current_group)

        if indicators:
            sheets.append({
                "excel_name": sheet_name,
                "display_name": title,
                "data_start_col": data_start_col,
                "indicators": indicators,
            })

    return {"period_config": period_config, "sheets": sheets}
