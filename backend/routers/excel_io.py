import uuid
import json
import io
from fastapi import APIRouter, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from backend.db import get_db
from backend.coord_key import (
    pack as _pack_coord,
    pack_sync as _pack_sync,
    to_uuid_coord_key as _ck_to_uuid,
)

router = APIRouter(prefix="/api/excel", tags=["excel"])

INDENT = 4  # spaces per hierarchy level


@router.get("/analytics/{analytic_id}/export")
async def export_analytic_records(analytic_id: str):
    db = get_db()
    fields = await db.execute_fetchall(
        "SELECT * FROM analytic_fields WHERE analytic_id = ? ORDER BY sort_order",
        (analytic_id,),
    )
    records = await db.execute_fetchall(
        "SELECT * FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
        (analytic_id,),
    )
    analytic = await db.execute_fetchall("SELECT name FROM analytics WHERE id = ?", (analytic_id,))
    name = analytic[0]["name"] if analytic else "export"

    fields = [dict(f) for f in fields]
    records = [dict(r) for r in records]

    # Build tree
    by_id = {r["id"]: r for r in records}
    children = {}
    roots = []
    for r in records:
        pid = r["parent_id"]
        if pid:
            children.setdefault(pid, []).append(r)
        else:
            roots.append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = name[:31]

    # Header
    headers = [f["name"] for f in fields]
    ws.append(headers)

    def write_rows(nodes, level):
        for node in nodes:
            data = json.loads(node["data_json"]) if isinstance(node["data_json"], str) else node["data_json"]
            row = []
            for i, f in enumerate(fields):
                val = data.get(f["code"], "")
                if i == 0 and level > 0:
                    val = " " * (level * INDENT) + str(val)
                row.append(val)
            ws.append(row)
            if node["id"] in children:
                write_rows(children[node["id"]], level + 1)

    write_rows(roots, 0)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{__import__('urllib.parse', fromlist=['quote']).quote(name + '.xlsx')}"},
    )


@router.post("/analytics/{analytic_id}/import")
async def import_analytic_records(analytic_id: str, file: UploadFile = File(...)):
    db = get_db()
    fields = await db.execute_fetchall(
        "SELECT * FROM analytic_fields WHERE analytic_id = ? ORDER BY sort_order",
        (analytic_id,),
    )
    fields = [dict(f) for f in fields]

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

    # Clear existing records
    await db.execute("DELETE FROM analytic_records WHERE analytic_id = ?", (analytic_id,))

    parent_stack: list[tuple[int, str]] = []  # (level, record_id)
    sort = 0

    for row_vals in rows:
        if not row_vals or all(v is None for v in row_vals):
            continue

        first_val = str(row_vals[0]) if row_vals[0] is not None else ""
        stripped = first_val.lstrip(" ")
        spaces = len(first_val) - len(stripped)
        level = spaces // max(INDENT, 1) if spaces > 0 else 0

        data = {}
        for i, f in enumerate(fields):
            val = row_vals[i] if i < len(row_vals) else None
            if i == 0 and val is not None:
                val = str(val).lstrip(" ")
            if val is not None:
                data[f["code"]] = val

        # Determine parent
        while parent_stack and parent_stack[-1][0] >= level:
            parent_stack.pop()
        parent_id = parent_stack[-1][1] if parent_stack else None

        rid = str(uuid.uuid4())
        await db.execute(
            "INSERT INTO analytic_records (id, analytic_id, parent_id, sort_order, data_json) VALUES (?,?,?,?,?)",
            (rid, analytic_id, parent_id, sort, json.dumps(data, ensure_ascii=False)),
        )
        parent_stack.append((level, rid))
        sort += 1

    await db.commit()
    records = await db.execute_fetchall(
        "SELECT * FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
        (analytic_id,),
    )
    return [dict(r) for r in records]


# ── Sheet data export/import ───────────────────────────────────────────────

def _build_record_tree(records):
    """Build ordered list with hierarchy info: [(record, data, level, has_children)]"""
    by_parent = {}
    for r in records:
        pid = r["parent_id"] or "__root__"
        by_parent.setdefault(pid, []).append(r)
    child_ids = set()
    for kids in by_parent.values():
        for k in kids:
            child_ids.add(k["id"])

    result = []
    def walk(pid, level):
        for r in by_parent.get(pid, []):
            data = json.loads(r["data_json"]) if isinstance(r["data_json"], str) else r["data_json"]
            has_children = r["id"] in by_parent
            result.append((r, data, level, has_children))
            walk(r["id"], level + 1)
    walk("__root__", 0)
    return result


@router.get("/sheets/{sheet_id}/export")
async def export_sheet_data(sheet_id: str):
    """Export sheet data as Excel: rows = indicators, cols = periods."""
    db = get_db()

    sheet = await db.execute_fetchall("SELECT * FROM sheets WHERE id = ?", (sheet_id,))
    if not sheet:
        return {"error": "sheet not found"}
    sheet_name = sheet[0]["name"]

    # Get bound analytics in order
    bindings = await db.execute_fetchall(
        """SELECT sa.analytic_id, a.name, a.is_periods
           FROM sheet_analytics sa JOIN analytics a ON a.id = sa.analytic_id
           WHERE sa.sheet_id = ? ORDER BY sa.sort_order""",
        (sheet_id,),
    )
    if len(bindings) < 2:
        return {"error": "sheet needs at least 2 analytics"}

    # First analytic = columns (periods), rest = rows
    col_analytic_id = bindings[0]["analytic_id"]
    row_analytic_ids = [b["analytic_id"] for b in bindings[1:]]

    # Load period records (columns) — leaf nodes only
    col_records = await db.execute_fetchall(
        "SELECT * FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
        (col_analytic_id,),
    )
    col_tree = _build_record_tree([dict(r) for r in col_records])
    # Only leaf periods (months)
    col_leaves = [(r, d) for r, d, lvl, has_ch in col_tree if not has_ch]

    # Load row records
    all_row_records = []
    for ra_id in row_analytic_ids:
        recs = await db.execute_fetchall(
            "SELECT * FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
            (ra_id,),
        )
        all_row_records.extend([dict(r) for r in recs])
    row_tree = _build_record_tree(all_row_records)

    # Load cells (DB stores seq_id form; rebuild as uuid form for in-memory lookup)
    cells_raw = await db.execute_fetchall(
        "SELECT coord_key, value FROM cell_data WHERE sheet_id = ?", (sheet_id,),
    )
    cells = {_ck_to_uuid(c["coord_key"]): c["value"] for c in cells_raw}

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # Header row: empty + period names
    header = ["Показатель"]
    for _, d in col_leaves:
        header.append(d.get("name", ""))
    ws.append(header)
    # Bold header
    for c in range(1, len(header) + 1):
        ws.cell(1, c).font = Font(bold=True)

    # Data rows
    col_rids = [r["id"] for r, _ in col_leaves]
    for rec, data, level, has_children in row_tree:
        name = data.get("name", "")
        if level > 0:
            name = "  " * level + name
        row = [name]
        for col_rid in col_rids:
            coord_key = f"{col_rid}|{rec['id']}"
            val = cells.get(coord_key, "")
            # Try to convert to number
            try:
                row.append(float(val))
            except (ValueError, TypeError):
                row.append(val)
        ws.append(row)
        # Bold groups
        if has_children:
            ws.cell(ws.max_row, 1).font = Font(bold=True)

    # Auto-width for first column
    ws.column_dimensions['A'].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{__import__('urllib.parse', fromlist=['quote']).quote(sheet_name + '.xlsx')}"},
    )


@router.get("/models/{model_id}/export")
async def export_model(model_id: str):
    """Export entire model as multi-sheet Excel workbook."""
    db = get_db()

    model = await db.execute_fetchall("SELECT name FROM models WHERE id = ?", (model_id,))
    if not model:
        return {"error": "model not found"}
    model_name = model[0]["name"]

    sheets = await db.execute_fetchall(
        "SELECT id, name FROM sheets WHERE model_id = ? ORDER BY sort_order, created_at",
        (model_id,),
    )
    if not sheets:
        return {"error": "no sheets"}

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for s in sheets:
        sheet_id = s["id"]
        sheet_name = s["name"]

        bindings = await db.execute_fetchall(
            """SELECT sa.analytic_id, a.name, a.is_periods
               FROM sheet_analytics sa JOIN analytics a ON a.id = sa.analytic_id
               WHERE sa.sheet_id = ? ORDER BY sa.sort_order""",
            (sheet_id,),
        )
        if len(bindings) < 2:
            continue

        col_analytic_id = bindings[0]["analytic_id"]
        row_analytic_ids = [b["analytic_id"] for b in bindings[1:]]

        # Columns (periods) — leaf nodes only
        col_records = await db.execute_fetchall(
            "SELECT * FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
            (col_analytic_id,),
        )
        col_tree = _build_record_tree([dict(r) for r in col_records])
        col_leaves = [(r, d) for r, d, lvl, has_ch in col_tree if not has_ch]

        # Rows (indicators)
        all_row_records = []
        for ra_id in row_analytic_ids:
            recs = await db.execute_fetchall(
                "SELECT * FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
                (ra_id,),
            )
            all_row_records.extend([dict(r) for r in recs])
        row_tree = _build_record_tree(all_row_records)

        # Cells
        cells_raw = await db.execute_fetchall(
            """SELECT cd.coord_key, cd.value, cd.rule,
                      COALESCE(NULLIF(cd.formula,''), f.text, '') AS formula
               FROM cell_data cd LEFT JOIN formulas f ON f.id = cd.formula_id
               WHERE cd.sheet_id = ?""",
            (sheet_id,),
        )
        cells = {_ck_to_uuid(c["coord_key"]): c for c in cells_raw}

        # Build worksheet
        # Truncate sheet name to 31 chars (Excel limit)
        ws_title = sheet_name[:31]
        # Ensure unique title
        existing = [ws.title for ws in wb.worksheets]
        if ws_title in existing:
            ws_title = ws_title[:28] + f"_{len(existing)}"
        ws = wb.create_sheet(title=ws_title)

        # Row 1: title
        ws.cell(1, 1, sheet_name)
        ws.cell(1, 1).font = Font(bold=True, size=12)

        # Row 3: period headers
        for ci, (_, d) in enumerate(col_leaves):
            ws.cell(3, ci + 2, d.get("name", ""))
            ws.cell(3, ci + 2).font = Font(bold=True)

        # Row 4+: data
        col_rids = [r["id"] for r, _ in col_leaves]
        data_row = 4
        for rec, data, level, has_children in row_tree:
            name = data.get("name", "")
            unit = data.get("unit", "")
            if level > 0:
                name = "  " * level + name
            ws.cell(data_row, 1, name)
            if has_children:
                ws.cell(data_row, 1).font = Font(bold=True)

            for ci, col_rid in enumerate(col_rids):
                coord_key = f"{col_rid}|{rec['id']}"
                cell_data = cells.get(coord_key)
                if cell_data:
                    val = cell_data["value"]
                    try:
                        ws.cell(data_row, ci + 2, float(val))
                    except (ValueError, TypeError):
                        ws.cell(data_row, ci + 2, val)

            data_row += 1

        ws.column_dimensions['A'].width = 45

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{model_name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{__import__('urllib.parse', fromlist=['quote']).quote(filename)}"},
    )


@router.put("/sheets/{sheet_id}/import")
async def import_sheet_data(sheet_id: str, file: UploadFile = File(...)):
    """Import cell values from Excel back into the sheet. Matches by row name and column period."""
    db = get_db()

    # Get bound analytics
    bindings = await db.execute_fetchall(
        "SELECT sa.analytic_id FROM sheet_analytics sa WHERE sa.sheet_id = ? ORDER BY sa.sort_order",
        (sheet_id,),
    )
    if len(bindings) < 2:
        return {"error": "sheet needs at least 2 analytics"}

    col_analytic_id = bindings[0]["analytic_id"]
    row_analytic_ids = [b["analytic_id"] for b in bindings[1:]]

    # Build period name -> record_id map (leaf nodes)
    col_records = await db.execute_fetchall(
        "SELECT * FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
        (col_analytic_id,),
    )
    col_tree = _build_record_tree([dict(r) for r in col_records])
    col_leaves = [(r, d) for r, d, lvl, has_ch in col_tree if not has_ch]

    # Build row name -> record_id map
    all_row_records = []
    for ra_id in row_analytic_ids:
        recs = await db.execute_fetchall(
            "SELECT * FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
            (ra_id,),
        )
        all_row_records.extend([dict(r) for r in recs])

    row_name_to_id = {}
    for r in all_row_records:
        data = json.loads(r["data_json"]) if isinstance(r["data_json"], str) else r["data_json"]
        name = data.get("name", "").strip()
        if name:
            row_name_to_id[name] = r["id"]

    # Read Excel
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    # Header = period names
    header = [ws.cell(1, c).value for c in range(2, ws.max_column + 1)]
    # Map header names to period record IDs
    col_name_to_rid = {}
    for rec, data in col_leaves:
        col_name_to_rid[data.get("name", "")] = rec["id"]

    col_rids = []
    for h in header:
        col_rids.append(col_name_to_rid.get(str(h).strip(), None) if h else None)

    # Read data rows
    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        row_name = ws.cell(row_idx, 1).value
        if row_name is None:
            continue
        row_name = str(row_name).strip()
        row_rid = row_name_to_id.get(row_name)
        if not row_rid:
            continue

        for col_idx, col_rid in enumerate(col_rids):
            if col_rid is None:
                continue
            val = ws.cell(row_idx, col_idx + 2).value
            if val is None:
                continue
            coord_key = await _pack_coord(db, [col_rid, row_rid])
            value_str = str(val)

            existing = await db.execute_fetchall(
                "SELECT id FROM cell_data WHERE sheet_id = ? AND coord_key = ?",
                (sheet_id, coord_key),
            )
            if existing:
                await db.execute(
                    "UPDATE cell_data SET value = ? WHERE sheet_id = ? AND coord_key = ?",
                    (value_str, sheet_id, coord_key),
                )
            else:
                cid = str(uuid.uuid4())
                await db.execute(
                    "INSERT INTO cell_data (id, sheet_id, coord_key, value, data_type, rule, formula) VALUES (?,?,?,?,?,?,?)",
                    (cid, sheet_id, coord_key, value_str, "sum", "manual", ""),
                )
            updated += 1

    await db.commit()
    return {"updated": updated}
