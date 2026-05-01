"""Excel → Pebble formula translator.

Pure function: translates Excel cell-reference formulas into Pebble
[indicator_name] formulas using row→record name mappings.

Key concepts:
- base_col: the column where this formula lives (e.g. 5 for col E = m2)
- If a cell reference points to a column LEFT of base_col → previous period
- If same column → current period
- Cross-sheet: ='SheetName'!E19 → [SheetDisplayName::indicator_name]
- First period (m1) refs to col before data_start → replace with 0
"""

import re
from openpyxl.utils import column_index_from_string


# ── Excel cell reference parser ────────────────────────────────────────────

# Matches: 'Sheet Name'!E19, Sheet!E19, E19, $E$19, E$19, $E19
# Note: \w with re.UNICODE matches Cyrillic and other Unicode word chars.
# Unquoted sheet names allow only word chars and '.' — Excel requires quoting
# names containing operators (+, -, space, etc), so we mustn't include those
# here or e.g. "A1+Sheet!B2" would capture "+Sheet" as the sheet name.
CELL_REF_RE = re.compile(
    r"(?:'([^']+)'|([\w.]+))!"                # optional sheet prefix (quoted or simple name, Unicode-aware)
    r"(\$?[A-Z]{1,3})(\$?\d+)"                # column + row
    r"|"                                        # OR
    r"(\$?[A-Z]{1,3})(\$?\d+)",               # bare column + row
    re.UNICODE,
)

# Matches Excel range like E19:E25 or 'Sheet'!E19:E25
RANGE_RE = re.compile(
    r"(?:(?:'([^']+)'|([\w.]+))!)?"
    r"(\$?[A-Z]{1,3})(\$?\d+)"
    r":"
    r"(\$?[A-Z]{1,3})(\$?\d+)",
    re.UNICODE,
)

# SUM/AVERAGE(range) pattern
SUM_RE = re.compile(r"(SUM|AVERAGE)\s*\(([^)]+)\)", re.IGNORECASE)


def _col_num(col_str: str) -> int:
    """Convert column letter (possibly with $) to 1-based number."""
    return column_index_from_string(col_str.replace("$", ""))


def _row_num(row_str: str) -> int:
    """Convert row string (possibly with $) to number."""
    return int(row_str.replace("$", ""))


# ── Core translator ───────────────────────────────────────────────────────

def translate_excel_formula(
    excel_formula: str,
    base_col: int,
    data_start_col: int,
    row_to_name: dict[int, str],
    sheet_row_maps: dict[str, dict[int, str]] | None = None,
    sheet_display_names: dict[str, str] | None = None,
    is_first_period: bool = False,
    sheet_data_starts: dict[str, int] | None = None,
    row_to_parent_names: dict[str, dict[int, str]] | None = None,
    pre_data_values: dict[int, float] | None = None,
    col_to_period_idx: dict[int, int] | None = None,
    sheet_col_to_period_idx: dict[str, dict[int, int]] | None = None,
    col_to_period_key: dict[int, str] | None = None,
    sheet_col_to_period_key: dict[str, dict[int, str]] | None = None,
) -> str:
    """Translate an Excel formula to Pebble formula syntax.

    Args:
        excel_formula: Excel formula string (with leading =)
        base_col: Column number where this formula lives (1-based)
        data_start_col: First data column number (1-based)
        row_to_name: {excel_row_number: pebble_indicator_name} for current sheet
        sheet_row_maps: {excel_sheet_name: {row: name}} for cross-sheet refs
        sheet_display_names: {excel_sheet_name: pebble_display_name} for cross-sheet refs
        is_first_period: True if this is the first period column
        sheet_data_starts: {excel_sheet_name: data_start_col} for cross-sheet period alignment
        row_to_parent_names: {sheet_name_or___self__: {row: parent_name}} for disambiguation

    Returns:
        Pebble formula string (without =)
    """
    if not excel_formula:
        return ""

    formula = excel_formula.lstrip("=").strip()
    if not formula:
        return ""

    sheet_row_maps = sheet_row_maps or {}
    sheet_display_names = sheet_display_names or {}
    sheet_data_starts = sheet_data_starts or {}

    # Step 0: Convert Excel percent literals (0.2% → 0.002)
    formula = re.sub(r'(\d+(?:\.\d+)?)%', lambda m: str(float(m.group(1)) / 100), formula)

    # Step 1: Expand SUM(range) into SUM(cell, cell, ...)
    formula = _expand_sum_ranges(formula, row_to_name, sheet_row_maps)

    # Step 2: Replace each cell reference with [indicator_name]
    result = _replace_cell_refs(
        formula, base_col, data_start_col, row_to_name,
        sheet_row_maps, sheet_display_names, is_first_period,
        sheet_data_starts, row_to_parent_names, pre_data_values,
        col_to_period_idx, sheet_col_to_period_idx,
        col_to_period_key, sheet_col_to_period_key,
    )

    return result


def _expand_sum_ranges(
    formula: str,
    row_to_name: dict[int, str],
    sheet_row_maps: dict[str, dict[int, str]],
) -> str:
    """Expand SUM/AVERAGE(E19:E25) into SUM/AVERAGE(E19,E20,...,E25) using row map to skip missing rows."""

    def expand_match(m):
        func_name = m.group(1).upper()  # SUM or AVERAGE
        inner = m.group(2)
        rm = RANGE_RE.match(inner.strip())
        if not rm:
            return m.group(0)

        sheet1 = rm.group(1) or rm.group(2) or None
        col1 = rm.group(3)
        row1 = _row_num(rm.group(4))
        col2 = rm.group(5)
        row2 = _row_num(rm.group(6))

        col1_clean = col1.replace("$", "")
        col2_clean = col2.replace("$", "")

        if col1_clean != col2_clean and row1 == row2:
            # Horizontal range (same row, different cols): D5:N5
            from openpyxl.utils import get_column_letter
            c1 = _col_num(col1)
            c2 = _col_num(col2)
            prefix = f"'{sheet1}'!" if sheet1 else ""
            refs = [f"{prefix}{get_column_letter(c)}{row1}" for c in range(c1, c2 + 1)]
            if refs:
                return f"{func_name}({','.join(refs)})"
            return m.group(0)

        if col1_clean != col2_clean:
            return m.group(0)  # Multi-column multi-row range — don't expand

        rmap = row_to_name
        if sheet1:
            rmap = sheet_row_maps.get(sheet1, row_to_name)

        prefix = f"'{sheet1}'!" if sheet1 else ""
        refs = []
        for r in range(row1, row2 + 1):
            if r in rmap:
                refs.append(f"{prefix}{col1}{r}")

        if refs:
            return f"{func_name}({','.join(refs)})"
        return m.group(0)

    return SUM_RE.sub(expand_match, formula)


def _replace_cell_refs(
    formula: str,
    base_col: int,
    data_start_col: int,
    row_to_name: dict[int, str],
    sheet_row_maps: dict[str, dict[int, str]],
    sheet_display_names: dict[str, str],
    is_first_period: bool,
    sheet_data_starts: dict[str, int] | None = None,
    row_to_parent_names: dict[str, dict[int, str]] | None = None,
    pre_data_values: dict[int, float] | None = None,
    col_to_period_idx: dict[int, int] | None = None,
    sheet_col_to_period_idx: dict[str, dict[int, int]] | None = None,
    col_to_period_key: dict[int, str] | None = None,
    sheet_col_to_period_key: dict[str, dict[int, str]] | None = None,
) -> str:
    """Replace all cell references in formula with [name] or [Sheet::name] tokens."""

    # We need to process references from right to left to preserve positions
    # First, find all references
    refs = []

    for m in CELL_REF_RE.finditer(formula):
        if m.group(1) is not None or m.group(2) is not None:
            # Sheet-prefixed reference
            sheet_name = m.group(1) or m.group(2)
            col_str = m.group(3)
            row_str = m.group(4)
        elif m.group(5) is not None:
            # Bare reference
            sheet_name = None
            col_str = m.group(5)
            row_str = m.group(6)
        else:
            continue

        refs.append({
            "start": m.start(),
            "end": m.end(),
            "sheet": sheet_name,
            "col": _col_num(col_str),
            "row": _row_num(row_str),
            "original": m.group(0),
        })

    # Process from right to left
    result = formula
    for ref in reversed(refs):
        replacement = _translate_ref(
            ref, base_col, data_start_col, row_to_name,
            sheet_row_maps, sheet_display_names, is_first_period,
            sheet_data_starts, row_to_parent_names, pre_data_values,
            col_to_period_idx, sheet_col_to_period_idx,
            col_to_period_key, sheet_col_to_period_key,
        )
        result = result[:ref["start"]] + replacement + result[ref["end"]:]

    return result


_PARENT_CHILD_SEP = "\x1f"  # ASCII unit separator — internal delimiter that can never appear in user-supplied names


def _format_ref(name: str, display: str | None = None) -> str:
    """Format an indicator reference with optional parent qualifier.

    parent\\x1fchild → [parent][child]
    plain name      → [name]
    With display (cross-sheet): display::parent\\x1fchild → [display::parent][child]
    """
    if _PARENT_CHILD_SEP in name:
        parent, child = name.split(_PARENT_CHILD_SEP, 1)
        if display:
            return f"[{display}::{parent}][{child}]"
        return f"[{parent}][{child}]"
    if display:
        return f"[{display}::{name}]"
    return f"[{name}]"


def _translate_ref(
    ref: dict,
    base_col: int,
    data_start_col: int,
    row_to_name: dict[int, str],
    sheet_row_maps: dict[str, dict[int, str]],
    sheet_display_names: dict[str, str],
    is_first_period: bool,
    sheet_data_starts: dict[str, int] | None = None,
    row_to_parent_names: dict[str, dict[int, str]] | None = None,
    pre_data_values: dict[int, float] | None = None,
    col_to_period_idx: dict[int, int] | None = None,
    sheet_col_to_period_idx: dict[str, dict[int, int]] | None = None,
    col_to_period_key: dict[int, str] | None = None,
    sheet_col_to_period_key: dict[str, dict[int, str]] | None = None,
) -> str:
    """Translate a single cell reference to Pebble [name] token."""
    sheet_name = ref["sheet"]
    col = ref["col"]
    row = ref["row"]
    sheet_data_starts = sheet_data_starts or {}
    row_to_parent_names = row_to_parent_names or {}
    pre_data_values = pre_data_values or {}

    # Determine indicator name
    if sheet_name:
        rmap = sheet_row_maps.get(sheet_name, {})
        display = sheet_display_names.get(sheet_name, sheet_name)
    else:
        rmap = row_to_name
        display = None

    name = rmap.get(row)
    if name is None:
        return ref["original"]  # Can't resolve — keep original

    # Check if name is duplicate in the row map — if so, disambiguate with
    # parent qualifier: [Parent][Indicator]. The calc engine resolves this
    # by filtering candidates whose parent record name matches.
    name_lower = name.lower()
    duplicates = sum(1 for r, n in rmap.items() if n.lower() == name_lower)
    if duplicates > 1:
        # Try parent name disambiguation
        pmap_key = sheet_name if sheet_name else "__self__"
        parent_map = row_to_parent_names.get(pmap_key, {}) if row_to_parent_names else {}
        parent_name = parent_map.get(row)
        if parent_name:
            # Use [parent][child] bracket format — encode internally with a
            # sentinel separator so parent_name or name may safely contain "/".
            name = f"{parent_name}{_PARENT_CHILD_SEP}{name}"
        else:
            name = f"{name}#row{row}"

    # Determine period modifier using period index alignment
    # Use col_to_period_idx mapping if available (skips total columns)
    # Otherwise fall back to simple column arithmetic
    col_to_period_idx = col_to_period_idx or {}
    sheet_col_to_period_idx = sheet_col_to_period_idx or {}
    col_to_period_key = col_to_period_key or {}
    sheet_col_to_period_key = sheet_col_to_period_key or {}

    # If the target column is NOT in the period_idx map (e.g., it's a total/aggregate column),
    # use absolute period key directly — period_diff computation would be wrong
    target_pidx_map = (sheet_col_to_period_idx.get(sheet_name) if sheet_name else None) or col_to_period_idx
    target_pk_map = (sheet_col_to_period_key.get(sheet_name) if sheet_name else None) or col_to_period_key

    if col not in target_pidx_map and col in target_pk_map:
        # Target is a total/aggregate column — use absolute period key
        ref_name = f"[{display}::{name}]" if display else f"[{name}]"
        return f'{ref_name}(периоды="{target_pk_map[col]}")'

    # Detect cross-level references (e.g. Y column referencing M column).
    # назад(N)/вперед(N) only works within the same period level because
    # the engine navigates the prev_period chain per level.
    def _pk_level(pk: str) -> str:
        if not pk:
            return "?"
        if re.match(r'\d{4}-\d{2}$', pk):
            return "M"
        if "-Q" in pk:
            return "Q"
        if "-H" in pk:
            return "H"
        if pk.endswith("-Y"):
            return "Y"
        return "?"

    source_pk = col_to_period_key.get(base_col, "")
    target_pk_val = target_pk_map.get(col, "")
    source_level = _pk_level(source_pk)
    target_level = _pk_level(target_pk_val)

    # Cross-level reference → always use absolute period key
    if source_level != target_level and target_pk_val and col != base_col:
        ref_name = f"[{display}::{name}]" if display else f"[{name}]"
        return f'{ref_name}(периоды="{target_pk_val}")'

    if base_col in col_to_period_idx:
        source_period_idx = col_to_period_idx[base_col]
    else:
        source_period_idx = base_col - data_start_col

    if sheet_name and sheet_name in sheet_col_to_period_idx:
        target_map = sheet_col_to_period_idx[sheet_name]
        if col in target_map:
            ref_period_idx = target_map[col]
        else:
            target_data_start = (sheet_data_starts or {}).get(sheet_name, data_start_col)
            ref_period_idx = col - target_data_start
    elif sheet_name and sheet_name in (sheet_data_starts or {}):
        target_data_start = sheet_data_starts[sheet_name]
        ref_period_idx = col - target_data_start
    elif col in col_to_period_idx:
        ref_period_idx = col_to_period_idx[col]
    else:
        ref_period_idx = col - data_start_col

    period_diff = ref_period_idx - source_period_idx

    if period_diff < 0:
        n_back = abs(period_diff)
        # Reference to column before data_start — always use starting value
        if ref_period_idx < 0:
            starting_val = pre_data_values.get(row)
            if starting_val is not None and starting_val != 0:
                s = f"{starting_val:.10f}".rstrip("0").rstrip(".")
                return s
            return "0"
        ref_name = _format_ref(name, display)
        if n_back == 1:
            return f'{ref_name}(периоды="предыдущий")'
        else:
            return f'{ref_name}(период=период.назад({n_back}))'
    elif period_diff > 0:
        # Forward reference — use absolute period key
        target_pk = target_pk_map.get(col)
        ref_name = _format_ref(name, display)
        if target_pk:
            return f'{ref_name}(периоды="{target_pk}")'
        else:
            # Fallback: use вперед(N) syntax
            return f'{ref_name}(период=период.вперед({period_diff}))'
    else:
        # Same period (period_diff == 0)
        return _format_ref(name, display)


# ── Batch translation ─────────────────────────────────────────────────────

def translate_sheet_formulas(
    ws_formulas,
    data_start_col: int,
    row_to_name: dict[int, str],
    sheet_row_maps: dict[str, dict[int, str]] | None = None,
    sheet_display_names: dict[str, str] | None = None,
    max_col: int | None = None,
) -> dict[int, dict]:
    """Translate all formulas on a sheet.

    Args:
        ws_formulas: openpyxl worksheet (with formulas, not data_only)
        data_start_col: first data column
        row_to_name: {row: indicator_name} for this sheet
        sheet_row_maps: {sheet_name: {row: name}} for cross-sheet
        sheet_display_names: {excel_name: display_name}
        max_col: max column to scan

    Returns:
        {row: {"formula": str, "formula_first": str or None}}
        formula = formula for m2+ periods
        formula_first = formula for m1 (if different), or None
    """
    if max_col is None:
        max_col = min(ws_formulas.max_column or 1, 200)

    results = {}

    for row in row_to_name:
        # Find m1 (first formula in data columns) and m2 (second formula)
        m1_formula = None
        m1_col = None
        m2_formula = None
        m2_col = None

        for c in range(data_start_col, max_col + 1):
            cell = ws_formulas.cell(row, c)
            val = cell.value
            if val is not None and isinstance(val, str) and val.startswith("="):
                if m1_formula is None:
                    m1_formula = val
                    m1_col = c
                elif m2_formula is None:
                    m2_formula = val
                    m2_col = c
                    break
            elif val is not None and m1_formula is None:
                # First period is a constant (manual input)
                m1_formula = str(val)
                m1_col = c
                continue

        if m1_formula is None:
            continue

        # Translate m2 (the "general" formula)
        if m2_formula and m2_col:
            formula = translate_excel_formula(
                m2_formula, m2_col, data_start_col, row_to_name,
                sheet_row_maps, sheet_display_names, is_first_period=False,
            )
        elif m1_formula.startswith("="):
            formula = translate_excel_formula(
                m1_formula, m1_col, data_start_col, row_to_name,
                sheet_row_maps, sheet_display_names, is_first_period=False,
            )
        else:
            continue  # Manual input only

        # Translate m1 if it's a formula and different from m2
        formula_first = None
        if m1_formula and m1_formula.startswith("="):
            pebble_m1 = translate_excel_formula(
                m1_formula, m1_col, data_start_col, row_to_name,
                sheet_row_maps, sheet_display_names, is_first_period=True,
            )
            if pebble_m1 != formula:
                formula_first = pebble_m1
        elif m1_formula and not m1_formula.startswith("="):
            # First period is a constant
            try:
                float(m1_formula)
                formula_first = m1_formula
            except ValueError:
                formula_first = "0"

        results[row] = {
            "formula": formula,
            "formula_first": formula_first,
        }

    return results
