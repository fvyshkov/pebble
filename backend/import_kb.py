"""Import Knowledge Base (KB) — pattern-based sheet hierarchy analysis.

Stores and applies structural patterns for Excel import:
- Title/unit row detection
- BOLD = root hierarchy
- Currency breakdown ("в нац. валюте" / "в ин. валюте")
- Category breakdown (КБ, МСБ, РБ, ИФЦ)
- "в т.ч." children pattern
- Indent-based nesting

Patterns are stored in the `import_kb` DB table and loaded at import time.
New patterns can be learned from user confirmations during interactive import.
"""

import json
import re
import uuid
import logging
from dataclasses import dataclass, field

log = logging.getLogger(__name__)

# ── Pattern types ─────────────────────────────────────────────────────────

PATTERN_TYPES = {
    "skip_title":         "Skip rows that match the sheet title",
    "skip_unit":          "Skip rows that are unit-of-measure markers",
    "skip_header":        "Skip header rows (Показатель, ЕИ, etc.)",
    "bold_root":          "BOLD rows are root-level group headers",
    "currency_breakdown": "Currency breakdown rows are children of preceding item",
    "vtch_children":      "\"в т.ч.\" rows start a children block",
    "itogo_group":        "\"Итого/Всего\" rows are group aggregates",
    "indent_nesting":     "Indent-based parent-child nesting",
    "sum_children":       "Groups with sum_children rule",
    "manual_input":       "Colored background = manual input",
}

# ── Default patterns (seeded on first run) ────────────────────────────────

DEFAULT_PATTERNS = [
    {
        "id": "kb-default-skip-title",
        "pattern_type": "skip_title",
        "pattern_key": "skip_title_row",
        "match_rule": {"type": "sheet_title_match",
                       "description": "First row containing the sheet name is not an indicator"},
        "action": {"type": "skip"},
        "confidence": 1.0,
        "source": "default",
    },
    {
        "id": "kb-default-skip-unit",
        "pattern_type": "skip_unit",
        "pattern_key": "skip_unit_markers",
        "match_rule": {"type": "regex",
                       "patterns": [
                           r"\(тыс\.?\s*сом\)",
                           r"\(млн\.?\s*сом\)",
                           r"\(тыс\.?\s*руб\)",
                           r"\(млн\.?\s*руб\)",
                           r"^\s*ед\.?\s*изм\.?\s*$",
                       ]},
        "action": {"type": "skip"},
        "confidence": 1.0,
        "source": "default",
    },
    {
        "id": "kb-default-skip-header",
        "pattern_type": "skip_header",
        "pattern_key": "skip_header_rows",
        "match_rule": {"type": "exact_lower",
                       "values": ["показатель", "показатели", "наименование",
                                  "наименование показателя", "ед. изм.", "ед.изм.",
                                  "еи", "ед. измерения", "отв.исп.", "ответственный",
                                  "примечание", "№ п/п", "№"]},
        "action": {"type": "skip"},
        "confidence": 1.0,
        "source": "default",
    },
    {
        "id": "kb-default-bold-root",
        "pattern_type": "bold_root",
        "pattern_key": "bold_is_root_group",
        "match_rule": {"type": "formatting", "condition": "bold"},
        "action": {"type": "make_group", "level": "root"},
        "confidence": 0.9,
        "source": "default",
    },
    {
        "id": "kb-default-currency",
        "pattern_type": "currency_breakdown",
        "pattern_key": "currency_nac_in_valuta",
        "match_rule": {"type": "text_contains_any",
                       "values": ["в национальной валюте", "в нац. валюте",
                                  "нац.валюта", "в иностранной валюте",
                                  "в ин. валюте", "ин.валюта",
                                  "в инвалюте", "в нацвалюте"]},
        "action": {"type": "make_child_of_previous_group"},
        "confidence": 1.0,
        "source": "default",
    },
    {
        "id": "kb-default-vtch",
        "pattern_type": "vtch_children",
        "pattern_key": "vtch_pattern",
        "match_rule": {"type": "text_ends_with",
                       "values": ["в т.ч.:", "в т.ч.", "в том числе:",
                                  "в том числе", "включая:", "включая"]},
        "action": {"type": "start_children_block"},
        "confidence": 1.0,
        "source": "default",
    },
    {
        "id": "kb-default-itogo",
        "pattern_type": "itogo_group",
        "pattern_key": "itogo_vsego_pattern",
        "match_rule": {"type": "text_starts_with_any",
                       "values": ["итого", "всего", "всего по ", "общее ",
                                  "общий ", "общая ", "суммарн", "сумма "]},
        "action": {"type": "mark_as_group_aggregate"},
        "confidence": 0.9,
        "source": "default",
    },
    {
        "id": "kb-default-indent",
        "pattern_type": "indent_nesting",
        "pattern_key": "indent_based_hierarchy",
        "match_rule": {"type": "indent_increase"},
        "action": {"type": "make_child"},
        "confidence": 1.0,
        "source": "default",
    },
    {
        "id": "kb-default-input",
        "pattern_type": "manual_input",
        "pattern_key": "colored_bg_is_manual",
        "match_rule": {"type": "formatting", "condition": "colored_bg"},
        "action": {"type": "set_rule", "rule": "manual"},
        "confidence": 0.8,
        "source": "default",
    },
]


# ── Data structures ───────────────────────────────────────────────────────

@dataclass
class SheetRow:
    """Parsed row from Excel sheet."""
    row_num: int
    name: str
    unit: str = ""
    is_bold: bool = False
    indent: int = 0
    outline_level: int = 0
    has_data: bool = False
    has_formula: bool = False
    is_input: bool = False
    bg_color: str | None = None
    formula1: str = ""
    formula2: str = ""
    label_col: int = 1  # which column has the label


@dataclass
class IndicatorNode:
    """An indicator in the hierarchy tree."""
    name: str
    unit: str = ""
    row: int = 0
    is_group: bool = False
    rule: str = "manual"
    formula: str = ""
    formula_first: str = ""
    children: list["IndicatorNode"] = field(default_factory=list)
    confidence: float = 1.0
    pattern_source: str = ""  # which pattern produced this classification
    _skip: bool = False

    def to_dict(self) -> dict:
        d = {
            "name": self.name,
            "unit": self.unit,
            "row": self.row,
            "is_group": self.is_group,
            "rule": self.rule,
            "children": [c.to_dict() for c in self.children if not c._skip],
        }
        if self.formula:
            d["formula"] = self.formula
        if self.formula_first:
            d["formula_first"] = self.formula_first
        return d


@dataclass
class KBQuestion:
    """A question to ask the user during import."""
    question_id: str
    sheet_name: str
    text: str
    context: str  # description of the ambiguous situation
    options: list[dict]  # [{value: str, label: str, description: str}]
    row_nums: list[int] = field(default_factory=list)  # affected rows
    pattern_type: str = ""


# ── KB class ──────────────────────────────────────────────────────────────

class ImportKB:
    """Knowledge Base for import pattern recognition."""

    def __init__(self):
        self.patterns: list[dict] = []
        self._session_patterns: dict[str, dict] = {}  # patterns confirmed in this session

    async def load(self, db) -> None:
        """Load patterns from DB."""
        rows = await db.execute_fetchall(
            "SELECT * FROM import_kb ORDER BY confidence DESC"
        )
        self.patterns = [dict(r) for r in rows]
        if not self.patterns:
            await self._seed_defaults(db)

    async def _seed_defaults(self, db) -> None:
        """Insert default patterns on first run."""
        for p in DEFAULT_PATTERNS:
            try:
                await db.execute(
                    """INSERT OR IGNORE INTO import_kb
                       (id, pattern_type, pattern_key, match_rule, action, confidence, source)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (p["id"], p["pattern_type"], p["pattern_key"],
                     json.dumps(p["match_rule"]), json.dumps(p["action"]),
                     p["confidence"], p["source"]),
                )
            except Exception as e:
                log.warning("Failed to seed KB pattern %s: %s", p["pattern_key"], e)
        await db.commit()
        rows = await db.execute_fetchall("SELECT * FROM import_kb ORDER BY confidence DESC")
        self.patterns = [dict(r) for r in rows]
        log.info("Seeded %d default KB patterns", len(self.patterns))

    async def save_pattern(self, db, pattern_type: str, pattern_key: str,
                           match_rule: dict, action: dict,
                           confidence: float = 1.0, source: str = "user_confirmed") -> str:
        """Save a new or updated pattern to DB."""
        pid = str(uuid.uuid4())
        try:
            await db.execute(
                """INSERT INTO import_kb (id, pattern_type, pattern_key, match_rule, action, confidence, source)
                   VALUES (?, ?, ?, ?, ?, ?, ?)
                   ON CONFLICT(pattern_key) DO UPDATE SET
                       match_rule=excluded.match_rule,
                       action=excluded.action,
                       confidence=excluded.confidence,
                       source=excluded.source,
                       updated_at=datetime('now')""",
                (pid, pattern_type, pattern_key,
                 json.dumps(match_rule), json.dumps(action),
                 confidence, source),
            )
            await db.commit()
        except Exception as e:
            log.warning("Failed to save KB pattern %s: %s", pattern_key, e)
        return pid

    def add_session_pattern(self, key: str, pattern: dict) -> None:
        """Add a pattern confirmed in this import session (used across sheets)."""
        self._session_patterns[key] = pattern

    def has_session_pattern(self, key: str) -> bool:
        return key in self._session_patterns

    def get_session_pattern(self, key: str) -> dict | None:
        return self._session_patterns.get(key)

    # ── Pattern matching helpers ──────────────────────────────────────────

    def _get_patterns_by_type(self, ptype: str) -> list[dict]:
        """Get all patterns of a specific type, highest confidence first."""
        return [p for p in self.patterns if p["pattern_type"] == ptype]

    def _match_rule(self, rule: dict | str, text: str, **ctx) -> bool:
        """Check if a text/context matches a pattern's match_rule."""
        if isinstance(rule, str):
            rule = json.loads(rule)
        rtype = rule.get("type", "")
        text_lower = text.strip().lower()

        if rtype == "exact_lower":
            return text_lower in [v.lower() for v in rule.get("values", [])]

        elif rtype == "regex":
            for pat in rule.get("patterns", []):
                if re.search(pat, text, re.IGNORECASE):
                    return True
            return False

        elif rtype == "text_contains_any":
            for v in rule.get("values", []):
                if v.lower() in text_lower:
                    return True
            return False

        elif rtype == "text_ends_with":
            for v in rule.get("values", []):
                if text_lower.rstrip().endswith(v.lower()):
                    return True
            return False

        elif rtype == "text_starts_with_any":
            for v in rule.get("values", []):
                if text_lower.startswith(v.lower()):
                    return True
            return False

        elif rtype == "sheet_title_match":
            sheet_name = ctx.get("sheet_name", "")
            return sheet_name.lower() in text_lower or text_lower in sheet_name.lower()

        elif rtype == "formatting":
            cond = rule.get("condition", "")
            if cond == "bold":
                return ctx.get("is_bold", False)
            elif cond == "colored_bg":
                return bool(ctx.get("bg_color"))
            return False

        elif rtype == "indent_increase":
            return ctx.get("indent_increased", False)

        return False


# ── Sheet parsing (from openpyxl worksheet, not from text) ────────────────

def extract_rows_from_worksheet(ws, sheet_name: str, max_rows: int = 500) -> list[SheetRow]:
    """Extract structured row data directly from openpyxl worksheet."""
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt

    max_col = min(ws.max_column or 1, 200)
    max_row = min(ws.max_row or 1, max_rows)
    rows: list[SheetRow] = []

    # Detect label column (col with most non-empty text values)
    col_scores = {}
    for c in range(1, min(6, max_col + 1)):
        count = 0
        for r in range(1, min(max_row + 1, 100)):
            v = ws.cell(r, c).value
            if v is not None and not isinstance(v, _dt) and str(v).strip():
                count += 1
        col_scores[c] = count
    label_col = max(col_scores, key=col_scores.get) if col_scores else 1

    for r in range(1, max_row + 1):
        # Get label from detected column (and fallback to other cols)
        name = ""
        unit = ""
        for c in range(1, min(6, max_col + 1)):
            v = ws.cell(r, c).value
            if v is not None and not isinstance(v, _dt):
                s = str(v).strip()
                if s and len(s) < 200:
                    if not name:
                        name = s
                    elif not unit and c != label_col:
                        # Second non-empty column might be unit
                        if len(s) < 20:
                            unit = s

        if not name:
            continue

        cell_a = ws.cell(r, label_col)
        is_bold = bool(cell_a.font and cell_a.font.bold)
        indent = int(cell_a.alignment.indent) if cell_a.alignment and cell_a.alignment.indent else 0
        outline = ws.row_dimensions[r].outline_level if hasattr(ws.row_dimensions[r], 'outline_level') else 0
        if indent == 0 and outline > 0:
            indent = outline

        has_data = False
        has_formula = False
        is_input = False
        bg_color = None
        formula1 = ""
        formula2 = ""

        for c in range(4, min(15, max_col + 1)):
            cv = ws.cell(r, c).value
            if cv is not None:
                has_data = True
                if isinstance(cv, str) and cv.startswith("="):
                    has_formula = True
                # Check bg color
                from backend.routers.import_excel import _get_cell_bg_color
                clr = _get_cell_bg_color(ws.cell(r, c))
                if clr:
                    is_input = True
                    bg_color = clr
                break

        if has_formula:
            for c in range(4, min(max_col + 1, 50)):
                cv = ws.cell(r, c).value
                if cv and isinstance(cv, str) and cv.startswith("="):
                    if not formula1:
                        formula1 = str(cv)[:120]
                    elif not formula2:
                        formula2 = str(cv)[:120]
                        break
                elif cv is not None and not formula1:
                    formula1 = str(cv)[:30]

        rows.append(SheetRow(
            row_num=r, name=name, unit=unit,
            is_bold=is_bold, indent=indent, outline_level=outline,
            has_data=has_data, has_formula=has_formula,
            is_input=is_input, bg_color=bg_color,
            formula1=formula1, formula2=formula2,
            label_col=label_col,
        ))

    return rows


def analyze_sheet_with_kb(kb: ImportKB, rows: list[SheetRow],
                          sheet_name: str, display_name: str = "",
                          data_start_col: int = 4) -> tuple[list[IndicatorNode], list[KBQuestion]]:
    """Analyze sheet rows using KB patterns.

    Returns:
        (indicators, questions) — indicators is the hierarchy tree,
        questions is a list of ambiguous cases needing user input.
    """
    if not display_name:
        display_name = sheet_name

    questions: list[KBQuestion] = []
    skip_patterns = kb._get_patterns_by_type("skip_title") + \
                    kb._get_patterns_by_type("skip_unit") + \
                    kb._get_patterns_by_type("skip_header")
    bold_patterns = kb._get_patterns_by_type("bold_root")
    currency_patterns = kb._get_patterns_by_type("currency_breakdown")
    vtch_patterns = kb._get_patterns_by_type("vtch_children")
    itogo_patterns = kb._get_patterns_by_type("itogo_group")
    input_patterns = kb._get_patterns_by_type("manual_input")

    # Phase 1: classify each row
    classified: list[dict] = []
    for row in rows:
        cls = {
            "row": row,
            "skip": False,
            "is_group": False,
            "is_currency_child": False,
            "is_vtch": False,
            "is_itogo": False,
            "rule": "manual",
            "confidence": 1.0,
            "pattern": "",
        }

        # Check skip patterns
        for p in skip_patterns:
            rule = p.get("match_rule", p) if isinstance(p.get("match_rule"), dict) else json.loads(p.get("match_rule", "{}"))
            if kb._match_rule(rule, row.name, sheet_name=sheet_name):
                cls["skip"] = True
                cls["pattern"] = p.get("pattern_key", "skip")
                break

        if cls["skip"]:
            classified.append(cls)
            continue

        # Check currency breakdown
        for p in currency_patterns:
            rule = p.get("match_rule", p) if isinstance(p.get("match_rule"), dict) else json.loads(p.get("match_rule", "{}"))
            if kb._match_rule(rule, row.name):
                cls["is_currency_child"] = True
                cls["pattern"] = p.get("pattern_key", "currency")
                break

        # Check "в т.ч." pattern
        for p in vtch_patterns:
            rule = p.get("match_rule", p) if isinstance(p.get("match_rule"), dict) else json.loads(p.get("match_rule", "{}"))
            if kb._match_rule(rule, row.name):
                cls["is_vtch"] = True
                cls["is_group"] = True
                cls["rule"] = "sum_children"
                cls["pattern"] = p.get("pattern_key", "vtch")
                break

        # Check itogo/vsego
        for p in itogo_patterns:
            rule = p.get("match_rule", p) if isinstance(p.get("match_rule"), dict) else json.loads(p.get("match_rule", "{}"))
            if kb._match_rule(rule, row.name):
                cls["is_itogo"] = True
                cls["is_group"] = True
                cls["rule"] = "sum_children"
                cls["pattern"] = p.get("pattern_key", "itogo")
                break

        # Check bold = group
        if row.is_bold and not cls["is_currency_child"]:
            for p in bold_patterns:
                rule = p.get("match_rule", p) if isinstance(p.get("match_rule"), dict) else json.loads(p.get("match_rule", "{}"))
                if kb._match_rule(rule, row.name, is_bold=True):
                    cls["is_group"] = True
                    cls["rule"] = "sum_children"
                    cls["confidence"] = min(cls["confidence"], p.get("confidence", 0.9))
                    cls["pattern"] = p.get("pattern_key", "bold")
                    break

        # Determine rule
        if not cls["is_group"] and not cls["is_currency_child"]:
            if row.has_formula:
                cls["rule"] = "formula"
            elif row.is_input:
                cls["rule"] = "manual"
            elif row.has_data:
                cls["rule"] = "manual"

        classified.append(cls)

    # Phase 2: Build hierarchy using indent + bold + patterns
    indicators: list[IndicatorNode] = []
    active_rows = [c for c in classified if not c["skip"]]

    if not active_rows:
        return indicators, questions

    # Build hierarchy from indent levels and bold status
    root_nodes: list[IndicatorNode] = []
    stack: list[tuple[int, IndicatorNode]] = []  # (indent_level, node)

    for i, cls in enumerate(active_rows):
        row = cls["row"]
        node = IndicatorNode(
            name=row.name,
            unit=row.unit,
            row=row.row_num,
            is_group=cls["is_group"],
            rule=cls["rule"],
            formula=row.formula1 if cls["rule"] == "formula" else "",
            confidence=cls["confidence"],
            pattern_source=cls["pattern"],
        )

        effective_indent = row.indent

        # Currency breakdown: attach as child of the most recent group
        if cls["is_currency_child"]:
            # Find the nearest preceding group
            attached = False
            for si in range(len(stack) - 1, -1, -1):
                parent_indent, parent_node = stack[si]
                if parent_node.is_group:
                    parent_node.children.append(node)
                    attached = True
                    break
            if not attached and root_nodes:
                # Make the previous root a group if it isn't
                prev = root_nodes[-1]
                prev.is_group = True
                prev.rule = "sum_children"
                prev.children.append(node)
                attached = True
            if not attached:
                root_nodes.append(node)
            continue

        # Normal hierarchy logic
        if cls["is_group"] and row.is_bold and row.indent == 0:
            # Bold root-level group — always goes to root
            root_nodes.append(node)
            stack = [(0, node)]
        elif effective_indent == 0 and not cls["is_currency_child"]:
            # Root-level item
            if stack and stack[-1][1].is_group and not row.is_bold:
                # Non-bold item after a bold group at same indent — might be child
                # Check: if there's a current group on stack at indent 0, add as child
                parent_indent, parent_node = stack[-1]
                if parent_indent == 0 and parent_node.is_group:
                    parent_node.children.append(node)
                    if cls["is_group"]:
                        stack.append((1, node))
                    continue
            root_nodes.append(node)
            stack = [(0, node)]
        else:
            # Indented item — find parent
            while stack and stack[-1][0] >= effective_indent:
                stack.pop()
            if stack:
                parent_indent, parent_node = stack[-1]
                parent_node.is_group = True
                if parent_node.rule == "manual":
                    parent_node.rule = "sum_children"
                parent_node.children.append(node)
            else:
                root_nodes.append(node)
            if cls["is_group"]:
                stack.append((effective_indent, node))

    return root_nodes, questions


async def log_qa(db, session_id: str, sheet_name: str,
                 question: str, answer: str, pattern_id: str | None = None) -> None:
    """Log a Q&A interaction."""
    await db.execute(
        """INSERT INTO import_kb_log (id, session_id, sheet_name, question, answer, pattern_id)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (str(uuid.uuid4()), session_id, sheet_name, question, answer, pattern_id),
    )
    await db.commit()
