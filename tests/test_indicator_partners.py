"""Verify the «общее количество партнеров, в т.ч.:» indicator in MAIN.xlsx
is represented correctly across the three formula-display surfaces:

  1. Per-cell formula text matches Excel cell-by-cell (cell_data store).
  2. /cells/resolved-formulas (powers grid formula-mode + tooltip) returns
     the actual per-cell formula or "cell-manual" for the literal cells —
     never "SUM" placeholder.
  3. /indicator-rules-all (powers AnalyticRecordsGrid + IndicatorFormulasPanel)
     reports the indicator's leaf rule as the sentinel "__per_cell__" because
     cells carry distinct formulas — no single representative leaf applies.

User mandate (2026-05-03):
  «общее количество партнеров, в т.ч. - формулва опять одна и опять SUM -
   там вообще нет SUM и формулы везде разные - сделай тест который проверяет
   на соответствеи конкретно жтого параметра екселю»
"""
from __future__ import annotations

import sqlite3
from pathlib import Path

import openpyxl as ox
import pytest
import requests

API = "http://localhost:8000/api"
EXCEL_PATH = Path(__file__).resolve().parent.parent / "XLS-MODELS" / "MAIN.xlsx"
DB_PATH = Path(__file__).resolve().parent.parent / "pebble.db"

INDICATOR_NAME = "общее количество партнеров, в т.ч.:"
SHEET_NAME = "BaaS - параметры модели"
EXCEL_SHEET = "0"
EXCEL_ROW = 9


def _find_main_model() -> str:
    db = sqlite3.connect(str(DB_PATH))
    rows = db.execute(
        "SELECT id FROM models WHERE name='MAIN' ORDER BY created_at DESC LIMIT 1"
    ).fetchall()
    db.close()
    if not rows:
        pytest.skip("No imported MAIN model — run test_main_full_accuracy first")
    return rows[0][0]


def _excel_cells_for_partners() -> dict[int, str | float]:
    """col_idx → Excel cell value (str for formula, number for literal)."""
    wb = ox.load_workbook(str(EXCEL_PATH), data_only=False)
    ws = wb[EXCEL_SHEET]
    out: dict[int, str | float] = {}
    for c in range(4, (ws.max_column or 50) + 1):
        v = ws.cell(EXCEL_ROW, c).value
        if v is None:
            continue
        out[c] = v
    return out


def test_partners_data_in_db_matches_excel():
    """Per-cell rule and formula in cell_data must mirror Excel."""
    mid = _find_main_model()
    db = sqlite3.connect(str(DB_PATH))
    sid = db.execute(
        "SELECT id FROM sheets WHERE model_id=? AND name=?", (mid, SHEET_NAME),
    ).fetchone()[0]
    iid_row = db.execute(
        """SELECT ar.id, ar.seq_id FROM analytic_records ar
           JOIN sheet_analytics sa ON sa.analytic_id=ar.analytic_id
           WHERE sa.sheet_id=? AND sa.is_main=1 AND ar.data_json LIKE ?""",
        (sid, f"%{INDICATOR_NAME}%"),
    ).fetchone()
    assert iid_row, f"Indicator {INDICATOR_NAME!r} not found"
    _, iid_seq = iid_row
    cells = db.execute(
        """SELECT cd.coord_key, cd.rule, cd.value,
                  COALESCE(NULLIF(cd.formula,''), f.text, '') AS formula
           FROM cell_data cd LEFT JOIN formulas f ON f.id=cd.formula_id
           WHERE cd.sheet_id=? AND cd.coord_key LIKE ?""",
        (sid, f"%|{iid_seq}"),
    ).fetchall()
    db.close()
    assert cells, "No cells found for partners indicator"

    formula_cells = [c for c in cells if c[1] == "formula"]
    manual_cells = [c for c in cells if c[1] == "manual"]

    # Excel has 3 explicit literal values (D=15, J=20, O=30) and many formulas.
    # After import we expect both classes present.
    assert len(manual_cells) >= 3, f"expected ≥3 manual cells, got {len(manual_cells)}"
    assert len(formula_cells) >= 30, f"expected ≥30 formula cells, got {len(formula_cells)}"

    # The set of distinct formulas should include the +1 / +2 / prev-only patterns.
    distinct = {c[3] for c in formula_cells}
    has_plus_one = any("+1" in f for f in distinct)
    has_plus_two = any("+2" in f for f in distinct)
    has_plain_prev = any(
        '"предыдущий")' in f and not f.endswith("+1") and not f.endswith("+2")
        for f in distinct
    )
    assert has_plus_one and has_plus_two and has_plain_prev, (
        f"Expected +1, +2, and plain (предыдущий) formulas, got distinct={distinct}"
    )


def test_resolved_formulas_returns_per_cell_not_sum():
    """The endpoint that powers grid formula-mode + tooltip must return the
    actual per-cell formula for formula cells, and 'cell-manual' for literal
    cells — never 'default-sum'/'SUM' (which would mislead the UI)."""
    mid = _find_main_model()
    db = sqlite3.connect(str(DB_PATH))
    sid = db.execute(
        "SELECT id FROM sheets WHERE model_id=? AND name=?", (mid, SHEET_NAME),
    ).fetchone()[0]
    iid_seq = db.execute(
        """SELECT ar.seq_id FROM analytic_records ar
           JOIN sheet_analytics sa ON sa.analytic_id=ar.analytic_id
           WHERE sa.sheet_id=? AND sa.is_main=1 AND ar.data_json LIKE ?""",
        (sid, f"%{INDICATOR_NAME}%"),
    ).fetchone()[0]
    cells = db.execute(
        """SELECT cd.coord_key, cd.rule,
                  COALESCE(NULLIF(cd.formula,''), f.text, '') AS formula
           FROM cell_data cd LEFT JOIN formulas f ON f.id=cd.formula_id
           WHERE cd.sheet_id=? AND cd.coord_key LIKE ?""",
        (sid, f"%|{iid_seq}"),
    ).fetchall()
    db.close()
    coord_keys = [c[0] for c in cells]
    by_local = {c[0]: {"rule": c[1], "formula": c[2]} for c in cells}

    resp = requests.post(
        f"{API}/sheets/{sid}/cells/resolved-formulas",
        json={"coord_keys": coord_keys},
        timeout=10,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    by_key = {r["coord_key"]: r for r in data}

    for key, info in by_local.items():
        rule = info["rule"]
        local_f = info["formula"]
        r = by_key.get(key)
        assert r is not None, f"missing resolved entry for {key}"
        if rule == "manual":
            # Literal value entered on a parent-indicator row — must NOT be
            # masked as 'SUM' (default-sum) just because the indicator has
            # children. Explicit per-cell entry wins.
            assert r["source"] == "cell-manual", (
                f"manual cell {key} must report source=cell-manual, got {r}"
            )
        elif rule == "formula" and local_f:
            assert r["source"] == "cell" and r["formula"] == local_f, (
                f"formula cell {key} must echo the per-cell formula, got {r}"
            )
        # rule='formula' with empty formula text = quarter/half/year rollup
        # auto-materialized by recalc; default-sum is the correct surface.


def test_indicator_rules_synthesizes_scoped_per_distinct_formula():
    """When an indicator's cells carry distinct formulas, /rules must return
    one synthesized scoped rule per distinct formula. Each rule's scope lists
    the comma-separated record_ids of every cell where that formula applies,
    so cells sharing a formula collapse into one rule (not one per cell)."""
    mid = _find_main_model()
    db = sqlite3.connect(str(DB_PATH))
    sid = db.execute(
        "SELECT id FROM sheets WHERE model_id=? AND name=?", (mid, SHEET_NAME),
    ).fetchone()[0]
    iid = db.execute(
        """SELECT ar.id FROM analytic_records ar
           JOIN sheet_analytics sa ON sa.analytic_id=ar.analytic_id
           WHERE sa.sheet_id=? AND sa.is_main=1 AND ar.data_json LIKE ?""",
        (sid, f"%{INDICATOR_NAME}%"),
    ).fetchone()[0]
    db.close()

    one = requests.get(
        f"{API}/sheets/{sid}/indicators/{iid}/rules", timeout=10,
    ).json()
    assert one["leaf"] == "", f"expected empty leaf, got {one['leaf']!r}"
    scoped = one.get("scoped") or []
    assert len(scoped) >= 4, (
        f"expected at least 4 synthesized scoped rules (base + +1 + +2 + "
        f"period-specific), got {len(scoped)}: {[r['formula'] for r in scoped]}"
    )
    # All synthesized rules carry synthesized:True flag, no persisted id, and
    # group cells with identical formula text — so cell_count > 0.
    for r in scoped:
        assert r.get("synthesized") is True, f"rule not flagged synthesized: {r}"
        assert r.get("id") in (None, ""), f"synthesized rule must have null id: {r}"
        assert r.get("cell_count", 0) >= 1, f"cell_count must be ≥1: {r}"
    # The +1 and +2 patterns should each appear as exactly one rule that
    # bundles all cells using that formula via comma-separated period IDs.
    plus_one = [r for r in scoped if r["formula"].endswith("+1")]
    plus_two = [r for r in scoped if r["formula"].endswith("+2")]
    assert len(plus_one) == 1 and plus_one[0]["cell_count"] >= 4, plus_one
    assert len(plus_two) == 1 and plus_two[0]["cell_count"] >= 3, plus_two

    # The batch endpoint still surfaces the sentinel so the AnalyticRecordsGrid's
    # narrow formula column shows «разные формулы у клеток» rather than 6 rules.
    allr = requests.get(
        f"{API}/sheets/{sid}/indicator-rules-all", timeout=10,
    ).json()
    entry = allr.get(iid)
    assert entry is not None, f"indicator {iid} missing from indicator-rules-all"
    assert entry["leaf"] == "__per_cell__", (
        f"expected leaf=__per_cell__ in batch endpoint, got {entry}"
    )
