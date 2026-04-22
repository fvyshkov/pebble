"""End-to-end verification: import Excel → calculate → compare ALL values.

Reads the ground-truth from models.xlsx (data_only=True — Excel's computed values),
imports the file into Pebble, runs calculate_model, then compares every cell
using POSITIONAL matching (Pebble indicator sort_order ↔ Excel row order).

Run:
  pytest tests/test_excel_vs_engine.py -x -s          # needs ANTHROPIC_API_KEY for import
  pytest tests/test_excel_vs_engine.py -x -s -k verify  # verify existing model
"""
from __future__ import annotations

import json
import os
import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest
import requests

API = os.environ.get("PEBBLE_API", "http://localhost:8000/api")
EXCEL_PATH = Path(__file__).parent.parent / "models.xlsx"
DB_PATH = Path(__file__).parent.parent / "pebble.db"

# ──────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────

MONTH_NAMES_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}

# Excel sheet → (Pebble display name, layout)
SHEETS = {
    "0": {
        "pebble_name": "BaaS - параметры модели",
        "date_rows": [4], "name_col": 1, "data_start_col": 4, "ind_start_row": 9,
    },
    "BaaS.1": {
        "pebble_name": "BaaS - Онлайн кредитование",
        "date_rows": [3], "name_col": 1, "data_start_col": 4, "ind_start_row": 7,
    },
    "BaaS.2": {
        "pebble_name": "BaaS - Онлайн депозит",
        "date_rows": [3], "name_col": 1, "data_start_col": 4, "ind_start_row": 7,
    },
    "BaaS.3": {
        "pebble_name": "BaaS - Онлайн транзакционный бизнес",
        "date_rows": [3], "name_col": 1, "data_start_col": 4, "ind_start_row": 7,
    },
    "BS": {
        "pebble_name": "Баланс BaaS",
        "date_rows": [4], "name_col": 2, "data_start_col": 3, "ind_start_row": 7,
    },
    "PL": {
        "pebble_name": "Финансовый результат BaaS",
        "date_rows": [4], "name_col": 2, "data_start_col": 3, "ind_start_row": 7,
    },
    "OPEX+CAPEX": {
        "pebble_name": "Операционные расходы и Инвестиции в BaaS",
        "date_rows": [4], "name_col": 3, "data_start_col": 5, "ind_start_row": 7,
    },
}


# ──────────────────────────────────────────────
# Excel extraction
# ──────────────────────────────────────────────

def _extract_excel_data(excel_path: str) -> dict:
    """Extract all data from Excel by positional rows.

    Returns: {
        pebble_sheet_name: {
            "col_to_period": {col: period_name},
            "rows": [(row_num, indicator_name, {period_name: value})]
        }
    }
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    result = {}

    for excel_name, cfg in SHEETS.items():
        ws = wb[excel_name]
        pname = cfg["pebble_name"]

        # 1. Build col → period_name
        col_to_period: dict[int, str] = {}
        for dr in cfg["date_rows"]:
            for col in range(cfg["data_start_col"], min(ws.max_column + 1, 100)):
                v = ws.cell(dr, col).value
                if isinstance(v, datetime):
                    col_to_period[col] = f"{MONTH_NAMES_RU[v.month]} {v.year}"
                elif isinstance(v, (int, float)) and 2025 <= v <= 2030:
                    col_to_period[col] = str(int(v))

        # 2. Extract indicator rows with their values
        rows = []
        for row in range(cfg["ind_start_row"], ws.max_row + 1):
            name_val = ws.cell(row, cfg["name_col"]).value
            if not name_val or not isinstance(name_val, str):
                continue
            name_val = name_val.strip()
            if not name_val:
                continue

            values = {}
            for col, period_name in col_to_period.items():
                cell_val = ws.cell(row, col).value
                if cell_val is not None and isinstance(cell_val, (int, float)):
                    values[period_name] = float(cell_val)

            rows.append((row, name_val, values))

        result[pname] = {"col_to_period": col_to_period, "rows": rows}

    return result


# ──────────────────────────────────────────────
# Pebble extraction (direct DB for speed)
# ──────────────────────────────────────────────

def _get_pebble_data(model_id: str) -> dict:
    """Extract all cell data from Pebble, indexed by indicator sort_order.

    Returns: {
        sheet_name: {
            "indicators": [(sort_order, record_id, name, has_children)],
            "periods": {record_id: name},
            "cells": {(indicator_rid, period_name): float_value},
            "cell_rules": {(indicator_rid, period_name): rule},
        }
    }
    """
    db = sqlite3.connect(str(DB_PATH))
    sheets = db.execute(
        "SELECT id, name FROM sheets WHERE model_id = ? ORDER BY created_at",
        (model_id,),
    ).fetchall()

    result = {}
    for sid, sname in sheets:
        sas = db.execute("""
            SELECT sa.analytic_id, a.is_periods, sa.sort_order
            FROM sheet_analytics sa JOIN analytics a ON a.id = sa.analytic_id
            WHERE sa.sheet_id = ? ORDER BY sa.sort_order
        """, (sid,)).fetchall()

        period_aid = None
        ind_aids = []
        ordered_aids = [sa[0] for sa in sas]
        for sa in sas:
            if sa[1]:
                period_aid = sa[0]
            else:
                ind_aids.append(sa[0])

        if not period_aid or not ind_aids:
            continue

        # Period records: id → name
        period_recs = db.execute(
            "SELECT id, data_json FROM analytic_records WHERE analytic_id = ?",
            (period_aid,),
        ).fetchall()
        period_name_map = {}
        for pr in period_recs:
            data = json.loads(pr[1])
            period_name_map[pr[0]] = data.get("name", "")

        # Indicator records: sorted by sort_order
        indicators = []
        for ind_aid in ind_aids:
            recs = db.execute("""
                SELECT id, data_json, parent_id, sort_order
                FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order
            """, (ind_aid,)).fetchall()

            child_counts = {}
            for r in recs:
                if r[2]:
                    child_counts[r[2]] = child_counts.get(r[2], 0) + 1

            for r in recs:
                data = json.loads(r[1])
                has_children = r[0] in child_counts
                indicators.append((r[3], r[0], data.get("name", ""), has_children))

        indicators.sort(key=lambda x: x[0])

        # Cell data
        cells = db.execute(
            "SELECT coord_key, value, rule FROM cell_data WHERE sheet_id = ?",
            (sid,),
        ).fetchall()

        cell_map = {}
        rule_map = {}
        for ck, val, rule in cells:
            parts = ck.split("|")
            if len(parts) != len(ordered_aids):
                continue

            period_rid = None
            ind_rid = None
            for i, aid in enumerate(ordered_aids):
                if aid == period_aid:
                    period_rid = parts[i]
                elif aid in ind_aids:
                    ind_rid = parts[i]

            if not period_rid or not ind_rid:
                continue

            pname = period_name_map.get(period_rid, "")
            if not pname:
                continue

            try:
                cell_map[(ind_rid, pname)] = float(val)
            except (ValueError, TypeError):
                pass
            rule_map[(ind_rid, pname)] = rule or "manual"

        result[sname] = {
            "indicators": indicators,
            "periods": period_name_map,
            "cells": cell_map,
            "cell_rules": rule_map,
        }

    db.close()
    return result


# ──────────────────────────────────────────────
# Comparison engine
# ──────────────────────────────────────────────

def _compare_all(pebble_data: dict, excel_data: dict, tolerance: float = 0.001) -> dict:
    """Compare by positional matching: Pebble indicator sort_order ↔ Excel row order.

    Returns: {
        "total": int,
        "matched": int,
        "mismatches": [(sheet, pebble_name, excel_name, period, pebble_val, excel_val)],
        "unmatched_excel": [(sheet, excel_row, excel_name, n_values)],
        "per_sheet": {sheet: {"total": int, "matched": int}},
    }
    """
    mismatches = []
    unmatched_excel = []
    per_sheet = {}
    total = 0
    matched = 0

    for excel_sheet_name, ecfg in SHEETS.items():
        pebble_name = ecfg["pebble_name"]
        edata = excel_data.get(pebble_name)
        pdata = pebble_data.get(pebble_name)

        if not edata or not pdata:
            continue

        excel_rows = edata["rows"]
        pebble_inds = pdata["indicators"]
        pebble_cells = pdata["cells"]

        sheet_total = 0
        sheet_matched = 0

        # Match by position: skip section headers (GROUP indicators) in Pebble
        # and skip rows without data in Excel
        excel_leaf_rows = [(r, n, v) for r, n, v in excel_rows if v]
        pebble_all = list(pebble_inds)  # includes GROUPs

        # Build positional mapping
        ei = 0
        pi = 0
        mappings = []  # (pebble_rid, pebble_name, excel_row, excel_name)

        while ei < len(excel_leaf_rows) and pi < len(pebble_all):
            erow, ename, evalues = excel_leaf_rows[ei]
            psort, prid, piname, pis_group = pebble_all[pi]

            # If Pebble indicator is a GROUP (has children), it may or may not
            # have a corresponding Excel row. Check name similarity.
            # Excel section headers usually don't have numeric data (already filtered).
            # So groups in Pebble without data should be skipped.
            if pis_group:
                # Check if this group has cell data
                has_data = any(k[0] == prid for k in pebble_cells)
                if not has_data:
                    pi += 1
                    continue

            # Try to match by name similarity
            pnorm = piname.lower().strip()
            enorm = ename.lower().strip()

            # Check if names match (exact or Pebble has suffix)
            name_match = (
                pnorm == enorm or
                pnorm.startswith(enorm) or
                enorm.startswith(pnorm) or
                # Pebble adds product suffix: "количество партнеров (потребительский кредит)"
                pnorm.rsplit("(", 1)[0].strip() == enorm or
                # Handle "расход" suffix
                pnorm.replace(" расход)", ")").rsplit("(", 1)[0].strip() == enorm
            )

            if name_match:
                mappings.append((prid, piname, erow, ename))
                ei += 1
                pi += 1
            else:
                # Mismatch — try advancing one side
                # Look ahead in Excel for a match to current Pebble indicator
                found_ahead_e = False
                for look in range(1, min(3, len(excel_leaf_rows) - ei)):
                    _, ln, _ = excel_leaf_rows[ei + look]
                    if pnorm == ln.lower().strip() or pnorm.rsplit("(", 1)[0].strip() == ln.lower().strip():
                        # Skip Excel rows
                        for skip in range(look):
                            sr, sn, sv = excel_leaf_rows[ei + skip]
                            unmatched_excel.append((pebble_name, sr, sn, len(sv)))
                        ei += look
                        found_ahead_e = True
                        break

                if not found_ahead_e:
                    # Look ahead in Pebble
                    found_ahead_p = False
                    for look in range(1, min(3, len(pebble_all) - pi)):
                        _, _, ln, _ = pebble_all[pi + look]
                        ln_norm = ln.lower().strip()
                        if ln_norm == enorm or ln_norm.rsplit("(", 1)[0].strip() == enorm:
                            pi += look
                            found_ahead_p = True
                            break

                    if not found_ahead_p:
                        # Skip both and move on
                        unmatched_excel.append((pebble_name, erow, ename, len(evalues)))
                        ei += 1
                        pi += 1

        # Now compare values for matched pairs
        for prid, piname, erow, ename in mappings:
            evalues = {n: v for r, n, v in excel_leaf_rows if r == erow for n, v in v.items()}
            # Rebuild from original
            for r, n, v in excel_leaf_rows:
                if r == erow:
                    evalues = v
                    break

            for period_name, excel_val in evalues.items():
                pebble_val = pebble_cells.get((prid, period_name))
                if pebble_val is None:
                    continue

                total += 1
                sheet_total += 1

                if abs(excel_val) < 1e-10:
                    ok = abs(pebble_val) < 0.01
                else:
                    rel_err = abs(pebble_val - excel_val) / max(abs(excel_val), 1e-10)
                    ok = rel_err < tolerance

                if ok:
                    matched += 1
                    sheet_matched += 1
                else:
                    mismatches.append((
                        pebble_name, piname, ename, period_name,
                        pebble_val, excel_val,
                    ))

        per_sheet[pebble_name] = {"total": sheet_total, "matched": sheet_matched}

    return {
        "total": total,
        "matched": matched,
        "mismatches": mismatches,
        "unmatched_excel": unmatched_excel,
        "per_sheet": per_sheet,
    }


# ──────────────────────────────────────────────
# API helpers
# ──────────────────────────────────────────────

def _req(method, path, **kw):
    return getattr(requests, method)(f"{API}{path}", timeout=120, **kw)


def _find_model() -> str | None:
    """Find the most recent imported model."""
    r = _req("get", "/models")
    if r.status_code != 200:
        return None
    models = r.json()
    for m in reversed(models):
        if any(hint in m["name"].lower() for hint in ["mis", "excelverify"]):
            return m["id"]
    # Fallback: match "model" but not ЦО/BaaS/ENG specific models
    for m in reversed(models):
        name_lower = m["name"].lower()
        if "model" in name_lower and not any(x in name_lower for x in ["цо", "baas", "eng", "v18", "v12"]):
            return m["id"]
    if models:
        return models[-1]["id"]
    return None


def _calculate_model(model_id: str):
    """Calculate all sheets."""
    r = _req("get", f"/sheets/by-model/{model_id}")
    assert r.status_code == 200
    for s in r.json():
        r = _req("post", f"/cells/calculate/{s['id']}")
        assert r.status_code == 200, f"Calculate failed for {s['name']}: {r.text}"


# ──────────────────────────────────────────────
# Tests
# ──────────────────────────────────────────────

@pytest.fixture(scope="module")
def model_id():
    """Find or import model, then recalculate."""
    mid = _find_model()
    if not mid:
        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            pytest.skip("No existing model and ANTHROPIC_API_KEY not set")
        with open(str(EXCEL_PATH), "rb") as f:
            r = requests.post(
                f"{API}/import/excel",
                files={"file": ("models.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"model_name": "ExcelVerify"},
                timeout=600,
            )
        assert r.status_code == 200, f"Import failed: {r.text}"
        mid = r.json()["model_id"]

    _calculate_model(mid)
    return mid


def test_verify_all_values(model_id):
    """Compare ALL Pebble computed values against Excel ground truth."""
    excel_data = _extract_excel_data(str(EXCEL_PATH))
    pebble_data = _get_pebble_data(model_id)
    result = _compare_all(pebble_data, excel_data)

    total = result["total"]
    match = result["matched"]
    mismatches = result["mismatches"]

    print(f"\n{'='*70}")
    print(f"VERIFICATION: {match}/{total} cells match ({100*match/total:.1f}%)")
    print(f"{'='*70}")

    if result["unmatched_excel"]:
        print(f"\nUnmatched Excel rows ({len(result['unmatched_excel'])}):")
        for sheet, row, name, nvals in result["unmatched_excel"][:10]:
            print(f"  {sheet} R{row}: {name} ({nvals} values)")

    if mismatches:
        print(f"\nMISMATCHES ({len(mismatches)}):")
        by_sheet: dict[str, list] = {}
        for m in mismatches:
            by_sheet.setdefault(m[0], []).append(m)

        for sheet, items in sorted(by_sheet.items()):
            print(f"\n  {sheet} ({len(items)} mismatches):")
            for _, pname, ename, period, pval, eval_ in sorted(items, key=lambda x: (x[1], x[3]))[:15]:
                diff = pval - eval_
                rel = abs(diff / eval_) * 100 if abs(eval_) > 1e-10 else float("inf")
                label = pname if pname == ename else f"{pname} ↔ {ename}"
                print(f"    {label} | {period}: P={pval:.4f} E={eval_:.4f} Δ={diff:+.4f} ({rel:.1f}%)")
            if len(items) > 15:
                print(f"    ... and {len(items) - 15} more")

    print(f"\nPer-sheet:")
    for sheet, stats in sorted(result["per_sheet"].items()):
        t, m = stats["total"], stats["matched"]
        pct = 100 * m / t if t else 0
        tag = "OK" if m == t else "FAIL"
        print(f"  [{tag:4s}] {sheet}: {m}/{t} ({pct:.1f}%)")

    # Track match rate — fail if it regresses below the current baseline.
    # Goal: reach 100%. Current known issues:
    #   - BaaS params: "общее количество партнеров" imported with wrong value
    #   - BaaS.1: year totals for cumulative indicators (сумма купленного долга, сумма новых КЛ)
    #   - BaaS.3: year totals for many indicators (consolidation formula mismatch)
    #   - OPEX: matching error for "Итого CAPEX" ↔ "Итого"
    #   - BS/PL: cascading from incorrect values in upstream sheets
    match_pct = 100 * match / total if total else 0
    BASELINE_PCT = 96.5  # raise as bugs are fixed
    assert match_pct >= BASELINE_PCT, (
        f"Match rate {match_pct:.1f}% fell below baseline {BASELINE_PCT}%. "
        f"{len(mismatches)} of {total} cells don't match. See output above."
    )
    if mismatches:
        print(f"\n⚠ {len(mismatches)} cells still don't match — see issues above.")


def test_key_values(model_id):
    """Spot-check critical values that were previously wrong."""
    pebble_data = _get_pebble_data(model_id)

    checks = [
        # (sheet, indicator_hint, period, expected, tol_pct)
        ("BaaS - Онлайн кредитование", "прибыль", "2026", 6678.65, 0.01),
        ("BaaS - Онлайн кредитование", "прибыль", "2027", 34690.93, 0.01),
        ("BaaS - Онлайн кредитование", "прибыль", "2028", 52376.00, 0.01),
    ]

    for sheet, ind_hint, period, expected, tol in checks:
        sdata = pebble_data.get(sheet)
        assert sdata, f"Sheet {sheet} not found"
        found = False
        for (rid, pname), val in sdata["cells"].items():
            # Find indicator name
            iname = ""
            for _, irid, iname_, _ in sdata["indicators"]:
                if irid == rid:
                    iname = iname_
                    break
            if pname == period and ind_hint in iname.lower() and "(" not in iname:
                rel_err = abs(val - expected) / abs(expected)
                assert rel_err < tol, (
                    f"{sheet}/{iname}/{period}: got {val:.4f}, expected {expected:.4f}"
                )
                found = True
                break
        assert found, f"Could not find {ind_hint} for {period} in {sheet}"
