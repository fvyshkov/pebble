"""Deterministic structure extractor for BaaS-style and OPEX+CAPEX-style sheets.
Reads MAIN.xlsx, prints inferred indicator tree for review.

Two detection profiles:
  A. "section header" — bold + dark fill row anchors a section; rows below
     until next anchor are its children.
  B. "A.B.C with indent" — name = (A, B, C) joined; bold C is a parent within
     the current product group, indent in C deepens the level.
"""
from __future__ import annotations
import openpyxl
import sys
from pathlib import Path

XL = Path(__file__).parent.parent / "XLS-MODELS" / "MAIN.xlsx"


def _is_dark_fill(cell) -> bool:
    fg = cell.fill.fgColor
    if fg is None:
        return False
    if fg.type == "rgb" and fg.rgb:
        try:
            r = int(fg.rgb[-6:-4], 16)
            g = int(fg.rgb[-4:-2], 16)
            b = int(fg.rgb[-2:], 16)
            return (r + g + b) < 200
        except ValueError:
            return False
    if fg.type == "theme":
        # Themes 1, 8 are typically dark
        return fg.theme in (1, 8)
    if fg.type == "indexed":
        return fg.indexed in (1, 8, 64)
    return False


def _row_label(ws, row: int) -> tuple[str, str, str]:
    return (
        (ws.cell(row, 1).value or "").strip() if isinstance(ws.cell(row, 1).value, str) else (str(ws.cell(row, 1).value) if ws.cell(row, 1).value is not None else ""),
        (ws.cell(row, 2).value or "").strip() if isinstance(ws.cell(row, 2).value, str) else (str(ws.cell(row, 2).value) if ws.cell(row, 2).value is not None else ""),
        (ws.cell(row, 3).value or "").strip() if isinstance(ws.cell(row, 3).value, str) else (str(ws.cell(row, 3).value) if ws.cell(row, 3).value is not None else ""),
    )


def _row_is_data(ws, row: int) -> bool:
    """A row is 'data' if any of cols 1-3 has content AND there's a numeric/formula in col 5+."""
    a, b, c = _row_label(ws, row)
    if not (a or b or c):
        return False
    for col in range(5, 12):
        v = ws.cell(row, col).value
        if v is not None:
            return True
    return False


def parse_section_anchored(ws, header_rows: int = 6) -> list[dict]:
    """Profile A: detect dark-fill bold rows as section roots, the rest as children.
    Suitable for BaaS - Онлайн депозит etc."""
    max_row = ws.max_row or 200
    out: list[dict] = []
    current_section: dict | None = None

    for row in range(header_rows + 1, max_row + 1):
        a, b, c = _row_label(ws, row)
        first_text = a or b or c
        if not first_text:
            continue
        cell = ws.cell(row, 1)
        if cell.value is None:
            cell = ws.cell(row, 2)
        is_anchor = (cell.font and cell.font.bold) and _is_dark_fill(cell)
        if is_anchor:
            current_section = {"row": row, "name": first_text, "children": []}
            out.append(current_section)
            continue
        # Data row under current section
        unit = b
        # Skip rows with no values and no formulas
        has_value = False
        for col in range(5, 12):
            if ws.cell(row, col).value is not None:
                has_value = True
                break
        if not has_value:
            continue
        leaf = {"row": row, "name": a or c, "unit": unit, "children": []}
        if current_section is None:
            out.append(leaf)
        else:
            current_section["children"].append(leaf)
    return out


def parse_abc_indent(ws, header_rows: int = 6) -> list[dict]:
    """Profile B: name = A.B.C; hierarchy via bold(C) + indent(C); product change in A starts new group.
    Suitable for OPEX+CAPEX."""
    max_row = ws.max_row or 200
    roots: list[dict] = []
    # stack[i] = parent at depth i
    stack: list[dict] = []

    last_a = ""
    last_b = ""

    for row in range(header_rows + 1, max_row + 1):
        a, b, c = _row_label(ws, row)
        if not (a or b or c):
            continue
        # Track propagated A/B (Excel often only fills A in the first row of the block)
        if a:
            last_a = a
        if b:
            last_b = b
        c_cell = ws.cell(row, 3)
        bold = bool(c_cell.font and c_cell.font.bold)
        indent = int(c_cell.alignment.indent or 0) if c_cell.alignment else 0
        # Build full name parts: include A if present, B if present, then C
        name_parts: list[str] = []
        if a:
            name_parts.append(a)
        if b:
            name_parts.append(b)
        if c:
            name_parts.append(c)
        if not name_parts:
            continue
        name = ". ".join(name_parts)
        # depth heuristic:
        #   - if bold(C) AND indent==0  → root within current product
        #   - else                       → child, depth = indent + (1 if bold above) ...
        #   simpler:
        #     - product-level group when A or B differs and bold
        #     - else depth = indent (nesting under nearest bold ancestor or product header)
        node = {"row": row, "name": name, "raw": (a, b, c), "bold": bold, "indent": indent, "children": []}

        if not stack:
            roots.append(node)
            stack.append(node)
            continue

        # Pop stack to find a parent whose "level" is shallower than this one's level.
        # level: bold and indent==0 → 0; otherwise indent+1 (so any non-bold goes under bold)
        def lvl(n):
            return 0 if (n["bold"] and n["indent"] == 0) else n["indent"] + 1
        my_lvl = lvl(node)
        while stack and lvl(stack[-1]) >= my_lvl:
            stack.pop()
        if not stack:
            roots.append(node)
        else:
            stack[-1]["children"].append(node)
        stack.append(node)
    return roots


def dump_tree(nodes: list[dict], indent: int = 0):
    for n in nodes:
        marker = "■" if indent == 0 else "·"
        unit = f" [{n.get('unit', '')}]" if n.get("unit") else ""
        bold = " *bold*" if n.get("bold") else ""
        ind_info = f" indent={n['indent']}" if "indent" in n and n["indent"] else ""
        print(f"  {' ' * indent}{marker} r{n['row']:3d} {n['name']}{unit}{bold}{ind_info}")
        if n.get("children"):
            dump_tree(n["children"], indent + 2)


def main():
    wb = openpyxl.load_workbook(XL, data_only=False)
    target = sys.argv[1] if len(sys.argv) > 1 else "BaaS.2"
    profile = sys.argv[2] if len(sys.argv) > 2 else "section"
    ws = wb[target]
    print(f"=== Sheet {target!r} (profile: {profile}) ===")
    if profile == "section":
        tree = parse_section_anchored(ws)
    else:
        tree = parse_abc_indent(ws)
    dump_tree(tree)


if __name__ == "__main__":
    main()
