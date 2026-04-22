"""Re-import and verify ENG model: delete old → import → recalc → compare with Excel."""
from __future__ import annotations
import json, os, re, sqlite3, sys, time, requests
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from pathlib import Path
from collections import defaultdict
import openpyxl

API = os.environ.get("PEBBLE_API", "http://localhost:8000/api")
EXCEL_PATH = Path("/Users/mac/pebble/XLS-MODELS/ANNEX 1 Simply Ecosystem FinModel 2025-2029_ENG Final.xlsx")
DB_PATH = Path("/Users/mac/pebble/pebble.db")
OLD_MODEL_ID = "a8fef989-e3d8-43dc-953e-2a90f6c6047c"

# ── Step 1: Delete old model via API ──
def delete_old():
    resp = requests.delete(f"{API}/models/{OLD_MODEL_ID}")
    if resp.status_code in (200, 204):
        print(f"Deleted model {OLD_MODEL_ID}")
    elif resp.status_code == 404:
        print("Model not found, skip delete")
    else:
        # Try direct DB delete
        conn = sqlite3.connect(DB_PATH)
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("DELETE FROM models WHERE id = ?", (OLD_MODEL_ID,))
        conn.commit()
        conn.close()
        print(f"Deleted model {OLD_MODEL_ID} via DB")

# ── Step 2: Import via streaming endpoint ──
def import_model():
    with open(EXCEL_PATH, "rb") as f:
        resp = requests.post(f"{API}/import/excel-stream", files={"file": (EXCEL_PATH.name, f)}, stream=True)
    model_id = None
    for line in resp.iter_lines(decode_unicode=True):
        if not line or not line.startswith("data: "):
            continue
        data = json.loads(line[6:])
        if data.get("type") == "progress":
            step = data.get("step", "")
            detail = data.get("detail", "")
            pct = data.get("percent", "")
            print(f"  [{step}] {detail} {pct}%")
        if data.get("done"):
            model_id = data.get("model_id")
            print(f"Import done. model_id={model_id}")
            break
    return model_id

# ── Step 3: Recalc (3 rounds) ──
def recalc(model_id):
    conn = sqlite3.connect(DB_PATH)
    sheet_ids = [r[0] for r in conn.execute("SELECT id FROM sheets WHERE model_id = ? ORDER BY sort_order", (model_id,)).fetchall()]
    conn.close()
    print(f"Sheets: {len(sheet_ids)}")
    for round_n in range(1, 4):
        for sid in sheet_ids:
            resp = requests.post(f"{API}/cells/calculate/{sid}")
            if resp.status_code != 200:
                print(f"  Recalc error sheet {sid}: {resp.status_code}")
        print(f"Recalc round {round_n} done")

# ── Step 4: Compare with Excel using positional matching ──
def compare(model_id):
    from backend.routers.import_excel import _detect_periods_from_headers

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    db = sqlite3.connect(str(DB_PATH))

    sheets = db.execute(
        "SELECT id, name, excel_code FROM sheets WHERE model_id = ? ORDER BY sort_order",
        (model_id,),
    ).fetchall()

    # Global Y0 detection across all sheets (same as import logic)
    _global_has_y0 = False
    for esn in wb.sheetnames:
        ws_scan = wb[esn]
        max_col_scan = min(ws_scan.max_column or 1, 200)
        for r in range(1, 21):
            for c2 in range(1, max_col_scan + 1):
                v = ws_scan.cell(r, c2).value
                if isinstance(v, str) and re.match(r'^[YГг]\s*0\b', v.strip()):
                    _global_has_y0 = True
                    break
            if _global_has_y0:
                break
        if _global_has_y0:
            break
    print(f"  Global Y0 detected: {_global_has_y0}")

    total_match = total_cells = 0
    mismatch_categories = defaultdict(int)
    mismatch_details = defaultdict(list)

    for sid, sname, excel_code in sheets:
        # Match by excel_code (original Excel sheet name)
        excel_sheet = None
        candidates = [excel_code, sname] if excel_code else [sname]
        for cand in candidates:
            for esn in wb.sheetnames:
                if esn.strip() == cand.strip():
                    excel_sheet = esn
                    break
            if excel_sheet:
                break
        if not excel_sheet:
            print(f"  SKIP {sname} (xl={excel_code}): no matching Excel sheet")
            continue

        ws = wb[excel_sheet]

        # Get sheet_analytics to find period and indicator analytics
        sas = db.execute("""
            SELECT sa.analytic_id, a.is_periods, sa.sort_order
            FROM sheet_analytics sa JOIN analytics a ON a.id = sa.analytic_id
            WHERE sa.sheet_id = ? ORDER BY sa.sort_order
        """, (sid,)).fetchall()

        period_aid = None
        ind_aids = []
        ordered_aids = [sa[0] for sa in sas]
        for sa in sas:
            if sa[1]:
                period_aid = sa[0]
            else:
                ind_aids.append(sa[0])

        if not period_aid or not ind_aids:
            continue

        # Period records: id → period_key
        period_recs = db.execute(
            "SELECT id, data_json FROM analytic_records WHERE analytic_id = ?",
            (period_aid,),
        ).fetchall()
        period_key_map = {}  # period_rid -> period_key
        for pr in period_recs:
            data = json.loads(pr[1])
            pk = data.get("period_key", "")
            if pk:
                period_key_map[pr[0]] = pk

        # Indicator records: with excel_row
        indicators = []
        for ind_aid in ind_aids:
            recs = db.execute("""
                SELECT id, data_json, excel_row, sort_order
                FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order
            """, (ind_aid,)).fetchall()
            for r in recs:
                if r[2] is not None:  # has excel_row
                    data = json.loads(r[1])
                    indicators.append({
                        "rid": r[0],
                        "name": data.get("name", ""),
                        "excel_row": r[2],
                    })

        # Cell data
        cells = db.execute(
            "SELECT coord_key, value, rule FROM cell_data WHERE sheet_id = ?",
            (sid,),
        ).fetchall()
        cell_map = {}  # (indicator_rid, period_rid) -> float
        for ck, val, rule in cells:
            parts = ck.split("|")
            if len(parts) != len(ordered_aids):
                continue
            period_rid = ind_rid = None
            for i, aid in enumerate(ordered_aids):
                if aid == period_aid:
                    period_rid = parts[i]
                elif aid in ind_aids:
                    ind_rid = parts[i]
            if period_rid and ind_rid and val is not None:
                try:
                    cell_map[(ind_rid, period_rid)] = float(val)
                except (ValueError, TypeError):
                    pass

        # Detect Excel periods — use global Y0 detection for consistent base_year
        max_col = min(ws.max_column or 1, 200)
        # Import uses: start_year from period_config + 1 if global _has_y0
        # For ENG model with Y0=2024: _nmes_base_year = 2024 + 1 = 2025
        base_year = 2025 if _global_has_y0 else 2024
        detected = _detect_periods_from_headers(ws, max_col, base_year=base_year)
        col_to_pk = {sp["col"]: sp["period_key"] for sp in detected}

        # Build period_key → period_rid lookup
        pk_to_rid = {}
        for prid, pk in period_key_map.items():
            pk_to_rid[pk] = prid

        sheet_match = sheet_total = 0
        for ind in indicators:
            row = ind["excel_row"]
            ind_rid = ind["rid"]
            for col, pk in col_to_pk.items():
                xl_val = ws.cell(row, col).value
                if xl_val is None:
                    continue
                try:
                    xl_num = float(xl_val)
                except (ValueError, TypeError):
                    continue

                p_rid = pk_to_rid.get(pk)
                if not p_rid:
                    continue

                peb_num = cell_map.get((ind_rid, p_rid), 0.0)

                total_cells += 1
                sheet_total += 1

                if abs(xl_num) < 1e-9 and abs(peb_num) < 1e-9:
                    total_match += 1
                    sheet_match += 1
                elif abs(xl_num) > 1e-9 and abs(peb_num - xl_num) / abs(xl_num) < 0.001:
                    total_match += 1
                    sheet_match += 1
                else:
                    if abs(peb_num) < 1e-9 and abs(xl_num) > 1e-9:
                        cat = "peb_zero"
                    elif abs(xl_num) < 1e-9 and abs(peb_num) > 1e-9:
                        cat = "peb_nonzero_xl_zero"
                    else:
                        cat = "both_nonzero_differ"
                    mismatch_categories[cat] += 1
                    if len(mismatch_details[sname]) < 3:
                        mismatch_details[sname].append(
                            (row, col, xl_num, peb_num, cat, ind["name"], pk))

        pct = (sheet_match / sheet_total * 100) if sheet_total > 0 else 100
        status = "OK" if pct > 99.9 else "  "
        print(f"  {status} {sname}: {sheet_match}/{sheet_total} ({pct:.1f}%)")

    db.close()

    pct = (total_match / total_cells * 100) if total_cells > 0 else 0
    print(f"\nTOTAL: {total_match}/{total_cells} ({pct:.1f}%)")
    print(f"Mismatches: {dict(mismatch_categories)}")

    if mismatch_details:
        print("\nSample mismatches:")
        for sn, details in sorted(mismatch_details.items()):
            if details:
                print(f"  {sn}:")
                for row, col, xl, peb, cat, name, pk in details:
                    print(f"    row={row} col={col} pk={pk} [{name}] xl={xl:.6f} peb={peb:.6f} ({cat})")


if __name__ == "__main__":
    if "--compare-only" in sys.argv:
        # Just compare existing model
        model_id = OLD_MODEL_ID
        # Check if model exists
        conn = sqlite3.connect(DB_PATH)
        r = conn.execute("SELECT id FROM models WHERE id = ?", (model_id,)).fetchone()
        conn.close()
        if not r:
            print(f"Model {model_id} not found")
            sys.exit(1)
        print(f"=== Comparing existing model {model_id} ===")
        compare(model_id)
    else:
        print("=== Deleting old ENG model ===")
        delete_old()
        print("\n=== Importing ENG model ===")
        model_id = import_model()
        if not model_id:
            print("IMPORT FAILED")
            sys.exit(1)
        print(f"\n=== Recalculating ({model_id}) ===")
        recalc(model_id)
        print("\n=== Comparing with Excel ===")
        compare(model_id)
