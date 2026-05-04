"""End-to-end accuracy test for MAIN.xlsx import + recalc.

This is the comprehensive regression test demanded by the user (2026-05-03):
  «дорабатай тест чтобы он проверял все ячейки после импорта и пересчета,
   в том числе и формулы. то есть текущая ситуация где в PL другие цифры -
   тест не должен проходить»

Steps performed by `test_full_accuracy`:
  1. Re-import MAIN.xlsx into a fresh model via the streaming API
  2. Run 3 rounds of full-model recalc on every sheet
  3. Compare every (excel_row × period_key) cell on every sheet against
     the original Excel (data_only) using the same logic as
     tests/compare_excel_exact.py — manual cells AND formula cells.
  4. Assert 100% match. If even one cell diverges, dump the first 30
     mismatches and FAIL.

Pre-requisite: backend running on http://localhost:8000 (start.py).
Run:  pytest tests/test_main_full_accuracy.py -s
"""
from __future__ import annotations

import json
import os
import sqlite3
import sys
import time
from pathlib import Path

import pytest
import requests

# Reuse exactly the comparison logic of compare_excel_exact.py — that script
# is the source of truth for "cell-for-cell match" and is the same logic the
# user runs interactively.
sys.path.insert(0, str(Path(__file__).resolve().parent))
import compare_excel_exact as cmp  # noqa: E402

API = os.environ.get("PEBBLE_API", "http://localhost:8000/api")
EXCEL_PATH = Path(__file__).resolve().parent.parent / "XLS-MODELS" / "MAIN.xlsx"
DB_PATH = Path(__file__).resolve().parent.parent / "pebble.db"


def _delete_existing_main():
    db = sqlite3.connect(str(DB_PATH))
    rows = db.execute("SELECT id FROM models WHERE name='MAIN'").fetchall()
    db.close()
    for (mid,) in rows:
        r = requests.delete(f"{API}/models/{mid}")
        print(f"[setup] deleted prior MAIN {mid}: HTTP {r.status_code}")


def _import_main() -> str:
    print(f"[import] streaming {EXCEL_PATH.name} → {API}/import/excel-stream")
    with open(EXCEL_PATH, "rb") as f:
        resp = requests.post(
            f"{API}/import/excel-stream",
            files={"file": (EXCEL_PATH.name, f)},
            stream=True,
            timeout=900,
        )
    model_id = None
    last_step = None
    for line in resp.iter_lines(decode_unicode=True):
        if not line or not line.startswith("data: "):
            continue
        data = json.loads(line[6:])
        if data.get("type") == "progress":
            step = data.get("step", "")
            if step != last_step:
                print(f"  [progress] {step}")
                last_step = step
        if data.get("type") == "error":
            pytest.fail(f"Import error: {data}")
        if data.get("done"):
            model_id = data.get("model_id")
            print(f"[import] done — model_id={model_id}")
            break
    assert model_id, "Import did not return a model_id"
    return model_id


def _recalc(model_id: str, rounds: int = 3):
    db = sqlite3.connect(str(DB_PATH))
    sheet_ids = [r[0] for r in db.execute(
        "SELECT id FROM sheets WHERE model_id=? ORDER BY sort_order", (model_id,)
    ).fetchall()]
    db.close()
    print(f"[recalc] {len(sheet_ids)} sheets × {rounds} rounds")
    for r in range(1, rounds + 1):
        t0 = time.perf_counter()
        for sid in sheet_ids:
            resp = requests.post(f"{API}/cells/calculate/{sid}", timeout=300)
            assert resp.status_code == 200, \
                f"recalc round {r} failed for sheet {sid}: {resp.status_code} {resp.text[:200]}"
        print(f"  round{r} done ({time.perf_counter() - t0:.1f}s)")


def _compare_all_cells(model_id: str, tolerance: float = 0.001) -> tuple[int, int, dict]:
    """Returns (matched, total, mismatches_by_sheet)."""
    excel = cmp.load_excel(str(EXCEL_PATH))
    pebble = cmp.load_pebble(model_id)
    sheet_map = cmp.match_sheets(pebble, excel)

    matched = total = 0
    mis: dict[str, list] = {}
    for psname, etitle in sheet_map.items():
        ws = excel[etitle]["ws"]
        period_cols = excel[etitle]["period_cols"]
        pmap: dict[tuple[int, str], tuple] = {}
        ind_names_by_row: dict[int, str] = {}
        for excel_row, pk, val, rule, _, ind_name in pebble[psname]:
            try:
                pv = float(val) if val not in ("", None) else None
            except (ValueError, TypeError):
                pv = None
            pmap[(excel_row, pk)] = (pv, rule or "", "", ind_name)
            ind_names_by_row.setdefault(excel_row, ind_name)

        sh_mis: list = []
        max_row = ws.max_row or 200
        for excel_row in range(2, max_row + 1):
            label_cells = [ws.cell(excel_row, c).value for c in (1, 2, 3)]
            label = next((str(v).strip() for v in label_cells
                          if isinstance(v, str) and v.strip()), "")
            if not label:
                continue
            for pk, ecol in period_cols.items():
                excel_val = ws.cell(excel_row, ecol).value
                if not isinstance(excel_val, (int, float)):
                    continue
                ev = float(excel_val)
                pv_tuple = pmap.get((excel_row, pk))
                if pv_tuple is None:
                    pebble_val = None
                    rule = "EMPTY" if excel_row in ind_names_by_row else "MISSING_IND"
                    ind_name = ind_names_by_row.get(excel_row, label)
                else:
                    pebble_val, rule, _, ind_name = pv_tuple
                    if pebble_val is None:
                        rule = rule or "EMPTY"
                total += 1
                if pebble_val is None:
                    sh_mis.append((ind_name, excel_row, pk, pebble_val, ev, rule))
                    continue
                if abs(ev) < 1e-10:
                    ok = abs(pebble_val) < 0.01
                else:
                    rel = abs(pebble_val - ev) / max(abs(ev), 1e-10)
                    ok = rel < tolerance
                if ok:
                    matched += 1
                else:
                    sh_mis.append((ind_name, excel_row, pk, pebble_val, ev, rule))
        if sh_mis:
            mis[psname] = sh_mis
    return matched, total, mis


def _compare_cell_rules(model_id: str) -> tuple[int, int, dict]:
    """For every Excel cell that holds a formula (=...), check that the
    matching Pebble cell has a rule mirroring it (formula / sum_children /
    consolidation), or — for total columns — an entry in
    indicator_formula_rules.

    Cells where Pebble has no record (header rows, date rows, non-indicator
    cells) are skipped — that's not a demotion, those rows simply aren't
    indicator data.
    Returns (rule_matched, formula_total, demotions_by_sheet).
    """
    import openpyxl as _ox
    wb_f = _ox.load_workbook(str(EXCEL_PATH), data_only=False)
    excel = cmp.load_excel(str(EXCEL_PATH))
    pebble = cmp.load_pebble(model_id)
    sheet_map = cmp.match_sheets(pebble, excel)

    db = sqlite3.connect(str(DB_PATH))
    consol_by_sheet_row: dict[tuple[str, int], bool] = {}
    for r in db.execute(
        """SELECT s.id AS sid, ar.excel_row
           FROM indicator_formula_rules ifr
           JOIN analytic_records ar ON ar.id = ifr.indicator_id
           JOIN sheets s ON s.id = ifr.sheet_id
           WHERE s.model_id = ? AND ifr.kind IN ('consolidation','formula')""",
        (model_id,),
    ).fetchall():
        consol_by_sheet_row[(r[0], r[1])] = True
    sheet_id_by_pname = {
        r[0]: r[1] for r in db.execute(
            "SELECT name, id FROM sheets WHERE model_id = ?", (model_id,),
        ).fetchall()
    }
    db.close()

    matched = formula_total = 0
    demotions: dict[str, list] = {}
    for psname, etitle in sheet_map.items():
        ws_data = excel[etitle]["ws"]
        sheetname = ws_data.title
        if sheetname not in wb_f.sheetnames:
            continue
        ws_f = wb_f[sheetname]
        period_cols = excel[etitle]["period_cols"]
        rules_by_key: dict[tuple[int, str], str] = {}
        ind_names: dict[int, str] = {}
        for excel_row, pk, val, rule, _, ind_name in pebble[psname]:
            rules_by_key[(excel_row, pk)] = rule or ""
            ind_names.setdefault(excel_row, ind_name)

        psid = sheet_id_by_pname.get(psname)
        sh_dem: list = []
        for excel_row in range(2, (ws_f.max_row or 200) + 1):
            for pk, ecol in period_cols.items():
                fval = ws_f.cell(excel_row, ecol).value
                if not (isinstance(fval, str) and fval.startswith("=")):
                    continue
                if (excel_row, pk) not in rules_by_key:
                    # Pebble doesn't import this row as an indicator cell —
                    # that's a header / date / non-data row, not a demotion.
                    continue
                formula_total += 1
                pebble_rule = rules_by_key[(excel_row, pk)]
                if pebble_rule in ("formula", "sum_children", "consolidation"):
                    matched += 1
                    continue
                # Total/year-rollup columns store their formula on the
                # indicator_formula_rules table instead of cell_data.rule.
                if psid and consol_by_sheet_row.get((psid, excel_row)):
                    matched += 1
                    continue
                sh_dem.append((excel_row, pk, fval[:80],
                               pebble_rule or "(missing)",
                               ind_names.get(excel_row, "")))
        if sh_dem:
            demotions[psname] = sh_dem
    return matched, formula_total, demotions


def test_cell_rules_match_excel_formulas():
    """For every cell where Excel has a formula, Pebble must mirror that
    with rule=formula (or sum_children/consolidation for parent/total rows).
    Demotions to 'manual' are bugs — they break propagation when dependencies
    change."""
    health = requests.get(f"{API.rstrip('/api')}/api/health", timeout=5)
    assert health.status_code == 200, "Backend not responding on /api/health"

    db = sqlite3.connect(str(DB_PATH))
    rows = db.execute(
        """SELECT m.id FROM models m JOIN sheets s ON s.model_id=m.id
           WHERE s.name='BaaS - параметры модели'
           ORDER BY m.created_at DESC LIMIT 1"""
    ).fetchall()
    db.close()
    if not rows:
        pytest.skip("No imported MAIN model — run test_full_accuracy first")
    mid = rows[0][0]

    matched, total, demotions = _compare_cell_rules(mid)
    pct = (matched / total * 100) if total else 0.0
    print(f"\nRULE RESULT: {matched}/{total} Excel-formula cells kept rule "
          f"in Pebble ({pct:.4f}%)")
    if demotions:
        print("Demotions to manual / missing:")
        shown = 0
        for psname, lst in demotions.items():
            if shown >= 30:
                break
            print(f"  [{psname}] {len(lst)} demotions:")
            for row, pk, f, r, name in lst[:8]:
                print(f"    r{row:3d} {pk}  {name[:40]:40s}  {f[:60]}  → {r}")
                shown += 1
                if shown >= 30:
                    break

    assert total > 0, "No Excel formula cells found — bad sheet mapping"
    assert matched == total, (
        f"{total - matched} Excel-formula cells got demoted to manual on "
        f"model {mid}. Each demotion breaks recalc propagation. Fix the "
        f"import path so these stay rule=formula."
    )


def test_full_accuracy_after_import_and_recalc():
    """Comprehensive: import MAIN.xlsx, recalc, every cell on every sheet
    must match Excel data_only — manual values AND computed formulas."""
    assert EXCEL_PATH.exists(), f"{EXCEL_PATH} not found"

    # Sanity: backend up
    health = requests.get(f"{API.rstrip('/api')}/api/health", timeout=5)
    assert health.status_code == 200, "Backend not responding on /api/health"

    _delete_existing_main()
    mid = _import_main()
    _recalc(mid, rounds=3)
    matched, total, mis = _compare_all_cells(mid)

    pct = (matched / total * 100) if total else 0.0
    print(f"\nRESULT: {matched}/{total} cells match ({pct:.4f}%)\n")
    if mis:
        print("Top mismatches:")
        shown = 0
        for psname, lst in mis.items():
            if shown >= 30:
                break
            print(f"  [{psname}] {len(lst)} mismatches:")
            for name, row, pk, pv, ev, rule in lst[:10]:
                pv_s = "(empty)" if pv is None else f"{pv:.4f}"
                print(f"    r{row:3d} {pk} {name[:60]:60s} | P={pv_s:>12} E={ev:14.4f} [{rule}]")
                shown += 1
                if shown >= 30:
                    break

    assert total > 0, "No cells compared — sheet mapping or period detection failed"
    assert matched == total, (
        f"{total - matched} cells diverge between Excel and Pebble on model "
        f"{mid}. Run `python3 tests/compare_excel_exact.py {EXCEL_PATH} {mid}` "
        f"for the full diff. Test must FAIL until 100% match is achieved."
    )
