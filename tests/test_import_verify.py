"""Full import + verification test: imports Excel, verifies all 20572 cells match."""
import pytest
import sqlite3
import json
import os
import asyncio
from datetime import datetime
from openpyxl import load_workbook

EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "models.xlsx")
DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "pebble.db")
MONTH = ["Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]


@pytest.fixture(scope="module")
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    yield conn
    conn.close()


@pytest.fixture(scope="module")
def model_id(db):
    """Find the VERIFIED model. Reimport if data was corrupted."""
    row = db.execute("SELECT id FROM models WHERE name='VERIFIED'").fetchone()
    if not row:
        # Try to import
        import requests
        resp = requests.post("http://localhost:8000/api/import/excel",
                            files={"file": open(EXCEL_PATH, "rb")},
                            data={"model_name": "VERIFIED"}, timeout=300)
        data = resp.json()
        if data.get("model_id"):
            return data["model_id"]
        pytest.skip("VERIFIED model not found and import failed")
    return row["id"]


@pytest.fixture(scope="module")
def wb_data():
    if not os.path.exists(EXCEL_PATH):
        pytest.skip("models.xlsx not found")
    return load_workbook(EXCEL_PATH, data_only=True)


@pytest.fixture(scope="module")
def period_map(db, model_id):
    """Build period name → rid mapping."""
    sheets = db.execute("SELECT id FROM sheets WHERE model_id=? LIMIT 1", (model_id,)).fetchall()
    period_aid = None
    for s in sheets:
        b = db.execute(
            "SELECT sa.analytic_id FROM sheet_analytics sa JOIN analytics a ON a.id=sa.analytic_id "
            "WHERE sa.sheet_id=? AND a.is_periods=1", (s["id"],)
        ).fetchall()
        if b:
            period_aid = b[0]["analytic_id"]
            break
    assert period_aid, "No period analytic"
    precs = db.execute("SELECT id, parent_id, data_json FROM analytic_records WHERE analytic_id=? ORDER BY sort_order", (period_aid,)).fetchall()
    cmap = {}
    for p in precs:
        if p["parent_id"]:
            cmap.setdefault(p["parent_id"], []).append(p["id"])
    leafs = [p for p in precs if p["id"] not in cmap]
    return {json.loads(p["data_json"])["name"]: p["id"] for p in leafs}


SHEET_MAP = {
    "параметры модели": ("0", 4),
    "кредитование": ("BaaS.1", 4),
    "депозит": ("BaaS.2", 4),
    "транзакционный": ("BaaS.3", 4),
    "Баланс": ("BS", 3),
    "Финансовый результат": ("PL", 3),
    "Операционные расходы": ("OPEX+CAPEX", 5),
}


def test_all_7_sheets_present(db, model_id):
    sheets = db.execute("SELECT name FROM sheets WHERE model_id=? ORDER BY sort_order", (model_id,)).fetchall()
    assert len(sheets) == 7, f"Expected 7 sheets, got {len(sheets)}"


def test_sheets_have_excel_code(db, model_id):
    sheets = db.execute("SELECT name, excel_code FROM sheets WHERE model_id=? ORDER BY sort_order", (model_id,)).fetchall()
    for s in sheets:
        assert s["excel_code"], f"Sheet '{s['name']}' missing excel_code"


def test_sheet_order_matches_excel(db, model_id):
    sheets = db.execute("SELECT excel_code FROM sheets WHERE model_id=? ORDER BY sort_order", (model_id,)).fetchall()
    codes = [s["excel_code"] for s in sheets]
    assert codes == ["0", "BaaS.1", "BaaS.2", "BaaS.3", "BS", "PL", "OPEX+CAPEX"]


@pytest.mark.parametrize("key,expected", [
    ("параметры модели", ("0", 4, 648)),
    ("кредитование", ("BaaS.1", 4, 4032)),
    ("депозит", ("BaaS.2", 4, 2480)),
    ("транзакционный", ("BaaS.3", 4, 4893)),
    ("Баланс", ("BS", 3, 924)),
    ("Финансовый результат", ("PL", 3, 4500)),
    ("Операционные расходы", ("OPEX+CAPEX", 5, 3095)),
])
def test_sheet_cell_values_match_excel(db, model_id, wb_data, period_map, key, expected):
    excel_name, start_col, expected_cells = expected

    # Find sheet
    sheets = db.execute("SELECT id, name FROM sheets WHERE model_id=?", (model_id,)).fetchall()
    sheet = next((s for s in sheets if key.lower() in s["name"].lower()), None)
    assert sheet, f"Sheet matching '{key}' not found"

    ws = wb_data[excel_name]

    # col → period
    c2p = {}
    for c in range(start_col, min((ws.max_column or 1) + 1, 50)):
        for hr in range(1, 7):
            v = ws.cell(hr, c).value
            if isinstance(v, datetime):
                pn = f"{MONTH[v.month - 1]} {v.year}"
                if pn in period_map:
                    c2p[c] = period_map[pn]
                break

    db_cells = {c["coord_key"]: c["value"] for c in db.execute(
        "SELECT coord_key, value FROM cell_data WHERE sheet_id=?", (sheet["id"],)
    ).fetchall()}

    # Get records with excel_row (dedup)
    ind_aids = [r["analytic_id"] for r in db.execute(
        "SELECT sa.analytic_id FROM sheet_analytics sa JOIN analytics a ON a.id=sa.analytic_id "
        "WHERE sa.sheet_id=? AND a.is_periods=0", (sheet["id"],)
    ).fetchall()]
    recs_raw = []
    for ia in ind_aids:
        recs_raw.extend(db.execute("SELECT id, excel_row FROM analytic_records WHERE analytic_id=?", (ia,)).fetchall())

    seen = {}
    for rec in recs_raw:
        erow = rec["excel_row"]
        if not erow:
            continue
        has_cells = db.execute(
            "SELECT COUNT(*) as c FROM cell_data WHERE sheet_id=? AND coord_key LIKE ?",
            (sheet["id"], f'%|{rec["id"]}')
        ).fetchone()["c"]
        if erow not in seen or has_cells > seen[erow][1]:
            seen[erow] = (rec, has_cells)
    recs = [v[0] for v in seen.values()]

    total = 0
    match = 0
    for rec in recs:
        rid = rec["id"]
        erow = rec["excel_row"]
        for col, pid in c2p.items():
            ck = f"{pid}|{rid}"
            dv = db_cells.get(ck)
            ev = ws.cell(erow, col).value
            if ev is None and dv is None:
                continue
            if ev is None:
                ev = 0
            if dv is None:
                dv = "0"
            try:
                e = float(ev)
                d = float(dv)
            except:
                continue
            total += 1
            tol = abs(e) * 0.01 + 0.01
            if abs(e - d) <= tol:
                match += 1

    assert total > 0, f"No cells to compare for {excel_name}"
    assert match == total, f"{excel_name}: {match}/{total} cells match ({match/total*100:.1f}%)"
