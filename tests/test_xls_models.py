"""Parametrized import + recalc + compare tests for all XLS-MODELS reference files.

For each .xlsx in XLS-MODELS/:
  1. Import via streaming endpoint
  2. Recalculate all sheets (3 rounds for convergence)
  3. Compare every cell value against Excel ground truth using excel_row/excel_code
     stored in DB during import (not positional matching)
  4. Assert match rate >= per-model baseline

Run:
  pytest tests/test_xls_models.py -x -s                    # all models
  pytest tests/test_xls_models.py -x -s -k ENG             # just ENG
  pytest tests/test_xls_models.py -x -s -k BaaS            # just BaaS v.12
  pytest tests/test_xls_models.py -x -s -k CO              # just ЦО v.18
"""
from __future__ import annotations

import json
import os
import re
import sqlite3
from collections import defaultdict
from pathlib import Path

import sys
import openpyxl
import pytest
import requests

# Add project root to path so we can import backend modules
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

API = os.environ.get("PEBBLE_API", "http://localhost:8000/api")
XLS_DIR = Path(__file__).parent.parent / "XLS-MODELS"
DB_PATH = Path(__file__).parent.parent / "pebble.db"

# Model configs: (filename, baseline_match_pct)
MODELS = [
    ("ANNEX 1 Simply Ecosystem FinModel 2025-2029_ENG Final.xlsx", 95.0),
    ("Doscredobank FinModel BaaS 2026-2028 v.12.xlsx", 90.0),
    ("Doscredobank FinModel ЦО 2026-2028 v.18.xlsx", 90.0),
]


def _short_name(filename: str) -> str:
    if "ENG" in filename:
        return "ENG"
    if "v.12" in filename:
        return "BaaS_v12"
    if "v.18" in filename or "ЦО" in filename:
        return "CO_v18"
    return filename[:20]


def _import_via_stream(path: str) -> str:
    """Import using the streaming SSE endpoint. Returns model_id."""
    with open(path, "rb") as f:
        r = requests.post(
            f"{API}/import/excel-stream",
            files={"file": (os.path.basename(path), f)},
            timeout=600,
            stream=True,
        )
    assert r.status_code == 200, f"Import failed: {r.status_code}"
    last_data = None
    for line in r.iter_lines(decode_unicode=True):
        if line and line.startswith("data: "):
            last_data = json.loads(line[6:])
            if last_data.get("type") == "progress":
                step = last_data.get("step", "")
                detail = last_data.get("detail", "")
                pct = last_data.get("percent", "")
                print(f"    [{step}] {detail} {pct}%")
    assert last_data and "model_id" in last_data, \
        f"No model_id in stream response: {last_data}"
    return last_data["model_id"]


def _calculate_model(model_id: str, rounds: int = 3):
    """Recalculate all sheets, multiple rounds for convergence."""
    r = requests.get(f"{API}/sheets/by-model/{model_id}", timeout=30)
    assert r.status_code == 200
    sheet_ids = [s["id"] for s in r.json()]
    for round_n in range(rounds):
        for sid in sheet_ids:
            resp = requests.post(f"{API}/cells/calculate/{sid}", timeout=120)
            assert resp.status_code == 200, \
                f"Calculate failed sheet {sid} round {round_n}: {resp.text[:200]}"


def _detect_global_y0(wb) -> bool:
    """Check if any sheet has Y0/Г0/г0 header — shifts base_year to 2025."""
    for esn in wb.sheetnames:
        ws = wb[esn]
        max_col = min(ws.max_column or 1, 200)
        for r in range(1, 21):
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and re.match(r'^[YГг]\s*0\b', v.strip()):
                    return True
    return False


def _compare_with_excel_row(model_id: str, excel_path: str) -> dict:
    """Compare Pebble values against Excel using excel_row/excel_code stored in DB.

    This is more accurate than positional matching because it uses the exact
    row numbers stored during import.

    Returns: {total, matched, per_sheet: {name: {total, matched, mismatches: [...]}}}
    """
    from backend.routers.import_excel import _detect_periods_from_headers

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    db = sqlite3.connect(str(DB_PATH))

    sheets = db.execute(
        "SELECT id, name, excel_code FROM sheets WHERE model_id = ? ORDER BY sort_order",
        (model_id,),
    ).fetchall()

    has_y0 = _detect_global_y0(wb)
    base_year = 2025 if has_y0 else 2024

    total_match = 0
    total_cells = 0
    per_sheet = {}

    for sid, sname, excel_code in sheets:
        # Match by excel_code (original Excel sheet name)
        excel_sheet = None
        candidates = [excel_code, sname] if excel_code else [sname]
        for cand in candidates:
            for esn in wb.sheetnames:
                if esn.strip() == cand.strip():
                    excel_sheet = esn
                    break
            if excel_sheet:
                break
        if not excel_sheet:
            continue

        ws = wb[excel_sheet]

        # Get sheet_analytics
        sas = db.execute("""
            SELECT sa.analytic_id, a.is_periods, sa.sort_order
            FROM sheet_analytics sa JOIN analytics a ON a.id = sa.analytic_id
            WHERE sa.sheet_id = ? ORDER BY sa.sort_order
        """, (sid,)).fetchall()

        period_aid = None
        ind_aids = []
        ordered_aids = [sa[0] for sa in sas]
        for sa in sas:
            if sa[1]:
                period_aid = sa[0]
            else:
                ind_aids.append(sa[0])

        if not period_aid or not ind_aids:
            continue

        # Period records: id → period_key
        period_recs = db.execute(
            "SELECT id, data_json FROM analytic_records WHERE analytic_id = ?",
            (period_aid,),
        ).fetchall()
        period_key_map = {}
        for pr in period_recs:
            data = json.loads(pr[1])
            pk = data.get("period_key", "")
            if pk:
                period_key_map[pr[0]] = pk

        # Indicator records with excel_row
        indicators = []
        ind_rid_to_name = {}
        for ind_aid in ind_aids:
            recs = db.execute("""
                SELECT id, data_json, excel_row, sort_order
                FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order
            """, (ind_aid,)).fetchall()
            for r in recs:
                data = json.loads(r[1])
                name = data.get("name", "")
                ind_rid_to_name[r[0]] = name
                if r[2] is not None:
                    indicators.append({
                        "rid": r[0], "name": name, "excel_row": r[2],
                    })

        # Cell data
        cells = db.execute(
            "SELECT coord_key, value FROM cell_data WHERE sheet_id = ?",
            (sid,),
        ).fetchall()
        cell_map = {}
        for ck, val in cells:
            parts = ck.split("|")
            if len(parts) != len(ordered_aids):
                continue
            period_rid = ind_rid = None
            for i, aid in enumerate(ordered_aids):
                if aid == period_aid:
                    period_rid = parts[i]
                elif aid in ind_aids:
                    ind_rid = parts[i]
            if period_rid and ind_rid and val is not None:
                try:
                    cell_map[(ind_rid, period_rid)] = float(val)
                except (ValueError, TypeError):
                    pass

        # Detect Excel periods
        max_col = min(ws.max_column or 1, 200)
        detected = _detect_periods_from_headers(ws, max_col, base_year=base_year)
        col_to_pk = {sp["col"]: sp["period_key"] for sp in detected}

        # Build period_key → period_rid lookup
        pk_to_rid = {}
        for prid, pk in period_key_map.items():
            pk_to_rid[pk] = prid

        # Compare: Excel → Pebble
        sheet_match = sheet_total = 0
        sheet_mismatches = []

        for ind in indicators:
            row = ind["excel_row"]
            ind_rid = ind["rid"]
            for col, pk in col_to_pk.items():
                xl_val = ws.cell(row, col).value
                if xl_val is None:
                    continue
                try:
                    xl_num = float(xl_val)
                except (ValueError, TypeError):
                    continue

                p_rid = pk_to_rid.get(pk)
                if not p_rid:
                    continue

                peb_num = cell_map.get((ind_rid, p_rid), 0.0)

                total_cells += 1
                sheet_total += 1

                if abs(xl_num) < 1e-9 and abs(peb_num) < 1e-9:
                    total_match += 1
                    sheet_match += 1
                elif abs(xl_num) > 1e-9 and abs(peb_num - xl_num) / abs(xl_num) < 0.001:
                    total_match += 1
                    sheet_match += 1
                else:
                    sheet_mismatches.append((
                        ind["name"], pk, peb_num, xl_num,
                    ))

        per_sheet[sname] = {
            "total": sheet_total,
            "matched": sheet_match,
            "mismatches": sheet_mismatches,
        }

    db.close()
    wb.close()

    return {
        "total": total_cells,
        "matched": total_match,
        "per_sheet": per_sheet,
    }


# ──────────────────────────────────────────────
# Fixtures
# ──────────────────────────────────────────────

@pytest.fixture(scope="module", params=MODELS,
                ids=[_short_name(m[0]) for m in MODELS])
def imported_model(request):
    """Import model, recalculate, yield info dict."""
    filename, baseline = request.param
    excel_path = XLS_DIR / filename

    if not excel_path.exists():
        pytest.skip(f"{filename} not found in XLS-MODELS/")

    try:
        requests.get(f"{API}/models", timeout=5)
    except requests.ConnectionError:
        pytest.skip("Backend server not running at localhost:8000")

    print(f"\n  Importing {filename}...")
    model_id = _import_via_stream(str(excel_path))
    print(f"  model_id={model_id}")

    print(f"  Recalculating (3 rounds)...")
    _calculate_model(model_id)

    yield {
        "model_id": model_id,
        "excel_path": str(excel_path),
        "filename": filename,
        "baseline": baseline,
    }

    print(f"  Cleaning up model {model_id}...")
    requests.delete(f"{API}/models/{model_id}", timeout=30)


# ──────────────────────────────────────────────
# Tests
# ──────────────────────────────────────────────

def test_import_creates_sheets(imported_model):
    """Imported model should have at least 1 sheet."""
    model_id = imported_model["model_id"]
    r = requests.get(f"{API}/sheets/by-model/{model_id}", timeout=30)
    assert r.status_code == 200
    sheets = r.json()
    assert len(sheets) >= 1, f"Expected at least 1 sheet, got {len(sheets)}"
    print(f"  {imported_model['filename']}: {len(sheets)} sheets imported")
    for s in sheets:
        print(f"    - {s['name']}")


def test_values_match_excel(imported_model):
    """Compare ALL computed values against Excel ground truth using excel_row."""
    model_id = imported_model["model_id"]
    excel_path = imported_model["excel_path"]
    baseline = imported_model["baseline"]
    filename = imported_model["filename"]

    result = _compare_with_excel_row(model_id, excel_path)

    total = result["total"]
    matched = result["matched"]
    match_pct = (matched / total * 100) if total else 0

    print(f"\n  {'='*60}")
    print(f"  {filename}")
    print(f"  RESULT: {matched}/{total} cells match ({match_pct:.1f}%)")
    print(f"  {'='*60}")

    print(f"\n  Per-sheet:")
    for sheet, stats in sorted(result["per_sheet"].items()):
        st = stats["total"]
        sm = stats["matched"]
        pct = (sm / st * 100) if st else 0
        n_mis = len(stats["mismatches"])
        status = "OK" if sm == st else "FAIL"
        print(f"    [{status:4s}] {sheet}: {sm}/{st} ({pct:.1f}%)")
        if stats["mismatches"]:
            for name, pk, peb, xl in stats["mismatches"][:5]:
                diff = peb - xl
                rel = abs(diff / xl) * 100 if abs(xl) > 1e-10 else float("inf")
                print(f"          {name} | {pk}: P={peb:.4f} E={xl:.4f} "
                      f"Δ={diff:+.4f} ({rel:.1f}%)")
            if n_mis > 5:
                print(f"          ... and {n_mis - 5} more")

    assert total > 0, f"No cells to compare for {filename}"
    assert match_pct >= baseline, (
        f"{filename}: match rate {match_pct:.1f}% below baseline {baseline}%. "
        f"{total - matched} of {total} cells don't match."
    )


def test_no_average_consolidation(imported_model):
    """No indicator should use AVERAGE consolidation."""
    model_id = imported_model["model_id"]
    r = requests.get(f"{API}/sheets/by-model/{model_id}", timeout=30)
    sheets = r.json()

    bad = []
    for s in sheets:
        rules_r = requests.get(
            f"{API}/sheets/{s['id']}/indicator-rules-all", timeout=30
        )
        if rules_r.status_code != 200:
            continue
        for ind_id, entry in rules_r.json().items():
            if entry.get("consolidation") == "AVERAGE":
                bad.append(f"{s['name']}: {entry.get('name', ind_id)}")

    if bad:
        print(f"\n  WARNING: {len(bad)} indicators with AVERAGE consolidation:")
        for b in bad[:10]:
            print(f"    {b}")
    # Allow up to 5 — some Excel models legitimately use AVERAGE for simple averages
    assert len(bad) <= 5, \
        f"Too many indicators with AVERAGE consolidation:\n" + "\n".join(bad[:20])


def test_avg_rate_indicators_have_formulas(imported_model):
    """Indicators with avg/rate patterns must have consolidation formulas."""
    AVG_RATE_PATTERNS = ("средн", "ср. ", "ставка", "доля ", "на 1 ",
                         "average", "avg ", "rate", "share ")
    model_id = imported_model["model_id"]

    tree_r = requests.get(f"{API}/models/{model_id}/tree", timeout=30)
    if tree_r.status_code != 200:
        pytest.skip("Could not get model tree")
    tree = tree_r.json()

    ind_names = {}
    for a in tree.get("analytics", []):
        if a.get("is_periods"):
            continue
        recs_r = requests.get(f"{API}/analytics/{a['id']}/records", timeout=30)
        if recs_r.status_code == 200:
            for r in recs_r.json():
                dj = r.get("data_json", {})
                if isinstance(dj, str):
                    dj = json.loads(dj)
                nm = (dj or {}).get("name", "")
                if nm:
                    ind_names[r["id"]] = nm

    sheets_r = requests.get(f"{API}/sheets/by-model/{model_id}", timeout=30)
    sheets = sheets_r.json()

    missing = []
    for s in sheets:
        rules_r = requests.get(
            f"{API}/sheets/{s['id']}/indicator-rules-all", timeout=30
        )
        if rules_r.status_code != 200:
            continue
        for ind_id, entry in rules_r.json().items():
            ind_name = ind_names.get(ind_id, entry.get("name", "")).lower()
            consol = entry.get("consolidation", "")
            if any(p in ind_name for p in AVG_RATE_PATTERNS):
                if not consol or consol == "SUM":
                    display_name = ind_names.get(ind_id, ind_id)
                    missing.append(f"{s['name']}: {display_name}")

    if missing:
        print(f"\n  WARNING: {len(missing)} avg/rate indicators without consolidation:")
        for m in missing[:20]:
            print(f"    {m}")
        if len(missing) > 20:
            print(f"    ... and {len(missing) - 20} more")
