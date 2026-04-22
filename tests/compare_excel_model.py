"""Generic Excel ↔ Pebble model comparison.

Given an Excel file path and a Pebble model ID, auto-detects sheet layouts,
extracts ground-truth values from Excel, extracts computed values from Pebble DB,
and compares them using positional matching.

Usage:
  python tests/compare_excel_model.py <excel_path> <model_id> [--recalc]
"""
from __future__ import annotations

import json
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import requests

DB_PATH = Path(__file__).parent.parent / "pebble.db"
API = "http://localhost:8000/api"

MONTH_NAMES_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}

MONTH_NAMES_EN = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}


def _detect_sheet_layout(ws) -> dict | None:
    """Auto-detect layout: date rows, name column, data start, indicator start."""
    max_row = min(ws.max_row or 100, 200)
    max_col = min(ws.max_column or 50, 100)

    # Find date rows (rows containing datetime values in columns 2-50)
    date_rows = []
    date_cols = {}  # col -> period_name
    name_col = None
    data_start_col = None

    for row in range(1, min(10, max_row + 1)):
        dates_in_row = {}
        for col in range(1, max_col + 1):
            v = ws.cell(row, col).value
            if isinstance(v, datetime):
                pname = f"{MONTH_NAMES_RU[v.month]} {v.year}"
                dates_in_row[col] = pname
            elif isinstance(v, (int, float)) and 2024 <= v <= 2035:
                dates_in_row[col] = str(int(v))
        if len(dates_in_row) >= 3:
            date_rows.append(row)
            date_cols.update(dates_in_row)

    if not date_rows:
        return None

    # Data starts at the first date column
    if date_cols:
        data_start_col = min(date_cols.keys())
        # Name column: pick the column before data_start_col with the MOST text
        best_nc = None
        best_count = 0
        for nc in range(1, data_start_col):
            text_count = 0
            for row in range(max(date_rows) + 1, min(max(date_rows) + 30, max_row + 1)):
                v = ws.cell(row, nc).value
                if isinstance(v, str) and v.strip():
                    text_count += 1
            if text_count > best_count:
                best_count = text_count
                best_nc = nc
        if best_nc and best_count >= 2:
            name_col = best_nc

    if not name_col:
        return None

    # Indicator start row = first row after date rows with text in name column
    ind_start_row = max(date_rows) + 1
    for row in range(max(date_rows) + 1, min(max(date_rows) + 10, max_row + 1)):
        v = ws.cell(row, name_col).value
        if isinstance(v, str) and v.strip():
            ind_start_row = row
            break

    return {
        "date_rows": date_rows,
        "name_col": name_col,
        "data_start_col": data_start_col,
        "ind_start_row": ind_start_row,
        "col_to_period": date_cols,
    }


def extract_excel_data(excel_path: str) -> dict:
    """Extract all data from Excel, auto-detecting layouts.

    Returns: {excel_sheet_name: {"col_to_period": {col: name}, "rows": [(row, name, {period: val})]}}
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        layout = _detect_sheet_layout(ws)
        if not layout:
            continue

        col_to_period = layout["col_to_period"]
        rows = []
        for row in range(layout["ind_start_row"], (ws.max_row or 0) + 1):
            name_val = ws.cell(row, layout["name_col"]).value
            if not name_val or not isinstance(name_val, str):
                continue
            name_val = name_val.strip()
            if not name_val:
                continue

            values = {}
            for col, period_name in col_to_period.items():
                cell_val = ws.cell(row, col).value
                if cell_val is not None and isinstance(cell_val, (int, float)):
                    values[period_name] = float(cell_val)

            if values:
                rows.append((row, name_val, values))

        if rows:
            result[sheet_name] = {"col_to_period": col_to_period, "rows": rows}

    wb.close()
    return result


def get_pebble_data(model_id: str) -> dict:
    """Extract all cell data from Pebble DB.

    Returns: {sheet_name: {"indicators": [(sort, rid, name, is_group)], "periods": {rid: name}, "cells": {(rid, period): val}}}
    """
    db = sqlite3.connect(str(DB_PATH))
    sheets = db.execute(
        "SELECT id, name FROM sheets WHERE model_id = ? ORDER BY created_at",
        (model_id,),
    ).fetchall()

    result = {}
    for sid, sname in sheets:
        sas = db.execute("""
            SELECT sa.analytic_id, a.is_periods, sa.sort_order
            FROM sheet_analytics sa JOIN analytics a ON a.id = sa.analytic_id
            WHERE sa.sheet_id = ? ORDER BY sa.sort_order
        """, (sid,)).fetchall()

        period_aid = None
        ind_aids = []
        ordered_aids = [sa[0] for sa in sas]
        for sa in sas:
            if sa[1]:
                period_aid = sa[0]
            else:
                ind_aids.append(sa[0])

        if not period_aid or not ind_aids:
            continue

        period_recs = db.execute(
            "SELECT id, data_json FROM analytic_records WHERE analytic_id = ?",
            (period_aid,),
        ).fetchall()
        period_name_map = {}
        for pr in period_recs:
            data = json.loads(pr[1])
            period_name_map[pr[0]] = data.get("name", "")

        indicators = []
        for ind_aid in ind_aids:
            recs = db.execute("""
                SELECT id, data_json, parent_id, sort_order
                FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order
            """, (ind_aid,)).fetchall()

            child_counts = {}
            for r in recs:
                if r[2]:
                    child_counts[r[2]] = child_counts.get(r[2], 0) + 1

            for r in recs:
                data = json.loads(r[1])
                has_children = r[0] in child_counts
                indicators.append((r[3], r[0], data.get("name", ""), has_children))

        indicators.sort(key=lambda x: x[0])

        cells = db.execute(
            "SELECT coord_key, value, rule FROM cell_data WHERE sheet_id = ?",
            (sid,),
        ).fetchall()

        cell_map = {}
        for ck, val, rule in cells:
            parts = ck.split("|")
            if len(parts) != len(ordered_aids):
                continue

            period_rid = None
            ind_rid = None
            for i, aid in enumerate(ordered_aids):
                if aid == period_aid:
                    period_rid = parts[i]
                elif aid in ind_aids:
                    ind_rid = parts[i]

            if not period_rid or not ind_rid:
                continue

            pname = period_name_map.get(period_rid, "")
            if not pname:
                continue

            try:
                cell_map[(ind_rid, pname)] = float(val)
            except (ValueError, TypeError):
                pass

        result[sname] = {
            "indicators": indicators,
            "periods": period_name_map,
            "cells": cell_map,
        }

    db.close()
    return result


def _name_match(pname: str, ename: str) -> bool:
    pn = pname.lower().strip()
    en = ename.lower().strip()
    return (
        pn == en
        or pn.startswith(en)
        or en.startswith(pn)
        or pn.rsplit("(", 1)[0].strip() == en
        or en.rsplit("(", 1)[0].strip() == pn
    )


def match_sheets(pebble_data: dict, excel_data: dict) -> dict[str, str]:
    """Match Pebble sheet names to Excel sheet names by indicator overlap."""
    mapping = {}
    # Keyword abbreviation map for fuzzy matching
    ABBREV = {
        "рб": "розничный бизнес", "мсб": "малый и средний бизнес",
        "baas": "baas", "bs": "баланс", "pl": "финансовый результат",
        "opex": "операционные расходы", "capex": "инвестиции",
    }

    def _name_hint_match(pname: str, ename: str) -> bool:
        """Check if abbreviation in Excel name matches Pebble name."""
        en = ename.lower().split(".")[0].replace("+", " ").strip()
        for abbr, full in ABBREV.items():
            if abbr in en and full in pname.lower():
                # Also check suffix number match
                esuffix = ename.lower().split(".")[-1] if "." in ename else ""
                psuffix = ""
                if "кредит" in pname.lower():
                    psuffix = "1"
                elif "счета" in pname.lower() or "депозит" in pname.lower():
                    psuffix = "2"
                elif "обслуж" in pname.lower():
                    psuffix = "3"
                if esuffix and psuffix:
                    return esuffix == psuffix
                return True
        return False

    # 1. Direct name match
    for pname in pebble_data:
        for ename in excel_data:
            if pname.lower() == ename.lower():
                mapping[pname] = ename
                break
        if pname not in mapping:
            for ename in excel_data:
                if ename.lower() in pname.lower() or pname.lower() in ename.lower():
                    mapping[pname] = ename
                    break
        # 1b. Abbreviation-based matching
        if pname not in mapping:
            for ename in excel_data:
                if ename not in mapping.values() and _name_hint_match(pname, ename):
                    mapping[pname] = ename
                    break

    # 2. Indicator name overlap matching for remaining
    #    Use global best-match (Hungarian-style greedy) to avoid misordering
    unmatched_p = [p for p in pebble_data if p not in mapping]
    unmatched_e = set(e for e in excel_data if e not in mapping.values())

    if unmatched_p and unmatched_e:
        # Score all pairs
        scores = []
        for pname in unmatched_p:
            pind_names = {ind[2].lower().strip() for ind in pebble_data[pname]["indicators"]}
            for ename in unmatched_e:
                eind_names = {row[1].lower().strip() for row in excel_data[ename]["rows"]}
                overlap = len(pind_names & eind_names)
                if overlap >= 3:
                    scores.append((overlap, pname, ename))
        # Greedy: pick highest overlap first
        scores.sort(reverse=True)
        used_p = set()
        used_e = set()
        for overlap, pname, ename in scores:
            if pname not in used_p and ename not in used_e:
                mapping[pname] = ename
                used_p.add(pname)
                used_e.add(ename)

    return mapping


def compare_all(pebble_data: dict, excel_data: dict, sheet_mapping: dict,
                tolerance: float = 0.001) -> dict:
    """Compare by positional matching."""
    mismatches = []
    per_sheet = {}
    total = 0
    matched = 0

    for pebble_sheet, excel_sheet in sheet_mapping.items():
        edata = excel_data[excel_sheet]
        pdata = pebble_data[pebble_sheet]

        excel_rows = edata["rows"]
        pebble_inds = pdata["indicators"]
        pebble_cells = pdata["cells"]

        sheet_total = 0
        sheet_matched = 0

        # Positional matching
        ei = 0
        pi = 0
        mappings = []

        while ei < len(excel_rows) and pi < len(pebble_inds):
            erow, ename, evalues = excel_rows[ei]
            psort, prid, piname, pis_group = pebble_inds[pi]

            if pis_group:
                has_data = any(k[0] == prid for k in pebble_cells)
                if not has_data:
                    pi += 1
                    continue

            if _name_match(piname, ename):
                mappings.append((prid, piname, erow, ename))
                ei += 1
                pi += 1
            else:
                found = False
                for look in range(1, min(3, len(excel_rows) - ei)):
                    _, ln, _ = excel_rows[ei + look]
                    if _name_match(piname, ln):
                        ei += look
                        found = True
                        break
                if not found:
                    for look in range(1, min(3, len(pebble_inds) - pi)):
                        _, _, ln, _ = pebble_inds[pi + look]
                        if _name_match(ln, ename):
                            pi += look
                            found = True
                            break
                if not found:
                    ei += 1
                    pi += 1

        # Compare values
        for prid, piname, erow, ename in mappings:
            evalues = {}
            for r, n, v in excel_rows:
                if r == erow:
                    evalues = v
                    break

            for period_name, excel_val in evalues.items():
                pebble_val = pebble_cells.get((prid, period_name))
                if pebble_val is None:
                    continue

                total += 1
                sheet_total += 1

                if abs(excel_val) < 1e-10:
                    ok = abs(pebble_val) < 0.01
                else:
                    rel_err = abs(pebble_val - excel_val) / max(abs(excel_val), 1e-10)
                    ok = rel_err < tolerance

                if ok:
                    matched += 1
                    sheet_matched += 1
                else:
                    mismatches.append((
                        pebble_sheet, piname, ename, period_name,
                        pebble_val, excel_val,
                    ))

        per_sheet[pebble_sheet] = {"total": sheet_total, "matched": sheet_matched}

    return {
        "total": total,
        "matched": matched,
        "mismatches": mismatches,
        "per_sheet": per_sheet,
    }


def recalc_model(model_id: str) -> int:
    """Recalculate all sheets, multiple rounds until stable."""
    db = sqlite3.connect(str(DB_PATH))
    sheets = db.execute(
        "SELECT id FROM sheets WHERE model_id = ?", (model_id,)
    ).fetchall()
    db.close()

    total_changes = 0
    for _ in range(5):
        round_changes = 0
        for (sid,) in sheets:
            r = requests.post(f"{API}/cells/calculate/{sid}", timeout=120)
            if r.status_code == 200:
                round_changes += r.json().get("computed", 0)
        total_changes += round_changes
        if round_changes == 0:
            break
    return total_changes


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("excel_path")
    parser.add_argument("model_id")
    parser.add_argument("--recalc", action="store_true")
    parser.add_argument("--tolerance", type=float, default=0.001)
    args = parser.parse_args()

    if args.recalc:
        print("Recalculating model...")
        changes = recalc_model(args.model_id)
        print(f"  {changes} cells recomputed")

    print(f"Extracting Excel data from {args.excel_path}...")
    excel_data = extract_excel_data(args.excel_path)
    print(f"  Found {len(excel_data)} sheets with data: {list(excel_data.keys())}")

    print(f"Extracting Pebble data for model {args.model_id}...")
    pebble_data = get_pebble_data(args.model_id)
    print(f"  Found {len(pebble_data)} sheets: {list(pebble_data.keys())}")

    sheet_mapping = match_sheets(pebble_data, excel_data)
    print(f"\nSheet mapping ({len(sheet_mapping)}):")
    for p, e in sheet_mapping.items():
        print(f"  {p} ↔ {e}")

    unmapped_p = set(pebble_data) - set(sheet_mapping)
    unmapped_e = set(excel_data) - set(sheet_mapping.values())
    if unmapped_p:
        print(f"  Unmapped Pebble sheets: {unmapped_p}")
    if unmapped_e:
        print(f"  Unmapped Excel sheets: {unmapped_e}")

    print("\nComparing values...")
    result = compare_all(pebble_data, excel_data, sheet_mapping, args.tolerance)

    total = result["total"]
    matched_count = result["matched"]
    match_pct = (matched_count / total * 100) if total else 0
    mismatches = result["mismatches"]

    print(f"\n{'='*60}")
    print(f"RESULT: {matched_count}/{total} cells match ({match_pct:.1f}%)")
    print(f"{'='*60}")

    print("\nPer-sheet:")
    for sheet, stats in sorted(result["per_sheet"].items()):
        st = stats["total"]
        sm = stats["matched"]
        pct = (sm / st * 100) if st else 0
        status = "OK" if sm == st else "FAIL"
        print(f"  [{status:4s}] {sheet}: {sm}/{st} ({pct:.1f}%)")

    if mismatches:
        print(f"\nMismatches ({len(mismatches)}):")
        by_sheet = {}
        for sheet, pn, en, period, pv, ev in mismatches:
            by_sheet.setdefault(sheet, []).append((pn, en, period, pv, ev))
        for sheet, items in sorted(by_sheet.items()):
            print(f"\n  {sheet} ({len(items)}):")
            for pn, en, period, pv, ev in items[:10]:
                delta = pv - ev
                pct = abs(delta / ev * 100) if abs(ev) > 1e-10 else 0
                label = f"{pn} ↔ {en}" if pn != en else pn
                print(f"    {label} | {period}: P={pv:.4f} E={ev:.4f} Δ={delta:+.4f} ({pct:.1f}%)")
            if len(items) > 10:
                print(f"    ... and {len(items) - 10} more")

    return 0 if match_pct >= 95 else 1


if __name__ == "__main__":
    sys.exit(main())
