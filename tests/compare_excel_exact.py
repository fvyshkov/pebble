"""Exact Excel ↔ Pebble cell comparison via excel_row + period_key.

Unlike compare_excel_model.py (which uses fuzzy name matching), this script
relies on the linkage stored at import time:
  - analytic_records.excel_row → original Excel row of every indicator record
  - analytic_records.data_json.period_key → period identifier (e.g. "2026-01")

For each Pebble cell on each sheet, we resolve:
  (sheet_title, excel_row, excel_col_for_period)
and compare against the matching Excel cell. This eliminates positional drift
around duplicate indicator names.

Usage: python tests/compare_excel_exact.py <excel_path> <model_id>
"""
from __future__ import annotations

import json
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

DB_PATH = Path(__file__).parent.parent / "pebble.db"


def _read_sheet_title(ws) -> str | None:
    for col in range(1, 7):
        v = ws.cell(1, col).value
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None


def _detect_period_columns(ws) -> dict[str, int]:
    """Find columns whose header (rows 1..6) is a date or "m{N}" key.

    Returns: {"2026-01": col_idx, "2026-Y": col_idx, ...}
    """
    period_cols: dict[str, int] = {}
    max_col = min(ws.max_column or 50, 100)
    for col in range(1, max_col + 1):
        for row in range(1, 8):
            v = ws.cell(row, col).value
            if isinstance(v, datetime):
                period_cols[f"{v.year}-{v.month:02d}"] = col
                break
            elif isinstance(v, (int, float)) and 2024 <= v <= 2035:
                period_cols[f"{int(v)}-Y"] = col
                break
    return period_cols


def load_excel(excel_path: str) -> dict:
    """Returns: {title: {"period_cols": {pk: col}, "ws": ws_obj}}"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    out = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        title = _read_sheet_title(ws) or sname
        out[title] = {"period_cols": _detect_period_columns(ws), "ws": ws}
    return out


def load_pebble(model_id: str) -> dict:
    """Returns: {sheet_title: [(excel_row, period_key, value, rule, formula, ind_name)]}"""
    db = sqlite3.connect(str(DB_PATH))
    db.row_factory = sqlite3.Row

    seq_to_uuid = {
        str(r["seq_id"]): r["id"]
        for r in db.execute(
            "SELECT id, seq_id FROM analytic_records WHERE seq_id IS NOT NULL"
        )
    }

    sheets = db.execute(
        "SELECT id, name FROM sheets WHERE model_id = ? ORDER BY created_at",
        (model_id,),
    ).fetchall()

    out: dict[str, list] = {}
    for sh in sheets:
        sid, sname = sh["id"], sh["name"]
        sas = db.execute(
            """SELECT sa.analytic_id, a.is_periods, sa.is_main
               FROM sheet_analytics sa JOIN analytics a ON a.id = sa.analytic_id
               WHERE sa.sheet_id = ? ORDER BY sa.sort_order""",
            (sid,),
        ).fetchall()
        if not sas:
            continue
        ordered_aids = [s["analytic_id"] for s in sas]
        period_aid = next((s["analytic_id"] for s in sas if s["is_periods"]), None)
        main_aid = next((s["analytic_id"] for s in sas if s["is_main"]), None)
        if not period_aid or not main_aid:
            continue

        # Extra analytics (Подразделения, Версии, etc) — when the user adds a
        # new analytic, sheets.py:_find_first_leaf migrates existing cell data
        # to (..., first_leaf_seq). Replicate that lookup here to project the
        # comparison onto the same terminal combination.
        extra_aids = [a for a in ordered_aids if a != period_aid and a != main_aid]
        first_leaf_seq_for: dict[str, str] = {}
        for aid in extra_aids:
            recs = db.execute(
                "SELECT id, parent_id, seq_id FROM analytic_records WHERE analytic_id = ? ORDER BY sort_order",
                (aid,),
            ).fetchall()
            if not recs:
                continue
            parent_ids = {r["id"] for r in recs if any(c["parent_id"] == r["id"] for c in recs)}
            leaf = next((r for r in recs if r["id"] not in parent_ids), recs[0])
            if leaf["seq_id"] is not None:
                first_leaf_seq_for[aid] = str(leaf["seq_id"])

        # Period uuid → period_key
        period_keys = {}
        for r in db.execute(
            "SELECT id, data_json FROM analytic_records WHERE analytic_id = ?",
            (period_aid,),
        ):
            d = json.loads(r["data_json"])
            pk = d.get("period_key") or ""
            if pk:
                period_keys[r["id"]] = pk

        # Indicator uuid → (excel_row, name, parent_id)
        ind_meta = {}
        for r in db.execute(
            "SELECT id, excel_row, data_json, parent_id FROM analytic_records WHERE analytic_id = ?",
            (main_aid,),
        ):
            n = json.loads(r["data_json"]).get("name", "")
            ind_meta[r["id"]] = (r["excel_row"], n, r["parent_id"])

        # parent_uuid → list[child_uuid] (only indicators with excel_row)
        children_of: dict[str, list[str]] = {}
        for uid, (_, _, par) in ind_meta.items():
            if par:
                children_of.setdefault(par, []).append(uid)

        cells = db.execute(
            "SELECT coord_key, value, rule FROM cell_data WHERE sheet_id = ?",
            (sid,),
        ).fetchall()

        # First pass: index existing cells by (period_uuid, ind_uuid).
        # Filter to cells whose extra-analytic positions match each analytic's
        # first leaf — that's where add_sheet_analytic migrated the original
        # value to.
        cell_by_pk_ind: dict[tuple[str, str], tuple[float | None, str]] = {}
        for c in cells:
            raw_parts = c["coord_key"].split("|")
            if len(raw_parts) != len(ordered_aids):
                continue
            skip = False
            period_uuid = ind_uuid = None
            for i, aid in enumerate(ordered_aids):
                raw = raw_parts[i]
                if aid == period_aid:
                    period_uuid = seq_to_uuid.get(raw, raw)
                elif aid == main_aid:
                    ind_uuid = seq_to_uuid.get(raw, raw)
                else:
                    expected = first_leaf_seq_for.get(aid)
                    if expected is not None and raw != expected:
                        skip = True
                        break
            if skip or not period_uuid or not ind_uuid:
                continue
            try:
                v = float(c["value"]) if c["value"] not in ("", None) else None
            except (ValueError, TypeError):
                v = None
            cell_by_pk_ind[(period_uuid, ind_uuid)] = (v, c["rule"] or "")

        # Compute sum_children aggregates for parents that have no persisted cell.
        # Mirrors backend/routers/cells.py _materialize_sums semantics: for each
        # period, sum descendant leaves into ancestor parents that aren't already
        # set. (sum_children rows are deleted at recalc — the engine recomputes
        # them on demand, so they may be absent from cell_data even when the UI
        # shows a value.)
        all_inds = list(ind_meta.keys())
        for period_uuid in period_keys:
            # Bottom-up walk: for each indicator, if no cell, sum children
            # Process leaves first by sorting by depth descending
            def depth(uid):
                d = 0
                cur = ind_meta.get(uid, (None, None, None))[2]
                while cur and d < 50:
                    d += 1
                    cur = ind_meta.get(cur, (None, None, None))[2]
                return d
            for uid in sorted(all_inds, key=depth, reverse=True):
                if (period_uuid, uid) in cell_by_pk_ind:
                    continue
                kids = children_of.get(uid, [])
                if not kids:
                    continue
                total = 0.0
                any_kid = False
                for k in kids:
                    pv = cell_by_pk_ind.get((period_uuid, k))
                    if pv and pv[0] is not None:
                        total += pv[0]
                        any_kid = True
                if any_kid:
                    cell_by_pk_ind[(period_uuid, uid)] = (total, "sum_children")

        recs = []
        for (period_uuid, ind_uuid), (v, rule) in cell_by_pk_ind.items():
            pk = period_keys.get(period_uuid)
            meta = ind_meta.get(ind_uuid)
            if not pk or not meta or meta[0] is None:
                continue
            recs.append((meta[0], pk, v, rule, "", meta[1]))
        out[sname] = recs

    db.close()
    return out


def _name_match(p: str, e: str) -> bool:
    return p.lower().strip() == e.lower().strip()


def match_sheets(pebble: dict, excel: dict) -> dict[str, str]:
    """Match by exact title — Excel row 1 holds the authoritative title."""
    used = set()
    mapping = {}
    for psname in pebble:
        for etitle in excel:
            if etitle in used:
                continue
            if _name_match(psname, etitle):
                mapping[psname] = etitle
                used.add(etitle)
                break
    return mapping


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("excel_path")
    ap.add_argument("model_id")
    ap.add_argument("--tolerance", type=float, default=0.001)
    ap.add_argument("--show", type=int, default=15, help="mismatches per sheet to print")
    args = ap.parse_args()

    print(f"Loading Excel from {args.excel_path}…")
    excel = load_excel(args.excel_path)
    print(f"  {len(excel)} sheets: {list(excel.keys())}")

    print(f"Loading Pebble model {args.model_id}…")
    pebble = load_pebble(args.model_id)
    print(f"  {len(pebble)} sheets: {list(pebble.keys())}")

    sheet_map = match_sheets(pebble, excel)
    print(f"\nSheet mapping ({len(sheet_map)}):")
    for p, e in sheet_map.items():
        print(f"  {p} ↔ {e}")
    unmapped_p = [p for p in pebble if p not in sheet_map]
    unmapped_e = [e for e in excel if e not in sheet_map.values()]
    if unmapped_p:
        print(f"  Unmapped Pebble: {unmapped_p}")
    if unmapped_e:
        print(f"  Unmapped Excel:  {unmapped_e}")

    total = matched = 0
    per_sheet = {}
    mismatches: dict[str, list] = {}

    for psname, etitle in sheet_map.items():
        pebble_recs = pebble[psname]
        ws = excel[etitle]["ws"]
        period_cols = excel[etitle]["period_cols"]

        # Index Pebble cells by (excel_row, period_key)
        pmap: dict[tuple[int, str], tuple[float | None, str, str, str]] = {}
        ind_names_by_row: dict[int, str] = {}
        for excel_row, pk, val, rule, _, ind_name in pebble_recs:
            try:
                pv = float(val) if val not in ("", None) else None
            except (ValueError, TypeError):
                pv = None
            pmap[(excel_row, pk)] = (pv, rule or "", "", ind_name)
            ind_names_by_row.setdefault(excel_row, ind_name)

        s_total = s_matched = 0
        sh_mis: list = []

        # Iterate over EXCEL as the source of truth: every Excel row that has a
        # text label in cols 1-3 (= an indicator row) and a numeric value in any
        # period column must have a matching Pebble cell. If the row has no
        # corresponding Pebble indicator at all, the cell is MISSING_IND. If the
        # indicator exists but no cell, it's EMPTY.
        max_row = ws.max_row or 200
        for excel_row in range(2, max_row + 1):
            label_cells = [ws.cell(excel_row, c).value for c in (1, 2, 3)]
            label = next((str(v).strip() for v in label_cells
                          if isinstance(v, str) and v.strip()), "")
            if not label:
                continue
            for pk, ecol in period_cols.items():
                excel_val = ws.cell(excel_row, ecol).value
                if not isinstance(excel_val, (int, float)):
                    continue
                ev = float(excel_val)
                pv_tuple = pmap.get((excel_row, pk))
                if pv_tuple is None:
                    pebble_val = None
                    if excel_row in ind_names_by_row:
                        rule = "EMPTY"
                        ind_name = ind_names_by_row[excel_row]
                    else:
                        rule = "MISSING_IND"
                        ind_name = label
                else:
                    pebble_val, rule, _, ind_name = pv_tuple
                    if pebble_val is None:
                        rule = rule or "EMPTY"

                total += 1
                s_total += 1
                if pebble_val is None:
                    ok = False
                elif abs(ev) < 1e-10:
                    ok = abs(pebble_val) < 0.01
                else:
                    rel = abs(pebble_val - ev) / max(abs(ev), 1e-10)
                    ok = rel < args.tolerance

                if ok:
                    matched += 1
                    s_matched += 1
                else:
                    sh_mis.append((ind_name, excel_row, pk, pebble_val, ev, rule))

        per_sheet[psname] = (s_total, s_matched)
        mismatches[psname] = sh_mis

    pct = (matched / total * 100) if total else 0.0
    print(f"\n{'='*60}\nRESULT: {matched}/{total} cells match ({pct:.2f}%)\n{'='*60}\n")
    print("Per-sheet:")
    for ps, (t, m) in sorted(per_sheet.items()):
        flag = "OK  " if t == m and t > 0 else "FAIL"
        sp = (m / t * 100) if t else 0.0
        print(f"  [{flag}] {ps}: {m}/{t} ({sp:.1f}%)")

    print("\nMismatches:")
    for ps, lst in sorted(mismatches.items()):
        if not lst:
            continue
        print(f"\n  {ps} ({len(lst)} mismatches, showing {min(args.show, len(lst))}):")
        for name, row, pk, pv, ev, rule in lst[:args.show]:
            if pv is None:
                pv_str = "      (empty)"
                d_str = f"Δ={-ev:+12.4f}"
                rp_str = "100.0%"
            else:
                d = pv - ev
                pv_str = f"{pv:14.4f}"
                d_str = f"Δ={d:+12.4f}"
                rp = abs(d) / max(abs(ev), 1e-10) * 100 if ev else float('inf')
                rp_str = f"{rp:.1f}%"
            print(f"    r{row:3d} {pk} {name[:50]:50s} | P={pv_str} E={ev:14.4f} {d_str} ({rp_str}) [{rule}]")


if __name__ == "__main__":
    main()
