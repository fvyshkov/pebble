"""End-to-end smoke for MAIN.xlsx — runs the 6 steps of the user scenario.

  1. Import MAIN.xlsx via streaming endpoint
  2. Recalc 3 rounds, compare every cell against Excel ground truth
  3. Pick one manual cell, change value, recalc, verify propagation
  4. Revert that cell, recalc, verify everything is back to original
  5. Add "Подразделения" analytic (standard branches reference) to every sheet,
     fill all manual leaf cells with original_value * (1 ± 20%) per branch leaf
  6. Recalc, verify HEAD totals are bigger than originals

Backend must be running at PEBBLE_API (default http://localhost:8000/api).

Usage:
    python scripts/main_full_flow.py [--keep] [--seed N]

  --keep   : skip teardown so the model stays visible in the UI
  --seed N : random seed for the ±20% noise (default 42)
"""
from __future__ import annotations

import argparse
import json
import os
import random
import sqlite3
import sys
import time
from collections import defaultdict
from pathlib import Path

import openpyxl
import requests

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

API = os.environ.get("PEBBLE_API", "http://localhost:8000/api")
XLS_PATH = ROOT / "XLS-MODELS" / "MAIN.xlsx"
DB_PATH = ROOT / "pebble.db"
BRANCHES_REF = ROOT / "docs" / "branches_reference.json"


# ───────────────────────────── helpers ─────────────────────────────

def _hr(title: str):
    print()
    print("═" * 72)
    print(f"  {title}")
    print("═" * 72)


def _step(n: int, title: str):
    print()
    print(f"── Step {n}: {title} " + "─" * (60 - len(title)))


def _ok(r, ctx=""):
    if r.status_code != 200:
        raise RuntimeError(f"{ctx}: HTTP {r.status_code} — {r.text[:300]}")
    return r


def _import_via_stream(path: Path) -> str:
    with open(path, "rb") as f:
        r = requests.post(
            f"{API}/import/excel-stream",
            files={"file": (path.name, f)},
            timeout=600,
            stream=True,
        )
    if r.status_code != 200:
        raise RuntimeError(f"Import failed: {r.status_code} {r.text[:200]}")
    last = None
    for line in r.iter_lines(decode_unicode=True):
        if line and line.startswith("data: "):
            last = json.loads(line[6:])
            if last.get("type") == "progress":
                step = last.get("step", "")
                detail = last.get("detail", "")
                pct = last.get("percent", "")
                print(f"   [{step}] {detail} {pct}%")
    if not last or "model_id" not in last:
        raise RuntimeError(f"No model_id in stream response: {last}")
    return last["model_id"]


def _calc_model(model_id: str, rounds: int = 3):
    sheets = _ok(requests.get(f"{API}/sheets/by-model/{model_id}", timeout=30),
                 "list sheets").json()
    for r in range(rounds):
        for s in sheets:
            _ok(requests.post(f"{API}/cells/calculate/{s['id']}", timeout=180),
                f"recalc round {r+1} sheet {s['name']}")


def _compare_all_cells(model_id: str, excel_path: Path) -> dict:
    """Returns per-sheet match stats. Reuses the logic from test_xls_models.py."""
    from backend.routers.import_excel import _detect_periods_from_headers

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    db = sqlite3.connect(str(DB_PATH))
    try:
        sheets = db.execute(
            "SELECT id, name, excel_code FROM sheets WHERE model_id = ? ORDER BY sort_order",
            (model_id,),
        ).fetchall()

        total = matched = 0
        per_sheet = {}

        for sid, sname, excel_code in sheets:
            xname = excel_code or sname
            if xname not in wb.sheetnames:
                continue
            ws = wb[xname]

            sas = db.execute("""
                SELECT sa.analytic_id, a.is_periods, sa.sort_order
                FROM sheet_analytics sa JOIN analytics a ON a.id = sa.analytic_id
                WHERE sa.sheet_id = ? ORDER BY sa.sort_order
            """, (sid,)).fetchall()
            period_aid = next((r[0] for r in sas if r[1]), None)
            ind_aids = [r[0] for r in sas if not r[1]]
            ordered_aids = [r[0] for r in sas]
            if not period_aid or not ind_aids:
                continue

            period_recs = db.execute(
                "SELECT id, data_json FROM analytic_records WHERE analytic_id = ?",
                (period_aid,),
            ).fetchall()
            pk_to_rid = {}
            for prid, dj in period_recs:
                d = json.loads(dj)
                pk = d.get("period_key")
                if pk:
                    pk_to_rid[pk] = prid

            indicators = []
            for aid in ind_aids:
                for r in db.execute(
                    "SELECT id, data_json, excel_row FROM analytic_records "
                    "WHERE analytic_id = ? ORDER BY sort_order", (aid,)
                ).fetchall():
                    if r[2] is None:
                        continue
                    nm = json.loads(r[1]).get("name", "")
                    indicators.append({"rid": r[0], "name": nm, "excel_row": r[2]})

            cells = db.execute(
                "SELECT coord_key, value FROM cell_data WHERE sheet_id = ?", (sid,)
            ).fetchall()
            # coord_key parts are seq_id strings — load mapping to translate
            seq_to_uuid = dict(db.execute(
                "SELECT CAST(seq_id AS TEXT), id FROM analytic_records WHERE seq_id IS NOT NULL"
            ).fetchall())
            cell_map = {}
            for ck, val in cells:
                parts = [seq_to_uuid.get(p, p) for p in ck.split("|")]
                if len(parts) != len(ordered_aids):
                    continue
                p_rid = ind_rid = None
                for i, aid in enumerate(ordered_aids):
                    if aid == period_aid:
                        p_rid = parts[i]
                    elif aid in ind_aids:
                        ind_rid = parts[i]
                if p_rid and ind_rid and val is not None:
                    try:
                        cell_map[(ind_rid, p_rid)] = float(val)
                    except (ValueError, TypeError):
                        pass

            max_col = min(ws.max_column or 1, 200)
            detected = _detect_periods_from_headers(ws, max_col, base_year=2024)
            col_to_pk = {sp["col"]: sp["period_key"] for sp in detected}

            sm = st = 0
            mismatches = []
            for ind in indicators:
                row = ind["excel_row"]
                for col, pk in col_to_pk.items():
                    xv = ws.cell(row, col).value
                    if xv is None:
                        continue
                    try:
                        xn = float(xv)
                    except (ValueError, TypeError):
                        continue
                    p_rid = pk_to_rid.get(pk)
                    if not p_rid:
                        continue
                    pv = cell_map.get((ind["rid"], p_rid), 0.0)
                    st += 1
                    total += 1
                    if (abs(xn) < 1e-9 and abs(pv) < 1e-9) or \
                       (abs(xn) > 1e-9 and abs(pv - xn) / abs(xn) < 0.001):
                        sm += 1
                        matched += 1
                    elif len(mismatches) < 5:
                        mismatches.append((ind["name"], pk, pv, xn))
            per_sheet[sname] = {"total": st, "matched": sm, "mismatches": mismatches}
        return {"total": total, "matched": matched, "per_sheet": per_sheet}
    finally:
        db.close()
        wb.close()


def _print_compare(report: dict, label: str):
    t, m = report["total"], report["matched"]
    pct = (m / t * 100) if t else 0.0
    print(f"\n   {label}: {m}/{t} cells match ({pct:.1f}%)")
    bad = [(name, st) for name, st in report["per_sheet"].items()
           if st["matched"] < st["total"]]
    if bad:
        print(f"   {len(bad)} sheet(s) with mismatches:")
        for name, st in bad[:8]:
            print(f"     · {name}: {st['matched']}/{st['total']}")
            for ind, pk, pv, xv in st["mismatches"][:3]:
                print(f"         {ind} | {pk}: pebble={pv:.4f} excel={xv:.4f}")
    return pct


def _all_cells(sheet_id: str) -> dict[str, dict]:
    r = _ok(requests.get(f"{API}/cells/by-sheet/{sheet_id}", timeout=120),
            f"get cells {sheet_id}")
    return {c["coord_key"]: c for c in r.json()}


def _put_cell(sheet_id: str, coord_key: str, value, rule="manual"):
    body = {
        "coord_key": coord_key,
        "value": str(value),
        "data_type": "number",
        "rule": rule,
    }
    _ok(requests.put(f"{API}/cells/by-sheet/{sheet_id}/single",
                     json=body, params={"no_recalc": "true"}, timeout=30),
        f"put cell {coord_key}")


def _put_cells_bulk(sheet_id: str, items: list[dict]):
    """Bulk save without auto-recalc."""
    _ok(requests.put(f"{API}/cells/by-sheet/{sheet_id}",
                     json={"cells": items},
                     params={"no_recalc": "true"},
                     timeout=120),
        f"bulk put {len(items)} cells")


# ───────────────────────────── steps ─────────────────────────────

def step_1_import() -> str:
    _step(1, "Import MAIN.xlsx")
    if not XLS_PATH.exists():
        raise RuntimeError(f"Not found: {XLS_PATH}")
    print(f"   File: {XLS_PATH} ({XLS_PATH.stat().st_size:,} bytes)")
    t0 = time.time()
    model_id = _import_via_stream(XLS_PATH)
    print(f"   model_id = {model_id}  ({time.time() - t0:.1f}s)")
    sheets = _ok(requests.get(f"{API}/sheets/by-model/{model_id}", timeout=30)).json()
    print(f"   {len(sheets)} sheet(s):")
    for s in sheets:
        print(f"     · {s['name']}")
    return model_id


def step_2_recalc_and_verify(model_id: str) -> dict:
    _step(2, "Recalc + verify all cells vs Excel")
    t0 = time.time()
    _calc_model(model_id, rounds=3)
    print(f"   recalc 3 rounds: {time.time() - t0:.1f}s")
    report = _compare_all_cells(model_id, XLS_PATH)
    pct = _print_compare(report, "RESULT")
    if pct < 90.0:
        print(f"   ⚠ match rate {pct:.1f}% below 90% baseline")
    return report


def _pick_test_cell(model_id: str):
    """Pick a manual leaf cell that has at least one downstream formula in the same sheet."""
    db = sqlite3.connect(str(DB_PATH))
    try:
        sheets = db.execute(
            "SELECT id, name FROM sheets WHERE model_id = ? ORDER BY sort_order",
            (model_id,),
        ).fetchall()
        for sid, sname in sheets:
            row = db.execute("""
                SELECT coord_key, value FROM cell_data
                WHERE sheet_id = ? AND rule = 'manual'
                  AND value IS NOT NULL AND value != '' AND value != '0'
                LIMIT 1
            """, (sid,)).fetchone()
            if not row:
                continue
            ck, val = row
            try:
                fval = float(val)
            except (ValueError, TypeError):
                continue
            if abs(fval) < 1e-9:
                continue
            return sid, sname, ck, fval
        raise RuntimeError("No suitable manual cell found")
    finally:
        db.close()


def step_3_change(model_id: str):
    _step(3, "Change one manual cell, verify propagation")
    sid, sname, ck, orig = _pick_test_cell(model_id)
    print(f"   sheet={sname}  coord={ck}  orig={orig}")

    before = _all_cells(sid)
    new_val = orig * 7.5  # big enough to ripple visibly
    _put_cell(sid, ck, new_val)
    _calc_model(model_id, rounds=2)
    after = _all_cells(sid)

    cur = float(after[ck]["value"])
    if abs(cur - new_val) > 1e-6:
        raise RuntimeError(f"Cell did not store new value: got {cur}, expected {new_val}")

    changed = 0
    for k, c in after.items():
        if k == ck:
            continue
        b = before.get(k)
        if not b:
            continue
        try:
            if abs(float(c["value"]) - float(b["value"])) > 1e-6:
                changed += 1
        except (ValueError, TypeError):
            continue
    print(f"   wrote {ck}: {orig} → {new_val}")
    print(f"   {changed} other cell(s) changed value")
    if changed == 0:
        print(f"   ⚠ no propagation detected — DAG/recalc may be broken")
    return sid, ck, orig, before


def step_4_revert(model_id: str, sid: str, ck: str, orig: float, before: dict):
    _step(4, "Revert cell, verify everything is back")
    _put_cell(sid, ck, orig)
    _calc_model(model_id, rounds=2)
    after = _all_cells(sid)

    diffs = []
    for k, b in before.items():
        a = after.get(k)
        if not a:
            diffs.append((k, "missing"))
            continue
        try:
            av = float(a["value"]); bv = float(b["value"])
        except (ValueError, TypeError):
            continue
        if abs(av - bv) > max(1e-6, abs(bv) * 1e-6):
            diffs.append((k, av, bv))
    if not diffs:
        print(f"   ✓ all {len(before)} cells restored")
    else:
        print(f"   ⚠ {len(diffs)} cell(s) did NOT return to original")
        for d in diffs[:5]:
            print(f"     {d}")


def _load_branches_tree() -> list[dict]:
    return json.loads(BRANCHES_REF.read_text())["structure"]


def _create_branches_analytic(model_id: str) -> tuple[str, list[str], str]:
    """Create branches analytic, return (analytic_id, leaf_rids, head_rid)."""
    aid = _ok(requests.post(f"{API}/analytics",
                            json={"model_id": model_id, "name": "Подразделения"},
                            timeout=30), "create branches analytic").json()["id"]

    leaves: list[str] = []
    head_rid: str | None = None

    def walk(node, parent_id=None):
        nonlocal head_rid
        rid = _ok(requests.post(f"{API}/analytics/{aid}/records",
                                json={"data_json": {"name": node["name"]},
                                      "parent_id": parent_id},
                                timeout=30),
                  f"add record {node['name']}").json()["id"]
        if parent_id is None:
            head_rid = rid
        kids = node.get("children") or []
        if not kids:
            leaves.append(rid)
        else:
            for k in kids:
                walk(k, rid)

    for top in _load_branches_tree():
        walk(top)
    return aid, leaves, head_rid


def _bind_branches_to_sheet(sheet_id: str, analytic_id: str):
    sas = _ok(requests.get(f"{API}/sheets/{sheet_id}/analytics", timeout=30)).json()
    next_order = max([s["sort_order"] for s in sas], default=-1) + 1
    _ok(requests.post(f"{API}/sheets/{sheet_id}/analytics",
                      json={"analytic_id": analytic_id, "sort_order": next_order},
                      timeout=60),
        f"bind branches to {sheet_id}")


def step_5_add_branches_and_populate(model_id: str, rng: random.Random) -> dict:
    _step(5, "Add Подразделения, populate manual cells per branch ±20%")

    aid, leaves, head_rid = _create_branches_analytic(model_id)
    print(f"   created analytic {aid}: 1 HEAD, {len(leaves)} leaves")

    sheets = _ok(requests.get(f"{API}/sheets/by-model/{model_id}", timeout=30)).json()

    db = sqlite3.connect(str(DB_PATH))
    try:
        # Capture baseline values BEFORE branches, but ONLY for cells with rule='manual'.
        # Formula / sum_children cells will recompute against the branch dimension and
        # may not follow the simple "HEAD ≈ N * orig" ratio (parent indicators that
        # consolidate children with mixed scales legitimately produce other ratios).
        baseline_by_sheet: dict[str, dict[str, float]] = {}
        for s in sheets:
            rows = db.execute(
                "SELECT coord_key, value FROM cell_data "
                "WHERE sheet_id = ? AND rule = 'manual'",
                (s["id"],),
            ).fetchall()
            bm = {}
            for ck, val in rows:
                if val in (None, ""):
                    continue
                try:
                    bm[ck] = float(val)
                except (ValueError, TypeError):
                    pass
            baseline_by_sheet[s["id"]] = bm

        # Bind branches analytic to each sheet (this migrates existing cells onto first leaf).
        for s in sheets:
            _bind_branches_to_sheet(s["id"], aid)
        print(f"   bound branches to {len(sheets)} sheet(s)")

        # The migration moves existing cells onto the FIRST leaf; we need to know
        # which leaf rid that is. We re-read leaves in DFS order — first leaf == leaves[0].
        first_leaf = leaves[0]

        # For every sheet: enumerate manual cells (now with first_leaf appended).
        # For each manual cell × each leaf: write original ± 20% noise.
        total_writes = 0
        for s in sheets:
            sid = s["id"]
            rows = db.execute(
                "SELECT coord_key, value FROM cell_data "
                "WHERE sheet_id = ? AND rule = 'manual' AND value IS NOT NULL AND value != ''",
                (sid,),
            ).fetchall()

            # Group by base coord (everything except the trailing first_leaf).
            base_to_val: dict[str, float] = {}
            for ck, val in rows:
                if not ck.endswith("|" + first_leaf):
                    continue
                base = ck[: -(len(first_leaf) + 1)]
                try:
                    base_to_val[base] = float(val)
                except (ValueError, TypeError):
                    continue

            if not base_to_val:
                continue

            items = []
            for base, orig in base_to_val.items():
                for lf in leaves:
                    noisy = orig * (1.0 + rng.uniform(-0.2, 0.2))
                    items.append({
                        "coord_key": f"{base}|{lf}",
                        "value": str(noisy),
                        "data_type": "number",
                        "rule": "manual",
                    })

            # Bulk write in chunks of 5000 to stay under limits.
            for i in range(0, len(items), 5000):
                _put_cells_bulk(sid, items[i:i + 5000])
            total_writes += len(items)
            print(f"   {s['name']}: wrote {len(items)} cells "
                  f"({len(base_to_val)} bases × {len(leaves)} leaves)")

        print(f"   total writes: {total_writes}")
        return {
            "analytic_id": aid,
            "head_rid": head_rid,
            "leaves": leaves,
            "first_leaf": first_leaf,
            "baseline_by_sheet": baseline_by_sheet,
            "sheets": sheets,
        }
    finally:
        db.close()


def step_6_verify_totals_grew(model_id: str, ctx: dict):
    _step(6, "Recalc and verify totals grew")
    t0 = time.time()
    _calc_model(model_id, rounds=3)
    print(f"   recalc: {time.time() - t0:.1f}s")

    head = ctx["head_rid"]
    leaves = ctx["leaves"]
    sheets = ctx["sheets"]
    baseline_by_sheet = ctx["baseline_by_sheet"]
    n_leaves = len(leaves)

    grew = 0
    same = 0
    shrunk = 0
    samples = []

    for s in sheets:
        sid = s["id"]
        baseline = baseline_by_sheet.get(sid, {})
        if not baseline:
            continue
        cells = _all_cells(sid)
        for base_ck, orig_val in baseline.items():
            if abs(orig_val) < 1e-6:
                continue
            head_ck = f"{base_ck}|{head}"
            c = cells.get(head_ck)
            if not c:
                continue
            try:
                new_val = float(c["value"])
            except (ValueError, TypeError):
                continue

            # Each leaf got orig*(1+noise), noise∈[-0.2,0.2]. Expected HEAD ≈ N*orig.
            ratio = new_val / orig_val
            if ratio > 1.05:
                grew += 1
            elif ratio < 0.95:
                shrunk += 1
            else:
                same += 1
            if len(samples) < 5:
                samples.append((s["name"], head_ck, orig_val, new_val, ratio))

    print(f"   summary across {sum(len(b) for b in baseline_by_sheet.values())} HEAD cells (originally manual):")
    print(f"     grew(+5%): {grew}   ~same(±5%): {same}   shrunk(-5%): {shrunk}")
    print(f"   expected ratio ≈ {n_leaves} (one leaf was original, "
          f"others orig×(1±0.2))")
    print(f"   sample HEAD cells:")
    for sname, ck, ov, nv, r in samples:
        print(f"     {sname}  {ck}: {ov:.2f} → {nv:.2f}  (×{r:.2f})")

    # Independent invariant: HEAD value must equal sum of leaf manual values for
    # every (period|indicator) base. This isolates the engine's analytic-axis
    # consolidation from any cross-indicator formula effects.
    print()
    print("   independent check: HEAD == Σ(leaf manuals) per base coord")
    leaf_set = set(leaves)
    inv_ok = inv_mis = 0
    inv_worst = []
    for s in sheets:
        sid = s["id"]
        cells = _all_cells(sid)
        leaf_sums = {}
        head_vals = {}
        for ck, c in cells.items():
            v = c.get("value")
            if v in (None, ""): continue
            try: vf = float(v)
            except (ValueError, TypeError): continue
            parts = ck.split("|")
            if len(parts) != 3: continue
            base = "|".join(parts[:2])
            br = parts[2]
            if br == head:
                head_vals[base] = vf
            elif br in leaf_set and c.get("rule") == "manual":
                leaf_sums[base] = leaf_sums.get(base, 0.0) + vf
        for base, lsum in leaf_sums.items():
            hv = head_vals.get(base)
            if hv is None: continue
            denom = max(abs(hv), abs(lsum), 1e-9)
            rel = abs(hv - lsum) / denom
            if rel < 1e-3:
                inv_ok += 1
            else:
                inv_mis += 1
                inv_worst.append((rel, s["name"], base, hv, lsum))
    print(f"     match: {inv_ok}   mismatch: {inv_mis}")
    if inv_mis:
        inv_worst.sort(reverse=True)
        for rel, sname, base, hv, lsum in inv_worst[:5]:
            print(f"       [{sname}] {base[:36]}…  HEAD={hv:.2f}  Σleaves={lsum:.2f}")

    if inv_mis == 0:
        print(f"   ✓ engine consolidation correct (HEAD = sum of branch leaves)")
    elif inv_mis < inv_ok * 0.05:
        print(f"   ✓ engine consolidation mostly correct ({inv_mis}/{inv_ok+inv_mis} mismatches)")
    else:
        print(f"   ⚠ engine consolidation issues — {inv_mis} cells out of {inv_ok+inv_mis}")


# ───────────────────────────── main ─────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--keep", action="store_true",
                    help="don't delete the imported model at the end")
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()

    rng = random.Random(args.seed)

    # Sanity: backend up?
    try:
        requests.get(f"{API}/models", timeout=5)
    except requests.ConnectionError:
        print(f"Backend not reachable at {API}. Start it: uvicorn backend.main:app")
        sys.exit(2)

    _hr("MAIN.xlsx full E2E flow")
    print(f"   API: {API}")
    print(f"   DB:  {DB_PATH}")

    model_id = step_1_import()
    try:
        step_2_recalc_and_verify(model_id)
        sid, ck, orig, before = step_3_change(model_id)
        step_4_revert(model_id, sid, ck, orig, before)
        ctx = step_5_add_branches_and_populate(model_id, rng)
        step_6_verify_totals_grew(model_id, ctx)
        _hr("DONE")
        print(f"   model_id = {model_id}")
        if args.keep:
            print(f"   Kept in DB. Open the UI to inspect.")
    finally:
        if not args.keep:
            try:
                requests.delete(f"{API}/models/{model_id}", timeout=120)
                print(f"   cleaned up {model_id}")
            except Exception as e:
                print(f"   cleanup warning: {e}")


if __name__ == "__main__":
    main()
