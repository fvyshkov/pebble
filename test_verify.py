#!/usr/bin/env python3
"""Verify imported model against Excel data_only values using excel_row mapping."""
import sqlite3, json, sys
from openpyxl import load_workbook
from datetime import datetime

MONTH = ["Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
SHEETS = {
    "параметры модели": ("0", 4),
    "кредитование": ("BaaS.1", 4),
    "депозит": ("BaaS.2", 4),
    "транзакционный": ("BaaS.3", 4),
    "Баланс": ("BS", 3),
    "Финансовый результат": ("PL", 3),
    "Операционные расходы": ("OPEX+CAPEX", 5),
}

db = sqlite3.connect('pebble.db')
db.row_factory = sqlite3.Row
wb = load_workbook('models.xlsx', data_only=True)

model_name = sys.argv[1] if len(sys.argv) > 1 else 'TEST_RUN1'
model = db.execute("SELECT id FROM models WHERE name=?", (model_name,)).fetchone()
if not model:
    print(f"Model '{model_name}' not found")
    sys.exit(1)

mid = model['id']
sheets = db.execute("SELECT id, name FROM sheets WHERE model_id=? ORDER BY sort_order", (mid,)).fetchall()

# Period mapping
period_aid = None
for s in sheets:
    b = db.execute("SELECT sa.analytic_id FROM sheet_analytics sa JOIN analytics a ON a.id=sa.analytic_id WHERE sa.sheet_id=? AND a.is_periods=1", (s['id'],)).fetchall()
    if b: period_aid = b[0]['analytic_id']; break

precs = db.execute("SELECT id, parent_id, data_json FROM analytic_records WHERE analytic_id=? ORDER BY sort_order", (period_aid,)).fetchall()
cmap = {}
for p in precs:
    if p['parent_id']: cmap.setdefault(p['parent_id'], []).append(p['id'])
leafs = [p for p in precs if p['id'] not in cmap]
pnames = {json.loads(p['data_json'])['name']: p['id'] for p in leafs}

grand_total = 0
grand_match = 0
grand_mismatch = 0
mismatches = []

for s in sheets:
    cfg = None
    for key, val in SHEETS.items():
        if key.lower() in s['name'].lower():
            cfg = val; break
    if not cfg: continue
    excel_name, start_col = cfg
    ws = wb[excel_name]
    
    # col → period_rid
    c2p = {}
    for c in range(start_col, min((ws.max_column or 1)+1, 50)):
        for hr in range(1, 7):
            v = ws.cell(hr, c).value
            if isinstance(v, datetime):
                pn = f"{MONTH[v.month-1]} {v.year}"
                if pn in pnames: c2p[c] = pnames[pn]
                break
    
    db_cells = {c['coord_key']: c['value'] for c in db.execute("SELECT coord_key, value FROM cell_data WHERE sheet_id=?", (s['id'],)).fetchall()}
    
    # Get records with excel_row
    ind_aids = [r['analytic_id'] for r in db.execute("SELECT sa.analytic_id FROM sheet_analytics sa JOIN analytics a ON a.id=sa.analytic_id WHERE sa.sheet_id=? AND a.is_periods=0", (s['id'],)).fetchall()]
    recs = []
    for ia in ind_aids:
        recs.extend(db.execute("SELECT id, excel_row, data_json FROM analytic_records WHERE analytic_id=?", (ia,)).fetchall())
    
    sheet_total = 0; sheet_match = 0; sheet_mismatch = 0
    
    for rec in recs:
        rid = rec['id']
        erow = rec['excel_row']
        if not erow: continue
        name = json.loads(rec['data_json'])['name']
        
        for col, pid in c2p.items():
            ck = f"{pid}|{rid}"
            dv = db_cells.get(ck)
            ev = ws.cell(erow, col).value
            
            if ev is None and dv is None: continue
            if ev is None: ev = 0
            if dv is None: dv = '0'
            
            try:
                e = float(ev)
                d = float(dv)
            except:
                continue
            
            sheet_total += 1
            tol = abs(e) * 0.01 + 0.01
            if abs(e - d) <= tol:
                sheet_match += 1
            else:
                sheet_mismatch += 1
                if len(mismatches) < 5:
                    mismatches.append(f"  {excel_name} R{erow} [{name[:30]}]: excel={e:.2f} db={d:.2f}")
    
    grand_total += sheet_total
    grand_match += sheet_match
    grand_mismatch += sheet_mismatch
    pct = sheet_match/sheet_total*100 if sheet_total else 0
    print(f"  {excel_name:12s}: {sheet_match}/{sheet_total} ({pct:.1f}%)")

pct = grand_match/grand_total*100 if grand_total else 0
print(f"\n  TOTAL: {grand_match}/{grand_total} ({pct:.1f}%)")
if mismatches:
    print(f"\n  Sample mismatches:")
    for m in mismatches: print(m)
